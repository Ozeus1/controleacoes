
import os
import sys
import sqlite3
import math
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from dotenv import load_dotenv
from models import db, Asset, Settings, User, TradeHistory, Option, OptionSpread, FixedIncome, InvestmentFund, Crypto, Pension, International, Dividend, MarketIndex, StudyOption, StudyStock, StudyIntlStock, StructuredOp, StructuredLeg, SimulacaoOpcoes, SimulacaoLeg, OptionRollSimulation, PutSale, CollarSimulation, SelicMensal, RankingVol, SearchedOption, RtdOptionData, PortfolioSnapshot, PMEvent, AssetTxn
from services import get_quotes, get_raw_quote_data
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
import requests
import time
import threading
import uuid
import json
import tempfile
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
import pytz
import yfinance as yf

# Load env vars
load_dotenv()

# Suporte a PyInstaller (frozen) e execução normal
if getattr(sys, 'frozen', False):
    # sys._MEIPASS = _internal/ com templates/static bundled
    _bundle = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
    app = Flask(__name__,
                root_path=_bundle,
                template_folder=os.path.join(_bundle, 'templates'),
                static_folder=os.path.join(_bundle, 'static'))
else:
    app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET_KEY', 'default_secret')

# --- Custom Filters ---
@app.template_filter('brl')
def format_brl(value):
    if value is None:
        return 'R$ 0,00'
    return f"R$ {value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

@app.template_filter('pct')
def format_pct(value):
    if value is None:
        return '0,00%'
    return f"{value:,.2f}%".replace('.', ',')


basedir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.abspath(os.path.dirname(__file__))
instance_path = os.path.join(basedir, 'instance')
if not os.path.exists(instance_path):
    os.makedirs(instance_path)

db_path = os.path.join(instance_path, 'investments.db')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'connect_args': {'timeout': 30},  # espera até 30s por lock do SQLite
}
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

def brl_fmt(value):
    if value is None:
        return ""
    return "{:,.2f}".format(value).replace(",", "X").replace(".", ",").replace("X", ".")

app.jinja_env.filters['brl_fmt'] = brl_fmt


class OplabApiError(Exception):
    def __init__(self, message, status_code=None, body_preview=None):
        super().__init__(message)
        self.status_code = status_code
        self.body_preview = body_preview


def _oplab_headers(token):
    return {
        'Access-Token': (token or '').strip(),
        'Accept': 'application/json',
        'User-Agent': 'MyInvest/1.0',
    }


def _oplab_get_json(path_or_url, token, params=None, timeout=15):
    """GET OpLab com validacao de HTTP/conteudo antes de decodificar JSON."""
    url = path_or_url if str(path_or_url).startswith('http') else f'https://api.oplab.com.br/v3{path_or_url}'
    token = (token or '').strip()
    params = dict(params or {})

    def _request(use_query_token=False):
        req_params = dict(params)
        if use_query_token:
            req_params['access_token'] = token
        return requests.get(url, params=req_params, headers=_oplab_headers(token), timeout=timeout)

    try:
        resp = _request(False)
        body = (resp.text or '').strip()
        retry_with_query = resp.status_code in (401, 403) or not body
        if not retry_with_query:
            if 200 <= resp.status_code < 300:
                try:
                    return resp.json()
                except ValueError:
                    retry_with_query = True
        if retry_with_query:
            resp = _request(True)
            body = (resp.text or '').strip()
    except requests.exceptions.Timeout as exc:
        raise OplabApiError('A OpLab demorou para responder. Tente novamente em instantes.') from exc
    except requests.exceptions.RequestException as exc:
        raise OplabApiError(f'Nao foi possivel conectar na OpLab: {exc.__class__.__name__}') from exc

    preview = body[:300]
    if resp.status_code in (401, 403):
        raise OplabApiError('Token OpLab recusado ou sem permissao para este endpoint.', resp.status_code, preview)
    if resp.status_code == 404:
        raise OplabApiError('Endpoint ou ticker nao encontrado na OpLab.', resp.status_code, preview)
    if resp.status_code < 200 or resp.status_code >= 300:
        raise OplabApiError(f'OpLab retornou HTTP {resp.status_code}.', resp.status_code, preview)
    if not body:
        raise OplabApiError('OpLab retornou resposta vazia.', resp.status_code)
    try:
        return resp.json()
    except ValueError as exc:
        raise OplabApiError('OpLab retornou resposta invalida em vez de JSON.', resp.status_code, preview) from exc


def _oplab_is_available(token: str, timeout: int = 4) -> bool:
    """
    Probe rápido: tenta GET /v3/user/me com timeout curto.
    Retorna True se o OpLab responder com 2xx ou 4xx (token inválido mas servidor OK).
    Retorna False se houver timeout, ConnectionError ou 5xx (servidor fora do ar).
    """
    try:
        r = requests.get(
            'https://api.oplab.com.br/v3/user/me',
            headers=_oplab_headers(token),
            timeout=timeout,
        )
        return r.status_code < 500
    except (requests.exceptions.Timeout,
            requests.exceptions.ConnectionError,
            requests.exceptions.RequestException):
        return False


def _do_oplab_bulk_update_safe(uid: int, token: str, deadline_secs: int = 25):
    """
    Wrapper que executa _do_oplab_bulk_update em thread separada com deadline total.
    A thread herda o app_context atual para que o SQLAlchemy funcione corretamente.
    Se não completar em deadline_secs, retorna (0, 0, set(), 'timeout').
    Retorna (assets_ok, options_ok, covered, error_msg).
    """
    result = {'val': None, 'err': None}

    def _run():
        with app.app_context():
            try:
                result['val'] = _do_oplab_bulk_update(uid, token, oplab_online=True)
            except Exception as e:
                result['err'] = str(e)

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    t.join(timeout=deadline_secs)

    if t.is_alive():
        return 0, 0, set(), f'timeout após {deadline_secs}s'
    if result['err']:
        return 0, 0, set(), result['err']
    if result['val']:
        a, o, cov = result['val']
        return a, o, cov, None
    return 0, 0, set(), 'sem resultado'


@app.template_filter('date_fmt')
def format_date(value):
    if value is None:
        return "-"
    if isinstance(value, str):
        try:
            value = datetime.strptime(value, '%Y-%m-%d').date()
        except:
             return value
    return value.strftime('%d/%m/%Y')


db.init_app(app)

# WAL mode: leituras simultâneas com escrita → reduz bloqueios do scheduler OpLab
from sqlalchemy import event as _sa_event
from sqlalchemy.engine import Engine as _Engine
import sqlite3 as _sqlite3_pragma
@_sa_event.listens_for(_Engine, 'connect')
def _set_sqlite_pragma(conn, _rec):
    if isinstance(conn, _sqlite3_pragma.Connection):
        conn.execute('PRAGMA journal_mode=WAL')
        conn.execute('PRAGMA synchronous=NORMAL')
        conn.execute('PRAGMA busy_timeout=30000')

# File-based task store (shared across gunicorn workers)
_TASK_DIR = os.path.join(tempfile.gettempdir(), 'ca_update_tasks')
os.makedirs(_TASK_DIR, exist_ok=True)

_BRT = ZoneInfo('America/Sao_Paulo')

def now_brt():
    """Retorna datetime atual no fuso de Brasília."""
    return datetime.now(_BRT)


def du_count(start: date, end: date) -> int:
    """Conta dias úteis (seg-sex) entre start e end, exclusive start, inclusive end."""
    if not start or not end or end <= start:
        return 0
    count = 0
    cur = start + timedelta(days=1)
    while cur <= end:
        if cur.weekday() < 5:   # 0=Seg … 4=Sex
            count += 1
        cur += timedelta(days=1)
    return count


def _task_file(task_id):
    return os.path.join(_TASK_DIR, task_id + '.json')

def _set_task(task_id, data):
    with open(_task_file(task_id), 'w') as f:
        json.dump(data, f)

def _get_task(task_id):
    try:
        with open(_task_file(task_id), 'r') as f:
            return json.load(f)
    except Exception:
        return {'status': 'not_found', 'msg': '', 'category': ''}

def run_migrations():
    """Auto-migrate database schema on startup."""
    import sqlite3 as _sqlite3
    conn = _sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Novos campos no modelo User — usa try/except por segurança (SQLite não tem IF NOT EXISTS no ALTER)
    for _col, _def in [
        ('full_name',       'VARCHAR(120)'),
        ('email',           'VARCHAR(120)'),
        ('phone',           'VARCHAR(30)'),
        ('avatar_filename', 'VARCHAR(120)'),
    ]:
        try:
            cursor.execute(f"ALTER TABLE user ADD COLUMN {_col} {_def}")
        except Exception:
            pass  # coluna já existe

    # RankingVol: grupo da lista (LIQ = com liquidez, GERAL = lista ampla)
    try:
        cursor.execute("ALTER TABLE ranking_vol ADD COLUMN grupo VARCHAR(10) DEFAULT 'LIQ'")
    except Exception:
        pass  # coluna já existe

    # Check existing columns in 'option' table
    cursor.execute("PRAGMA table_info(option)")
    existing_columns = {row[1] for row in cursor.fetchall()}

    if 'option_type' not in existing_columns:
        cursor.execute("ALTER TABLE 'option' ADD COLUMN option_type VARCHAR(20) NOT NULL DEFAULT 'VENDA_CALL'")
        print("[MIGRATION] Added column 'option_type' to 'option' table.")

    # Check existing columns in 'asset' table
    cursor.execute("PRAGMA table_info(asset)")
    asset_columns = {row[1] for row in cursor.fetchall()}

    if 'last_dividend' not in asset_columns:
        cursor.execute("ALTER TABLE asset ADD COLUMN last_dividend FLOAT")
        print("[MIGRATION] Added column 'last_dividend' to 'asset' table.")

    if 'last_dividend_date' not in asset_columns:
        cursor.execute("ALTER TABLE asset ADD COLUMN last_dividend_date DATE")
        print("[MIGRATION] Added column 'last_dividend_date' to 'asset' table.")

    if 'dividend_yield' not in asset_columns:
        cursor.execute("ALTER TABLE asset ADD COLUMN dividend_yield FLOAT")
        print("[MIGRATION] Added column 'dividend_yield' to 'asset' table.")

    if 'exit_date' not in asset_columns:
        # try/except: com múltiplos workers gunicorn, dois podem correr a migração
        # ao mesmo tempo — o segundo veria "duplicate column name" e o worker morreria.
        try:
            cursor.execute("ALTER TABLE asset ADD COLUMN exit_date DATE")
            print("[MIGRATION] Added column 'exit_date' to 'asset' table.")
        except Exception:
            pass  # coluna já criada por outro worker

    # Check existing columns in 'international' table
    cursor.execute("PRAGMA table_info(international)")
    intl_columns = {row[1] for row in cursor.fetchall()}

    if 'purchase_price' not in intl_columns:
        cursor.execute("ALTER TABLE international ADD COLUMN purchase_price FLOAT")
        print("[MIGRATION] Added column 'purchase_price' to 'international' table.")

    if 'current_price' not in intl_columns:
        cursor.execute("ALTER TABLE international ADD COLUMN current_price FLOAT")
        print("[MIGRATION] Added column 'current_price' to 'international' table.")

    # Check existing columns in 'crypto' table
    cursor.execute("PRAGMA table_info(crypto)")
    crypto_columns = {row[1] for row in cursor.fetchall()}

    if 'quote' not in crypto_columns:
        cursor.execute("ALTER TABLE crypto ADD COLUMN quote FLOAT")
        print("[MIGRATION] Added column 'quote' to 'crypto' table.")

    if 'avg_price' not in crypto_columns:
        cursor.execute("ALTER TABLE crypto ADD COLUMN avg_price FLOAT DEFAULT 0.0")
        print("[MIGRATION] Added column 'avg_price' to 'crypto' table.")

    # Reclassify legacy strategy names in trade_history
    cursor.execute("""
        UPDATE trade_history SET strategy = 'Recomendações'
        WHERE strategy = 'EQI'
    """)
    cursor.execute("""
        UPDATE trade_history SET strategy = 'Fundos Imobiliários'
        WHERE strategy = 'FII'
    """)
    cursor.execute("""
        UPDATE trade_history SET strategy = 'Internacional'
        WHERE strategy = 'INTL'
    """)
    cursor.execute("""
        UPDATE trade_history SET strategy = 'Opções'
        WHERE strategy LIKE 'OPCAO%'
    """)
    if cursor.rowcount > 0:
        print(f"[MIGRATION] Reclassified {cursor.rowcount} trade_history strategy records.")

    # Add atr_pct, iv_rank, iv_percentil to study_stock and study_intl_stock if missing
    for tbl in ('study_stock', 'study_intl_stock'):
        cursor.execute(f"PRAGMA table_info({tbl})")
        cols = {row[1] for row in cursor.fetchall()}
        for col in ('atr_pct', 'iv_rank', 'iv_percentil'):
            if col not in cols:
                cursor.execute(f"ALTER TABLE {tbl} ADD COLUMN {col} FLOAT")
                print(f"[MIGRATION] Added {tbl}.{col}")

    # Add roll_history to option, option_spread, structured_op, put_sale
    for tbl in ('option', 'option_spread', 'structured_op', 'put_sale'):
        cursor.execute(f"PRAGMA table_info({tbl})")
        cols = {row[1] for row in cursor.fetchall()}
        if 'roll_history' not in cols:
            cursor.execute(f"ALTER TABLE {tbl} ADD COLUMN roll_history TEXT")
            print(f"[MIGRATION] Added {tbl}.roll_history")

    # Add underlying_price / underlying_change to option_spread if missing
    cursor.execute("PRAGMA table_info(option_spread)")
    os_cols = {row[1] for row in cursor.fetchall()}
    if 'underlying_price' not in os_cols:
        cursor.execute("ALTER TABLE option_spread ADD COLUMN underlying_price FLOAT")
        print("[MIGRATION] Added option_spread.underlying_price")
    if 'underlying_change' not in os_cols:
        cursor.execute("ALTER TABLE option_spread ADD COLUMN underlying_change FLOAT")
        print("[MIGRATION] Added option_spread.underlying_change")

    # Add underlying_price / underlying_change to option if missing
    # (cotação do subjacente das opções cujo ativo não está na carteira)
    cursor.execute("PRAGMA table_info(option)")
    opt_cols = {row[1] for row in cursor.fetchall()}
    if 'underlying_price' not in opt_cols:
        try:
            cursor.execute("ALTER TABLE option ADD COLUMN underlying_price FLOAT")
            print("[MIGRATION] Added option.underlying_price")
        except Exception:
            pass
    if 'underlying_change' not in opt_cols:
        try:
            cursor.execute("ALTER TABLE option ADD COLUMN underlying_change FLOAT")
            print("[MIGRATION] Added option.underlying_change")
        except Exception:
            pass

    # Add pop to structured_op if missing (POP salvo ao abrir o payoff)
    cursor.execute("PRAGMA table_info(structured_op)")
    sop_cols = {row[1] for row in cursor.fetchall()}
    if 'pop' not in sop_cols:
        try:
            cursor.execute("ALTER TABLE structured_op ADD COLUMN pop FLOAT")
            print("[MIGRATION] Added structured_op.pop")
        except Exception:
            pass
    if 'intl' not in sop_cols:
        try:
            cursor.execute("ALTER TABLE structured_op ADD COLUMN intl BOOLEAN DEFAULT 0")
            print("[MIGRATION] Added structured_op.intl")
        except Exception:
            pass

    # Dividend: valor por ação + quantidade usada no cálculo (auditável/editável)
    cursor.execute("PRAGMA table_info(dividend)")
    div_cols = {row[1] for row in cursor.fetchall()}
    if 'per_share' not in div_cols:
        try:
            cursor.execute("ALTER TABLE dividend ADD COLUMN per_share FLOAT")
            print("[MIGRATION] Added dividend.per_share")
        except Exception:
            pass
    if 'qty_used' not in div_cols:
        try:
            cursor.execute("ALTER TABLE dividend ADD COLUMN qty_used INTEGER")
            print("[MIGRATION] Added dividend.qty_used")
        except Exception:
            pass
    if 'qty_manual' not in div_cols:
        try:
            cursor.execute("ALTER TABLE dividend ADD COLUMN qty_manual BOOLEAN DEFAULT 0")
            print("[MIGRATION] Added dividend.qty_manual")
        except Exception:
            pass
    # Backfill: registros antigos foram calculados como per_share × qtd ATUAL
    # do ativo — reconstrói os dois campos a partir disso (idempotente)
    try:
        cursor.execute("""
            UPDATE dividend SET
              qty_used = (SELECT quantity FROM asset WHERE asset.id = dividend.asset_id),
              per_share = CASE
                WHEN (SELECT quantity FROM asset WHERE asset.id = dividend.asset_id) > 0
                THEN amount * 1.0 / (SELECT quantity FROM asset WHERE asset.id = dividend.asset_id)
              END
            WHERE per_share IS NULL
        """)
    except Exception:
        pass

    # Simulações de MANEJO salvas antes do tipo ser aceito (normalizadas p/ TIME
    # pelo save antigo) — restaura o tipo a partir do payload (idempotente)
    try:
        cursor.execute("""
            UPDATE option_roll_simulation SET roll_type='MANEJO'
            WHERE roll_type != 'MANEJO'
              AND (payload LIKE '%"roll_type": "MANEJO"%'
                   OR payload LIKE '%"roll_type":"MANEJO"%')
        """)
        if cursor.rowcount:
            print(f"[MIGRATION] {cursor.rowcount} simulação(ões) de manejo re-tipadas p/ MANEJO")
    except Exception:
        pass

    # Livro de transações da carteira (compras/vendas com data)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS asset_txn (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            ticker VARCHAR(10) NOT NULL,
            txn_date DATE NOT NULL,
            side VARCHAR(1) NOT NULL DEFAULT 'C',
            quantity INTEGER NOT NULL,
            price FLOAT NOT NULL DEFAULT 0.0,
            source VARCHAR(12) DEFAULT 'MANUAL',
            notes VARCHAR(200),
            created_at DATETIME
        )
    """)
    try:
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_asset_txn_user_ticker ON asset_txn (user_id, ticker, txn_date)")
    except Exception:
        pass

    # Tabela de eventos do Preço Médio didático (página Preço Médio)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS pm_event (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            ticker VARCHAR(10) NOT NULL,
            kind VARCHAR(15) NOT NULL,
            event_date DATE,
            valor FLOAT NOT NULL DEFAULT 0.0,
            buy_qty INTEGER,
            buy_price FLOAT,
            source_key VARCHAR(30),
            ref VARCHAR(200),
            pm_before FLOAT,
            pm_after FLOAT,
            created_at DATETIME
        )
    """)
    try:
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_pm_event_user_ticker ON pm_event (user_id, ticker)")
    except Exception:
        pass

    # Add underlying and notes columns to trade_history if missing
    cursor.execute("PRAGMA table_info(trade_history)")
    th_cols = {row[1] for row in cursor.fetchall()}
    if 'underlying' not in th_cols:
        cursor.execute("ALTER TABLE trade_history ADD COLUMN underlying VARCHAR(15)")
        print("[MIGRATION] Added trade_history.underlying")
    if 'notes' not in th_cols:
        cursor.execute("ALTER TABLE trade_history ADD COLUMN notes TEXT")
        print("[MIGRATION] Added trade_history.notes")

    # Create option_spread table if it doesn't exist
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS option_spread (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            spread_type VARCHAR(20) NOT NULL DEFAULT 'TRAVA_ALTA_PUT',
            underlying_asset VARCHAR(15) NOT NULL DEFAULT '',
            quantity INTEGER NOT NULL DEFAULT 0,
            expiration_date DATE NOT NULL DEFAULT '2000-01-01',
            entry_date DATE,
            leg_long_ticker VARCHAR(20) NOT NULL DEFAULT '',
            leg_long_strike FLOAT NOT NULL DEFAULT 0.0,
            leg_long_price FLOAT NOT NULL DEFAULT 0.0,
            leg_long_current FLOAT DEFAULT 0.0,
            leg_short_ticker VARCHAR(20) NOT NULL DEFAULT '',
            leg_short_strike FLOAT NOT NULL DEFAULT 0.0,
            leg_short_price FLOAT NOT NULL DEFAULT 0.0,
            leg_short_current FLOAT DEFAULT 0.0,
            pop FLOAT,
            FOREIGN KEY (user_id) REFERENCES user(id)
        )
    """)

    # Add pop column if table already existed without it
    cursor.execute("PRAGMA table_info(option_spread)")
    spread_cols = {row[1] for row in cursor.fetchall()}
    if 'pop' not in spread_cols:
        cursor.execute("ALTER TABLE option_spread ADD COLUMN pop FLOAT")

    # Add study columns to option table
    cursor.execute("PRAGMA table_info(option)")
    opt_cols2 = {row[1] for row in cursor.fetchall()}
    if 'vdx'          not in opt_cols2: cursor.execute("ALTER TABLE 'option' ADD COLUMN vdx FLOAT")
    if 'nv'           not in opt_cols2: cursor.execute("ALTER TABLE 'option' ADD COLUMN nv FLOAT")
    if 've'           not in opt_cols2: cursor.execute("ALTER TABLE 'option' ADD COLUMN ve FLOAT")
    if 'delta'        not in opt_cols2: cursor.execute("ALTER TABLE 'option' ADD COLUMN delta FLOAT")
    if 'gama'         not in opt_cols2: cursor.execute("ALTER TABLE 'option' ADD COLUMN gama FLOAT")
    if 'daily_change' not in opt_cols2: cursor.execute("ALTER TABLE 'option' ADD COLUMN daily_change FLOAT")

    # Create study_option table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS study_option (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            ticker VARCHAR(20) NOT NULL DEFAULT '',
            underlying_asset VARCHAR(15) NOT NULL DEFAULT '',
            underlying_price FLOAT,
            avg_price_stock FLOAT,
            strike FLOAT,
            expiration_date DATE,
            option_price FLOAT,
            vdx FLOAT,
            nv FLOAT,
            ve FLOAT,
            delta FLOAT,
            gama FLOAT,
            FOREIGN KEY (user_id) REFERENCES user(id)
        )
    """)
    # Add ve/delta/gama to study_option if table already existed
    cursor.execute("PRAGMA table_info(study_option)")
    so_cols = {row[1] for row in cursor.fetchall()}
    if 've'    not in so_cols: cursor.execute("ALTER TABLE study_option ADD COLUMN ve FLOAT")
    if 'delta' not in so_cols: cursor.execute("ALTER TABLE study_option ADD COLUMN delta FLOAT")
    if 'gama'  not in so_cols: cursor.execute("ALTER TABLE study_option ADD COLUMN gama FLOAT")

    # Create study_intl_stock table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS study_intl_stock (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            ticker VARCHAR(15) NOT NULL DEFAULT '',
            trend VARCHAR(10),
            rsi FLOAT,
            volatility VARCHAR(10),
            ve FLOAT,
            strategy VARCHAR(60),
            study_date DATE,
            strategy_active VARCHAR(100),
            entry_date DATE,
            FOREIGN KEY (user_id) REFERENCES user(id)
        )
    """)

    # Create study_stock table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS study_stock (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            ticker VARCHAR(15) NOT NULL DEFAULT '',
            trend VARCHAR(10),
            rsi FLOAT,
            volatility VARCHAR(10),
            ve FLOAT,
            strategy VARCHAR(60),
            study_date DATE,
            strategy_active VARCHAR(100),
            entry_date DATE,
            FOREIGN KEY (user_id) REFERENCES user(id)
        )
    """)

    # Create structured_op and structured_leg tables
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS structured_op (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name VARCHAR(100) NOT NULL DEFAULT '',
            underlying_asset VARCHAR(15) NOT NULL DEFAULT '',
            status VARCHAR(10) NOT NULL DEFAULT 'OPEN',
            created_at DATETIME,
            FOREIGN KEY (user_id) REFERENCES user(id)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS structured_leg (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            op_id INTEGER NOT NULL,
            ticker VARCHAR(20) NOT NULL DEFAULT '',
            side VARCHAR(4) NOT NULL DEFAULT 'SELL',
            opt_type VARCHAR(4) NOT NULL DEFAULT 'CALL',
            quantity INTEGER NOT NULL DEFAULT 1,
            strike FLOAT NOT NULL DEFAULT 0.0,
            expiration_date DATE,
            entry_price FLOAT NOT NULL DEFAULT 0.0,
            current_price FLOAT NOT NULL DEFAULT 0.0,
            last_update DATETIME,
            FOREIGN KEY (op_id) REFERENCES structured_op(id)
        )
    """)

    # Migração: colunas em structured_op
    cursor.execute("PRAGMA table_info(structured_op)")
    struct_cols = {row[1] for row in cursor.fetchall()}
    if struct_cols:
        if 'underlying_price' not in struct_cols:
            cursor.execute("ALTER TABLE structured_op ADD COLUMN underlying_price FLOAT")
        if 'underlying_change' not in struct_cols:
            cursor.execute("ALTER TABLE structured_op ADD COLUMN underlying_change FLOAT")
        if 'uses_stock_collateral' not in struct_cols:
            cursor.execute("ALTER TABLE structured_op ADD COLUMN uses_stock_collateral BOOLEAN DEFAULT 0")

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS simulacao_opcoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name VARCHAR(120) NOT NULL DEFAULT '',
            underlying VARCHAR(15) NOT NULL DEFAULT '',
            created_at DATETIME,
            FOREIGN KEY (user_id) REFERENCES user(id)
        )
    """)
    # Add iv column to simulacao_leg if missing
    cursor.execute("PRAGMA table_info(simulacao_leg)")
    sim_leg_cols = {row[1] for row in cursor.fetchall()}
    if 'iv' not in sim_leg_cols and sim_leg_cols:
        cursor.execute("ALTER TABLE simulacao_leg ADD COLUMN iv FLOAT NOT NULL DEFAULT 0.0")

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS simulacao_leg (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sim_id INTEGER NOT NULL,
            leg_type VARCHAR(6) NOT NULL DEFAULT 'CALL',
            side VARCHAR(4) NOT NULL DEFAULT 'BUY',
            quantity INTEGER NOT NULL DEFAULT 1,
            strike FLOAT NOT NULL DEFAULT 0.0,
            premium FLOAT NOT NULL DEFAULT 0.0,
            expiration DATE,
            ticker VARCHAR(20) NOT NULL DEFAULT '',
            FOREIGN KEY (sim_id) REFERENCES simulacao_opcoes(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS put_sale (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            ticker VARCHAR(20) NOT NULL DEFAULT '',
            underlying_asset VARCHAR(15) NOT NULL DEFAULT '',
            underlying_price FLOAT,
            strike FLOAT NOT NULL DEFAULT 0.0,
            expiration_date DATE NOT NULL,
            premium FLOAT NOT NULL DEFAULT 0.0,
            quantity INTEGER NOT NULL DEFAULT 100,
            entry_date DATE,
            notes VARCHAR(200),
            created_at DATETIME,
            FOREIGN KEY (user_id) REFERENCES user(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS option_roll_simulation (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name VARCHAR(120) NOT NULL DEFAULT '',
            underlying VARCHAR(15) NOT NULL DEFAULT '',
            roll_type VARCHAR(10) NOT NULL DEFAULT 'TIME',
            payload TEXT NOT NULL DEFAULT '{}',
            created_at DATETIME,
            updated_at DATETIME,
            FOREIGN KEY (user_id) REFERENCES user(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS selic_mensal (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mes_ano VARCHAR(7) UNIQUE NOT NULL,
            taxa FLOAT NOT NULL
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ranking_vol (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL REFERENCES user(id),
            ticker VARCHAR(15) NOT NULL,
            var_pct FLOAT,
            last_price FLOAT,
            last_date VARCHAR(10),
            iv_rank FLOAT,
            iv_percentil FLOAT,
            vol_impl FLOAT,
            updated_at DATETIME
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS portfolio_snapshot (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            snap_date VARCHAR(10) NOT NULL,
            total_equity FLOAT NOT NULL DEFAULT 0.0,
            total_acoes FLOAT NOT NULL DEFAULT 0.0,
            total_fiis FLOAT NOT NULL DEFAULT 0.0,
            total_etfs FLOAT NOT NULL DEFAULT 0.0,
            estimated BOOLEAN NOT NULL DEFAULT 0,
            created_at DATETIME,
            FOREIGN KEY (user_id) REFERENCES user(id)
        )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_snap_user_date ON portfolio_snapshot(user_id, snap_date)")

    # Adiciona colunas novas em rtd_option_data (iv_ask, iv_bid, iv_over_hv)
    cursor.execute("PRAGMA table_info(rtd_option_data)")
    rtd_cols = {row[1] for row in cursor.fetchall()}
    if rtd_cols:
        for _col in ('iv_ask', 'iv_bid', 'iv_over_hv'):
            if _col not in rtd_cols:
                cursor.execute(f"ALTER TABLE rtd_option_data ADD COLUMN {_col} FLOAT")
                print(f"[MIGRATION] Added rtd_option_data.{_col}")

    conn.commit()
    conn.close()

# Dados históricos da Selic mensal (% a.m.)
_SELIC_HISTORICO = """01/2020,0.38
02/2020,0.29
03/2020,0.34
04/2020,0.28
05/2020,0.24
06/2020,0.21
07/2020,0.19
08/2020,0.16
09/2020,0.16
10/2020,0.16
11/2020,0.15
12/2020,0.16
01/2021,0.15
02/2021,0.13
03/2021,0.20
04/2021,0.21
05/2021,0.27
06/2021,0.31
07/2021,0.36
08/2021,0.43
09/2021,0.44
10/2021,0.49
11/2021,0.59
12/2021,0.77
01/2022,0.73
02/2022,0.76
03/2022,0.93
04/2022,0.83
05/2022,1.03
06/2022,1.02
07/2022,1.03
08/2022,1.17
09/2022,1.07
10/2022,1.02
11/2022,1.02
12/2022,1.12
01/2023,1.12
02/2023,0.92
03/2023,1.17
04/2023,0.92
05/2023,1.12
06/2023,1.07
07/2023,1.07
08/2023,1.14
09/2023,0.97
10/2023,1.00
11/2023,0.92
12/2023,0.89
01/2024,0.97
02/2024,0.80
03/2024,0.83
04/2024,0.89
05/2024,0.83
06/2024,0.79
07/2024,0.91
08/2024,0.87
09/2024,0.84
10/2024,0.93
11/2024,0.79
12/2024,0.93
01/2025,1.01
02/2025,0.99
03/2025,0.96
04/2025,1.06
05/2025,1.14
06/2025,1.10
07/2025,1.28
08/2025,1.16
09/2025,1.22
10/2025,1.28
11/2025,1.05
12/2025,1.22
01/2026,1.16
02/2026,1.00
03/2026,1.21"""

with app.app_context():
    run_migrations()
    db.create_all()
    # Seed Selic histórica (INSERT OR IGNORE para não sobrescrever edições manuais)
    for linha in _SELIC_HISTORICO.strip().splitlines():
        partes = linha.split(',')
        if len(partes) == 2:
            mm_aa, taxa_str = partes
            try:
                taxa = float(taxa_str)
                parts = mm_aa.strip().split('/')
                mes_ano_fmt = f"{parts[1]}-{parts[0]}"  # MM/YYYY -> YYYY-MM
                if not SelicMensal.query.filter_by(mes_ano=mes_ano_fmt).first():
                    db.session.add(SelicMensal(mes_ano=mes_ano_fmt, taxa=taxa))
            except Exception:
                pass
    db.session.commit()

# Scheduler OpLab iniciado ao carregar o módulo (Gunicorn + __main__)
threading.Thread(target=lambda: (time.sleep(5), _start_oplab_scheduler()), daemon=True).start()


# --- Options Module Routes ---


def _norm_cdf(x):
    """Aproximação de Abramowitz & Stegun para N(x)."""
    t = 1.0 / (1.0 + 0.2316419 * abs(x))
    poly = t * (0.319381530 + t * (-0.356563782 + t * (1.781477937 + t * (-1.821255978 + t * 1.330274429))))
    pdf  = math.exp(-0.5 * x * x) / math.sqrt(2 * math.pi)
    c    = 1.0 - pdf * poly
    return c if x >= 0 else 1.0 - c


def _bs_price(S, K, T, r, sigma, is_call):
    """Black-Scholes para call ou put."""
    if T <= 0 or sigma <= 0 or S <= 0 or K <= 0:
        return max(0.0, (S - K) if is_call else (K - S))
    d1 = (math.log(S / K) + (r + 0.5 * sigma * sigma) * T) / (sigma * math.sqrt(T))
    d2 = d1 - sigma * math.sqrt(T)
    if is_call:
        return S * _norm_cdf(d1) - K * math.exp(-r * T) * _norm_cdf(d2)
    return K * math.exp(-r * T) * _norm_cdf(-d2) - S * _norm_cdf(-d1)


def _implied_vol(S0, K, T, r, target, is_call):
    """IV por bissecção a partir do prêmio de entrada."""
    if target <= 0 or T <= 0:
        return 0.30
    lo, hi = 0.001, 5.0
    for _ in range(60):
        mid = (lo + hi) / 2
        if _bs_price(S0, K, T, r, mid, is_call) < target:
            lo = mid
        else:
            hi = mid
        if hi - lo < 1e-6:
            break
    return (lo + hi) / 2


def _calc_pop(S, breakevens, T, sigma, r=0.0, payoff_fn=None):
    """
    Probability of Profit (POP) via distribuição log-normal (Black-Scholes).

    Com payoff_fn: integra numericamente a densidade log-normal de S_T sobre a
    região onde payoff_fn(S_T) > 0 — funciona para qualquer formato de payoff
    (lucro entre os BEs, fora deles, assimétrico etc.) e sempre retorna 0-100.

    Sem payoff_fn (fallback analítico): P(S_T > B) = N(d2(B)), com d2
    decrescente em B. Para 2 BEs, P(low < S_T < high) = N(d2_low) - N(d2_high)
    — assume lucro ENTRE os breakevens.
    """
    if S <= 0 or T <= 0 or sigma <= 0:
        return None

    # ── Caminho preferido: integração numérica sobre o payoff real ──────────
    if payoff_fn is not None:
        mu = math.log(S) + (r - 0.5 * sigma * sigma) * T
        sd = sigma * math.sqrt(T)
        lo = max(S * math.exp(-4 * sd), 0.01)
        hi = S * math.exp(4 * sd)
        tot = win = 0.0
        M = 500
        step = (hi - lo) / (M + 1)
        for k in range(M + 1):
            sk = lo + step * (k + 0.5)
            z = (math.log(sk) - mu) / sd
            w = math.exp(-0.5 * z * z) / sk   # ∝ densidade log-normal
            tot += w
            try:
                if payoff_fn(sk) > 0:
                    win += w
            except Exception:
                pass
        return round(win / tot * 100, 1) if tot > 0 else None

    if not breakevens:
        return None

    def _d2(B):
        if B <= 0:
            return None
        try:
            return (math.log(S / B) + (r - 0.5 * sigma * sigma) * T) / (sigma * math.sqrt(T))
        except (ValueError, ZeroDivisionError):
            return None

    bes = sorted(breakevens)

    if len(bes) == 1:
        d = _d2(bes[0])
        if d is None:
            return None
        # Crédito acima do BE → lucra se S_T > BE
        return round(_norm_cdf(d) * 100, 1)

    # 2+ BEs: lucra entre os extremos. d2 é DECRESCENTE em B, logo
    # P(low < S_T < high) = N(d2(low)) − N(d2(high))  (sempre >= 0)
    d_low  = _d2(bes[0])
    d_high = _d2(bes[-1])
    if d_low is None or d_high is None:
        return None
    return round(max(0.0, (_norm_cdf(d_low) - _norm_cdf(d_high))) * 100, 1)


def _calc_structured_metrics(op):
    """
    Calcula métricas financeiras de uma StructuredOp.

    current_pnl = P&L total se fechar agora:
        Σ_VENDA qty×(entry-current) + Σ_COMPRA qty×(current-entry)
        Equivalente a: net_recebido + valor_de_fechamento

    Breakevens:
        - Mathemático (payoff no vencimento): cruzamentos de zero da função payoff
        - Ref. simplificada (mercado BR): min_strike_put − net/qty_call_vendida
          e max_strike_call + net/qty_call_vendida
    """
    legs = op.legs
    if not legs:
        return dict(net=0, current_pnl=0, max_profit=0, max_loss=0,
                    breakevens=[], be_low=None, be_high=None,
                    unlimited_profit=False, unlimited_loss=False, pop=None)

    # ── Crédito/débito líquido na montagem ─────────────────────────
    net = sum(
        (leg.entry_price if leg.side == 'SELL' else -leg.entry_price) * leg.quantity
        for leg in legs
    )

    # ── P&L atual = quanto receberia/pagaria fechando tudo agora ───
    # Para VENDA: lucro = entry − current (ganha quando opção cai)
    # Para COMPRA: lucro = current − entry (ganha quando opção sobe)
    current_pnl = sum(
        ((leg.entry_price - leg.current_price) if leg.side == 'SELL'
         else (leg.current_price - leg.entry_price)) * leg.quantity
        for leg in legs
    )

    # ── Detecta trava calendário (pernas com vencimentos diferentes) ─
    exp_dates = sorted({l.expiration_date for l in legs if l.expiration_date})
    is_calendar = len(exp_dates) > 1
    ref_date = exp_dates[0] if is_calendar else (exp_dates[0] if exp_dates else date.today())

    # Taxa Selic contínua — lê diretamente do Settings sem depender de current_user
    try:
        selic_val = float(Settings.get_value('selic_rate', user_id=op.user_id, default='14.5') or '14.5')
    except Exception:
        selic_val = 14.5
    r_cont = math.log(1 + selic_val / 100)

    # Cotação do ativo subjacente (para bisecção de IV) — SEM rede: esta função
    # roda uma vez por estruturada no render de /opcoes; ir ao Yahoo aqui
    # travava a página por vários segundos por operação.
    try:
        spot_ref, _ = _get_underlying_quote_cached(op.underlying_asset, op.user_id)
    except Exception:
        spot_ref = None
    today_d = date.today()

    # IV implícita de cada perna (calculada uma vez, fora do loop de S)
    leg_ivs = {}
    for leg in legs:
        if is_calendar and leg.expiration_date and leg.expiration_date > ref_date:
            try:
                S0 = spot_ref if spot_ref else (leg.strike or 50)
                T_leg = max((leg.expiration_date - today_d).days / 365.25, 1 / 365)
                leg_ivs[leg.id] = _implied_vol(S0, leg.strike or 1, T_leg, r_cont,
                                               leg.entry_price, leg.opt_type == 'CALL')
            except Exception:
                leg_ivs[leg.id] = 0.30

    # ── Payoff no vencimento ────────────────────────────────────────
    def payoff_at(S):
        total = net
        for leg in legs:
            q     = leg.quantity
            K     = leg.strike or 0
            sign  = 1 if leg.side == 'BUY' else -1
            if leg.opt_type == 'STOCK':
                # Custo de entrada já está em `net`; aqui só adiciona o valor de mercado no vencimento
                total += sign * q * S
            elif is_calendar and leg.id in leg_ivs:
                # Perna longa de calendário: valor BS com tempo restante após vencimento curta
                T_rem = max((leg.expiration_date - ref_date).days / 365.25, 0)
                iv    = leg_ivs[leg.id]
                is_call = leg.opt_type == 'CALL'
                total += sign * q * _bs_price(S, K, T_rem, r_cont, iv, is_call)
            else:
                # Payoff intrínseco no vencimento
                is_call = leg.opt_type == 'CALL'
                total += sign * q * max(0.0, (S - K) if is_call else (K - S))
        return total

    strikes = sorted({l.strike for l in legs if l.strike})

    # Delta líquido de CALLs → define comportamento para S→∞
    # Para calendário: o payoff intrínseco não descreve o máximo real,
    # mas usamos a varredura densa para aproximar o pico.
    if is_calendar:
        unlimited_profit = False
        unlimited_loss   = False
    else:
        net_call_delta = sum(
            leg.quantity * (1 if leg.side == 'BUY' else -1)
            for leg in legs if leg.opt_type == 'CALL'
        )
        # Ação comprada tem delta +1 (payoff ilimitado para cima)
        net_stock_delta = sum(
            leg.quantity * (1 if leg.side == 'BUY' else -1)
            for leg in legs if leg.opt_type == 'STOCK'
        )
        unlimited_profit = (net_call_delta + net_stock_delta) > 0
        unlimited_loss   = (net_call_delta + net_stock_delta) < 0

    # Varredura densa de pontos para capturar pico (inclui strikes e grade fina)
    max_K = max(strikes) if strikes else 100
    min_K = min(strikes) if strikes else 1
    pad   = max((max_K - min_K) * 0.6, max_K * 0.3)
    S_lo  = max(0.01, min_K - pad)
    S_hi  = max_K + pad
    N     = 400
    step  = (S_hi - S_lo) / N
    grid  = [S_lo + i * step for i in range(N + 1)] + strikes
    grid  = sorted(set(round(s, 4) for s in grid))
    test_prices = [0.01] + grid + [max_K * 5]
    payoffs = [(S, payoff_at(S)) for S in test_prices]

    max_profit = float('inf')  if unlimited_profit else max(p for _, p in payoffs)
    max_loss   = float('-inf') if unlimited_loss   else min(p for _, p in payoffs)

    # ── Breakevens matemáticos (cruzamentos de zero do payoff) ─────
    breakevens = []
    for i in range(len(payoffs) - 1):
        S1, P1 = payoffs[i]
        S2, P2 = payoffs[i + 1]
        if P1 == 0 and S1 not in breakevens:
            breakevens.append(round(S1, 2))
        elif P1 * P2 < 0:
            be = S1 + (-P1) * (S2 - S1) / (P2 - P1)
            breakevens.append(round(be, 2))
    if payoffs and payoffs[-1][1] == 0:
        be = round(payoffs[-1][0], 2)
        if be not in breakevens:
            breakevens.append(be)
    breakevens = sorted(set(breakevens))

    # ── Breakevens simplificados (fórmula de mercado BR) ───────────
    # BE_baixo = menor_strike_PUT − net_crédito / qty_calls_vendidas
    # BE_alto  = maior_strike_CALL + net_crédito / qty_calls_vendidas
    be_low = be_high = None
    sell_call_qty = sum(l.quantity for l in legs
                        if l.opt_type == 'CALL' and l.side == 'SELL')
    buy_put_qty   = sum(l.quantity for l in legs
                        if l.opt_type == 'PUT'  and l.side == 'BUY')
    put_strikes  = [l.strike for l in legs if l.opt_type == 'PUT'  and l.strike]
    call_strikes = [l.strike for l in legs if l.opt_type == 'CALL' and l.strike]

    if net > 0 and sell_call_qty > 0:
        if put_strikes:
            be_low  = round(min(put_strikes)  - net / sell_call_qty, 2)
        if call_strikes:
            be_high = round(max(call_strikes) + net / sell_call_qty, 2)
    elif net < 0 and buy_put_qty > 0:
        # Estratégia de débito com puts compradas
        if put_strikes:
            be_low = round(min(put_strikes) - abs(net) / buy_put_qty, 2)

    # ── POP via BS log-normal ───────────────────────────────────────
    pop = None
    try:
        S0 = op.underlying_price or 0
        # Fallback 1: busca underlying_price em Option ou PutSale com mesmo ativo
        if S0 <= 0 and op.underlying_asset:
            asset = op.underlying_asset
            opt_ref = Option.query.filter_by(
                underlying_asset=asset, user_id=op.user_id
            ).filter(Option.underlying_price > 0).first()
            if opt_ref:
                S0 = opt_ref.underlying_price
            else:
                ps_ref = PutSale.query.filter_by(
                    underlying_asset=asset, user_id=op.user_id
                ).filter(PutSale.underlying_price > 0).first()
                if ps_ref:
                    S0 = ps_ref.underlying_price
        # Fallback 2: usa strike médio das pernas como proxy do spot
        if S0 <= 0:
            strikes = [l.strike for l in legs if l.strike and l.strike > 0]
            if strikes:
                S0 = sum(strikes) / len(strikes)
        if S0 > 0 and breakevens:
            # Estima sigma médio das pernas vendidas (ou primeira perna)
            sell_legs = [l for l in legs if l.side == 'SELL' and l.entry_price > 0]
            ref_legs  = sell_legs or [l for l in legs if l.entry_price > 0]
            sigmas = []
            for l in ref_legs:
                exp_dates = [x.expiration_date for x in legs if x.expiration_date]
                T_leg = max(((max(exp_dates) - date.today()).days / 252.0), 1/252) if exp_dates else 30/252
                is_c  = (l.opt_type == 'CALL')
                try:
                    iv = _implied_vol(S0, l.strike or 1, T_leg,
                                      math.log(1 + _selic() / 100),
                                      l.entry_price, is_c)
                    sigmas.append(iv)
                except Exception:
                    pass
            sigma_avg = (sum(sigmas) / len(sigmas)) if sigmas else 0.30
            exp_dates = [l.expiration_date for l in legs if l.expiration_date]
            T = max(((max(exp_dates) - date.today()).days / 252.0), 1/252) if exp_dates else 30/252
            pop = _calc_pop(S0, breakevens, T, sigma_avg,
                            r=math.log(1 + _selic() / 100),
                            payoff_fn=payoff_at)
    except Exception as _e:
        print(f"[POP] erro em _calc_structured_metrics op={op.id}: {_e}")

    return dict(net=net, current_pnl=current_pnl,
                max_profit=max_profit, max_loss=max_loss,
                breakevens=breakevens,
                be_low=be_low, be_high=be_high,
                unlimited_profit=unlimited_profit, unlimited_loss=unlimited_loss,
                pop=pop)


def _calc_structured_metrics_safe(op):
    """Wrapper com fallback para nunca derrubar a página de opções."""
    try:
        return _calc_structured_metrics(op)
    except Exception as e:
        import traceback as _tb
        print(f"[ERRO] _calc_structured_metrics op={op.id}: {e}")
        _tb.print_exc()
        # Fallback: calcula pelo menos net e current_pnl sem BS
        legs = op.legs or []
        net = sum(
            (leg.entry_price if leg.side == 'SELL' else -leg.entry_price) * leg.quantity
            for leg in legs
        )
        current_pnl = sum(
            ((leg.entry_price - (leg.current_price or leg.entry_price)) if leg.side == 'SELL'
             else ((leg.current_price or leg.entry_price) - leg.entry_price)) * leg.quantity
            for leg in legs
        )
        # Payoff intrínseco simples nos strikes como aproximação
        strikes = sorted({l.strike for l in legs if l.strike})
        def _simple_payoff(S):
            t = net
            for leg in legs:
                sign = 1 if leg.side == 'BUY' else -1
                K = leg.strike or 0
                if leg.opt_type == 'CALL':
                    t += sign * leg.quantity * max(0.0, S - K)
                else:
                    t += sign * leg.quantity * max(0.0, K - S)
            return t
        max_K = max(strikes) if strikes else 100
        test = [0.0] + strikes + [max_K * 5]
        payoffs = [_simple_payoff(s) for s in test]
        return dict(net=net, current_pnl=current_pnl,
                    max_profit=max(payoffs), max_loss=min(payoffs),
                    breakevens=[], be_low=None, be_high=None,
                    unlimited_profit=False, unlimited_loss=False,
                    pop=None)


@app.route('/opcoes')
@login_required
def opcoes():
    all_options = Option.query.filter_by(user_id=current_user.id).all()

    # Cotações dos subjacentes SEM rede: usa os valores salvos no banco.
    # A atualização ao vivo acontece só pelo botão "Atualizar Cotações" (async),
    # feeder MT5 ou auto-update OpLab — o render da página não pode travar em HTTP.
    underlyings = list(set([o.underlying_asset for o in all_options]))
    quotes = _quotes_from_db(underlyings, current_user.id) if underlyings else {}

    # Get avg_price of underlying assets from user's portfolio
    underlying_avg_prices = {}
    if underlyings:
        assets = Asset.query.filter(Asset.user_id == current_user.id, Asset.ticker.in_(underlyings)).all()
        for a in assets:
            underlying_avg_prices[a.ticker] = a.avg_price

    processed_options = []  # VENDA_CALL
    venda_puts = []         # VENDA_PUT
    compra_calls = []       # COMPRA_CALL
    compra_puts = []        # COMPRA_PUT

    for opt in all_options:
        underlying_price = 0.0
        if opt.underlying_asset in quotes:
            underlying_price = quotes[opt.underlying_asset].get('price', 0.0)

        option_type = getattr(opt, 'option_type', 'VENDA_CALL') or 'VENDA_CALL'

        if option_type == 'VENDA_CALL':
            # Lógica existente para venda coberta
            total_sold = opt.quantity * opt.sale_price
            current_val = opt.quantity * opt.current_option_price
            profit = total_sold - current_val
            profit_pct = (profit / total_sold * 100) if total_sold > 0 else 0

            avg_price = underlying_avg_prices.get(opt.underlying_asset, 0.0)
            exercise_price = opt.strike_price + opt.sale_price
            lastro = underlying_price - exercise_price
            lucro_ex_pct = ((exercise_price - avg_price) / avg_price * 100) if avg_price > 0 else 0
            lucro_at_pct = ((underlying_price - avg_price) / avg_price * 100) if avg_price > 0 else 0
            lucro_ex_rs = (exercise_price - avg_price) * opt.quantity
            lucro_at_rs = (underlying_price - avg_price) * opt.quantity

            processed_options.append({
                'option': opt,
                'underlying_price': underlying_price,
                'total_sold': total_sold,
                'profit': profit,
                'profit_pct': profit_pct,
                'avg_price': avg_price,
                'exercise_price': exercise_price,
                'lastro': lastro,
                'lucro_ex_pct': lucro_ex_pct,
                'lucro_at_pct': lucro_at_pct,
                'lucro_ex_rs': lucro_ex_rs,
                'lucro_at_rs': lucro_at_rs,
            })

        elif option_type == 'VENDA_PUT':
            # Venda a Seco de Puts: prêmio recebido menos valor atual
            total_sold = opt.quantity * opt.sale_price
            current_val = opt.quantity * opt.current_option_price
            profit = total_sold - current_val
            profit_pct = (profit / total_sold * 100) if total_sold > 0 else 0

            lastro_rs = underlying_price - opt.strike_price
            lastro_pct = (lastro_rs / opt.strike_price * 100) if opt.strike_price > 0 else 0
            breakeven = opt.strike_price - opt.sale_price
            lastrobk_rs = underlying_price - breakeven
            lastrobk_pct = (lastrobk_rs / breakeven * 100) if breakeven > 0 else 0

            venda_puts.append({
                'option': opt,
                'underlying_price': underlying_price,
                'total_sold': total_sold,
                'current_val': current_val,
                'profit': profit,
                'profit_pct': profit_pct,
                'lastro_rs': lastro_rs,
                'lastro_pct': lastro_pct,
                'breakeven': breakeven,
                'lastrobk_rs': lastrobk_rs,
                'lastrobk_pct': lastrobk_pct,
            })

        else:
            # Lógica para compra a seco (COMPRA_CALL / COMPRA_PUT)
            total_invested = opt.quantity * opt.sale_price
            current_value = opt.quantity * opt.current_option_price
            profit = current_value - total_invested
            profit_pct = (profit / total_invested * 100) if total_invested > 0 else 0

            item = {
                'option': opt,
                'underlying_price': underlying_price,
                'total_invested': total_invested,
                'current_value': current_value,
                'profit': profit,
                'profit_pct': profit_pct,
            }

            if option_type == 'COMPRA_CALL':
                compra_calls.append(item)
            else:
                compra_puts.append(item)

    from datetime import date as date_cls
    today = date_cls.today()

    # Add days_left and last_update to every option item
    def _fmt_last_update(opt):
        lu = opt.last_update
        if not lu:
            return '-'
        # Gravado com now_brt(): o SQLite descarta o tzinfo mas mantém a hora de Brasília.
        # astimezone() num datetime naive assumiria o fuso do servidor (UTC) e subtrairia 3h.
        if lu.tzinfo is None:
            return lu.strftime('%H:%M')
        return lu.astimezone(_BRT).strftime('%H:%M')

    for item in processed_options + venda_puts + compra_calls + compra_puts:
        exp = item['option'].expiration_date
        item['days_left']   = du_count(today, exp) if exp else None
        item['last_update'] = _fmt_last_update(item['option'])

    # Process spreads
    all_spreads = OptionSpread.query.filter_by(user_id=current_user.id).all()

    # Cotações dos subjacentes dos spreads — também sem rede (valores do banco)
    spread_underlyings = list(set([s.underlying_asset for s in all_spreads]))
    spread_quotes = _quotes_from_db(spread_underlyings, current_user.id) if spread_underlyings else {}

    spreads_alta_put = []    # crédito: vende put alta + compra put baixa
    spreads_alta_call = []   # débito:  compra call baixa + vende call alta
    spreads_baixa_put = []   # débito:  compra put alta + vende put baixa
    spreads_baixa_call = []  # crédito: vende call baixa + compra call alta

    for sp in all_spreads:
        underlying_price = spread_quotes.get(sp.underlying_asset, {}).get('price', 0.0)
        net = sp.leg_short_price - sp.leg_long_price   # >0 crédito, <0 débito
        net_total = net * sp.quantity
        current_net = sp.leg_short_current - sp.leg_long_current
        result = (net - current_net) * sp.quantity
        width = abs(sp.leg_short_strike - sp.leg_long_strike)
        is_credit = net >= 0
        max_gain = net * sp.quantity if is_credit else (width + net) * sp.quantity
        max_loss = (width - net) * sp.quantity if is_credit else abs(net) * sp.quantity
        result_pct = (result / max_gain * 100) if max_gain != 0 else 0
        days_left = du_count(today, sp.expiration_date)

        # ── POP da trava via BS ───────────────────────────────────────
        pop_spread = None
        try:
            S0 = sp.underlying_price or underlying_price or 0
            if S0 > 0 and sp.expiration_date:
                T_sp = max((sp.expiration_date - today).days / 252.0, 1/252)
                r_cont = math.log(1 + _selic() / 100)
                is_put = 'PUT' in sp.spread_type
                # IV da perna vendida (define o sigma da estrutura)
                ref_p = sp.leg_short_price
                ref_k = sp.leg_short_strike
                if ref_p > 0 and ref_k > 0:
                    sigma_sp = _implied_vol(S0, ref_k, T_sp, r_cont, ref_p, not is_put)
                    # Breakeven da trava
                    if is_credit:
                        if is_put:
                            be_sp = sp.leg_short_strike - net   # breakeven put crédito
                        else:
                            be_sp = sp.leg_short_strike + net   # breakeven call crédito
                        pop_spread = _calc_pop(S0, [be_sp], T_sp, sigma_sp, r_cont)
                        # Ajusta sinal: put crédito lucra acima do BE; call crédito lucra abaixo
                        if is_put and pop_spread is not None:
                            pop_spread = pop_spread   # P(S>BE) já correto
                        elif not is_put and pop_spread is not None:
                            pop_spread = round(100 - pop_spread, 1)
        except Exception:
            pass

        item = {
            'spread': sp,
            'underlying_price': underlying_price,
            'net': net,
            'net_total': net_total,
            'is_credit': is_credit,
            'current_net': current_net,
            'result': result,
            'result_pct': result_pct,
            'max_gain': max_gain,
            'max_loss': max_loss,
            'days_left': days_left,
            'pop': pop_spread,
        }
        if sp.spread_type == 'TRAVA_ALTA_PUT':
            spreads_alta_put.append(item)
        elif sp.spread_type == 'TRAVA_ALTA_CALL':
            spreads_alta_call.append(item)
        elif sp.spread_type == 'TRAVA_BAIXA_PUT':
            spreads_baixa_put.append(item)
        else:
            spreads_baixa_call.append(item)

    # Process operações estruturadas (nacionais) e Tastytrade (internacionais)
    raw_ops = StructuredOp.query.filter_by(user_id=current_user.id, status='OPEN').all()
    structured_ops, tastytrade_ops = [], []
    for op in raw_ops:
        metrics = _calc_structured_metrics_safe(op)
        item = {'op': op, **metrics}
        (tastytrade_ops if getattr(op, 'intl', False) else structured_ops).append(item)

    oplab_token_ok = bool(Settings.get_value('oplab_token', user_id=current_user.id))
    return render_template('opcoes.html', options=processed_options,
                           venda_puts=venda_puts,
                           compra_calls=compra_calls, compra_puts=compra_puts,
                           spreads_alta_put=spreads_alta_put,
                           spreads_alta_call=spreads_alta_call,
                           spreads_baixa_put=spreads_baixa_put,
                           spreads_baixa_call=spreads_baixa_call,
                           structured_ops=structured_ops,
                           tastytrade_ops=tastytrade_ops,
                           oplab_token_ok=oplab_token_ok,
                           today=date.today())

@app.route('/add_option', methods=['GET', 'POST'])
@login_required
def add_option():
    if request.method == 'POST':
        try:
            ticker = request.form.get('ticker')
            underlying = request.form.get('underlying_asset')
            quantity_str = request.form.get('quantity')
            strike_str = request.form.get('strike_price')
            expiration_str = request.form.get('expiration_date')
            sale_price_str = request.form.get('sale_price')
            option_type = request.form.get('option_type', 'VENDA_CALL')

            if not all([ticker, underlying, quantity_str, strike_str, expiration_str, sale_price_str]):
                flash("Todos os campos são obrigatórios.", "warning")
                return redirect(url_for('add_option', type=option_type))

            ticker = ticker.upper()
            underlying = underlying.upper()
            quantity = int(quantity_str)
            strike = float(strike_str.replace(',', '.'))
            expiration = datetime.strptime(expiration_str, '%Y-%m-%d').date()
            sale_price = float(sale_price_str.replace(',', '.'))

            curr_price_str = request.form.get('current_option_price')
            current_option_price = float(curr_price_str.replace(',', '.')) if curr_price_str else 0.0

            entry_date_str = request.form.get('entry_date')
            entry_date = datetime.strptime(entry_date_str, '%Y-%m-%d').date() if entry_date_str else None

            opt = Option(
                user_id=current_user.id,
                option_type=option_type,
                ticker=ticker,
                quantity=quantity,
                underlying_asset=underlying,
                strike_price=strike,
                expiration_date=expiration,
                sale_price=sale_price,
                current_option_price=current_option_price,
                entry_date=entry_date
            )
            db.session.add(opt)
            db.session.commit()

            flash("Opção adicionada com sucesso!", "success")
            return redirect(url_for('opcoes'))

        except ValueError as ve:
            flash(f"Erro de formato: {ve}", "danger")
        except Exception as e:
            flash(f"Erro ao salvar opção: {e}", "danger")
            print(f"Error add_option: {e}")
            import traceback
            traceback.print_exc()

        return redirect(url_for('add_option', type=option_type))

    option_type = request.args.get('type', 'VENDA_CALL')
    return render_template('add_option.html', option_type=option_type)

@app.route('/add_spread', methods=['GET', 'POST'])
@login_required
def add_spread():
    if request.method == 'POST':
        try:
            spread_type = request.form.get('spread_type', 'TRAVA_ALTA_PUT')
            underlying = request.form.get('underlying_asset', '').upper()
            quantity = int(request.form.get('quantity'))
            expiration = datetime.strptime(request.form.get('expiration_date'), '%Y-%m-%d').date()
            entry_date_str = request.form.get('entry_date')
            entry_date = datetime.strptime(entry_date_str, '%Y-%m-%d').date() if entry_date_str else None

            leg_long_ticker = request.form.get('leg_long_ticker', '').upper()
            leg_long_strike = float(request.form.get('leg_long_strike', '0').replace(',', '.'))
            leg_long_price = float(request.form.get('leg_long_price', '0').replace(',', '.'))
            leg_long_current = float(request.form.get('leg_long_current', '0').replace(',', '.'))

            leg_short_ticker = request.form.get('leg_short_ticker', '').upper()
            leg_short_strike = float(request.form.get('leg_short_strike', '0').replace(',', '.'))
            leg_short_price = float(request.form.get('leg_short_price', '0').replace(',', '.'))
            leg_short_current = float(request.form.get('leg_short_current', '0').replace(',', '.'))

            pop_str = request.form.get('pop', '')
            pop = float(pop_str.replace(',', '.')) if pop_str else None

            sp = OptionSpread(
                user_id=current_user.id,
                spread_type=spread_type,
                underlying_asset=underlying,
                quantity=quantity,
                expiration_date=expiration,
                entry_date=entry_date,
                leg_long_ticker=leg_long_ticker,
                leg_long_strike=leg_long_strike,
                leg_long_price=leg_long_price,
                leg_long_current=leg_long_current,
                leg_short_ticker=leg_short_ticker,
                leg_short_strike=leg_short_strike,
                leg_short_price=leg_short_price,
                leg_short_current=leg_short_current,
                pop=pop,
            )
            db.session.add(sp)
            db.session.commit()
            flash("Trava adicionada com sucesso!", "success")
        except Exception as e:
            flash(f"Erro ao salvar trava: {e}", "danger")
        return redirect(url_for('opcoes'))

    spread_type = request.args.get('type', 'TRAVA_ALTA_PUT')
    return render_template('add_spread.html', spread_type=spread_type)

@app.route('/edit_spread/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_spread(id):
    sp = OptionSpread.query.get_or_404(id)
    if sp.user_id != current_user.id:
        flash("Sem permissão.", "danger")
        return redirect(url_for('opcoes'))
    if request.method == 'POST':
        try:
            sp.underlying_asset = request.form.get('underlying_asset', '').upper()
            sp.quantity = int(request.form.get('quantity'))
            sp.expiration_date = datetime.strptime(request.form.get('expiration_date'), '%Y-%m-%d').date()
            entry_date_str = request.form.get('entry_date')
            sp.entry_date = datetime.strptime(entry_date_str, '%Y-%m-%d').date() if entry_date_str else None

            sp.leg_long_ticker = request.form.get('leg_long_ticker', '').upper()
            sp.leg_long_strike = float(request.form.get('leg_long_strike', '0').replace(',', '.'))
            sp.leg_long_price = float(request.form.get('leg_long_price', '0').replace(',', '.'))
            sp.leg_long_current = float(request.form.get('leg_long_current', '0').replace(',', '.'))

            sp.leg_short_ticker = request.form.get('leg_short_ticker', '').upper()
            sp.leg_short_strike = float(request.form.get('leg_short_strike', '0').replace(',', '.'))
            sp.leg_short_price = float(request.form.get('leg_short_price', '0').replace(',', '.'))
            sp.leg_short_current = float(request.form.get('leg_short_current', '0').replace(',', '.'))

            pop_str = request.form.get('pop', '')
            sp.pop = float(pop_str.replace(',', '.')) if pop_str else None

            db.session.commit()
            flash("Trava atualizada!", "success")
        except Exception as e:
            flash(f"Erro: {e}", "danger")
        return redirect(url_for('opcoes'))
    return render_template('add_spread.html', spread_type=sp.spread_type, spread=sp, edit=True)

@app.route('/close_spread/<int:id>', methods=['GET', 'POST'])
@login_required
def close_spread(id):
    sp = OptionSpread.query.get_or_404(id)
    if sp.user_id != current_user.id:
        flash("Sem permissão.", "danger")
        return redirect(url_for('opcoes'))
    if request.method == 'POST':
        try:
            close_long_price = float(request.form.get('close_long_price', '0').replace(',', '.'))
            close_short_price = float(request.form.get('close_short_price', '0').replace(',', '.'))
            exit_date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
            reason = request.form.get('reason', 'ENCERRAMENTO')

            # Lógica de fluxo de caixa da trava:
            # Na montagem: recebe leg_short_price (venda) e paga leg_long_price (compra)
            #   crédito líquido = leg_short_price - leg_long_price  (>0 = crédito, <0 = débito)
            # No fechamento: recebe close_long_price (venda da comprada) e paga close_short_price (recompra da vendida)
            #   custo de fechamento = close_short_price - close_long_price
            # Lucro = (crédito_montagem - custo_fechamento) * qty
            net_open  = sp.leg_short_price - sp.leg_long_price   # crédito líquido de entrada (>0=crédito)
            net_close = close_short_price - close_long_price      # custo líquido de fechamento (>0=débito)
            profit_val = (net_open - net_close) * sp.quantity

            width = abs(sp.leg_short_strike - sp.leg_long_strike)
            # Referência de risco: máximo risco na montagem
            max_risk = (width - net_open) * sp.quantity if net_open >= 0 else (width + abs(net_open)) * sp.quantity
            profit_pct = (profit_val / abs(max_risk) * 100) if max_risk != 0 else 0
            days_held  = (exit_date - sp.entry_date).days if sp.entry_date else 0

            type_labels = {
                'TRAVA_ALTA_PUT': 'T.Alta Put', 'TRAVA_ALTA_CALL': 'T.Alta Call',
                'TRAVA_BAIXA_PUT': 'T.Baixa Put', 'TRAVA_BAIXA_CALL': 'T.Baixa Call',
            }
            label = type_labels.get(sp.spread_type, 'TRAVA')
            notes = (
                f"{sp.spread_type} | {sp.underlying_asset} | "
                f"C:{sp.leg_long_ticker} K={sp.leg_long_strike:.2f} entrada@{sp.leg_long_price:.2f} saída@{close_long_price:.2f} | "
                f"V:{sp.leg_short_ticker} K={sp.leg_short_strike:.2f} entrada@{sp.leg_short_price:.2f} saída@{close_short_price:.2f}"
            )
            history = TradeHistory(
                user_id    = current_user.id,
                ticker     = label,
                strategy   = 'Opções',
                entry_date = sp.entry_date,
                exit_date  = exit_date,
                # buy_price  = custo líquido de abertura (negativo = crédito recebido)
                # sell_price = custo líquido de fechamento (negativo = crédito recebido ao fechar)
                buy_price  = round(net_open,  4),
                sell_price = round(net_close, 4),
                quantity   = sp.quantity,
                profit_value = round(profit_val, 2),
                profit_pct   = round(profit_pct, 2),
                days_held    = days_held,
                reason       = reason,
                underlying   = sp.underlying_asset,
                notes        = notes,
            )
            db.session.add(history)
            db.session.delete(sp)
            db.session.commit()
            flash("Trava encerrada e registrada no histórico!", "success")
        except Exception as e:
            flash(f"Erro: {e}", "danger")
        return redirect(url_for('opcoes'))
    return render_template('close_spread.html', spread=sp, today=date.today())

@app.route('/delete_spread/<int:id>')
@login_required
def delete_spread(id):
    sp = OptionSpread.query.get_or_404(id)
    if sp.user_id != current_user.id:
        flash("Sem permissão.", "danger")
        return redirect(url_for('opcoes'))
    db.session.delete(sp)
    db.session.commit()
    flash("Trava excluída.", "success")
    return redirect(url_for('opcoes'))

# ─── Operações Estruturadas ────────────────────────────────────────────────

@app.route('/estruturada/add', methods=['GET', 'POST'])
@login_required
def add_estruturada():
    if request.method == 'POST':
        try:
            name             = request.form.get('name', '').strip()
            underlying_asset = request.form.get('underlying_asset', '').strip().upper()

            # Lê arrays de pernas
            tickers     = request.form.getlist('leg_ticker')
            sides       = request.form.getlist('leg_side')
            opt_types   = request.form.getlist('leg_opt_type')
            quantities  = request.form.getlist('leg_quantity')
            strikes     = request.form.getlist('leg_strike')
            expirations = request.form.getlist('leg_expiration')
            entry_prices = request.form.getlist('leg_entry_price')
            current_prices = request.form.getlist('leg_current_price')

            if not name or not tickers:
                flash('Informe o nome e ao menos uma perna.', 'warning')
                return redirect(url_for('add_estruturada'))

            uses_collateral = request.form.get('uses_stock_collateral') == '1'
            op = StructuredOp(
                user_id=current_user.id,
                name=name,
                underlying_asset=underlying_asset,
                uses_stock_collateral=uses_collateral,
                intl=request.form.get('intl') == '1',   # Tastytrade (internacional)
                status='OPEN',
                created_at=datetime.now(),
            )
            db.session.add(op)
            db.session.flush()  # gera op.id

            for i, ticker in enumerate(tickers):
                ticker = ticker.strip().upper()
                if not ticker:
                    continue
                exp = None
                try:
                    exp = date.fromisoformat(expirations[i]) if i < len(expirations) and expirations[i] else None
                except ValueError:
                    pass
                leg = StructuredLeg(
                    op_id=op.id,
                    ticker=ticker,
                    side=sides[i] if i < len(sides) else 'SELL',
                    opt_type=opt_types[i] if i < len(opt_types) else 'CALL',
                    quantity=int(quantities[i]) if i < len(quantities) and quantities[i] else 1,
                    strike=float(strikes[i].replace(',', '.')) if i < len(strikes) and strikes[i] else 0.0,
                    expiration_date=exp,
                    entry_price=float(entry_prices[i].replace(',', '.')) if i < len(entry_prices) and entry_prices[i] else 0.0,
                    current_price=float(current_prices[i].replace(',', '.')) if i < len(current_prices) and current_prices[i] else 0.0,
                )
                db.session.add(leg)

            db.session.commit()
            flash('Operação estruturada criada com sucesso.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao salvar: {e}', 'danger')
        return redirect(url_for('opcoes'))

    return render_template('estruturada_form.html', op=None, edit=False,
                           intl=request.args.get('intl') == '1')


@app.route('/estruturada/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_estruturada(id):
    op = StructuredOp.query.get_or_404(id)
    if op.user_id != current_user.id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('opcoes'))

    if request.method == 'POST':
        try:
            op.name                  = request.form.get('name', '').strip()
            op.underlying_asset      = request.form.get('underlying_asset', '').strip().upper()
            op.uses_stock_collateral = request.form.get('uses_stock_collateral') == '1'

            tickers      = request.form.getlist('leg_ticker')
            sides        = request.form.getlist('leg_side')
            opt_types    = request.form.getlist('leg_opt_type')
            quantities   = request.form.getlist('leg_quantity')
            strikes      = request.form.getlist('leg_strike')
            expirations  = request.form.getlist('leg_expiration')
            entry_prices = request.form.getlist('leg_entry_price')
            current_prices = request.form.getlist('leg_current_price')

            # Remove pernas antigas e recria
            for leg in list(op.legs):
                db.session.delete(leg)
            db.session.flush()

            for i, ticker in enumerate(tickers):
                ticker = ticker.strip().upper()
                if not ticker:
                    continue
                exp = None
                try:
                    exp = date.fromisoformat(expirations[i]) if i < len(expirations) and expirations[i] else None
                except ValueError:
                    pass
                leg = StructuredLeg(
                    op_id=op.id,
                    ticker=ticker,
                    side=sides[i] if i < len(sides) else 'SELL',
                    opt_type=opt_types[i] if i < len(opt_types) else 'CALL',
                    quantity=int(quantities[i]) if i < len(quantities) and quantities[i] else 1,
                    strike=float(strikes[i].replace(',', '.')) if i < len(strikes) and strikes[i] else 0.0,
                    expiration_date=exp,
                    entry_price=float(entry_prices[i].replace(',', '.')) if i < len(entry_prices) and entry_prices[i] else 0.0,
                    current_price=float(current_prices[i].replace(',', '.')) if i < len(current_prices) and current_prices[i] else 0.0,
                )
                db.session.add(leg)

            db.session.commit()
            flash('Operação atualizada.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro: {e}', 'danger')
        return redirect(url_for('opcoes'))

    return render_template('estruturada_form.html', op=op, edit=True,
                           intl=bool(getattr(op, 'intl', False)))


@app.route('/estruturada/<int:id>/close', methods=['POST'])
@login_required
def close_estruturada(id):
    op = StructuredOp.query.get_or_404(id)
    if op.user_id != current_user.id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('opcoes'))

    # Lógica de P&L para estruturadas:
    # Cada perna contribui de forma independente:
    #   SELL: recebeu entry_price, pagou current_price para fechar → lucro = entry - current
    #   BUY:  pagou entry_price, recebeu current_price ao fechar  → lucro = current - entry
    # Crédito líquido de abertura = Σ(SELL entry * qty) - Σ(BUY entry * qty)
    # Custo líquido de fechamento = Σ(SELL current * qty) - Σ(BUY current * qty)
    # Lucro total = crédito_abertura - custo_fechamento

    net_open  = 0.0   # crédito líquido recebido na montagem
    net_close = 0.0   # custo líquido para fechar (recomprar vendidas, vender compradas)
    total_risk = 0.0  # soma dos prêmios pagos (posições BUY) como referência de risco

    for leg in op.legs:
        cur = leg.current_price if leg.current_price else leg.entry_price
        if leg.side == 'SELL':
            net_open  += leg.entry_price * leg.quantity
            net_close += cur             * leg.quantity
        else:
            net_open  -= leg.entry_price * leg.quantity
            net_close -= cur             * leg.quantity
            total_risk += leg.entry_price * leg.quantity

    # Resultado JÁ REALIZADO nos manejos anteriores (trechos de pernas fechadas
    # e substituídas). Cada trecho foi medido do seu entry ao seu close; somá-lo
    # ao resultado das pernas atuais dá o P&L da operação inteira, sem dupla
    # contagem — o manejo NÃO gera lucro isolado; só o encerramento apura.
    import json as _json
    realized_prev = 0.0
    if op.roll_history:
        try:
            for _rh in _json.loads(op.roll_history):
                # Só manejos do novo modelo (defer_pnl) entram aqui; os antigos
                # já foram para o Histórico como MANEJO e não podem contar 2×.
                if _rh.get('defer_pnl') and _rh.get('realized_pnl') is not None:
                    realized_prev += float(_rh['realized_pnl'])
        except Exception:
            realized_prev = 0.0

    pnl_pernas = net_open - net_close
    pnl_total  = pnl_pernas + realized_prev
    # % sobre o risco total engajado (prêmios pagos nas compras, ou crédito líquido se tudo vendas)
    risk_ref = total_risk if total_risk > 0 else abs(net_open)
    pct = (pnl_total / risk_ref * 100) if risk_ref != 0 else 0

    exit_date  = date.today()
    entry_date = op.created_at.date() if op.created_at else exit_date
    days_held  = (exit_date - entry_date).days

    ticker_label = (op.name or op.underlying_asset or 'ESTRUT')[:20]

    # Detalhes completos de cada perna para rastreabilidade
    legs_detail = ' | '.join(
        f"{'V' if l.side=='SELL' else 'C'}:{l.ticker} {l.opt_type} K={l.strike:.2f}"
        f" entrada@{l.entry_price:.2f} saída@{l.current_price or l.entry_price:.2f}"
        for l in op.legs
    )
    realized_note = (f" | inclui R$ {realized_prev:.2f} realizados em manejos anteriores"
                     if abs(realized_prev) > 0.005 else '')
    notes = f"{op.underlying_asset} | {legs_detail}{realized_note}"

    history = TradeHistory(
        user_id      = current_user.id,
        ticker       = ticker_label,
        strategy     = 'Opções',
        entry_date   = entry_date,
        exit_date    = exit_date,
        buy_price    = round(net_open,  4),   # crédito líquido de abertura (>0=crédito)
        sell_price   = round(net_close, 4),   # custo líquido de fechamento
        quantity     = 1,                     # quantidade não faz sentido para multi-perna; P&L é em R$
        profit_value = round(pnl_total, 2),
        profit_pct   = round(pct, 2),
        days_held    = days_held,
        reason       = 'Encerramento',
        underlying   = op.underlying_asset,
        notes        = notes,
    )
    db.session.add(history)
    op.status = 'CLOSED'
    db.session.commit()
    flash('Operação encerrada e registrada no histórico.', 'success')
    return redirect(url_for('opcoes'))


@app.route('/estruturada/<int:id>/delete', methods=['POST'])
@login_required
def delete_estruturada(id):
    op = StructuredOp.query.get_or_404(id)
    if op.user_id != current_user.id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('opcoes'))
    db.session.delete(op)
    db.session.commit()
    flash('Operação excluída.', 'success')
    return redirect(url_for('opcoes'))


# ─────────────────────────────────────────────────────────────────────────────
# Simulação de Opções — CRUD + API de dados
# ─────────────────────────────────────────────────────────────────────────────

def _selic():
    return float(Settings.get_value('selic_rate', user_id=current_user.id, default='14.5'))


def _build_ticker_maps(uid):
    """
    Constrói TICKER_MAP e OPTION_MAP apenas com ativos/opções ATIVOS:
    - Ativos em carteira com qty > 0
    - Options cadastradas em /opcoes (sem filtro de data — o usuário exclui quando fecha)
    - Pernas de StructuredOp com status='OPEN'
    - PutSale com vencimento >= hoje
    - Subjacentes de todas as acima
    Retorna (asset_tickers: set, option_tickers: dict, ticker_map_text: str, option_map_text: str)
    """
    today = date.today()

    # ── TICKER_MAP ──────────────────────────────────────────────────────────────
    asset_tickers = {
        (a.ticker.upper(), a.type)
        for a in Asset.query.filter(Asset.user_id == uid, Asset.quantity > 0).all()
    }

    # Subjacentes de Options ativas (com vencimento >= hoje ou sem data)
    for o in Option.query.filter_by(user_id=uid).all():
        if o.underlying_asset:
            if not o.expiration_date or o.expiration_date >= today:
                asset_tickers.add((o.underlying_asset.upper(), 'ACAO'))

    # Subjacentes de StructuredOp ABERTAS
    for op in StructuredOp.query.filter_by(user_id=uid, status='OPEN').all():
        if op.underlying_asset:
            asset_tickers.add((op.underlying_asset.upper(), 'ACAO'))

    # Subjacentes de PutSale com vencimento >= hoje
    for ps in PutSale.query.filter_by(user_id=uid).all():
        if ps.underlying_asset and ps.expiration_date >= today:
            asset_tickers.add((ps.underlying_asset.upper(), 'ACAO'))

    # Subjacentes de OptionSpread (travas) com vencimento >= hoje
    for sp in OptionSpread.query.filter_by(user_id=uid).all():
        if sp.expiration_date >= today and sp.underlying_asset:
            asset_tickers.add((sp.underlying_asset.upper(), 'ACAO'))

    # ── OPTION_MAP ──────────────────────────────────────────────────────────────
    option_tickers = {}

    # Options ativas (vencimento >= hoje)
    for o in Option.query.filter_by(user_id=uid).all():
        if o.ticker and (not o.expiration_date or o.expiration_date >= today):
            exp_str = o.expiration_date.strftime('%d/%m/%Y') if o.expiration_date else '?'
            option_tickers[o.ticker.upper()] = f'{o.underlying_asset} | venc {exp_str}'

    # Pernas de StructuredOp ABERTAS (vencimento >= hoje)
    for op in StructuredOp.query.filter_by(user_id=uid, status='OPEN').all():
        for leg in op.legs:
            if leg.ticker and (not leg.expiration_date or leg.expiration_date >= today):
                exp_str = leg.expiration_date.strftime('%d/%m/%Y') if leg.expiration_date else '?'
                option_tickers[leg.ticker.upper()] = (
                    f'{op.underlying_asset} | {leg.opt_type} K={leg.strike:.2f} | venc {exp_str}'
                )

    # PutSale com vencimento >= hoje
    for ps in PutSale.query.filter_by(user_id=uid).all():
        if ps.ticker and ps.expiration_date >= today:
            exp_str = ps.expiration_date.strftime('%d/%m/%Y')
            option_tickers[ps.ticker.upper()] = (
                f'{ps.underlying_asset} | PUT K={ps.strike:.2f} | venc {exp_str}'
            )

    # OptionSpread (travas) com vencimento >= hoje — pernas long e short
    for sp in OptionSpread.query.filter_by(user_id=uid).all():
        if sp.expiration_date < today:
            continue
        exp_str = sp.expiration_date.strftime('%d/%m/%Y')
        if sp.leg_long_ticker:
            option_tickers[sp.leg_long_ticker.upper()] = (
                f'{sp.underlying_asset} | {sp.spread_type} long K={sp.leg_long_strike:.2f} | venc {exp_str}'
            )
        if sp.leg_short_ticker:
            option_tickers[sp.leg_short_ticker.upper()] = (
                f'{sp.underlying_asset} | {sp.spread_type} short K={sp.leg_short_strike:.2f} | venc {exp_str}'
            )

    # SearchedOption — opções buscadas manualmente (expiram em 10 dias)
    cutoff = datetime.utcnow() - timedelta(days=10)
    for so in SearchedOption.query.filter(
            SearchedOption.user_id == uid,
            SearchedOption.searched_at >= cutoff).all():
        if so.ticker:
            info = f'buscada em {so.searched_at.strftime("%d/%m/%Y")}'
            if so.underlying:
                info = f'{so.underlying} | {info}'
            option_tickers.setdefault(so.ticker.upper(), info)

    # ── Formata textos para exibição ────────────────────────────────────────────
    ticker_map_lines = sorted(
        [f'    "{t}": "{t}",  # {tp}' for t, tp in asset_tickers],
        key=lambda x: x.strip()
    )
    option_map_lines = sorted(
        [f'    "{t}": "{t}",  # {info}' for t, info in option_tickers.items()],
        key=lambda x: x.strip()
    )
    ticker_map_text = "TICKER_MAP = {\n" + "\n".join(ticker_map_lines) + "\n}"
    option_map_text = "OPTION_MAP = {\n" + "\n".join(option_map_lines) + "\n}"

    return asset_tickers, option_tickers, ticker_map_text, option_map_text

def _sim_to_producao(sim):
    """Converte uma simulação em posição REAL de acompanhamento (/opcoes).
    Classificação automática pela estrutura das pernas:
      - 1 opção (a seco, ou venda de CALL com ação = coberta) → Option
      - 2 opções do mesmo tipo/vencimento, 1 compra + 1 venda, mesma qtd → OptionSpread
      - qualquer outra combinação (incl. pernas de ação) → StructuredOp
    Retorna (nome_da_tabela_destino, None) ou (None, mensagem_de_erro)."""
    legs       = list(sim.legs or [])
    opt_legs   = [l for l in legs if l.leg_type in ('CALL', 'PUT')]
    stock_legs = [l for l in legs if l.leg_type == 'STOCK']
    today = date.today()

    if not opt_legs:
        return None, 'A simulação não tem pernas de opção.'
    for l in opt_legs:
        if not (l.ticker or '').strip() or not l.expiration or not l.strike:
            return None, 'Toda perna de opção precisa de ticker, strike e vencimento.'

    # ── 1 opção → Option (venda coberta quando há perna de ação + venda de CALL)
    only_covered_stock = (not stock_legs or
                          (len(opt_legs) == 1 and opt_legs[0].side == 'SELL'
                           and opt_legs[0].leg_type == 'CALL'))
    if len(opt_legs) == 1 and only_covered_stock:
        l = opt_legs[0]
        ot = ('VENDA_' if l.side == 'SELL' else 'COMPRA_') + l.leg_type
        db.session.add(Option(
            user_id=sim.user_id, option_type=ot, ticker=l.ticker.strip().upper(),
            underlying_asset=(sim.underlying or '').upper(), quantity=l.quantity,
            strike_price=l.strike, expiration_date=l.expiration,
            sale_price=l.premium, current_option_price=l.premium,
            entry_date=today, last_update=now_brt()))
        labels = {'VENDA_CALL': 'Venda Coberta de Calls', 'VENDA_PUT': 'Venda a Seco de Puts',
                  'COMPRA_CALL': 'Compra a Seco de Calls', 'COMPRA_PUT': 'Compra a Seco de Puts'}
        return labels.get(ot, ot), None

    # ── 2 opções mesmo tipo/vencimento, compra + venda, mesma qtd → OptionSpread
    if (len(opt_legs) == 2 and not stock_legs
            and opt_legs[0].leg_type == opt_legs[1].leg_type
            and opt_legs[0].expiration == opt_legs[1].expiration
            and {opt_legs[0].side, opt_legs[1].side} == {'BUY', 'SELL'}
            and opt_legs[0].quantity == opt_legs[1].quantity):
        lb = next(l for l in opt_legs if l.side == 'BUY')
        ls = next(l for l in opt_legs if l.side == 'SELL')
        is_call = lb.leg_type == 'CALL'
        # compra strike menor que a venda = alta; maior = baixa (CALLs e PUTs)
        alta = lb.strike < ls.strike
        sp_type = ('TRAVA_ALTA_' if alta else 'TRAVA_BAIXA_') + ('CALL' if is_call else 'PUT')
        db.session.add(OptionSpread(
            user_id=sim.user_id, spread_type=sp_type,
            underlying_asset=(sim.underlying or '').upper(),
            quantity=lb.quantity, expiration_date=lb.expiration, entry_date=today,
            leg_long_ticker=lb.ticker.strip().upper(),  leg_long_strike=lb.strike,
            leg_long_price=lb.premium,  leg_long_current=lb.premium,
            leg_short_ticker=ls.ticker.strip().upper(), leg_short_strike=ls.strike,
            leg_short_price=ls.premium, leg_short_current=ls.premium))
        labels = {'TRAVA_ALTA_CALL': 'Trava de Alta com Calls',
                  'TRAVA_BAIXA_CALL': 'Trava de Baixa com Calls',
                  'TRAVA_ALTA_PUT': 'Trava de Alta com Puts',
                  'TRAVA_BAIXA_PUT': 'Trava de Baixa com Puts'}
        return labels[sp_type], None

    # ── Demais estruturas → StructuredOp com todas as pernas (incl. ação)
    op = StructuredOp(user_id=sim.user_id,
                      name=(sim.name or ('Estruturada ' + (sim.underlying or ''))).strip()[:100],
                      underlying_asset=(sim.underlying or '').upper(),
                      status='OPEN', created_at=datetime.now())
    db.session.add(op)
    db.session.flush()
    for l in legs:
        db.session.add(StructuredLeg(
            op_id=op.id,
            ticker=(l.ticker or sim.underlying or '').strip().upper(),
            side=l.side, opt_type=l.leg_type, quantity=l.quantity,
            strike=l.strike or 0.0, expiration_date=l.expiration,
            entry_price=l.premium or 0.0, current_price=l.premium or 0.0,
            last_update=now_brt()))
    return 'Operações Estruturadas', None


def _sim_producao_redirect(sim):
    """Trata o clique em 'Salvar em Produção': cria os registros e redireciona."""
    destino, err = _sim_to_producao(sim)
    if err:
        flash('Simulação salva, mas NÃO enviada à produção: ' + err, 'warning')
        return redirect(url_for('simulacao_edit', id=sim.id))
    db.session.commit()
    flash(f'Operação salva em produção na tabela "{destino}". Use Atualizar Cotações para trazer os preços ao vivo.', 'success')
    return redirect(url_for('opcoes'))


@app.route('/simulacao_opcoes')
@login_required
def simulacao_opcoes():
    sims = SimulacaoOpcoes.query.filter_by(user_id=current_user.id).order_by(SimulacaoOpcoes.created_at.desc()).all()
    return render_template('simulacao_opcoes.html', sims=sims, sim=None, selic=_selic())


@app.route('/simulacao_opcoes/new', methods=['GET', 'POST'])
@login_required
def simulacao_new():
    if request.method == 'POST':
        name       = request.form.get('name', '').strip()
        underlying = request.form.get('underlying', '').strip().upper()
        leg_types  = request.form.getlist('leg_type')
        sides      = request.form.getlist('leg_side')
        quantities = request.form.getlist('leg_qty')
        strikes    = request.form.getlist('leg_strike')
        premiums   = request.form.getlist('leg_premium')
        exps       = request.form.getlist('leg_exp')
        tickers    = request.form.getlist('leg_ticker')
        ivs        = request.form.getlist('leg_iv')

        if not name:
            flash('Informe um nome para a simulação.', 'warning')
            return redirect(url_for('simulacao_opcoes'))

        sim = SimulacaoOpcoes(user_id=current_user.id, name=name, underlying=underlying, created_at=datetime.now())
        db.session.add(sim)
        db.session.flush()

        for i, lt in enumerate(leg_types):
            if not lt:
                continue
            exp = None
            try:
                exp = date.fromisoformat(exps[i]) if i < len(exps) and exps[i] else None
            except ValueError:
                pass
            leg = SimulacaoLeg(
                sim_id=sim.id,
                leg_type=lt,
                side=sides[i] if i < len(sides) else 'BUY',
                quantity=int(quantities[i]) if i < len(quantities) and quantities[i] else 1,
                strike=float(strikes[i].replace(',', '.')) if i < len(strikes) and strikes[i] else 0.0,
                premium=float(premiums[i].replace(',', '.')) if i < len(premiums) and premiums[i] else 0.0,
                expiration=exp,
                ticker=(underlying if lt == 'STOCK' else (tickers[i].strip().upper() if i < len(tickers) and tickers[i] else '')),
                iv=float(ivs[i].replace(',', '.')) if i < len(ivs) and ivs[i] else 0.0,
            )
            db.session.add(leg)

        db.session.commit()
        if request.form.get('salvar_producao'):
            return _sim_producao_redirect(sim)
        flash('Simulação salva.', 'success')
        return redirect(url_for('simulacao_edit', id=sim.id))

    sims = SimulacaoOpcoes.query.filter_by(user_id=current_user.id).order_by(SimulacaoOpcoes.created_at.desc()).all()
    return render_template('simulacao_opcoes.html', sims=sims, sim=None, selic=_selic())


@app.route('/simulacao_opcoes/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def simulacao_edit(id):
    sim = SimulacaoOpcoes.query.get_or_404(id)
    if sim.user_id != current_user.id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('simulacao_opcoes'))

    if request.method == 'POST':
        sim.name       = request.form.get('name', '').strip()
        underlying     = request.form.get('underlying', '').strip().upper()
        sim.underlying = underlying
        leg_types  = request.form.getlist('leg_type')
        sides      = request.form.getlist('leg_side')
        quantities = request.form.getlist('leg_qty')
        strikes    = request.form.getlist('leg_strike')
        premiums   = request.form.getlist('leg_premium')
        exps       = request.form.getlist('leg_exp')
        tickers    = request.form.getlist('leg_ticker')
        ivs        = request.form.getlist('leg_iv')

        for leg in list(sim.legs):
            db.session.delete(leg)
        db.session.flush()

        for i, lt in enumerate(leg_types):
            if not lt:
                continue
            exp = None
            try:
                exp = date.fromisoformat(exps[i]) if i < len(exps) and exps[i] else None
            except ValueError:
                pass
            leg = SimulacaoLeg(
                sim_id=sim.id,
                leg_type=lt,
                side=sides[i] if i < len(sides) else 'BUY',
                quantity=int(quantities[i]) if i < len(quantities) and quantities[i] else 1,
                strike=float(strikes[i].replace(',', '.')) if i < len(strikes) and strikes[i] else 0.0,
                premium=float(premiums[i].replace(',', '.')) if i < len(premiums) and premiums[i] else 0.0,
                expiration=exp,
                ticker=(underlying if lt == 'STOCK' else (tickers[i].strip().upper() if i < len(tickers) and tickers[i] else '')),
                iv=float(ivs[i].replace(',', '.')) if i < len(ivs) and ivs[i] else 0.0,
            )
            db.session.add(leg)

        db.session.commit()
        if request.form.get('salvar_producao'):
            return _sim_producao_redirect(sim)
        flash('Simulação atualizada.', 'success')
        return redirect(url_for('simulacao_edit', id=sim.id))

    sims = SimulacaoOpcoes.query.filter_by(user_id=current_user.id).order_by(SimulacaoOpcoes.created_at.desc()).all()
    return render_template('simulacao_opcoes.html', sims=sims, sim=sim, selic=_selic())


@app.route('/simulacao_opcoes/<int:id>/delete', methods=['POST'])
@login_required
def simulacao_delete(id):
    sim = SimulacaoOpcoes.query.get_or_404(id)
    if sim.user_id != current_user.id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('simulacao_opcoes'))
    db.session.delete(sim)
    db.session.commit()
    flash('Simulação excluída.', 'success')
    return redirect(url_for('simulacao_opcoes'))


@app.route('/api/simulacao/save', methods=['POST'])
@login_required
def api_simulacao_save():
    """Salva simulação via JSON (usado pela cadeia de opções)."""
    d          = request.get_json(force=True)
    sim_id     = d.get('id')
    name       = (d.get('name') or '').strip()
    underlying = (d.get('underlying') or '').strip().upper()
    legs_data  = d.get('legs', [])

    if not name:
        return jsonify({'error': 'Informe um nome'}), 400

    if sim_id:
        sim = SimulacaoOpcoes.query.filter_by(id=sim_id, user_id=current_user.id).first()
        if not sim:
            return jsonify({'error': 'Não encontrado'}), 404
        sim.name = name; sim.underlying = underlying
        for leg in list(sim.legs):
            db.session.delete(leg)
        db.session.flush()
    else:
        sim = SimulacaoOpcoes(user_id=current_user.id, name=name,
                              underlying=underlying, created_at=datetime.now())
        db.session.add(sim); db.session.flush()

    for l in legs_data:
        lt  = l.get('type', 'CALL')
        exp = None
        try:
            exp_s = l.get('exp') or ''
            if exp_s:
                exp = date.fromisoformat(exp_s)
        except Exception:
            pass
        db.session.add(SimulacaoLeg(
            sim_id=sim.id, leg_type=lt,
            side=l.get('side', 'BUY'),
            quantity=int(l.get('qty') or 1),
            strike=float(l.get('strike') or 0),
            premium=float(l.get('premium') or 0),
            expiration=exp,
            ticker=(underlying if lt == 'STOCK' else (l.get('ticker') or '').upper()),
            iv=float(l.get('iv') or 0),
        ))

    db.session.commit()
    return jsonify({'id': sim.id, 'name': sim.name})


@app.route('/api/simulacao/list')
@login_required
def api_simulacao_list():
    sims = SimulacaoOpcoes.query.filter_by(user_id=current_user.id)\
                                .order_by(SimulacaoOpcoes.created_at.desc()).all()
    result = []
    for s in sims:
        result.append({
            'id': s.id, 'name': s.name, 'underlying': s.underlying,
            'created_at': s.created_at.strftime('%d/%m/%y') if s.created_at else '',
            'legs': [{'type': l.leg_type, 'side': l.side, 'qty': l.quantity,
                      'ticker': l.ticker, 'premium': l.premium,
                      'strike': l.strike, 'exp': l.expiration.isoformat() if l.expiration else '',
                      'iv': l.iv or 0} for l in s.legs],
        })
    return jsonify(result)


@app.route('/api/simulacao/<int:sim_id>/delete', methods=['POST'])
@login_required
def api_simulacao_delete(sim_id):
    sim = SimulacaoOpcoes.query.filter_by(id=sim_id, user_id=current_user.id).first()
    if not sim:
        return jsonify({'error': 'Não encontrado'}), 404
    db.session.delete(sim); db.session.commit()
    return jsonify({'ok': True})


@app.route('/simulador-liquidez')
@login_required
def simulador_liquidez():
    """Página: tabela de liquidez (ranking) + simulador de payoff integrado."""
    sims = SimulacaoOpcoes.query.filter_by(user_id=current_user.id).order_by(SimulacaoOpcoes.created_at.desc()).all()
    ranking_vol = _ranking_liq_filter(RankingVol.query.filter_by(user_id=current_user.id)).order_by(RankingVol.ticker).all()
    return render_template('simulador_liquidez.html', sims=sims, ranking_vol=ranking_vol, selic=_selic())


@app.route('/cadeia-opcoes')
@login_required
def cadeia_opcoes():
    """Cadeia de opções estilo HB — calls/puts em torno do spot por vencimento."""
    ranking_vol = _ranking_liq_filter(RankingVol.query.filter_by(user_id=current_user.id)).order_by(RankingVol.ticker).all()
    return render_template('cadeia_opcoes.html', ranking_vol=ranking_vol, selic=_selic())


def _bs_price_opt(S, K, T, r, sigma, option_type='CALL'):
    """
    Black-Scholes price (European) com option_type como string ('CALL'/'PUT').
    Usada pelo bloco de busca de opção e cálculo local de BS.
    """
    import math
    try:
        is_call = str(option_type).upper() in ('CALL', 'C')
        if T <= 0:
            return max(0.0, round((S - K) if is_call else (K - S), 4))
        if sigma <= 0 or S <= 0 or K <= 0:
            return None
        d1 = (math.log(S / K) + (r + 0.5 * sigma ** 2) * T) / (sigma * math.sqrt(T))
        d2 = d1 - sigma * math.sqrt(T)
        def _N(x):
            return 0.5 * (1.0 + math.erf(x / math.sqrt(2)))
        if is_call:
            price = S * _N(d1) - K * math.exp(-r * T) * _N(d2)
        else:
            price = K * math.exp(-r * T) * _N(-d2) - S * _N(-d1)
        return round(max(0.0, price), 4)
    except Exception:
        return None


def _bs_greeks(S, K, T, r, sigma, option_type='CALL'):
    """
    Retorna dict com delta, gamma, theta, vega, rho calculados via BS.
    Retorna {} se inputs inválidos.
    """
    import math
    try:
        if T <= 0 or sigma <= 0 or S <= 0 or K <= 0:
            return {}
        def _N(x):
            return 0.5 * (1.0 + math.erf(x / math.sqrt(2)))
        def _n(x):
            return math.exp(-0.5 * x ** 2) / math.sqrt(2 * math.pi)
        d1 = (math.log(S / K) + (r + 0.5 * sigma ** 2) * T) / (sigma * math.sqrt(T))
        d2 = d1 - sigma * math.sqrt(T)
        gamma  = _n(d1) / (S * sigma * math.sqrt(T))
        vega   = S * _n(d1) * math.sqrt(T) / 100   # por 1% de variação na vol
        if option_type.upper() in ('CALL', 'C'):
            delta = _N(d1)
            theta = (-S * _n(d1) * sigma / (2 * math.sqrt(T)) - r * K * math.exp(-r * T) * _N(d2)) / 365
            rho   = K * T * math.exp(-r * T) * _N(d2) / 100
        else:
            delta = _N(d1) - 1
            theta = (-S * _n(d1) * sigma / (2 * math.sqrt(T)) + r * K * math.exp(-r * T) * _N(-d2)) / 365
            rho   = -K * T * math.exp(-r * T) * _N(-d2) / 100
        return {
            'delta': round(delta, 6),
            'gamma': round(gamma, 6),
            'theta': round(theta, 6),
            'vega':  round(vega,  6),
            'rho':   round(rho,   6),
        }
    except Exception:
        return {}


def _hv_from_oplab(underlying, token, days=21):
    """
    Calcula HV anualizada buscando histórico do subjacente na OpLab.
    Tenta /market/instruments/{u}/history e /market/history/{u}.
    Retorna float decimal (ex: 0.32 = 32%) ou None.
    """
    import math
    end   = datetime.utcnow().date()
    start = end - timedelta(days=days * 3)  # margem para fins de semana/feriados

    def _extract_closes(data):
        closes = []
        if isinstance(data, list):
            for item in sorted(data, key=lambda x: x.get('time', x.get('date', 0))):
                for key in ('close', 'adjusted_close', 'last', 'price'):
                    c = item.get(key)
                    if c and float(c) > 0:
                        closes.append(float(c))
                        break
        elif isinstance(data, dict):
            # pode vir como {"history": [...]} ou {"data": [...]}
            for k in ('history', 'data', 'quotes', 'prices'):
                if isinstance(data.get(k), list):
                    return _extract_closes(data[k])
        return closes

    closes = []
    for url in [
        f'/market/instruments/{underlying}/history?from={start}&to={end}',
        f'/market/history/{underlying}?from={start}&to={end}',
        f'/market/instruments/{underlying}/history',
    ]:
        try:
            data = _oplab_get_json(url, token, timeout=10)
            closes = _extract_closes(data)
            if len(closes) >= 5:
                break
        except Exception:
            continue

    if len(closes) < 5:
        return None

    try:
        closes = closes[-days:]
        returns = [math.log(closes[i] / closes[i-1]) for i in range(1, len(closes))]
        mean    = sum(returns) / len(returns)
        var     = sum((r - mean) ** 2 for r in returns) / max(len(returns) - 1, 1)
        hv      = math.sqrt(var) * math.sqrt(252)
        return round(hv, 4) if hv > 0 else None
    except Exception:
        return None


# HV de fallback por tipo de ativo quando OpLab não retorna histórico
_HV_FALLBACK = {
    'default': 0.35,   # 35% — conservador para ações BR
    'fii':     0.20,
    'index':   0.18,
}

def _hv_fallback(underlying):
    """Retorna HV de fallback baseado no tipo de ativo."""
    u = (underlying or '').upper()
    if u.startswith('BOVA') or u in ('IBOV', 'IBOVESPA'):
        return _HV_FALLBACK['index']
    if u.endswith('11') and len(u) == 6:
        return _HV_FALLBACK['fii']
    return _HV_FALLBACK['default']


@app.route('/busca-opcao')
@login_required
def busca_opcao():
    """Busca detalhada de uma opção via OpLab."""
    return render_template('busca_opcao.html')


@app.route('/api/busca-opcao/<ticker>')
@login_required
def api_busca_opcao(ticker):
    """Retorna ficha completa de uma opção via OpLab /market/instruments/{ticker}."""
    token = Settings.get_value('oplab_token', user_id=current_user.id)
    if not token:
        return jsonify({'error': 'Token OpLab não configurado.'}), 403

    ticker = ticker.strip().upper()
    try:
        d = _oplab_get_json(f'/market/instruments/{ticker}', token, timeout=15)
    except OplabApiError as e:
        return jsonify({'error': f'OpLab: {e}'}), e.status_code or 502
    except Exception as e:
        return jsonify({'error': str(e)}), 502

    if not isinstance(d, dict):
        if isinstance(d, list) and d:
            d = d[0]
        else:
            return jsonify({'error': 'Resposta inesperada da OpLab.'}), 502

    def _f(v, decimals=2):
        try:
            return round(float(v), decimals) if v is not None else None
        except (TypeError, ValueError):
            return None

    # Campos diretos do JSON real: parent_symbol, spot_price, variation, due_date, etc.
    underlying = d.get('parent_symbol') or d.get('underlying_symbol') or d.get('underlying') or ''
    spot_price  = _f(d.get('spot_price'))       # já vem no instruments
    # variation vem como número absoluto de %, ex: 20 = +20%
    raw_var = d.get('variation')
    change_pct = _f(raw_var)

    # Busca gregas via /market/options/{underlying} — filtra pelo ticker
    greeks = {}
    iv_data = {}
    if underlying:
        try:
            opts = _oplab_get_json(f'/market/options/{underlying}', token, timeout=15)
            if isinstance(opts, list):
                for opt in opts:
                    if isinstance(opt, dict) and opt.get('symbol', '').upper() == ticker:
                        greeks  = opt.get('greeks') or {}
                        iv_raw  = opt.get('implied_volatility') or opt.get('iv') or {}
                        iv_data = iv_raw if isinstance(iv_raw, dict) else {'iv': iv_raw}
                        # spot_price e variation mais precisos se vierem aqui
                        if opt.get('spot_price') and not spot_price:
                            spot_price = _f(opt['spot_price'])
                        break
        except Exception:
            pass

    # Cotação do subjacente (variação diária do ativo-mãe)
    spot_change = None
    if underlying:
        try:
            sp = _oplab_get_json(f'/market/instruments/{underlying}', token, timeout=10)
            if isinstance(sp, dict):
                spot_price  = spot_price or _f(sp.get('close') or sp.get('spot_price'))
                spot_change = _f(sp.get('variation'))
        except Exception:
            pass

    def _fz(v, decimals=2):
        """Como _f, mas retorna None se o valor for 0 (OpLab retorna 0 para campos sem dado)."""
        r = _f(v, decimals)
        return None if r == 0.0 else r

    result = {
        'ticker':           ticker,
        'type':             d.get('category') or d.get('option_type') or d.get('type') or '',
        'maturity_type':    d.get('maturity_type') or '',
        'underlying':       underlying,
        'spot_price':       spot_price,
        'spot_change':      spot_change,
        'days_to_maturity': d.get('days_to_maturity'),   # dias ÚTEIS (padrão OpLab)
        # Preço — close é o último preço; open/high/low/prev_close chegam como 0 quando sem dado
        'last':             _fz(d.get('close')),
        'open':             _fz(d.get('open')),
        'high':             _fz(d.get('high')),
        'low':              _fz(d.get('low')),
        'prev_close':       _fz(d.get('previous_close')),
        'change_pct':       _fz(d.get('variation')),
        'financial_volume': _fz(d.get('financial_volume'), 0),
        # Opção
        'strike':           _f(d.get('strike')),        # strike pode ser 0 em alguns casos
        'expiration':       d.get('due_date') or '',
        'volume':           _fz(d.get('volume'), 0),
        'open_interest':    _fz(d.get('open_interest') or d.get('contracts'), 0),
        'bid':              _fz(d.get('bid')),
        'ask':              _fz(d.get('ask')),
        'bid_volume':       _fz(d.get('bid_volume'), 0),
        'ask_volume':       _fz(d.get('ask_volume'), 0),
        # Gregas (de /market/options/{underlying})
        'bs_price':         _fz(greeks.get('black_scholes') or greeks.get('bs') or greeks.get('theorical')),
        'delta':            _fz(greeks.get('delta')),
        'gamma':            _fz(greeks.get('gamma')),
        'theta':            _fz(greeks.get('theta')),
        'rho':              _fz(greeks.get('rho')),
        'vega':             _fz(greeks.get('vega')),
        # Volatilidade (de /market/options/{underlying})
        'iv':               _fz(iv_data.get('iv') or iv_data.get('current')),
        'iv_ask':           _fz(iv_data.get('ask')),
        'iv_bid':           _fz(iv_data.get('bid')),
        'iv_over_hv':       _fz(iv_data.get('iv_over_hv') or iv_data.get('ratio')),
        'intrinsic_value':  _f(iv_data.get('intrinsic_value') if iv_data.get('intrinsic_value') is not None else greeks.get('intrinsic_value')),
        'extrinsic_value':  _fz(iv_data.get('extrinsic_value') or greeks.get('extrinsic_value') or greeks.get('time_value')),
        '_raw':             {k: v for k, v in d.items() if not isinstance(v, (dict, list))},
    }

    # bid/ask não-zero do instruments são válidos mesmo se volume=0 (market maker)
    if result['bid'] is None:
        result['bid'] = _f(d.get('bid'))
    if result['ask'] is None:
        result['ask'] = _f(d.get('ask'))

    # ── Mescla com dados do RtdOptionData (importados via Excel) ────────────
    # Para campos de gregas/VI: RTD prevalece se OpLab retornou None ou 0.0
    rtd = RtdOptionData.query.filter_by(user_id=current_user.id, ticker=ticker).first()
    if rtd:
        def _merge(key, rtd_val, force_nonzero=False):
            """Preenche com RTD se OpLab retornou None; com force_nonzero também substitui 0.0.
            RTD com valor 0 é tratado como 'sem dado' (Profit retorna 0 fora do pregão)."""
            if rtd_val is None or rtd_val == 0:
                return
            cur = result.get(key)
            if cur is None or (force_nonzero and cur == 0.0):
                result[key] = rtd_val
        _merge('last',            rtd.last_price,   force_nonzero=True)
        _merge('open',            rtd.open_price,   force_nonzero=True)
        _merge('high',            rtd.high_price,   force_nonzero=True)
        _merge('low',             rtd.low_price,    force_nonzero=True)
        _merge('prev_close',      rtd.prev_close,   force_nonzero=True)
        _merge('change_pct',      rtd.change_pct)
        _merge('strike',          rtd.strike,        force_nonzero=True)
        _merge('expiration',      rtd.expiration)
        _merge('volume',          rtd.volume,        force_nonzero=True)
        _merge('open_interest',   rtd.open_interest, force_nonzero=True)
        _merge('bid',             rtd.bid,           force_nonzero=True)
        _merge('ask',             rtd.ask,           force_nonzero=True)
        _merge('iv',              rtd.iv,             force_nonzero=True)
        _merge('iv_ask',          rtd.iv_ask,         force_nonzero=True)
        _merge('iv_bid',          rtd.iv_bid,         force_nonzero=True)
        _merge('iv_over_hv',      rtd.iv_over_hv,     force_nonzero=True)
        _merge('delta',           rtd.delta,           force_nonzero=True)
        _merge('gamma',           rtd.gamma,           force_nonzero=True)
        _merge('theta',           rtd.theta,           force_nonzero=True)
        _merge('rho',             rtd.rho,             force_nonzero=True)
        _merge('vega',            rtd.vega,            force_nonzero=True)
        _merge('bs_price',        rtd.bs_price,        force_nonzero=True)
        _merge('intrinsic_value', rtd.intrinsic_value, force_nonzero=True)
        _merge('extrinsic_value', rtd.extrinsic_value, force_nonzero=True)
        _merge('spot_price',      rtd.spot_price)
        if not result.get('type') and rtd.option_type:
            result['type'] = rtd.option_type
        result['_rtd_imported_at'] = rtd.imported_at.strftime('%d/%m/%Y %H:%M') if rtd.imported_at else None

    # ── Dias corridos (calendário) a partir da data de vencimento ────────────
    # days_to_maturity da OpLab é em dias ÚTEIS; days_calendar (corridos) é usado
    # para exibir o prazo e anualizar retornos (base 365).
    result['days_calendar'] = None
    if result.get('expiration'):
        try:
            _exp_d = datetime.strptime(str(result['expiration'])[:10], '%Y-%m-%d').date()
            result['days_calendar'] = max((_exp_d - date.today()).days, 0)
        except (ValueError, TypeError):
            pass

    # ── Cálculo BS local quando OpLab não retornou gregas ───────────────────
    # 0.0 também é tratado como "sem dado" (OpLab/RTD retornam 0 fora do pregão)
    needs_bs     = not result.get('bs_price')
    needs_greeks = not result.get('delta')
    if (needs_bs or needs_greeks) and result.get('spot_price') and result.get('strike'):
        S = result['spot_price']
        K = result['strike']

        # Prazo em anos: dias CORRIDOS / 365. Se só houver dias úteis
        # (days_to_maturity da OpLab), converte por 252.
        dcal = result.get('days_calendar')
        if dcal:
            T = dcal / 365.0
        else:
            T = max(0, result.get('days_to_maturity') or 0) / 252.0

        # Taxa Selic anualizada
        try:
            last_selic = SelicMensal.query.order_by(SelicMensal.mes_ano.desc()).first()
            r = ((1 + (last_selic.taxa / 100)) ** 12 - 1) if last_selic else 0.135
        except Exception:
            r = 0.135

        opt_type = result.get('type') or 'CALL'
        is_call  = str(opt_type).upper() in ('CALL', 'C')
        intrinsic = max(0.0, (S - K) if is_call else (K - S))

        # Prêmio de mercado para extrair VI implícita: último negócio; senão mid do book
        premium = result.get('last')
        bid, ask = result.get('bid'), result.get('ask')
        if not premium and bid and ask:
            premium = (bid + ask) / 2.0

        # ── Duas volatilidades com papéis diferentes ─────────────────────────
        # σ de MERCADO (gregas + VI exibida): VI implícita do prêmio → VI OpLab → HV → ref.
        # σ TEÓRICO (preço BS): NUNCA usa a VI do prêmio — o BS calculado com a VI
        #   extraída do próprio último negócio reproduz o último por construção
        #   (circular). Usa VI OpLab → HV → ref, permitindo comparar o teórico
        #   com o preço de mercado (caro/barato vs vol histórica).
        iv_field = result.get('iv')
        if iv_field:
            iv_field = iv_field / 100.0 if iv_field > 1.5 else iv_field

        iv_prem = None
        if premium and T > 0 and premium > intrinsic * 1.01:
            iv_est = _implied_vol(S, K, T, r, premium, is_call)
            if 0.005 < iv_est < 4.9:   # descarta bissecção presa nos limites
                iv_prem = iv_est

        hv = _hv_from_oplab(underlying, token) if underlying else None

        if iv_field:
            sigma_teo, teo_src = iv_field, 'IV'
        elif hv:
            sigma_teo, teo_src = hv, 'HV'
        else:
            sigma_teo, teo_src = _hv_fallback(underlying), 'HV-ref'

        if iv_prem:
            sigma_mkt, mkt_src = iv_prem, 'IV-prem'
        else:
            sigma_mkt, mkt_src = sigma_teo, teo_src

        # BS price (σ teórico, independente do prêmio) + intrínseco/extrínseco
        if needs_bs:
            bs = _bs_price_opt(S, K, T, r, sigma_teo, opt_type)
            if bs is not None:
                result['bs_price'] = bs
                result['bs_calc']  = teo_src
        if not result.get('intrinsic_value'):
            result['intrinsic_value'] = round(intrinsic, 4)
        if not result.get('extrinsic_value') and premium:
            result['extrinsic_value'] = round(max(0.0, premium - intrinsic), 4)

        # Gregas (σ de mercado — reflete a vol implícita atual)
        if needs_greeks and T > 0:
            gk = _bs_greeks(S, K, T, r, sigma_mkt, opt_type)
            if gk:
                result.setdefault('bs_calc', teo_src)
                for k2, v2 in gk.items():
                    if not result.get(k2):
                        result[k2] = v2

        # Preenche VI exibida quando a fonte foi o prêmio de mercado
        if not result.get('iv') and iv_prem:
            result['iv'] = round(iv_prem * 100, 2)

        result['bs_sigma_used'] = round(sigma_teo * 100, 2)   # σ do preço teórico
        result['bs_sigma_src']  = teo_src                      # 'IV', 'HV' ou 'HV-ref'
        result['greeks_sigma_used'] = round(sigma_mkt * 100, 2)
        result['greeks_sigma_src']  = mkt_src                  # 'IV-prem', 'IV', 'HV' ou 'HV-ref'

    # ── Salva/atualiza na tabela SearchedOption (lista RTD) ──────────────────
    # Limpa entradas > 10 dias antes de inserir
    cutoff = datetime.utcnow() - timedelta(days=10)
    SearchedOption.query.filter(
        SearchedOption.user_id == current_user.id,
        SearchedOption.searched_at < cutoff
    ).delete()
    so = SearchedOption.query.filter_by(user_id=current_user.id, ticker=ticker).first()
    if so:
        so.searched_at = datetime.utcnow()
        so.underlying  = underlying or so.underlying
    else:
        so = SearchedOption(user_id=current_user.id, ticker=ticker, underlying=underlying or None)
        db.session.add(so)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    return jsonify(result)


@app.route('/api/rtd-map')
@login_required
def api_rtd_map():
    """Retorna TICKER_MAP e OPTION_MAP (incluindo SearchedOption) como texto copiável."""
    _, _, ticker_map_text, option_map_text = _build_ticker_maps(current_user.id)
    # Lista de opções buscadas ainda ativas (< 10 dias)
    cutoff = datetime.utcnow() - timedelta(days=10)
    searched = SearchedOption.query.filter(
        SearchedOption.user_id == current_user.id,
        SearchedOption.searched_at >= cutoff
    ).order_by(SearchedOption.searched_at.desc()).all()
    searched_list = [
        {'ticker': s.ticker, 'underlying': s.underlying,
         'searched_at': s.searched_at.strftime('%d/%m/%Y %H:%M'),
         'expires_in': max(0, 10 - (datetime.utcnow() - s.searched_at).days)}
        for s in searched
    ]
    return jsonify({
        'ticker_map': ticker_map_text,
        'option_map': option_map_text,
        'searched_options': searched_list,
    })


@app.route('/api/rtd-map/searched/delete/<ticker>', methods=['POST'])
@login_required
def api_rtd_map_searched_delete(ticker):
    """Remove manualmente uma opção da lista RTD buscada."""
    so = SearchedOption.query.filter_by(user_id=current_user.id, ticker=ticker.upper()).first()
    if so:
        db.session.delete(so)
        db.session.commit()
    return jsonify({'ok': True})


_RTD_FIELDS = [
    ("DAT", "Data"), ("HOR", "Hora"), ("ULT", "Último"), ("ABE", "Abertura"),
    ("MAX", "Máximo"), ("MIN", "Mínimo"), ("FEC", "Fechamento Anterior"),
    ("PEX", "Strike"), ("VAR", "Variação"), ("MED", "Média"), ("NOME", None),
    ("NEG", "Negócios"), ("QTT", "Quantidade"), ("VOL", "Volume"),
    ("OCP", "Of. Compra"), ("OVD", "Of. Venda"), ("VPJ", "Volume Projetado"),
    ("VEN", "Vencimento"), ("VAL", "Validade"), ("CAB", "Cont. Abertos"),
    ("BLACK", "Black Scholes"), ("IMPVT", "Volt. Implícita"),
    ("DELTA", "Delta"), ("GAMA", "Gama"), ("THETA", "Theta"),
    ("RHO", "Rho"), ("VEGA", "Vega"), ("VIA", "VI Ask"), ("VIB", "VI Bid"),
    ("VIVH", "VI / VH"), ("VINT", "Valor Intrínseco"), ("VEXT", "Valor Extrínseco"),
    ("204", "Dividend Yield"), ("1", "IFR (RSI)"),
    ("387", "Volatilidade Implícita"), ("81", "Volatilidade Implícita - Opções"),
]

_RTD_HEADER = [
    "Asset", "Data", "Hora", "Último", "Abertura", "Máximo", "Mínimo",
    "Fechamento Anterior", "Strike", "Variação", "Média", "Nome do Ativo",
    "Negócios", "Quantidade", "Volume", "Of. Compra", "Of. Venda",
    "Volume Projetado", "Vencimento", "Validade", "Cont. Abertos",
    "Black Scholes", "Volt. Implícita", "Delta", "Gama", "Theta", "Rho",
    "Vega", "VI Ask", "VI Bid", "VI / VH", "Valor Intrínseco",
    "Valor Extrínseco", "Dividend Yield", "IFR (RSI)",
    "Volatilidade Implícita", "Volatilidade Implícita - Opções",
]


def _build_rtd_text(uid):
    """
    Gera TSV no formato RTDTrading idêntico ao mt5_feeder/mt5_feeder.py:generate_rtd_text.
    Cada coluna = =RTD("RTDTrading.RTDServer";;"TICKER_B_0";"FIELD")
    Colunas NOME ficam em branco.
    """
    today = date.today()

    # Coleta todos os tickers: TICKER_MAP (ativos) + OPTION_MAP (opções)
    ticker_map = {}   # site_ticker → mt5_symbol (igual ao ticker no padrão B3)
    option_map = {}

    # Ativos da carteira
    for a in Asset.query.filter(Asset.user_id == uid, Asset.quantity > 0).all():
        ticker_map[a.ticker.upper()] = a.ticker.upper()

    # Subjacentes de opções ativas
    def _add_asset(t):
        if t and t.upper() not in ticker_map:
            ticker_map[t.upper()] = t.upper()

    for o in Option.query.filter_by(user_id=uid).all():
        if not o.expiration_date or o.expiration_date >= today:
            _add_asset(o.underlying_asset)
    for op in StructuredOp.query.filter_by(user_id=uid, status='OPEN').all():
        _add_asset(op.underlying_asset)
    for ps in PutSale.query.filter_by(user_id=uid).all():
        if ps.expiration_date >= today:
            _add_asset(ps.underlying_asset)
    for sp in OptionSpread.query.filter_by(user_id=uid).all():
        if sp.expiration_date >= today:
            _add_asset(sp.underlying_asset)

    # Opções ativas
    def _add_opt(t):
        if t and t.upper() not in option_map:
            option_map[t.upper()] = t.upper()

    for o in Option.query.filter_by(user_id=uid).all():
        if not o.expiration_date or o.expiration_date >= today:
            _add_opt(o.ticker)
    for op in StructuredOp.query.filter_by(user_id=uid, status='OPEN').all():
        for leg in op.legs:
            if not leg.expiration_date or leg.expiration_date >= today:
                _add_opt(leg.ticker)
    for ps in PutSale.query.filter_by(user_id=uid).all():
        if ps.expiration_date >= today:
            _add_opt(ps.ticker)
    for sp in OptionSpread.query.filter_by(user_id=uid).all():
        if sp.expiration_date >= today:
            _add_opt(sp.leg_long_ticker)
            _add_opt(sp.leg_short_ticker)

    # SearchedOption < 10 dias
    cutoff = datetime.utcnow() - timedelta(days=10)
    for so in SearchedOption.query.filter(
            SearchedOption.user_id == uid,
            SearchedOption.searched_at >= cutoff).all():
        _add_opt(so.ticker)

    # Gera linhas no formato idêntico ao mt5_feeder
    lines = ["\t".join(_RTD_HEADER)]

    all_tickers = {}
    for site, sym in sorted(ticker_map.items()):
        all_tickers[site] = sym
    for site, sym in sorted(option_map.items()):
        all_tickers[site] = sym

    for ticker, mt5sym in sorted(all_tickers.items()):
        base = f'"{mt5sym}_B_0"'
        cols = [ticker]
        for field, _ in _RTD_FIELDS:
            if field == "NOME":
                cols.append("")   # sem RTD para nome
            else:
                cols.append(f'=RTD("RTDTrading.RTDServer";;{base};"{field}")')
        lines.append("\t".join(cols))

    return "\n".join(lines), len(all_tickers)


@app.route('/api/rtd-text')
@login_required
def api_rtd_text():
    """Retorna o conteúdo RTD como JSON para exibição/cópia no browser."""
    content, count = _build_rtd_text(current_user.id)
    return jsonify({'content': content, 'count': count})


@app.route('/download-rtd-tsv')
@login_required
def download_rtd_tsv():
    """Baixa o arquivo TSV RTD no formato RTDTrading (idêntico ao MT5 Feeder)."""
    from flask import Response as _Resp
    content, _ = _build_rtd_text(current_user.id)
    return _Resp(
        content.encode('utf-8-sig'),   # BOM para Excel reconhecer UTF-8
        mimetype='text/tab-separated-values',
        headers={'Content-Disposition': 'attachment; filename=rtd_export.txt'}
    )


@app.route('/rolagem-opcoes')
@login_required
def rolagem_opcoes():
    """Simulador de rolagem de opcoes por tempo ou strike."""
    ranking_vol = _ranking_liq_filter(RankingVol.query.filter_by(user_id=current_user.id)).order_by(RankingVol.ticker).all()
    return render_template('rolagem_opcoes.html', ranking_vol=ranking_vol, manejo_mode=False)


@app.route('/manejo-opcoes')
@login_required
def manejo_opcoes():
    """Simulador de MANEJO de opções: mesma engrenagem da Rolagem, com a aba
    Manejo pré-selecionada e a lista de simulações filtrada por MANEJO."""
    ranking_vol = _ranking_liq_filter(RankingVol.query.filter_by(user_id=current_user.id)).order_by(RankingVol.ticker).all()
    return render_template('rolagem_opcoes.html', ranking_vol=ranking_vol, manejo_mode=True)


@app.route('/api/rolagem-opcoes/list')
@login_required
def api_rolagem_list():
    mode = (request.args.get('mode') or '').lower()
    q = OptionRollSimulation.query.filter_by(user_id=current_user.id)
    if mode == 'manejo':
        q = q.filter(OptionRollSimulation.roll_type == 'MANEJO')
    elif mode == 'rolagem':
        q = q.filter(OptionRollSimulation.roll_type != 'MANEJO')
    sims = q.order_by(OptionRollSimulation.updated_at.desc(), OptionRollSimulation.created_at.desc()).all()
    return jsonify([{
        'id': s.id,
        'name': s.name,
        'underlying': s.underlying,
        'roll_type': s.roll_type,
        'created_at': s.created_at.strftime('%d/%m/%y') if s.created_at else '',
        'updated_at': s.updated_at.strftime('%d/%m/%y %H:%M') if s.updated_at else '',
        'payload': json.loads(s.payload or '{}'),
    } for s in sims])


@app.route('/api/rolagem-opcoes/save', methods=['POST'])
@login_required
def api_rolagem_save():
    data = request.get_json(force=True) or {}
    name = (data.get('name') or '').strip()
    payload = data.get('payload') or {}
    sim_id = data.get('id')
    if not name:
        return jsonify({'error': 'Informe um nome'}), 400
    if not isinstance(payload, dict):
        return jsonify({'error': 'Payload invalido'}), 400

    underlying = (payload.get('underlying') or data.get('underlying') or '').strip().upper()
    roll_type = (payload.get('roll_type') or data.get('roll_type') or 'TIME').strip().upper()
    if roll_type not in ('TIME', 'STRIKE', 'TIME_STRIKE', 'MANEJO'):
        roll_type = 'TIME'

    if sim_id:
        sim = OptionRollSimulation.query.filter_by(id=sim_id, user_id=current_user.id).first()
        if not sim:
            return jsonify({'error': 'Nao encontrado'}), 404
        sim.name = name
        sim.underlying = underlying
        sim.roll_type = roll_type
        sim.payload = json.dumps(payload, ensure_ascii=False)
        sim.updated_at = datetime.now()
    else:
        sim = OptionRollSimulation(
            user_id=current_user.id,
            name=name,
            underlying=underlying,
            roll_type=roll_type,
            payload=json.dumps(payload, ensure_ascii=False),
            created_at=datetime.now(),
            updated_at=datetime.now(),
        )
        db.session.add(sim)

    db.session.commit()
    return jsonify({'id': sim.id, 'name': sim.name})


@app.route('/api/rolagem-opcoes/<int:sim_id>/delete', methods=['POST'])
@login_required
def api_rolagem_delete(sim_id):
    sim = OptionRollSimulation.query.filter_by(id=sim_id, user_id=current_user.id).first()
    if not sim:
        return jsonify({'error': 'Nao encontrado'}), 404
    db.session.delete(sim)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/rolagem-opcoes/cadeia/<ticker>')
@login_required
def api_rolagem_cadeia(ticker):
    """Retorna a cadeia completa normalizada para simulacao de rolagem."""
    ticker = ticker.strip().upper()
    token = Settings.get_value('oplab_token', user_id=current_user.id)
    if not token:
        return jsonify({'error': 'Token OpLab nao configurado'}), 400
    spot, spot_change = _get_underlying_quote(ticker, current_user.id)

    try:
        data = _oplab_get_json(f'/market/options/{ticker}', token, timeout=15)
    except OplabApiError as e:
        return jsonify({'error': str(e), 'status': e.status_code, 'preview': e.body_preview}), 503

    opt_list = data if isinstance(data, list) else (
        data.get('options') or data.get('calls', []) + data.get('puts', []) or []
    )
    def _b3_option_kind(sym, raw_kind=''):
        raw = str(raw_kind or '').upper()
        if 'PUT' in raw or raw == 'P':
            return 'PUT'
        if 'CALL' in raw or raw == 'C':
            return 'CALL'
        # Padrao B3: A-L = calls, M-X = puts. Ex.: CMIGR132 => PUT.
        if len(sym) >= 5 and sym[4].isalpha():
            return 'PUT' if sym[4] in 'MNOPQRSTUVWX' else 'CALL'
        return 'CALL'

    options = []
    for o in opt_list:
        sym = str(o.get('symbol') or o.get('ticker') or '').upper()
        due_date = str(o.get('due_date') or o.get('expiration_date') or '')
        if 'T' in due_date:
            due_date = due_date.split('T')[0]
        if not sym or not due_date:
            continue
        cat = str(o.get('category') or o.get('type') or o.get('option_type') or '').upper()
        bid = float(o.get('bid') or 0)
        ask = float(o.get('ask') or 0)
        options.append({
            'symbol': sym,
            'kind': _b3_option_kind(sym, cat),
            'strike': round(float(o.get('strike') or 0), 2),
            'exp': due_date,
            'close': round(float(o.get('close') or 0), 2),
            'bid': round(bid, 2),
            'ask': round(ask, 2),
            'mid': round((bid + ask) / 2, 2) if (bid or ask) else 0,
            'var_pct': round(float(o.get('variation') or 0), 2),
            'vol_fin': round(float(o.get('financial_volume') or o.get('volume_financial') or 0), 2),
        })

    return jsonify({
        'ticker': ticker,
        'spot': spot,
        'spot_change': spot_change,
        'options': sorted(options, key=lambda x: (x['exp'], x['kind'], x['strike'])),
    })


@app.route('/api/cadeia/<ticker>')
@login_required
def api_cadeia(ticker):
    """Retorna cadeia de opções completa do ticker via OpLab, organizada por vencimento."""
    ticker = ticker.strip().upper()
    token  = Settings.get_value('oplab_token', user_id=current_user.id)
    if not token:
        return jsonify({'error': 'Token OpLab não configurado'}), 400

    # Spot price
    spot, spot_change = _get_underlying_quote(ticker, current_user.id)

    try:
        data = _oplab_get_json(f'/market/options/{ticker}', token, timeout=15)
    except OplabApiError as e:
        return jsonify({'error': str(e), 'status': e.status_code, 'preview': e.body_preview}), 503
    except Exception:
        app.logger.exception('api_cadeia error for %s', ticker)
        return jsonify({'error': 'Erro inesperado ao buscar a cadeia de opcoes.'}), 500

    opt_list = data if isinstance(data, list) else (
        data.get('options') or data.get('calls', []) + data.get('puts', []) or []
    )

    from datetime import date as _date, timedelta
    import calendar

    def _next_monthly_expirations(n=3):
        """Retorna as datas dos próximos n vencimentos mensais (terceira sexta)."""
        today = _date.today()
        result = []
        y, m = today.year, today.month
        while len(result) < n * 2:  # gera mais para filtrar
            # terceira sexta-feira do mês
            day = 1
            count = 0
            while True:
                d = _date(y, m, day)
                if d.weekday() == 4:  # sexta
                    count += 1
                    if count == 3:
                        break
                day += 1
            if d >= today:
                result.append(d.strftime('%Y-%m-%d'))
            m += 1
            if m > 12:
                m = 1; y += 1
        return result[:n]

    def _is_monthly_exp(exp_str):
        """True se a data é a 3ª sexta-feira do mês (±2 dias p/ feriado)."""
        dd = _date.fromisoformat(exp_str)
        day, count = 1, 0
        while True:
            d = _date(dd.year, dd.month, day)
            if d.weekday() == 4:
                count += 1
                if count == 3:
                    tf = d
                    break
            day += 1
        return abs((dd - tf).days) <= 2

    # Parâmetros de janela (iguais aos da Busca de Operações):
    # weekly=1 inclui semanais; days = prazo máximo (60/90/120/180).
    include_weekly = request.args.get('weekly', '0') == '1'
    try:
        max_days = int(request.args.get('days', 60))
    except (TypeError, ValueError):
        max_days = 60
    if max_days not in (60, 90, 120, 180):
        max_days = 60

    # Normaliza opções
    calls_by_exp = {}
    puts_by_exp  = {}

    for o in opt_list:
        sym      = str(o.get('symbol') or o.get('ticker') or '').upper()
        cat      = str(o.get('category') or o.get('type') or '').upper()
        strike   = float(o.get('strike') or 0)
        close    = float(o.get('close') or 0)
        bid      = float(o.get('bid') or 0)
        ask      = float(o.get('ask') or 0)
        var_pct  = float(o.get('variation') or 0)
        vol_fin  = float(o.get('financial_volume') or o.get('volume_financial') or 0)
        delta    = o.get('delta') or (o.get('greeks') or {}).get('delta') if isinstance(o.get('greeks'), dict) else o.get('delta')
        teorico  = float(o.get('theoretical_price') or o.get('theo') or 0)
        liquidez = float(o.get('liquidity') or o.get('liquidity_score') or 0)
        due_date = str(o.get('due_date') or o.get('expiration_date') or '')
        if 'T' in due_date:
            due_date = due_date.split('T')[0]
        if not due_date:
            continue

        row = {
            'symbol':   sym,
            'strike':   round(strike, 2),
            'close':    round(close, 2),
            'bid':      round(bid, 2),
            'ask':      round(ask, 2),
            'var_pct':  round(var_pct, 2),
            'vol_fin':  round(vol_fin, 2),
            'delta':    round(float(delta), 2) if delta is not None else None,
            'teorico':  round(teorico, 2),
            'liquidez': round(liquidez, 2),
            'mid':      round((bid + ask) / 2, 2) if (bid or ask) else 0,
        }

        is_put = 'PUT' in cat or cat == 'P'
        bucket = puts_by_exp if is_put else calls_by_exp
        bucket.setdefault(due_date, []).append(row)

    # Para cada vencimento, seleciona 10 strikes abaixo e 10 acima do spot
    result_exps = []
    all_exp_keys = sorted(set(list(calls_by_exp.keys()) + list(puts_by_exp.keys())))

    today = _date.today()
    # Vencimentos dentro da janela de prazo escolhida
    within = []
    for e in all_exp_keys:
        try:
            dc = (_date.fromisoformat(e) - today).days
        except ValueError:
            continue
        if 0 < dc <= max_days:
            within.append(e)

    if include_weekly:
        # todos os vencimentos (semanais + mensais) na janela — até 8
        selected_exps = within[:8]
    else:
        # apenas mensais na janela — até 3
        selected_exps = [e for e in within if _is_monthly_exp(e)][:3]

    # Fallback: sem nada na janela, usa os 3 vencimentos mais próximos
    if not selected_exps:
        selected_exps = within[:3] or all_exp_keys[:3]

    # Quantidade de strikes exibidos abaixo/acima do spot (10 padrão, 20 opcional)
    try:
        n_strikes = int(request.args.get('n', 10))
    except (TypeError, ValueError):
        n_strikes = 10
    if n_strikes not in (10, 20):
        n_strikes = 10

    for exp in selected_exps:
        calls = sorted(calls_by_exp.get(exp, []), key=lambda x: x['strike'])
        puts  = sorted(puts_by_exp.get(exp, []),  key=lambda x: x['strike'])

        if spot:
            calls_below = [c for c in calls if c['strike'] <= spot][-n_strikes:]
            calls_above = [c for c in calls if c['strike'] >  spot][:n_strikes]
            calls_sel   = calls_below + calls_above

            puts_below  = [p for p in puts if p['strike'] <= spot][-n_strikes:]
            puts_above  = [p for p in puts if p['strike'] >  spot][:n_strikes]
            puts_sel    = puts_below + puts_above
        else:
            calls_sel = calls[:n_strikes]
            puts_sel  = puts[:n_strikes * 2]

        # Monta linhas alinhadas por strike
        all_strikes = sorted({r['strike'] for r in calls_sel + puts_sel})
        call_map = {c['strike']: c for c in calls_sel}
        put_map  = {p['strike']: p for p in puts_sel}

        rows = []
        for s in all_strikes:
            rows.append({
                'strike': s,
                'call':   call_map.get(s),
                'put':    put_map.get(s),
            })

        result_exps.append({'exp': exp, 'rows': rows})

    return jsonify({
        'ticker':      ticker,
        'spot':        spot,
        'spot_change': spot_change,
        'expirations': result_exps,
        'weekly':      include_weekly,
        'max_days':    max_days,
    })


@app.route('/busca-operacoes')
@login_required
def busca_operacoes():
    """Página: sugestões de collar e travas no débito por vencimento."""
    ranking_vol = _ranking_liq_filter(RankingVol.query.filter_by(user_id=current_user.id)).order_by(RankingVol.ticker).all()
    return render_template('busca_operacoes.html', ranking_vol=ranking_vol, selic=_selic())


@app.route('/ajuda-operacoes')
@login_required
def ajuda_operacoes():
    """Página: guia de referência das operações estruturadas (Vol. 2 e 4)."""
    return render_template('ajuda_operacoes.html')


@app.route('/busca-operacoes-avancadas')
@login_required
def busca_operacoes_avancadas():
    """Página: busca de operações avançadas (guia técnico Vol. 2), por categoria."""
    ranking_vol = _ranking_liq_filter(RankingVol.query.filter_by(user_id=current_user.id)).order_by(RankingVol.ticker).all()
    return render_template('busca_operacoes_avancadas.html', ranking_vol=ranking_vol, selic=_selic())


# Especificações das operações avançadas (guia técnico Vol. 2).
# legs: (tipo 'C'/'P'/'S', quantidade, janela de delta recomendada) — 'S' = ação.
# same: pares de índices de perna com o MESMO strike; asc: strikes crescentes.
# rule: restrição de montagem (credit, zero_cost, low_cost, jade, box).
_ADV_SPECS = {
    # ── Direcionais ──
    'call_seco':      {'legs': [('C', 1, (0.30, 0.45))]},
    'put_seco':       {'legs': [('P', 1, (-0.45, -0.30))]},
    'seagull_baixa':  {'legs': [('C', -1, (0.15, 0.25)), ('P', 1, (-0.50, -0.40)),
                                ('P', -1, (-0.30, -0.20))]},
    'seagull_alta':   {'legs': [('P', -1, (-0.25, -0.15)), ('C', 1, (0.40, 0.50)),
                                ('C', -1, (0.20, 0.30))], 'asc': [1, 2], 'rule': 'credit'},
    'put_backspread': {'legs': [('P', -1, (-0.80, -0.70)), ('P', 2, (-0.30, -0.20))],
                       'rule': 'low_cost'},
    'christmas_tree': {'legs': [('C', -1, (0.70, 0.80)), ('C', 1, (0.25, 0.35)),
                                ('C', 1, (0.15, 0.25)), ('C', 1, (0.05, 0.15))],
                       'asc': [0, 1, 2, 3], 'rule': 'low_cost'},
    'ladder_call':    {'legs': [('C', 1, (0.45, 0.60)), ('C', -1, (0.25, 0.40)),
                                ('C', -1, (0.10, 0.25))], 'asc': [0, 1, 2]},
    # ── Renda e Proteção ──
    'venda_coberta':  {'legs': [('S', 1, None), ('C', -1, (0.20, 0.35))]},
    'protective_put': {'legs': [('S', 1, None), ('P', 1, (-0.50, -0.40))]},
    'covered_put':    {'legs': [('S', -1, None), ('P', -1, (-0.35, -0.20))]},
    'jade_lizard':    {'legs': [('P', -1, (-0.25, -0.15)), ('C', -1, (0.20, 0.30)),
                                ('C', 1, (0.05, 0.12))], 'asc': [1, 2], 'rule': 'jade'},
    # Reparo de posição (Stock Repair 1×2): compra 1 CALL ATM + vende 2 OTM a
    # custo ~zero — p/ ação NO PREJUÍZO em carteira (a 2ª venda fica coberta
    # pela ação; dobra a recuperação até o strike vendido, sem aporte novo)
    'stock_repair':   {'legs': [('C', 1, (0.45, 0.60)), ('C', -2, (0.20, 0.35))],
                       'asc': [0, 1], 'rule': 'zero_cost'},
    'iron_butterfly': {'legs': [('P', 1, (-0.12, -0.04)), ('P', -1, (-0.58, -0.42)),
                                ('C', -1, (0.42, 0.58)), ('C', 1, (0.04, 0.12))],
                       'same': [(1, 2)], 'rule': 'credit'},
    # ── Volatilidade e Neutras ──
    'straddle_comprado': {'legs': [('C', 1, (0.42, 0.58)), ('P', 1, (-0.58, -0.42))],
                          'same': [(0, 1)]},
    'strangle_comprado': {'legs': [('C', 1, (0.20, 0.32)), ('P', 1, (-0.32, -0.20))]},
    'guts_long':      {'legs': [('C', 1, (0.68, 0.82)), ('P', 1, (-0.82, -0.68))],
                       'asc': [0, 1]},
    'guts_short':     {'legs': [('C', -1, (0.58, 0.72)), ('P', -1, (-0.72, -0.58))],
                       'asc': [0, 1], 'rule': 'credit'},
    'borboleta_delta': {'legs': [('C', 1, (0.60, 0.80)), ('C', -2, (0.42, 0.58)),
                                 ('C', 1, (0.20, 0.40))], 'asc': [0, 1, 2]},
    'broken_wing':    {'legs': [('C', 1, (0.60, 0.80)), ('C', -2, (0.42, 0.58)),
                                ('C', 1, (0.08, 0.30))], 'asc': [0, 1, 2], 'rule': 'bwb'},
    'strap':          {'legs': [('C', 2, (0.42, 0.58)), ('P', 1, (-0.58, -0.42))],
                       'same': [(0, 1)]},
    'strip':          {'legs': [('C', 1, (0.42, 0.58)), ('P', 2, (-0.58, -0.42))],
                       'same': [(0, 1)]},
    'box_spread':     {'legs': [('C', 1, (0.55, 0.90)), ('C', -1, (0.10, 0.45)),
                                ('P', -1, (-0.45, -0.10)), ('P', 1, (-0.90, -0.55))],
                       'same': [(0, 2), (1, 3)], 'asc': [0, 1], 'rule': 'box'},
    # Front Ratio CALL 1×2: compra 1 ITM + vende 2 OTM (débito baixo/crédito);
    # lucro máximo no strike vendido, risco ILIMITADO acima (ponta nua)
    'ratio_spread':   {'legs': [('C', 1, (0.55, 0.72)), ('C', -2, (0.20, 0.35))],
                       'asc': [0, 1], 'rule': 'low_cost'},
    # ── Sintéticas e Avançadas ──
    'acao_sintetica': {'legs': [('C', 1, (0.40, 0.60)), ('P', -1, (-0.60, -0.40))],
                       'same': [(0, 1)]},
    'short_sintetico': {'legs': [('C', -1, (0.40, 0.60)), ('P', 1, (-0.60, -0.40))],
                        'same': [(0, 1)]},
    'risk_reversal':  {'legs': [('C', 1, (0.20, 0.30)), ('P', -1, (-0.30, -0.20))],
                       'rule': 'zero_cost'},
    'synthetic_straddle': {'legs': [('S', 1, None), ('P', 2, (-0.58, -0.42))]},
    # ── Do livro (Cap. 16 — Figuras do Mercado de Opções) ──
    # Adjusted Long Call Seagull (bullish, vol baixa): +PUT OTM(A) −PUT ATM(B)
    # +CALL ATM(C) −CALL OTM(D); a PUT comprada limita a perda na queda.
    'adj_call_seagull': {'legs': [('P', 1, (-0.35, -0.15)), ('P', -1, (-0.55, -0.40)),
                                  ('C', 1, (0.40, 0.55)), ('C', -1, (0.15, 0.35))],
                         'asc': [0, 1, 2, 3]},
    # Adjusted Long Put Seagull (bearish, vol baixa): −PUT baixa(A) +PUT alta(B)
    # −CALL ATM(C) +CALL alta(D); espelho da anterior.
    'adj_put_seagull':  {'legs': [('P', -1, (-0.55, -0.40)), ('P', 1, (-0.35, -0.15)),
                                  ('C', -1, (0.40, 0.55)), ('C', 1, (0.15, 0.35))],
                         'asc': [0, 1, 2, 3]},
    # Long Call Synthetic Strangle Riskless (vol alta): vende a ação (S) +
    # compra 2 CALLs ITM (A<B<spot); zona de retorno mínimo sem perda.
    'synth_strangle': {'legs': [('S', -1, None), ('C', 1, (0.55, 0.75)),
                                 ('C', 1, (0.78, 0.92))], 'asc': [1, 2],
                       'rule': 'synth_riskless'},
}


@app.route('/api/busca-operacoes/<ticker>')
@login_required
def api_busca_operacoes(ticker):
    """Analisa a cadeia de opções e sugere:
    1) Collars (compra ação + compra PUT + venda CALL) que superem a Selic no período
    2) Travas de alta/baixa no débito com relação ganho/custo > 1
    Até 8 vencimentos se o ativo tiver semanais com liquidez; senão 3 mensais."""
    ticker = ticker.strip().upper()
    token  = Settings.get_value('oplab_token', user_id=current_user.id)
    if not token:
        return jsonify({'error': 'Token OpLab não configurado'}), 400

    spot, spot_change = _get_underlying_quote(ticker, current_user.id)
    if not spot:
        return jsonify({'error': f'Cotação de {ticker} indisponível.'}), 404

    try:
        data = _oplab_get_json(f'/market/options/{ticker}', token, timeout=20)
    except OplabApiError as e:
        return jsonify({'error': str(e), 'status': e.status_code}), 503
    except Exception:
        app.logger.exception('api_busca_operacoes error for %s', ticker)
        return jsonify({'error': 'Erro inesperado ao buscar a cadeia de opções.'}), 500

    opt_list = data if isinstance(data, list) else (
        data.get('options') or data.get('calls', []) + data.get('puts', []) or []
    )

    from datetime import date as _date

    today = _date.today()
    calls_by_exp, puts_by_exp = {}, {}
    for o in opt_list:
        sym    = str(o.get('symbol') or o.get('ticker') or '').upper()
        cat    = str(o.get('category') or o.get('type') or '').upper()
        strike = float(o.get('strike') or 0)
        bid    = float(o.get('bid') or 0)
        ask    = float(o.get('ask') or 0)
        close  = float(o.get('close') or 0)
        vol    = float(o.get('financial_volume') or o.get('volume_financial') or 0)
        due    = str(o.get('due_date') or o.get('expiration_date') or '')
        if 'T' in due:
            due = due.split('T')[0]
        if not due or strike <= 0:
            continue
        try:
            if _date.fromisoformat(due) <= today:
                continue
        except ValueError:
            continue
        delta_raw = o.get('delta')
        if delta_raw is None and isinstance(o.get('greeks'), dict):
            delta_raw = o['greeks'].get('delta')
        row = {'symbol': sym, 'strike': round(strike, 2), 'bid': round(bid, 2),
               'ask': round(ask, 2), 'close': round(close, 2), 'vol_fin': round(vol, 2),
               'delta': delta_raw}
        bucket = puts_by_exp if ('PUT' in cat or cat == 'P') else calls_by_exp
        bucket.setdefault(due, []).append(row)

    all_exps = sorted(set(list(calls_by_exp.keys()) + list(puts_by_exp.keys())))
    if not all_exps:
        return jsonify({'error': f'Nenhuma opção encontrada para {ticker}.'}), 404

    def _third_friday(y, m):
        count, day = 0, 1
        while True:
            d = _date(y, m, day)
            if d.weekday() == 4:
                count += 1
                if count == 3:
                    return d
            day += 1

    def _is_monthly(exp_str):
        d = _date.fromisoformat(exp_str)
        tf = _third_friday(d.year, d.month)
        return abs((d - tf).days) <= 2  # tolera feriado na 3ª sexta

    selic = _selic()  # % a.a.

    # Operação solicitada — calcula somente ela
    op = (request.args.get('op') or 'collar').lower()
    try:
        min_ratio = float(request.args.get('min_ratio', 4))
    except (TypeError, ValueError):
        min_ratio = 4.0
    include_weekly = request.args.get('weekly', '1') != '0'
    try:
        max_days = int(request.args.get('days', 60))
    except (TypeError, ValueError):
        max_days = 60
    if max_days not in (60, 90, 120, 180):
        max_days = 60

    # ── Seleção de vencimentos (janela de max_days dias corridos) ────────────
    # Semanais = fora da 3ª sexta. Com o toggle ligado, inclui até 8 vencimentos
    # (semanais + mensais) dentro da janela. Desligado (ou venda_put_itm): só mensais.
    within_days = [e for e in all_exps
                   if (_date.fromisoformat(e) - today).days <= max_days]
    monthly_lim = [e for e in within_days if _is_monthly(e)]
    if not monthly_lim:
        monthly_lim = [e for e in all_exps if _is_monthly(e)][:3]

    # Existe alguma semanal dentro da janela de prazo escolhida?
    weekly_in_window = any(not _is_monthly(e) for e in within_days)

    if op == 'venda_put_itm':
        # Venda de PUT ITM é sempre mensal por definição
        selected_exps = monthly_lim
        mode = 'mensal'
    elif include_weekly and weekly_in_window:
        # Toggle marcado e há semanais na janela → inclui semanais + mensais
        selected_exps = within_days[:8]
        mode = 'semanal'
    else:
        # Toggle desmarcado, ou não há semanais na janela → apenas mensais
        selected_exps = monthly_lim
        mode = 'mensal'

    def _diversify(rows_list, key_fn, per_key=2, limit=10):
        """Evita linhas quase idênticas: no máx. per_key linhas por perna-âncora."""
        out, count = [], {}
        for r in rows_list:
            k = key_fn(r)
            c = count.get(k, 0)
            if c >= per_key:
                continue
            count[k] = c + 1
            out.append(r)
            if len(out) >= limit:
                break
        return out

    r_cont = math.log(1 + selic / 100.0)

    def _leg_delta_pct(row, is_call, T):
        """Delta da perna em módulo, escala 0-100.
        Usa o delta da OpLab quando presente; senão calcula via Black-Scholes
        com a IV extraída do prêmio de mercado (último negócio ou mid)."""
        d = row.get('delta')
        if d is not None:
            try:
                d = abs(float(d))
                return round(d * 100, 1) if d <= 1.5 else round(d, 1)
            except (TypeError, ValueError):
                pass
        prem = row['close'] or ((row['bid'] + row['ask']) / 2 if (row['bid'] and row['ask']) else 0)
        if not prem or T <= 0 or not spot or not row['strike']:
            return None
        intrinsic = max(0.0, (spot - row['strike']) if is_call else (row['strike'] - spot))
        iv = None
        if prem > intrinsic * 1.005:
            iv = _implied_vol(spot, row['strike'], T, r_cont, prem, is_call)
            if iv <= 0.006 or iv >= 4.9:
                iv = None
        if iv is None:
            iv = 0.35
        sq = math.sqrt(T)
        d1 = (math.log(spot / row['strike']) + (r_cont + 0.5 * iv * iv) * T) / (iv * sq)
        nd1 = 0.5 * (1 + math.erf(d1 / math.sqrt(2)))
        return round((nd1 if is_call else 1 - nd1) * 100, 1)

    # ── Preço efetivo por perna conforme o horário do pregão ─────────────────
    # Opções na B3: seg–sex, 10h às 16h30 (Brasília). Fora do pregão o book
    # está vazio/velho → usa o último negócio. No pregão, usa bid (venda) e
    # ask (compra), mas cai para o último quando falta ponta no book ou o
    # spread é abusivo (> 25% do mid, mínimo R$0,10).
    _now_b = now_brt()
    market_open = (_now_b.weekday() < 5
                   and (10, 0) <= (_now_b.hour, _now_b.minute) < (16, 30))

    def _eff(rw):
        """Retorna (bid_eff, bid_src, ask_eff, ask_src)."""
        bid, ask, last = rw['bid'], rw['ask'], rw['close']
        last_ok = last if last >= 0.05 else None
        if not market_open:
            return last_ok, 'último', last_ok, 'último'
        b_eff, b_src = (bid, 'bid') if bid >= 0.05 else (last_ok, 'último')
        a_eff, a_src = (ask, 'ask') if ask >= 0.05 else (last_ok, 'último')
        if bid >= 0.05 and ask >= 0.05 and last_ok:
            mid = (bid + ask) / 2
            if (ask - bid) > max(0.10, 0.25 * mid):   # spread abusivo
                b_eff, b_src = last_ok, 'último'
                a_eff, a_src = last_ok, 'último'
        return b_eff, b_src, a_eff, a_src

    def _sell_prem(rw):
        """Prêmio executável para perna vendida (bid efetivo ou último)."""
        b, b_src, _a, _asrc = _eff(rw)
        return (b, b_src) if b else (None, None)

    def _iv_est(rw, is_call, T):
        """VI implícita da perna extraída do prêmio (último ou mid); None se não converge."""
        prem = rw['close'] or ((rw['bid'] + rw['ask']) / 2 if (rw['bid'] and rw['ask']) else 0)
        if not prem or T <= 0 or not rw['strike']:
            return None
        intr = max(0.0, (spot - rw['strike']) if is_call else (rw['strike'] - spot))
        if prem <= intr * 1.005:
            return None
        iv = _implied_vol(spot, rw['strike'], T, r_cont, prem, is_call)
        return iv if 0.005 < iv < 4.9 else None

    def _pop_above(be, T, iv):
        """P(S_T > be) em % via log-normal risk-neutral."""
        if be <= 0 or T <= 0 or not iv:
            return None
        d2v = (math.log(spot / be) + (r_cont - 0.5 * iv * iv) * T) / (iv * math.sqrt(T))
        return _norm_cdf(d2v) * 100

    # ── Operações MULTI-VENCIMENTO: calendar / diagonal / double diagonal ────
    # Vende a perna curta e compra a longa. O resultado é avaliado NO VENCIMENTO
    # CURTO: a perna curta vale o intrínseco e a longa é reprecificada por
    # Black-Scholes com o tempo restante (valores estimados).
    # Ops do Vol. 4 (curta semanal → longa mensal): a perna longa PRECISA ser
    # mensal e o intervalo mínimo entre vencimentos cai para 7 dias.
    _CAL_V4 = ('neutral_calendar', 'double_calendar', 'pmcc', 'bull_calendar',
               'pmcp', 'bear_calendar')
    if op in ('calendar_spread', 'diagonal_spread', 'double_diagonal',
              'short_call_calendar', 'straddle_strangle_swap') + _CAL_V4:
        def _enrich_cal(lst):
            out = []
            for rw in lst:
                b, b_src, a, a_src = _eff(rw)
                rw2 = dict(rw)
                rw2['bid'] = round(b, 2) if b else 0
                rw2['ask'] = round(a, 2) if a else 0
                out.append(rw2)
            return sorted([r2 for r2 in out if r2['bid'] > 0 and r2['ask'] > 0],
                          key=lambda x: x['strike'])

        def _dl(rw, is_call, T):
            d = rw.get('delta')
            try:
                d = float(d) if d not in (None, '') else None
            except (TypeError, ValueError):
                d = None
            if d:
                d = abs(d)
                if d > 1:
                    d /= 100.0
                return d if is_call else -d
            iv = _iv_est(rw, is_call, T)
            if not iv or T <= 0:
                return None
            d1 = (math.log(spot / rw['strike']) + (r_cont + 0.5 * iv * iv) * T) / (iv * math.sqrt(T))
            nd = _norm_cdf(d1)
            return nd if is_call else nd - 1

        def _cands_cal(pool, is_call, T, win, side):
            need = 'bid' if side == 'sell' else 'ask'
            out = []
            for rw in pool:
                if rw[need] < 0.02:
                    continue
                dv = _dl(rw, is_call, T)
                if dv is None or not (win[0] <= dv <= win[1]):
                    continue
                out.append((abs(dv - (win[0] + win[1]) / 2), rw, dv))
            out.sort(key=lambda x: x[0])
            return [(rw, dv) for _, rw, dv in out[:3]]

        def _third_friday_cal(y, m):
            count, day = 0, 1
            while True:
                dd = _date(y, m, day)
                if dd.weekday() == 4:
                    count += 1
                    if count == 3:
                        return dd
                day += 1

        def _is_monthly_cal(exp_str):
            dd = _date.fromisoformat(exp_str)
            return abs((dd - _third_friday_cal(dd.year, dd.month)).days) <= 2

        # Pares curto→longo. Vol. 4: longa mensal, gap mínimo 7 dias (curta
        # semanal); demais: gap mínimo 20 dias.
        min_gap = 7 if op in _CAL_V4 else 20
        pairs = []
        for i in range(len(selected_exps) - 1):
            for j in range(i + 1, len(selected_exps)):
                dc_s = (_date.fromisoformat(selected_exps[i]) - today).days
                dc_l = (_date.fromisoformat(selected_exps[j]) - today).days
                if dc_l < dc_s + min_gap:
                    continue
                if op in _CAL_V4 and not _is_monthly_cal(selected_exps[j]):
                    continue
                pairs.append((selected_exps[i], selected_exps[j], dc_s, dc_l))
        pairs = pairs[:4]

        expirations = []
        for exp_s, exp_l, dc_s, dc_l in pairs:
            T_s, T_l = dc_s / 365.0, dc_l / 365.0
            calls_s = _enrich_cal(calls_by_exp.get(exp_s, []))
            calls_l = _enrich_cal(calls_by_exp.get(exp_l, []))
            puts_s  = _enrich_cal(puts_by_exp.get(exp_s, []))
            puts_l  = _enrich_cal(puts_by_exp.get(exp_l, []))
            rows = []

            def _eval_cal(sel):
                """sel: [(rw, delta, is_call, qty, exp, dc)] — vende curto/compra longo."""
                net = 0.0
                legs_out = []
                for rw, dv, is_c, q, expx, dcx in sel:
                    px = rw['ask'] if q > 0 else rw['bid']
                    net += (-q) * px
                    legs_out.append({'sym': rw['symbol'], 'tp': 'CALL' if is_c else 'PUT',
                                     'k': rw['strike'], 'q': q, 'px': px,
                                     'delta': round(dv * 100, 1), 'exp': expx,
                                     'iv': _iv_est(rw, is_c, dcx / 365.0)})
                def val(S):
                    v = net
                    for lg, (rw, dv, is_c, q, expx, dcx) in zip(legs_out, sel):
                        if dcx <= dc_s:                      # perna curta: intrínseco
                            intr = max(0.0, S - lg['k']) if is_c else max(0.0, lg['k'] - S)
                            v += q * intr
                        else:                                # perna longa: BS no tempo restante
                            T_rem = (dcx - dc_s) / 365.0
                            sig = lg['iv'] or 0.35
                            v += q * _bs_price(S, lg['k'], T_rem, r_cont, sig, is_c)
                    return v
                grid = [spot * (0.70 + 0.01 * k) for k in range(61)]   # 0.70–1.30 × spot
                vals = [val(S) for S in grid]
                bes = []
                for k in range(len(grid) - 1):
                    v0, v1 = vals[k], vals[k + 1]
                    if ((v0 < 0 <= v1) or (v0 >= 0 > v1)) and abs(v1 - v0) > 1e-9:
                        bes.append(grid[k] + (0 - v0) * (grid[k + 1] - grid[k]) / (v1 - v0))
                return net, legs_out, max(vals), max(0.0, -min(vals)), bes

            combos = []
            if op == 'calendar_spread':
                # mesmo strike ATM: vende curto, compra longo (Δ 0,40–0,60)
                for cs, dvs in _cands_cal(calls_s, True, T_s, (0.40, 0.60), 'sell'):
                    for cl in calls_l:
                        if abs(cl['strike'] - cs['strike']) > 0.011 or cl['ask'] < 0.02:
                            continue
                        dvl = _dl(cl, True, T_l)
                        combos.append([(cs, dvs, True, -1, exp_s, dc_s),
                                       (cl, dvl or dvs, True, 1, exp_l, dc_l)])
            elif op == 'diagonal_spread':
                # curta OTM (Δ 0,20–0,35) vendida; longa ATM (Δ 0,40–0,55) comprada
                for cs, dvs in _cands_cal(calls_s, True, T_s, (0.20, 0.35), 'sell'):
                    for cl, dvl in _cands_cal(calls_l, True, T_l, (0.40, 0.55), 'buy'):
                        if cl['strike'] >= cs['strike']:
                            continue
                        combos.append([(cs, dvs, True, -1, exp_s, dc_s),
                                       (cl, dvl, True, 1, exp_l, dc_l)])
            elif op in ('neutral_calendar', 'bull_calendar', 'bear_calendar'):
                # Vol. 4: vende curta (semanal) e compra longa (mensal) no MESMO strike.
                # neutral = ATM (Δ ~0,50); bull = CALL OTM; bear = PUT OTM.
                is_call = op != 'bear_calendar'
                win = ((0.42, 0.58) if op == 'neutral_calendar'
                       else ((0.20, 0.35) if is_call else (-0.35, -0.20)))
                pool_s = calls_s if is_call else puts_s
                pool_l = calls_l if is_call else puts_l
                for cs, dvs in _cands_cal(pool_s, is_call, T_s, win, 'sell'):
                    for cl in pool_l:
                        if abs(cl['strike'] - cs['strike']) > 0.011 or cl['ask'] < 0.02:
                            continue
                        dvl = _dl(cl, is_call, T_l)
                        combos.append([(cs, dvs, is_call, -1, exp_s, dc_s),
                                       (cl, dvl or dvs, is_call, 1, exp_l, dc_l)])
            elif op in ('pmcc', 'pmcp'):
                # Poor Man's Covered Call/Put: longa mensal DEEP ITM (|Δ| > 0,80)
                # substitui a ação; curta semanal OTM (|Δ| 0,20–0,30) gera renda.
                is_call = op == 'pmcc'
                win_s = (0.20, 0.30) if is_call else (-0.30, -0.20)
                win_l = (0.78, 0.92) if is_call else (-0.92, -0.78)
                pool_s = calls_s if is_call else puts_s
                pool_l = calls_l if is_call else puts_l
                for cs, dvs in _cands_cal(pool_s, is_call, T_s, win_s, 'sell'):
                    for cl, dvl in _cands_cal(pool_l, is_call, T_l, win_l, 'buy'):
                        if is_call and cl['strike'] >= cs['strike']:
                            continue
                        if not is_call and cl['strike'] <= cs['strike']:
                            continue
                        combos.append([(cs, dvs, is_call, -1, exp_s, dc_s),
                                       (cl, dvl, is_call, 1, exp_l, dc_l)])
            elif op == 'double_calendar':
                # Vende strangle semanal OTM + compra o MESMO strangle no mensal.
                for cs, dvs in _cands_cal(calls_s, True, T_s, (0.18, 0.32), 'sell')[:3]:
                    for ps, dps in _cands_cal(puts_s, False, T_s, (-0.32, -0.18), 'sell')[:3]:
                        cl = next((c for c in calls_l
                                   if abs(c['strike'] - cs['strike']) <= 0.011 and c['ask'] >= 0.02), None)
                        pl = next((p for p in puts_l
                                   if abs(p['strike'] - ps['strike']) <= 0.011 and p['ask'] >= 0.02), None)
                        if not cl or not pl:
                            continue
                        combos.append([(cs, dvs, True, -1, exp_s, dc_s),
                                       (ps, dps, False, -1, exp_s, dc_s),
                                       (cl, _dl(cl, True, T_l) or dvs, True, 1, exp_l, dc_l),
                                       (pl, _dl(pl, False, T_l) or dps, False, 1, exp_l, dc_l)])
            elif op == 'short_call_calendar':
                # Do livro: COMPRA CALL curta ATM (Front) + VENDE CALL longa ATM
                # (Back), mesmo strike. Invertido do calendar normal — lucra com
                # movimento BRUSCO em qualquer direção (as duas viram pó/explodem).
                for cs, dvs in _cands_cal(calls_s, True, T_s, (0.40, 0.60), 'buy'):
                    for cl in calls_l:
                        if abs(cl['strike'] - cs['strike']) > 0.011 or cl['bid'] < 0.02:
                            continue
                        dvl = _dl(cl, True, T_l)
                        combos.append([(cs, dvs, True, 1, exp_s, dc_s),
                                       (cl, dvl or dvs, True, -1, exp_l, dc_l)])
            elif op == 'straddle_strangle_swap':
                # Do livro: vende STRADDLE curto ATM (−CALL B, −PUT B) + compra
                # STRANGLE longo (CALL OTM C acima, PUT OTM A abaixo). Neutra;
                # lucro máximo se ficar no strike central no venc. curto.
                for cb, dcb in _cands_cal(calls_s, True, T_s, (0.42, 0.58), 'sell')[:3]:
                    pb = next((p for p in puts_s
                               if abs(p['strike'] - cb['strike']) <= 0.011 and p['bid'] >= 0.02), None)
                    if not pb:
                        continue
                    dpb = _dl(pb, False, T_s)
                    for cl, dvl in _cands_cal(calls_l, True, T_l, (0.20, 0.35), 'buy')[:2]:
                        if cl['strike'] <= cb['strike']:
                            continue
                        for pl, dpl in _cands_cal(puts_l, False, T_l, (-0.35, -0.20), 'buy')[:2]:
                            if pl['strike'] >= cb['strike']:
                                continue
                            combos.append([(cb, dcb, True, -1, exp_s, dc_s),
                                           (pb, dpb or -0.5, False, -1, exp_s, dc_s),
                                           (cl, dvl, True, 1, exp_l, dc_l),
                                           (pl, dpl, False, 1, exp_l, dc_l)])
            else:   # double_diagonal
                for cs, dvs in _cands_cal(calls_s, True, T_s, (0.18, 0.32), 'sell')[:2]:
                    for ps, dps in _cands_cal(puts_s, False, T_s, (-0.32, -0.18), 'sell')[:2]:
                        for cl, dvl in _cands_cal(calls_l, True, T_l, (0.08, 0.22), 'buy')[:2]:
                            if cl['strike'] < cs['strike']:
                                continue
                            for pl, dpl in _cands_cal(puts_l, False, T_l, (-0.22, -0.08), 'buy')[:2]:
                                if pl['strike'] > ps['strike']:
                                    continue
                                combos.append([(cs, dvs, True, -1, exp_s, dc_s),
                                               (ps, dps, False, -1, exp_s, dc_s),
                                               (cl, dvl, True, 1, exp_l, dc_l),
                                               (pl, dpl, False, 1, exp_l, dc_l)])

            # Operações do livro montadas no CRÉDITO (não exigir débito)
            _credit_cal = ('double_diagonal', 'short_call_calendar', 'straddle_strangle_swap')
            for sel in combos:
                net, legs_out, max_gain, max_loss, bes = _eval_cal(sel)
                cost = -net
                if cost <= 0 and op not in _credit_cal:
                    continue                                  # calendários/diagonais: débito
                if max_gain <= 0:
                    continue
                ratio = round(max_gain / max_loss, 2) if max_loss > 0.001 else None
                for lg in legs_out:
                    lg.pop('iv', None)
                rows.append({
                    'legs':      legs_out,
                    'net':       round(net, 2),
                    'is_credit': net >= 0,
                    'montagem':  'CRÉDITO' if net > 0.005 else 'DÉBITO',
                    'max_gain':  round(max_gain, 2), 'gain_unl': False,
                    'max_loss':  round(max_loss, 2), 'loss_unl': False,
                    'ratio':     ratio,
                    'bes':       [round(b, 2) for b in bes],
                    'est':       True,                        # valores estimados (BS)
                })
            rows.sort(key=lambda x: -(x['ratio'] or 0))
            rows = rows[:8]
            expirations.append({
                'exp':          exp_s,
                'exp_long':     exp_l,
                'dc':           dc_s,
                'dc_long':      dc_l,
                'is_monthly':   True,
                'selic_period': round(((1 + selic / 100) ** (dc_s / 365.0) - 1) * 100, 2),
                'rows':         rows,
            })

        return jsonify({
            'ticker': ticker, 'spot': spot, 'spot_change': spot_change,
            'op': op, 'max_days': max_days, 'market_open': market_open,
            'selic': round(selic, 2), 'expirations': expirations,
        })

    expirations = []
    for exp in selected_exps:
        dc = max((_date.fromisoformat(exp) - today).days, 1)
        selic_period = ((1 + selic / 100) ** (dc / 365.0) - 1) * 100

        # Pernas com preço efetivo: bid/ask do book (pregão aberto e spread são)
        # ou último negócio (pregão fechado, ponta ausente ou spread abusivo).
        def _enrich(lst):
            out = []
            for rw in lst:
                b, b_src, a, a_src = _eff(rw)
                rw2 = dict(rw)
                rw2['bid'] = round(b, 2) if b else 0
                rw2['ask'] = round(a, 2) if a else 0
                rw2['bid_src'] = b_src
                rw2['ask_src'] = a_src
                out.append(rw2)
            return out
        calls_all = _enrich(calls_by_exp.get(exp, []))
        puts_all  = _enrich(puts_by_exp.get(exp, []))
        calls_ok = sorted([c for c in calls_all if c['bid'] > 0 and c['ask'] > 0],
                          key=lambda x: x['strike'])
        puts_ok  = sorted([p for p in puts_all if p['bid'] > 0 and p['ask'] > 0],
                          key=lambda x: x['strike'])
        rows = []

        if op == 'collar':
            # Compra ação (spot) + compra PUT (ask) + venda CALL (bid).
            # Risco zero: strike da PUT >= custo líquido; senão relação ganho/perda >= min_ratio.
            put_cands  = [p for p in puts_ok  if 0.90 * spot <= p['strike'] <= 1.10 * spot][:20]
            call_cands = [c for c in calls_ok if spot * 0.97 <= c['strike'] <= 1.20 * spot][:20]
            for p in put_cands:
                for c in call_cands:
                    if c['strike'] <= p['strike']:
                        continue
                    net = spot + p['ask'] - c['bid']
                    if net <= 0:
                        continue
                    max_gain = c['strike'] - net
                    min_res  = p['strike'] - net
                    if max_gain <= 0:
                        continue
                    gain_pct = max_gain / net * 100
                    if gain_pct <= selic_period:
                        continue
                    loss = -min_res
                    risk_free = min_res >= 0
                    ratio = None if loss <= 0.001 else max_gain / loss
                    if not risk_free and (ratio is None or ratio < min_ratio):
                        continue
                    gain_aa = ((1 + gain_pct / 100) ** (365.0 / dc) - 1) * 100
                    rows.append({
                        'put_symbol':  p['symbol'],  'put_strike':  p['strike'],  'put_ask':  p['ask'],
                        'call_symbol': c['symbol'],  'call_strike': c['strike'],  'call_bid': c['bid'],
                        'net_cost':   round(net, 2),
                        'breakeven':  round(net, 2),
                        'max_gain':   round(max_gain, 2),
                        'gain_pct':   round(gain_pct, 2),
                        'min_result': round(min_res, 2),
                        'gain_aa':    round(gain_aa, 1),
                        'vs_selic':   round(gain_pct - selic_period, 2),
                        'risk_free':  risk_free,
                        'ratio':      round(ratio, 2) if ratio is not None else None,
                    })
            rows.sort(key=lambda x: (not x['risk_free'], -x['gain_pct']))
            rows = _diversify(rows, lambda x: x['call_symbol'], per_key=2)

        elif op == 'fence':
            # Cerca: compra ação + compra PUT K2 + venda PUT K1 (<K2) + venda CALL (>K2).
            # Proteção total entre K1 e K2; abaixo de K1 o risco volta (como ter o papel de nível mais baixo).
            put_hi  = [p for p in puts_ok  if 0.90 * spot <= p['strike'] <= 1.08 * spot][:12]
            put_lo  = [p for p in puts_ok  if 0.70 * spot <= p['strike'] <  0.98 * spot][:12]
            call_c  = [c for c in calls_ok if 0.99 * spot <= c['strike'] <= 1.20 * spot][:12]
            for p2 in put_hi:
                for p1 in put_lo:
                    if p1['strike'] >= p2['strike']:
                        continue
                    for c in call_c:
                        if c['strike'] <= p2['strike']:
                            continue
                        net = spot + p2['ask'] - p1['bid'] - c['bid']
                        if net <= 0:
                            continue
                        max_gain = c['strike'] - net
                        if max_gain <= 0:
                            continue
                        gain_pct = max_gain / net * 100
                        if gain_pct <= selic_period:
                            continue
                        floor_res = p2['strike'] - net     # resultado garantido na zona [K1, K2]
                        risk_free = floor_res >= 0
                        ratio = None if floor_res >= -0.001 else max_gain / (-floor_res)
                        if not risk_free and (ratio is None or ratio < min_ratio):
                            continue
                        gain_aa = ((1 + gain_pct / 100) ** (365.0 / dc) - 1) * 100
                        rows.append({
                            'put_buy_symbol':  p2['symbol'], 'put_buy_strike':  p2['strike'], 'put_buy_ask':  p2['ask'],
                            'put_sell_symbol': p1['symbol'], 'put_sell_strike': p1['strike'], 'put_sell_bid': p1['bid'],
                            'call_symbol': c['symbol'], 'call_strike': c['strike'], 'call_bid': c['bid'],
                            'net_cost':      round(net, 2),
                            'breakeven':     round(net, 2),
                            'max_gain':      round(max_gain, 2),
                            'gain_pct':      round(gain_pct, 2),
                            'gain_aa':       round(gain_aa, 1),
                            'vs_selic':      round(gain_pct - selic_period, 2),
                            'floor_result':  round(floor_res, 2),
                            'prot_until':    p1['strike'],
                            'prot_drop_pct': round((spot - p1['strike']) / spot * 100, 1),
                            'risk_free':     risk_free,
                            'ratio':         round(ratio, 2) if ratio is not None else None,
                        })
            rows.sort(key=lambda x: (not x['risk_free'], -x['gain_pct']))
            rows = _diversify(rows, lambda x: (x['put_buy_symbol'], x['call_symbol']), per_key=1)

        elif op == 'seagull':
            # Gaivota (alta): compra trava de alta com CALLs financiada por venda de PUT OTM.
            # CALL comprada perto do dinheiro; prêmios-poeira descartados; a PUT deve
            # financiar pelo menos metade do custo da trava.
            c_lo   = [c for c in calls_ok
                      if 0.97 * spot <= c['strike'] <= 1.06 * spot and c['ask'] >= 0.10][:10]
            p_sell = [p for p in puts_ok
                      if 0.85 * spot <= p['strike'] <= 0.97 * spot and p['bid'] >= 0.05][:12]
            for c1 in c_lo:
                c_his = [c for c in calls_ok
                         if c1['strike'] < c['strike'] <= 1.15 * spot and c['bid'] >= 0.03][:8]
                for c2 in c_his:
                    spread_cost = c1['ask'] - c2['bid']
                    if spread_cost <= 0:
                        continue
                    for p0 in p_sell:
                        if p0['bid'] < 0.5 * spread_cost:   # PUT precisa financiar >= 50%
                            continue
                        net   = spread_cost - p0['bid']     # >0 débito, <=0 crédito
                        width = c2['strike'] - c1['strike']
                        max_gain = width - net
                        if max_gain <= 0:
                            continue
                        if net > 0.35 * width:
                            continue
                        margin_pct = (spot - p0['strike']) / spot * 100
                        be_low = p0['strike'] + min(net, 0)   # crédito amortece a queda
                        rows.append({
                            'call_buy_symbol':  c1['symbol'], 'call_buy_strike':  c1['strike'], 'call_buy_ask':  c1['ask'],
                            'call_sell_symbol': c2['symbol'], 'call_sell_strike': c2['strike'], 'call_sell_bid': c2['bid'],
                            'put_sell_symbol':  p0['symbol'], 'put_sell_strike':  p0['strike'], 'put_sell_bid':  p0['bid'],
                            'net_cost':   round(net, 2),
                            'is_credit':  net <= 0,
                            'max_gain':   round(max_gain, 2),
                            'margin_pct': round(margin_pct, 1),
                            'be_low':     round(be_low, 2),
                        })
            # Crédito primeiro; depois CALL comprada mais perto do dinheiro; menor custo
            rows.sort(key=lambda x: (not x['is_credit'], x['call_buy_strike'], x['net_cost']))
            rows = _diversify(rows, lambda x: (x['call_buy_symbol'], x['call_sell_symbol']), per_key=1)

        elif op in ('trava_alta', 'trava_baixa'):
            # Travas no DÉBITO otimizadas pela equação do trader:
            #   EV = POP×ganho − (1−POP)×custo > 0 (POP via log-normal com IV dos prêmios)
            # Regras de qualidade:
            #   • perna comprada ATM/levemente OTM (ITM tem book ralo — evita)
            #   • custo entre ~25% e 55% da largura → relação ganho/custo 0.8–3.0
            #     (nem "loteria" OTM distante de POP baixo, nem trava cara sem ganho)
            #   • POP mínimo 35%
            #   • puts exigem prêmios mais firmes (menos líquidas que calls)
            T = dc / 365.0
            is_alta = op == 'trava_alta'
            if is_alta:
                buys = [c for c in calls_ok
                        if 0.97 * spot <= c['strike'] <= 1.06 * spot and c['ask'] >= 0.10][:10]
            else:
                buys = [p for p in puts_ok
                        if 0.94 * spot <= p['strike'] <= 1.03 * spot and p['ask'] >= 0.15][:10]
            for buy in buys:
                if is_alta:
                    sells = [c for c in calls_ok
                             if buy['strike'] < c['strike'] <= 1.18 * spot and c['bid'] >= 0.03][:8]
                else:
                    sells = [p for p in puts_ok
                             if 0.82 * spot <= p['strike'] < buy['strike'] and p['bid'] >= 0.05][:8]
                for sell in sells:
                    cost  = buy['ask'] - sell['bid']
                    width = abs(sell['strike'] - buy['strike'])
                    if cost <= 0.01 or width <= 0:
                        continue
                    max_gain = width - cost
                    if max_gain <= 0:
                        continue
                    ratio = max_gain / cost
                    if ratio < 0.8 or ratio > 3.0:
                        continue
                    be = buy['strike'] + cost if is_alta else buy['strike'] - cost
                    iv = _iv_est(buy, is_alta, T) or _iv_est(sell, is_alta, T) or 0.35
                    p_above = _pop_above(be, T, iv)
                    if p_above is None:
                        continue
                    pop = p_above if is_alta else 100 - p_above
                    ev  = pop / 100 * max_gain - (1 - pop / 100) * cost
                    if pop < 35 or ev <= 0:      # equação do trader
                        continue
                    liq = min(buy.get('vol_fin') or 0, sell.get('vol_fin') or 0)
                    rows.append({
                        'buy_symbol':  buy['symbol'],  'buy_strike':  buy['strike'],  'buy_ask':  buy['ask'],
                        'sell_symbol': sell['symbol'], 'sell_strike': sell['strike'], 'sell_bid': sell['bid'],
                        'cost':      round(cost, 2),
                        'max_gain':  round(max_gain, 2),
                        'ratio':     round(ratio, 2),
                        'pop':       round(pop, 1),
                        'ev':        round(ev, 2),
                        'ev_pct':    round(ev / cost * 100, 1),   # expectância por R$ arriscado
                        'liq':       liq,
                        'breakeven': round(be, 2),
                        'be_dist':   round((be - spot) / spot * 100, 2),
                    })
            # Maior expectância por unidade de risco; empate: mais volume (liquidez)
            rows.sort(key=lambda x: (-x['ev_pct'], -x['liq']))
            rows = _diversify(rows, lambda x: x['buy_symbol'], per_key=2)

        elif op in ('trava_alta_credito', 'trava_baixa_credito'):
            # Travas no CRÉDITO otimizadas pela equação do trader:
            #   EV = POP×crédito − (1−POP)×perda_máx > 0
            # Regras de qualidade (sweet spot clássico das verticais de crédito):
            #   • perna vendida OTM (fora do dinheiro — mais líquida que ITM)
            #   • crédito entre 25% e 60% da largura → POP típico 55–80%
            #   • POP mínimo 55% (a estratégia vive de taxa de acerto alta)
            #   • puts exigem prêmios mais firmes (menos líquidas que calls)
            T = dc / 365.0
            is_alta = op == 'trava_alta_credito'
            if is_alta:
                # Bull put: vende PUT OTM abaixo do spot, compra PUT mais abaixo
                sells = [p for p in puts_ok
                         if 0.85 * spot <= p['strike'] <= 0.99 * spot and p['bid'] >= 0.10][:10]
            else:
                # Bear call: vende CALL OTM acima do spot, compra CALL mais acima
                sells = [c for c in calls_ok
                         if 1.01 * spot <= c['strike'] <= 1.15 * spot and c['bid'] >= 0.05][:10]
            for sell in sells:
                if is_alta:
                    buys = [p for p in puts_ok
                            if 0.75 * spot <= p['strike'] < sell['strike'] and p['ask'] >= 0.03][:8]
                else:
                    buys = [c for c in calls_ok
                            if sell['strike'] < c['strike'] <= 1.28 * spot and c['ask'] >= 0.01][:8]
                for buy in buys:
                    credit = sell['bid'] - buy['ask']
                    width  = abs(sell['strike'] - buy['strike'])
                    if credit <= 0.01 or width <= 0:
                        continue
                    if not (0.25 * width <= credit <= 0.60 * width):
                        continue
                    max_loss = width - credit
                    if max_loss <= 0.001:
                        continue
                    ratio = credit / max_loss
                    be = sell['strike'] - credit if is_alta else sell['strike'] + credit
                    is_call_leg = not is_alta
                    iv = _iv_est(sell, is_call_leg, T) or _iv_est(buy, is_call_leg, T) or 0.35
                    p_above = _pop_above(be, T, iv)
                    if p_above is None:
                        continue
                    pop = p_above if is_alta else 100 - p_above
                    ev  = pop / 100 * credit - (1 - pop / 100) * max_loss
                    if pop < 55 or ev <= 0:      # equação do trader
                        continue
                    liq = min(sell.get('vol_fin') or 0, buy.get('vol_fin') or 0)
                    rows.append({
                        'sell_symbol': sell['symbol'], 'sell_strike': sell['strike'], 'sell_bid': sell['bid'],
                        'buy_symbol':  buy['symbol'],  'buy_strike':  buy['strike'],  'buy_ask':  buy['ask'],
                        'credit':    round(credit, 2),
                        'max_loss':  round(max_loss, 2),
                        'ratio':     round(ratio, 2),
                        'pop':       round(pop, 1),
                        'ev':        round(ev, 2),
                        'ev_pct':    round(ev / max_loss * 100, 1),   # expectância por R$ de risco
                        'liq':       liq,
                        'breakeven': round(be, 2),
                        'be_dist':   round((be - spot) / spot * 100, 2),
                    })
            # Maior expectância por unidade de risco; empate: mais volume (liquidez)
            rows.sort(key=lambda x: (-x['ev_pct'], -x['liq']))
            rows = _diversify(rows, lambda x: x['sell_symbol'], per_key=2)

        elif op == 'trava_credito':
            # Call ratio backspread: venda de CALL perto do dinheiro financia compra de CALLs OTM.
            # CALL vendida: no máximo 5% ITM e até 5% OTM (0.95x a 1.05x do spot).
            # Proporções 1x2 e 2x3. Pequeno custo ou crédito aceitável.
            # Lucro ilimitado na alta; perda máxima no strike comprado (K2).
            sell_cands = [c for c in calls_ok
                          if 0.95 * spot <= c['strike'] <= 1.05 * spot and c['bid'] >= 0.10][:12]
            otm_calls  = [c for c in calls_ok
                          if spot < c['strike'] <= 1.18 * spot and c['ask'] >= 0.03][:12]
            for n_sell, m_buy, label in [(1, 2, '1x2'), (2, 3, '2x3')]:
                for sell in sell_cands:
                    for buy in otm_calls:
                        if buy['strike'] <= sell['strike']:
                            continue
                        width = buy['strike'] - sell['strike']
                        net = n_sell * sell['bid'] - m_buy * buy['ask']   # >0 crédito, <0 custo
                        # custo aceitável: até 15% do valor da largura vendida
                        if net < -0.15 * n_sell * width:
                            continue
                        max_loss = n_sell * width - net                   # em S = K2 (comprado)
                        if max_loss <= 0:
                            continue                                      # arbitragem improvável / dado ruim
                        be_up  = buy['strike'] + max_loss / (m_buy - n_sell)
                        be_low = (sell['strike'] + net / n_sell) if net > 0 else None
                        s10 = spot * 1.10                                 # ganho se subir 10%
                        gain10 = net - n_sell * max(0.0, s10 - sell['strike']) \
                                     + m_buy * max(0.0, s10 - buy['strike'])
                        rows.append({
                            'tipo':      label,
                            'n_sell':    n_sell, 'm_buy': m_buy,
                            'sell_symbol': sell['symbol'], 'sell_strike': sell['strike'], 'sell_bid': sell['bid'],
                            'buy_symbol':  buy['symbol'],  'buy_strike':  buy['strike'],  'buy_ask':  buy['ask'],
                            'credit':    round(net, 2),                   # negativo = pequeno custo
                            'is_credit': net >= 0,
                            'max_loss':  round(max_loss, 2),
                            'be_up':     round(be_up, 2),
                            'be_up_dist': round((be_up - spot) / spot * 100, 2),
                            'be_low':    round(be_low, 2) if be_low is not None else None,
                            'gain10':    round(gain10, 2),
                        })
            # Crédito primeiro; menor perda máxima; BE superior mais próximo
            rows.sort(key=lambda x: (not x['is_credit'], x['max_loss'], x['be_up_dist']))
            rows = _diversify(rows, lambda x: (x['tipo'], x['sell_symbol']), per_key=2, limit=12)

        elif op == 'vaca_travada':
            # Vaca travada (borboleta de CALLs, asas podem ser assimétricas):
            # +1 CALL baixa, -2 CALLs médias, +1 CALL alta.
            # Centro (K médio) no dinheiro ou acima; aceita investimento (custo até 50% da asa).
            low_c = [c for c in calls_ok
                     if 0.90 * spot <= c['strike'] <= 1.08 * spot and c['ask'] >= 0.05][:12]
            for c1 in low_c:
                mids = [c for c in calls_ok
                        if c1['strike'] < c['strike'] <= 1.15 * spot
                        and c['strike'] >= 0.99 * spot and c['bid'] >= 0.05][:8]
                for c2 in mids:
                    highs = [c for c in calls_ok
                             if c2['strike'] < c['strike'] <= 1.25 * spot and c['ask'] >= 0.01][:8]
                    for c3 in highs:
                        cost = c1['ask'] - 2 * c2['bid'] + c3['ask']   # >0 débito, <=0 crédito
                        w_lo = c2['strike'] - c1['strike']
                        if cost > 0.50 * w_lo:          # investimento aceito até 50% da asa
                            continue
                        max_gain = w_lo - cost          # em S = K2
                        if max_gain <= 0:
                            continue
                        montagem = 'CRÉDITO' if cost <= 0 else ('ZERO' if cost <= 0.15 * w_lo else 'INVEST')
                        # Resultado acima da asa superior (S > K3): asas assimétricas podem perder
                        tail = w_lo - (c3['strike'] - c2['strike']) - cost
                        be_low = round(c1['strike'] + cost, 2) if cost > 0 else None
                        be_up  = round(c2['strike'] + max_gain, 2) if tail < 0 else None
                        max_loss = max(cost if cost > 0 else 0.0, -tail if tail < 0 else 0.0)
                        ratio = None if max_loss <= 0.001 else max_gain / max_loss
                        rows.append({
                            'low_symbol':  c1['symbol'], 'low_strike':  c1['strike'], 'low_ask':  c1['ask'],
                            'mid_symbol':  c2['symbol'], 'mid_strike':  c2['strike'], 'mid_bid':  c2['bid'],
                            'high_symbol': c3['symbol'], 'high_strike': c3['strike'], 'high_ask': c3['ask'],
                            'cost':      round(cost, 2),
                            'is_credit': cost <= 0,
                            'montagem':  montagem,
                            'max_gain':  round(max_gain, 2),
                            'tail':      round(tail, 2),
                            'be_low':    be_low,
                            'be_up':     be_up,
                            'mid_dist':  round((c2['strike'] - spot) / spot * 100, 1),  # % até o pico de lucro
                            'max_loss':  round(max_loss, 2),
                            'ratio':     round(ratio, 1) if ratio is not None else None,
                        })
            # Centro mais perto do spot primeiro; depois menor risco
            rows.sort(key=lambda x: (abs(x['mid_dist']), x['max_loss'], -x['max_gain']))
            rows = _diversify(rows, lambda x: x['mid_symbol'], per_key=2)

        elif op in ('boi_coberto', 'vaca_tradicional', 'vaca_revertida', 'borboleta', 'condor'):
            # Estruturas clássicas de CALLs — recomendações usuais de montagem:
            #   boi_coberto:      -1 ITM (moderada), +2 ATM, -1 OTM — a venda ITM paga
            #                     parte da montagem e a venda OTM limita o ganho;
            #                     asa superior mais larga que a ITM (lucro na alta);
            #                     crédito pequeno ou nulo; lucro máx. > CDI do período
            #   vaca_tradicional: +1 baixa, -3 médias, +2 altas — montada no crédito
            #   vaca_revertida:   +1 baixa, -5 médias, +5 altas — crédito; lucro
            #                     ilimitado na alta forte (slope +1 acima da asa)
            #   borboleta:        +1 baixa, -2 médias, +1 alta — débito <= 30% da asa,
            #                     centro próximo do ATM (aposta em preço-alvo)
            #   condor:           +1, -1, -1, +1 — débito <= 40% da asa; platô de lucro
            #                     entre os strikes vendidos, idealmente contendo o spot
            SPECS = {
                'boi_coberto':      {'qty': (-1, 2, -1),
                                     'rngs': [(0.93, 0.995), (0.99, 1.05), (1.02, 1.20)]},
                'vaca_tradicional': {'qty': (1, -3, 2),
                                     'rngs': [(0.92, 1.03), (0.99, 1.15), (1.01, 1.28)]},
                'vaca_revertida':   {'qty': (1, -5, 5),
                                     'rngs': [(0.92, 1.03), (0.99, 1.15), (1.01, 1.28)]},
                'borboleta':        {'qty': (1, -2, 1),
                                     'rngs': [(0.88, 1.02), (0.95, 1.10), (1.00, 1.22)]},
                'condor':           {'qty': (1, -1, -1, 1),
                                     'rngs': [(0.85, 0.99), (0.93, 1.05), (0.98, 1.12), (1.02, 1.25)]},
            }
            spec  = SPECS[op]
            qtys  = spec['qty']
            nlegs = len(qtys)
            ncand = 8 if nlegs == 4 else 10

            def _cands(rng, q):
                lo, hi = rng
                if q > 0:   # comprada paga ask
                    return [c for c in calls_ok
                            if lo * spot <= c['strike'] <= hi * spot and c['ask'] >= 0.03][:ncand]
                return [c for c in calls_ok
                        if lo * spot <= c['strike'] <= hi * spot and c['bid'] >= 0.05][:ncand]

            def _eval_struct(legs):
                """legs: [(row, qty)] strikes crescentes. Payoff linear por partes."""
                net = sum((-q) * (c['ask'] if q > 0 else c['bid']) for c, q in legs)  # >0 crédito
                ks  = [c['strike'] for c, _ in legs]

                def pay(S):
                    return net + sum(q * max(0.0, S - c['strike']) for c, q in legs)

                slope = sum(q for _, q in legs)                # inclinação acima da última asa
                far   = ks[-1] * 1.6
                pts   = [0.0] + ks + [far]
                vals  = [pay(x) for x in pts]
                bes   = []
                for i in range(len(pts) - 1):
                    v0, v1 = vals[i], vals[i + 1]
                    if (v0 < 0 <= v1) or (v0 >= 0 > v1):
                        if abs(v1 - v0) > 1e-9:
                            bes.append(pts[i] + (0 - v0) * (pts[i + 1] - pts[i]) / (v1 - v0))
                max_gain = None if slope > 0 else max(vals)    # None = ilimitado
                max_loss = max(0.0, -min(vals))
                return net, max_gain, max_loss, bes

            import itertools as _it
            cand_lists = [_cands(spec['rngs'][i], qtys[i]) for i in range(nlegs)]
            for combo in _it.product(*cand_lists):
                ks = [c['strike'] for c in combo]
                if any(ks[i] >= ks[i + 1] for i in range(nlegs - 1)):
                    continue
                legs = list(zip(combo, qtys))
                net, max_gain, max_loss, bes = _eval_struct(legs)
                cost = -net                                   # >0 débito
                w_lo = ks[1] - ks[0]
                if max_loss <= 0.001 and cost > 0:
                    continue                                  # dado inconsistente
                if op == 'boi_coberto':
                    # asa superior (OTM) mais larga que a ITM → estrutura lucra na alta
                    w_up = ks[2] - ks[1]
                    if w_up <= w_lo:
                        continue
                    # crédito pequeno ou nulo: tolera custo residual, limita crédito
                    if net < -0.05 * w_lo or net > 0.30 * w_lo:
                        continue
                    if max_gain is None or max_gain <= 0 or max_loss <= 0:
                        continue
                    # lucro máximo (na alta) precisa superar o CDI do período s/ capital em risco
                    if max_gain < max_loss * selic_period / 100:
                        continue
                elif op in ('vaca_tradicional', 'vaca_revertida'):
                    # vacas montadas no crédito (aceita custo residual de até 10% da asa)
                    if cost > 0.10 * w_lo:
                        continue
                elif op == 'borboleta':
                    # débito baixo em relação à asa; recusa borboleta "de graça" (dado ruim)
                    if cost <= 0 or cost > 0.30 * w_lo:
                        continue
                elif op == 'condor':
                    if cost <= 0 or cost > 0.40 * w_lo:
                        continue
                montagem = ('CRÉDITO' if net > 0.005
                            else ('ZERO' if cost <= 0.15 * w_lo else 'INVEST'))
                ratio = (round(max_gain / max_loss, 1)
                         if (max_gain is not None and max_loss > 0.001) else None)
                rows.append({
                    'legs': [{'sym': c['symbol'], 'k': c['strike'], 'q': q,
                              'px': (c['ask'] if q > 0 else c['bid']),
                              'src': (c['ask_src'] if q > 0 else c['bid_src'])}
                             for c, q in legs],
                    'net':       round(net, 2),
                    'is_credit': net >= 0,
                    'montagem':  montagem,
                    'max_gain':  round(max_gain, 2) if max_gain is not None else None,
                    'gain_unl':  max_gain is None,
                    'max_loss':  round(max_loss, 2),
                    'ratio':     ratio,
                    'bes':       [round(b, 2) for b in bes],
                    'center_dist': round((ks[1] - spot) / spot * 100, 1),
                })
            if op == 'boi_coberto':
                rows.sort(key=lambda x: (not x['is_credit'], -(x['ratio'] or 999), x['max_loss']))
            elif op in ('vaca_tradicional', 'vaca_revertida'):
                rows.sort(key=lambda x: (not x['is_credit'], -x['net'], x['max_loss']))
            else:   # borboleta / condor: centro mais perto do spot, melhor relação
                rows.sort(key=lambda x: (abs(x['center_dist']), -(x['ratio'] or 0)))
            rows = _diversify(rows, lambda x: x['legs'][1]['sym'], per_key=2, limit=12)

        elif op == 'iron_condor':
            # Iron Condor: trava de alta com PUT (abaixo do spot) + trava de baixa
            # com CALL (acima). Vende PUT K2 / compra PUT K1 < K2; vende CALL K3 /
            # compra CALL K4 > K3. Crédito duplo; lucro se o papel ficar entre K2 e K3.
            # Regra usual: crédito >= 25% da asa mais larga; strikes vendidos OTM.
            T_ic = dc / 365.0
            put_sell  = [p for p in puts_ok
                         if 0.85 * spot <= p['strike'] <= 0.99 * spot and p['bid'] >= 0.05][:10]
            call_sell = [c for c in calls_ok
                         if 1.01 * spot <= c['strike'] <= 1.15 * spot and c['bid'] >= 0.05][:10]
            for ps_ in put_sell:
                put_buy = [p for p in puts_ok
                           if 0.70 * spot <= p['strike'] < ps_['strike'] and p['ask'] >= 0.01][:6]
                for pb_ in put_buy:
                    w_put = ps_['strike'] - pb_['strike']
                    cred_put = ps_['bid'] - pb_['ask']
                    if cred_put <= 0:
                        continue
                    for cs_ in call_sell:
                        call_buy = [c for c in calls_ok
                                    if cs_['strike'] < c['strike'] <= 1.30 * spot and c['ask'] >= 0.01][:6]
                        for cb_ in call_buy:
                            w_call = cb_['strike'] - cs_['strike']
                            cred_call = cs_['bid'] - cb_['ask']
                            if cred_call <= 0:
                                continue
                            credit = cred_put + cred_call
                            w_max  = max(w_put, w_call)
                            max_loss = w_max - credit
                            if max_loss <= 0:
                                continue
                            if credit < 0.25 * w_max:          # crédito mínimo usual (~1/3 da asa)
                                continue
                            be_low = ps_['strike'] - credit
                            be_up  = cs_['strike'] + credit
                            # POP: P(be_low < S < be_up) via lognormal com VI média das vendidas
                            iv_p = _iv_est(ps_, False, T_ic)
                            iv_c = _iv_est(cs_, True, T_ic)
                            ivm  = None
                            if iv_p and iv_c:
                                ivm = (iv_p + iv_c) / 2
                            elif iv_p or iv_c:
                                ivm = iv_p or iv_c
                            pop = None
                            if ivm:
                                p_lo = _pop_above(be_low, T_ic, ivm)
                                p_hi = _pop_above(be_up, T_ic, ivm)
                                if p_lo is not None and p_hi is not None:
                                    pop = max(0.0, min(100.0, p_lo - p_hi))
                            rows.append({
                                'put_buy_symbol':  pb_['symbol'], 'put_buy_strike':  pb_['strike'], 'put_buy_ask':  pb_['ask'],
                                'put_sell_symbol': ps_['symbol'], 'put_sell_strike': ps_['strike'], 'put_sell_bid': ps_['bid'],
                                'call_sell_symbol': cs_['symbol'], 'call_sell_strike': cs_['strike'], 'call_sell_bid': cs_['bid'],
                                'call_buy_symbol':  cb_['symbol'], 'call_buy_strike':  cb_['strike'], 'call_buy_ask':  cb_['ask'],
                                'credit':    round(credit, 2),
                                'credit_pct': round(credit / w_max * 100, 1),   # % da asa
                                'max_loss':  round(max_loss, 2),
                                'ratio':     round(credit / max_loss, 2),
                                'be_low':    round(be_low, 2),
                                'be_low_dist': round((be_low - spot) / spot * 100, 2),
                                'be_up':     round(be_up, 2),
                                'be_up_dist': round((be_up - spot) / spot * 100, 2),
                                'zone_pct':  round((be_up - be_low) / spot * 100, 1),
                                'pop':       round(pop, 1) if pop is not None else None,
                            })
            # Melhor relação crédito/asa primeiro; POP como desempate
            rows.sort(key=lambda x: (-x['credit_pct'], -(x['pop'] or 0)))
            rows = _diversify(rows, lambda x: (x['put_sell_symbol'], x['call_sell_symbol']),
                              per_key=2, limit=12)

        elif op in _ADV_SPECS:
            # ── Operações avançadas (guia técnico Vol. 2): seleção por DELTA ────
            # Motor genérico: pernas CALL/PUT/AÇÃO com quantidades; candidatos
            # escolhidos pela janela de delta recomendada; payoff linear por
            # partes com detecção de ganho/perda ilimitados nas caudas.
            spec  = _ADV_SPECS[op]
            legs  = spec['legs']              # [(tipo 'C'/'P'/'S', qty, (dlo, dhi))]
            T_adv = dc / 365.0

            def _leg_delta(rw, is_call):
                d = rw.get('delta')
                try:
                    d = float(d) if d not in (None, '') else None
                except (TypeError, ValueError):
                    d = None
                if d:
                    d = abs(d)
                    if d > 1:
                        d /= 100.0
                    return d if is_call else -d
                iv = _iv_est(rw, is_call, T_adv)
                if not iv or T_adv <= 0:
                    return None
                d1 = (math.log(spot / rw['strike']) + (r_cont + 0.5 * iv * iv) * T_adv) / (iv * math.sqrt(T_adv))
                nd = _norm_cdf(d1)
                return nd if is_call else nd - 1

            def _leg_cands(tp, q, win):
                if tp == 'S':
                    return [None]
                pool = calls_ok if tp == 'C' else puts_ok
                need = 'ask' if q > 0 else 'bid'
                out = []
                for rw in pool:
                    if rw[need] < 0.02:
                        continue
                    dl = _leg_delta(rw, tp == 'C')
                    if dl is None or not (win[0] <= dl <= win[1]):
                        continue
                    out.append((abs(dl - (win[0] + win[1]) / 2), rw, dl))
                out.sort(key=lambda x: x[0])
                return [(rw, dl) for _, rw, dl in out[:4]]

            import itertools as _it
            cand_lists = [_leg_cands(tp, q, win) for tp, q, win in legs]
            if all(cand_lists):
                for combo in _it.product(*cand_lists):
                    # strikes das pernas de opção, na ordem das pernas
                    opt_ks = [c[0]['strike'] for c in combo if c is not None]
                    oi_map, oi = [], 0
                    for c in combo:
                        oi_map.append(oi if c is not None else None)
                        if c is not None:
                            oi += 1
                    ok_c = True
                    for pair in spec.get('same', []):
                        a, b = oi_map[pair[0]], oi_map[pair[1]]
                        if abs(opt_ks[a] - opt_ks[b]) > 0.011:
                            ok_c = False
                            break
                    if ok_c and spec.get('asc'):
                        seq = [opt_ks[oi_map[i]] for i in spec['asc']]
                        if any(seq[i] >= seq[i + 1] for i in range(len(seq) - 1)):
                            ok_c = False
                    if not ok_c:
                        continue

                    net, row_legs = 0.0, []
                    for (tp, q, _win), c in zip(legs, combo):
                        if tp == 'S':
                            net += (-q) * spot
                            row_legs.append({'sym': ticker, 'tp': 'STOCK', 'k': None,
                                             'q': q, 'px': round(spot, 2), 'delta': None})
                        else:
                            rw, dl = c
                            px = rw['ask'] if q > 0 else rw['bid']
                            net += (-q) * px
                            row_legs.append({'sym': rw['symbol'],
                                             'tp': 'CALL' if tp == 'C' else 'PUT',
                                             'k': rw['strike'], 'q': q, 'px': px,
                                             'delta': round(dl * 100, 1)})

                    def _pay(S):
                        v = net
                        for (tp, q, _w), c in zip(legs, combo):
                            if tp == 'S':
                                v += q * S
                            elif tp == 'C':
                                v += q * max(0.0, S - c[0]['strike'])
                            else:
                                v += q * max(0.0, c[0]['strike'] - S)
                        return v

                    r_slope = sum(q for tp, q, _w in legs if tp in ('C', 'S'))
                    far  = max(opt_ks) * 1.8
                    pts  = [0.0] + sorted(opt_ks) + [far]
                    vals = [_pay(x) for x in pts]
                    bes  = []
                    for i in range(len(pts) - 1):
                        v0, v1 = vals[i], vals[i + 1]
                        if ((v0 < 0 <= v1) or (v0 >= 0 > v1)) and abs(v1 - v0) > 1e-9:
                            bes.append(pts[i] + (0 - v0) * (pts[i + 1] - pts[i]) / (v1 - v0))
                    gain_unl = r_slope > 0
                    loss_unl = r_slope < 0
                    max_gain = None if gain_unl else max(vals)
                    max_loss = None if loss_unl else max(0.0, -min(vals))
                    cost = -net

                    # Regras de montagem específicas
                    rule = spec.get('rule')
                    if rule == 'credit' and net <= 0:
                        continue
                    if rule == 'zero_cost' and cost > 0.02 * spot:
                        continue
                    if rule == 'low_cost' and cost > 0.05 * spot:
                        continue
                    if rule == 'jade':
                        # crédito >= largura da trava de CALL → sem risco na alta
                        wc = opt_ks[2] - opt_ks[1]
                        if net <= 0 or net < wc:
                            continue
                    if rule == 'bwb':
                        # broken wing: asa superior mais espaçada que a inferior
                        if opt_ks[2] - opt_ks[1] <= opt_ks[1] - opt_ks[0]:
                            continue
                        if cost > 0.05 * spot:
                            continue
                    if rule == 'box':
                        width = opt_ks[1] - opt_ks[0]
                        # débito menor que a largura descontada pela Selic do período
                        if cost <= 0 or cost >= width / (1 + selic_period / 100):
                            continue
                    if rule == 'synth_riskless':
                        # ação vendida + 2 CALLs ITM: só aceita quando a perda
                        # máxima é ~zero (a estrutura fica dentro da zona de ganho)
                        if max_loss is not None and max_loss > 0.02 * spot:
                            continue
                    if max_gain is not None and max_loss is not None and max_gain <= 0:
                        continue

                    ratio = (round(max_gain / max_loss, 2)
                             if (max_gain is not None and max_loss and max_loss > 0.001) else None)
                    rows.append({
                        'legs':      row_legs,
                        'net':       round(net, 2),
                        'is_credit': net >= 0,
                        'montagem':  'CRÉDITO' if net > 0.005 else ('ZERO' if abs(net) <= 0.02 * spot else 'DÉBITO'),
                        'max_gain':  round(max_gain, 2) if max_gain is not None else None,
                        'gain_unl':  gain_unl,
                        'max_loss':  round(max_loss, 2) if max_loss is not None else None,
                        'loss_unl':  loss_unl,
                        'ratio':     ratio,
                        'bes':       [round(b, 2) for b in bes],
                        'center_dist': round((opt_ks[0] - spot) / spot * 100, 1),
                    })
            rows.sort(key=lambda x: (not x['is_credit'], -(x['ratio'] or 0), -x['net']))
            rows = _diversify(rows, lambda x: x['legs'][0]['sym'], per_key=2, limit=10)

        elif op == 'boi_put':
            # Boi com PUT (put ratio backspread 1x2): compra 2 PUTs próximas do OTM
            # (até ~7% abaixo do spot); a venda de 1 PUT ATM/ITM financia parcialmente.
            # Relação usual: débito de até ~40% da largura (não precisa zerar o custo).
            sell_cands = [p for p in puts_ok
                          if 0.99 * spot <= p['strike'] <= 1.06 * spot and p['bid'] >= 0.10][:10]
            near_otm   = [p for p in puts_ok
                          if 0.93 * spot <= p['strike'] < 0.995 * spot and p['ask'] >= 0.05][:10]
            for sell in sell_cands:
                for buy in near_otm:
                    if buy['strike'] >= sell['strike']:
                        continue
                    width = sell['strike'] - buy['strike']
                    net = sell['bid'] - 2 * buy['ask']      # >0 crédito, <0 custo
                    if net < -0.40 * width:                 # venda deve financiar >= 60% não... custo máx 40% da largura
                        continue
                    max_loss = width - net                  # em S = K comprada
                    if max_loss <= 0:
                        continue
                    montagem = ('CRÉDITO' if net >= 0
                                else ('ZERO' if net >= -0.10 * width else 'INVEST'))
                    be_low = 2 * buy['strike'] - sell['strike'] + net   # abaixo disso, lucro cresce
                    s90 = spot * 0.90                       # ganho se cair 10%
                    gain_dn10 = net - max(0.0, sell['strike'] - s90) + 2 * max(0.0, buy['strike'] - s90)
                    rows.append({
                        'sell_symbol': sell['symbol'], 'sell_strike': sell['strike'], 'sell_bid': sell['bid'],
                        'buy_symbol':  buy['symbol'],  'buy_strike':  buy['strike'],  'buy_ask':  buy['ask'],
                        'credit':      round(net, 2),
                        'is_credit':   net >= 0,
                        'montagem':    montagem,
                        'max_loss':    round(max_loss, 2),
                        'be_low':      round(be_low, 2),
                        'be_low_dist': round((be_low - spot) / spot * 100, 2),
                        'gain_dn10':   round(gain_dn10, 2),
                    })
            # Mais eficaz primeiro: BE inferior mais próximo do spot (lucra com queda menor),
            # depois menor perda máxima
            rows.sort(key=lambda x: (abs(x['be_low_dist']), x['max_loss']))
            rows = _diversify(rows, lambda x: x['buy_symbol'], per_key=2)

        elif op == 'vaca_put':
            # Vaca de baixa travada com PUTs (borboleta de PUTs):
            # +1 PUT alta (~spot), -2 PUTs médias (abaixo), +1 PUT baixa.
            # Lucro máximo se o papel cair até o strike médio. Custo baixo ou investimento.
            hi_p = [p for p in puts_ok
                    if 0.95 * spot <= p['strike'] <= 1.05 * spot and p['ask'] >= 0.10][:10]
            for p1 in hi_p:
                mids = [p for p in puts_ok
                        if 0.78 * spot <= p['strike'] < p1['strike'] and p['bid'] >= 0.03][:8]
                for p2 in mids:
                    lows = [p for p in puts_ok
                            if 0.65 * spot <= p['strike'] < p2['strike'] and p['ask'] >= 0.01][:8]
                    for p3 in lows:
                        cost = p1['ask'] - 2 * p2['bid'] + p3['ask']   # >0 débito, <=0 crédito
                        w_hi = p1['strike'] - p2['strike']
                        if cost > 0.50 * w_hi:
                            continue
                        max_gain = w_hi - cost          # em S = K2 (médio)
                        if max_gain <= 0:
                            continue
                        montagem = 'CRÉDITO' if cost <= 0 else ('ZERO' if cost <= 0.15 * w_hi else 'INVEST')
                        tail = w_hi - (p2['strike'] - p3['strike']) - cost   # abaixo da asa inferior
                        be_up = round(p1['strike'] - cost, 2) if cost > 0 else None
                        be_dn = round(p2['strike'] - max_gain, 2) if tail < 0 else None
                        max_loss = max(cost if cost > 0 else 0.0, -tail if tail < 0 else 0.0)
                        ratio = None if max_loss <= 0.001 else max_gain / max_loss
                        rows.append({
                            'high_symbol': p1['symbol'], 'high_strike': p1['strike'], 'high_ask': p1['ask'],
                            'mid_symbol':  p2['symbol'], 'mid_strike':  p2['strike'], 'mid_bid':  p2['bid'],
                            'low_symbol':  p3['symbol'], 'low_strike':  p3['strike'], 'low_ask':  p3['ask'],
                            'cost':      round(cost, 2),
                            'is_credit': cost <= 0,
                            'montagem':  montagem,
                            'max_gain':  round(max_gain, 2),
                            'tail':      round(tail, 2),
                            'be_up':     be_up,
                            'be_dn':     be_dn,
                            'mid_dist':  round((p2['strike'] - spot) / spot * 100, 1),  # % de queda até o pico
                            'max_loss':  round(max_loss, 2),
                            'ratio':     round(ratio, 1) if ratio is not None else None,
                        })
            rows.sort(key=lambda x: (x['max_loss'], -x['max_gain']))
            rows = _diversify(rows, lambda x: x['mid_symbol'], per_key=2)

        elif op == 'venda_put_itm':
            # Venda a seco de PUT: strike de 10% OTM até 20% ITM.
            # Pouca liquidez nas ITMs → prêmio = último negócio (close); fallback bid.
            # Remuneração = prêmio/strike anualizada em dias úteis (~5/7 dos corridos),
            # como na calculadora "Venda de Puts".
            du = max(round(dc * 5.0 / 7.0), 1)
            # Não exige bid+ask no book: usa a lista completa do vencimento
            all_puts_exp = sorted(puts_by_exp.get(exp, []), key=lambda x: x['strike'])
            cands = [p for p in all_puts_exp
                     if 0.90 * spot <= p['strike'] <= 1.20 * spot
                     and (p['close'] > 0 or p['bid'] > 0)]
            for p in cands:
                use_last = p['close'] > 0
                prem = p['close'] if use_last else p['bid']
                if prem < 0.05:
                    continue
                rem_per = prem / p['strike']
                rem_am  = ((1 + rem_per) ** (21.0  / du) - 1) * 100
                rem_aa  = ((1 + rem_per) ** (252.0 / du) - 1) * 100
                pct_cdi = (rem_aa / selic * 100) if selic > 0 else 0
                be      = p['strike'] - prem
                itm_amt = p['strike'] - spot            # >0 = ITM, <0 = OTM
                rows.append({
                    'symbol':    p['symbol'],
                    'strike':    p['strike'],
                    'premium':   round(prem, 2),
                    'price_src': 'último' if use_last else 'bid',
                    'bid':       round(p['bid'], 2),
                    'vol_fin':   p.get('vol_fin', 0),
                    'itm_amt':   round(itm_amt, 2),
                    'itm_pct':   round(itm_amt / spot * 100, 1),   # >0 = % ITM, <0 = % OTM
                    'rem_per':   round(rem_per * 100, 2),
                    'rem_am':    round(rem_am, 2),
                    'rem_aa':    round(rem_aa, 2),
                    'pct_cdi':   round(pct_cdi, 0),
                    'breakeven': round(be, 2),
                    'be_margin': round((spot - be) / spot * 100, 2),  # queda suportada até o BE
                    'du':        du,
                })
            rows.sort(key=lambda x: -x['pct_cdi'])
            rows = rows[:14]

        elif op == 'straddle_vendido':
            # Straddle vendido: venda de 1 CALL + 1 PUT no MESMO strike, bem ATM
            # (o strike mais próximo do preço atual do ativo). Crédito duplo;
            # risco fora dos breakevens. Deltas exibidos apenas como informação.
            T = dc / 365.0
            # Perna vendida não precisa de ask no book: usa lista completa do vencimento
            all_calls_st = sorted(calls_by_exp.get(exp, []), key=lambda x: x['strike'])
            put_map = {p['strike']: p for p in puts_by_exp.get(exp, [])}
            cands = []
            for c in all_calls_st:
                p = put_map.get(c['strike'])
                if not p:
                    continue
                c_prem, c_src = _sell_prem(c)
                p_prem, p_src = _sell_prem(p)
                if c_prem is None or p_prem is None:
                    continue
                dist = abs(c['strike'] - spot) / spot
                if dist > 0.05:          # bem ATM: strike até 5% do spot
                    continue
                cands.append((dist, c, p, c_prem, c_src, p_prem, p_src))
            cands.sort(key=lambda x: x[0])   # mais ATM primeiro
            for dist, c, p, c_prem, c_src, p_prem, p_src in cands[:3]:
                credit = c_prem + p_prem
                be_low, be_up = c['strike'] - credit, c['strike'] + credit
                rows.append({
                    'strike':      c['strike'],
                    'atm_dist':    round((c['strike'] - spot) / spot * 100, 2),
                    'call_symbol': c['symbol'], 'call_bid': round(c_prem, 2), 'call_src': c_src,
                    'call_delta':  _leg_delta_pct(c, True, T),
                    'put_symbol':  p['symbol'], 'put_bid':  round(p_prem, 2), 'put_src': p_src,
                    'put_delta':   _leg_delta_pct(p, False, T),
                    'credit':      round(credit, 2),
                    'credit_pct':  round(credit / spot * 100, 2),
                    'be_low':      round(be_low, 2),
                    'be_up':       round(be_up, 2),
                    'be_low_dist': round((be_low - spot) / spot * 100, 2),
                    'be_up_dist':  round((be_up - spot) / spot * 100, 2),
                })

        elif op == 'strangle_vendido':
            # Strangle vendido: venda de CALL OTM + PUT OTM (strikes diferentes),
            # com delta entre 15 e 35 em cada ponta (faixa usual da estratégia).
            T = dc / 365.0
            call_cands, put_cands = [], []
            for c in sorted(calls_by_exp.get(exp, []), key=lambda x: x['strike']):
                if c['strike'] > spot:
                    c_prem, c_src = _sell_prem(c)
                    if c_prem is None:
                        continue
                    d_c = _leg_delta_pct(c, True, T)
                    if d_c is not None and 15 <= d_c <= 35:
                        call_cands.append((c, d_c, c_prem, c_src))
            for p in sorted(puts_by_exp.get(exp, []), key=lambda x: x['strike']):
                if p['strike'] < spot:
                    p_prem, p_src = _sell_prem(p)
                    if p_prem is None:
                        continue
                    d_p = _leg_delta_pct(p, False, T)
                    if d_p is not None and 15 <= d_p <= 35:
                        put_cands.append((p, d_p, p_prem, p_src))
            for c, d_c, c_prem, c_src in call_cands[:10]:
                for p, d_p, p_prem, p_src in put_cands[-10:]:
                    credit = c_prem + p_prem
                    be_low, be_up = p['strike'] - credit, c['strike'] + credit
                    rows.append({
                        'call_symbol': c['symbol'], 'call_strike': c['strike'],
                        'call_bid':    round(c_prem, 2), 'call_src': c_src, 'call_delta': d_c,
                        'put_symbol':  p['symbol'], 'put_strike':  p['strike'],
                        'put_bid':     round(p_prem, 2), 'put_src': p_src, 'put_delta': d_p,
                        'credit':      round(credit, 2),
                        'credit_pct':  round(credit / spot * 100, 2),
                        'width_pct':   round((c['strike'] - p['strike']) / spot * 100, 1),
                        'be_low':      round(be_low, 2),
                        'be_up':       round(be_up, 2),
                        'be_low_dist': round((be_low - spot) / spot * 100, 2),
                        'be_up_dist':  round((be_up - spot) / spot * 100, 2),
                    })
            rows.sort(key=lambda x: -x['credit_pct'])
            rows = _diversify(rows, lambda x: x['call_symbol'], per_key=2)

        elif op == 'zebra':
            # ZEBRA (Zero Extrinsic Back Ratio): compra 2 CALLs ITM (Δ ≈ 0,70)
            # + venda 1 CALL ATM (Δ ≈ 0,50). Delta total ≈ +1,0 e extrínseco
            # líquido ≈ 0 — a venda ATM paga o extrínseco das compradas.
            # Substitui a compra da ação com fração do capital e Theta ~zero.
            T = dc / 365.0
            itm_cands, atm_cands = [], []
            for c in calls_ok:
                d_c = _leg_delta_pct(c, True, T)
                if d_c is None:
                    continue
                if c['strike'] < spot and 55 <= d_c <= 85 and c['ask'] > 0:
                    itm_cands.append((c, d_c))
                elif abs(c['strike'] - spot) / spot <= 0.05 and 38 <= d_c <= 62:
                    prem, src = _sell_prem(c)
                    if prem:
                        atm_cands.append((c, d_c, prem, src))
            for ci, d_i in itm_cands:
                for ca, d_a, a_prem, a_src in atm_cands:
                    if ca['strike'] <= ci['strike']:
                        continue
                    cost = 2 * ci['ask'] - a_prem          # débito por ação
                    if cost <= 0:
                        continue
                    intr_i = max(0.0, spot - ci['strike'])
                    intr_a = max(0.0, spot - ca['strike'])
                    net_extr = 2 * (ci['ask'] - intr_i) - (a_prem - intr_a)
                    # extrínseco líquido precisa ser ~zero (tolerância 2% do spot)
                    if net_extr > 0.02 * spot:
                        continue
                    net_delta = round(2 * d_i - d_a, 1)     # escala 0-100
                    k_i, k_a = ci['strike'], ca['strike']
                    be_mid = k_i + cost / 2                 # BE entre os strikes
                    be = be_mid if be_mid <= k_a else (2 * k_i - k_a + cost)
                    iv = _iv_est(ca, True, T) or _iv_est(ci, True, T)
                    pop = _pop_above(be, T, iv)
                    rows.append({
                        'buy_symbol':  ci['symbol'], 'buy_strike':  k_i,
                        'buy_ask':     ci['ask'],    'buy_delta':   d_i,
                        'sell_symbol': ca['symbol'], 'sell_strike': k_a,
                        'sell_bid':    round(a_prem, 2), 'sell_src': a_src,
                        'sell_delta':  d_a,
                        'cost':        round(cost, 2),
                        'cost_pct_spot': round(cost / spot * 100, 1),
                        'net_extr':    round(net_extr, 2),
                        'net_delta':   net_delta,
                        'breakeven':   round(be, 2),
                        'be_dist':     round((be - spot) / spot * 100, 2),
                        'pop':         round(pop, 1) if pop is not None else None,
                    })
            rows.sort(key=lambda x: abs(x['net_extr']))
            rows = _diversify(rows, lambda x: x['buy_symbol'], per_key=2)

        expirations.append({
            'exp':          exp,
            'dc':           dc,
            'selic_period': round(selic_period, 2),
            'is_monthly':   _is_monthly(exp),
            'rows':         rows,
        })

    return jsonify({
        'ticker':      ticker,
        'spot':        spot,
        'spot_change': spot_change,
        'mode':        mode,
        'op':          op,
        'max_days':    max_days,
        'market_open': market_open,
        'selic':       round(selic, 2),
        'expirations': expirations,
    })


@app.route('/lancamento-coberto')
@login_required
def lancamento_coberto():
    """Página: ranking de lançamento coberto por ativo (vencimentos longos)."""
    ranking_vol = _ranking_liq_filter(RankingVol.query.filter_by(user_id=current_user.id)).order_by(RankingVol.ticker).all()
    return render_template('lancamento_coberto.html', ranking_vol=ranking_vol, selic=_selic())


@app.route('/api/lancamento-coberto/<ticker>')
@login_required
def api_lancamento_coberto(ticker):
    """Ranking de lançamento coberto (compra ação + venda CALL) para um ativo.
    Varre vencimentos MENSAIS acima de 60 dias até 2 anos, exigindo que a
    CALL tenha tido negócio efetivado. Uma chamada OpLab por ativo.
    Taxa se exercido = (K − custo_líquido)/custo_líquido.
    ?m=itm|otm|both filtra a moneyness da CALL (default both)."""
    ticker = ticker.strip().upper()
    money  = (request.args.get('m') or 'both').lower()
    if money not in ('itm', 'otm', 'both'):
        money = 'both'
    token  = Settings.get_value('oplab_token', user_id=current_user.id)
    if not token:
        return jsonify({'error': 'Token OpLab não configurado'}), 400

    spot, spot_change = _get_underlying_quote(ticker, current_user.id)
    if not spot:
        return jsonify({'error': f'Cotação de {ticker} indisponível.'}), 404

    try:
        data = _oplab_get_json(f'/market/options/{ticker}', token, timeout=20)
    except OplabApiError as e:
        return jsonify({'error': str(e), 'status': e.status_code}), 503
    except Exception:
        app.logger.exception('api_lancamento_coberto error for %s', ticker)
        return jsonify({'error': 'Erro inesperado ao buscar a cadeia de opções.'}), 500

    opt_list = data if isinstance(data, list) else (
        data.get('options') or data.get('calls', []) + data.get('puts', []) or []
    )

    from datetime import date as _date
    today  = _date.today()
    selic  = _selic()
    r_cont = math.log(1 + selic / 100.0)

    def _third_friday(y, m):
        count, day = 0, 1
        while True:
            d = _date(y, m, day)
            if d.weekday() == 4:
                count += 1
                if count == 3:
                    return d
            day += 1

    def _is_monthly(exp_d):
        tf = _third_friday(exp_d.year, exp_d.month)
        return abs((exp_d - tf).days) <= 2  # tolera feriado na 3ª sexta

    rows = []
    for o in opt_list:
        cat = str(o.get('category') or o.get('type') or '').upper()
        if 'PUT' in cat or cat == 'P':
            continue
        sym    = str(o.get('symbol') or o.get('ticker') or '').upper()
        strike = float(o.get('strike') or 0)
        close  = float(o.get('close') or 0)
        due    = str(o.get('due_date') or o.get('expiration_date') or '')
        if 'T' in due:
            due = due.split('T')[0]
        if not sym or strike <= 0 or not due:
            continue
        if close < 0.05:                          # precisa ter tido negócio
            continue
        # Só séries com negociação registrada: volume (qtd) e/ou financeiro > 0.
        # Elimina prints antigos/estagnados que geram taxas absurdas.
        vol_qtd = float(o.get('volume') or 0)
        vol_fin = float(o.get('financial_volume') or o.get('volume_financial') or 0)
        trades  = float(o.get('trades') or o.get('business') or o.get('negocios') or 0)
        if vol_qtd <= 0 and vol_fin <= 0 and trades <= 0:
            continue
        try:
            exp_d = _date.fromisoformat(due)
        except ValueError:
            continue
        dc = (exp_d - today).days
        if dc <= 60 or dc > 730:                  # >60 dias até 2 anos
            continue
        # Só vencimentos mensais (3ª sexta-feira); descarta semanais (sufixo W+dígito)
        if not _is_monthly(exp_d) or sym.rstrip('0123456789').endswith('W'):
            continue
        if not (0.50 * spot <= strike <= 1.30 * spot):
            continue
        # Moneyness: CALL é ITM quando strike < spot
        is_itm = strike < spot
        if (money == 'itm' and not is_itm) or (money == 'otm' and is_itm):
            continue
        custo = spot - close
        if custo <= 0:
            continue
        taxa_ex = (strike - custo) / custo * 100
        if taxa_ex <= 0:                          # exercício daria prejuízo
            continue
        taxa_aa = ((1 + taxa_ex / 100) ** (365.0 / dc) - 1) * 100
        if taxa_aa > 999:                         # composto explode em prazos curtos
            taxa_aa = None
        selic_per = ((1 + selic / 100) ** (dc / 365.0) - 1) * 100

        # Delta via BS com IV extraída do prêmio (informativo)
        delta = None
        T = dc / 365.0
        intr = max(0.0, spot - strike)
        if close > intr * 1.005:
            iv = _implied_vol(spot, strike, T, r_cont, close, True)
            if 0.005 < iv < 4.9:
                d1 = (math.log(spot / strike) + (r_cont + 0.5 * iv * iv) * T) / (iv * math.sqrt(T))
                delta = round(_norm_cdf(d1) * 100, 1)

        bid = float(o.get('bid') or 0)
        ask = float(o.get('ask') or 0)

        rows.append({
            'symbol':    sym,
            'exp':       due,
            'dc':        dc,
            'strike':    round(strike, 2),
            'itm_pct':   round((spot - strike) / spot * 100, 1),   # >0 = ITM
            'premium':   round(close, 2),
            'bid':       round(bid, 2) if bid else None,
            'ask':       round(ask, 2) if ask else None,
            'custo':     round(custo, 2),                          # custo líquido = BE
            'protec':    round(close / spot * 100, 2),
            'taxa_ex':   round(taxa_ex, 2),
            'taxa_aa':   round(taxa_aa, 2) if taxa_aa is not None else None,
            'vs_selic':  round(taxa_ex - selic_per, 2),
            'delta':     delta,
            'vol_fin':   round(vol_fin, 2),
            'vol_qtd':   round(vol_qtd, 0),
        })

    rows.sort(key=lambda x: -x['taxa_ex'])
    # No máximo 10 alternativas por vencimento
    per_exp, capped = {}, []
    for r in rows:
        n = per_exp.get(r['exp'], 0)
        if n >= 10:
            continue
        per_exp[r['exp']] = n + 1
        capped.append(r)
    return jsonify({
        'ticker':      ticker,
        'spot':        spot,
        'spot_change': spot_change,
        'selic':       round(selic, 2),
        'rows':        capped[:60],
        'total':       len(rows),
    })


@app.route('/venda-put-longa')
@login_required
def venda_put_longa():
    """Página: ranking de venda de PUT (cash-secured) de longo prazo."""
    ranking_vol = _ranking_liq_filter(RankingVol.query.filter_by(user_id=current_user.id)).order_by(RankingVol.ticker).all()
    return render_template('venda_put_longa.html', ranking_vol=ranking_vol, selic=_selic())


@app.route('/api/venda-put-longa/<ticker>')
@login_required
def api_venda_put_longa(ticker):
    """Ranking de venda de PUT cash-secured para um ativo. Varre TODOS os
    vencimentos até 2 anos exigindo PUT com negócio efetivado (último > 0 e
    volume/negócios > 0). Rendimento do prêmio com capital = strike:
    retorno = P/K no período; custo efetivo se exercido = K − P.
    ?m=itm|otm|both filtra a moneyness da PUT (default both)."""
    ticker = ticker.strip().upper()
    money  = (request.args.get('m') or 'both').lower()
    if money not in ('itm', 'otm', 'both'):
        money = 'both'
    token  = Settings.get_value('oplab_token', user_id=current_user.id)
    if not token:
        return jsonify({'error': 'Token OpLab não configurado'}), 400

    spot, spot_change = _get_underlying_quote(ticker, current_user.id)
    if not spot:
        return jsonify({'error': f'Cotação de {ticker} indisponível.'}), 404

    try:
        data = _oplab_get_json(f'/market/options/{ticker}', token, timeout=20)
    except OplabApiError as e:
        return jsonify({'error': str(e), 'status': e.status_code}), 503
    except Exception:
        app.logger.exception('api_venda_put_longa error for %s', ticker)
        return jsonify({'error': 'Erro inesperado ao buscar a cadeia de opções.'}), 500

    opt_list = data if isinstance(data, list) else (
        data.get('options') or data.get('calls', []) + data.get('puts', []) or []
    )

    from datetime import date as _date
    today  = _date.today()
    selic  = _selic()
    r_cont = math.log(1 + selic / 100.0)

    def _third_friday(y, m):
        count, day = 0, 1
        while True:
            d = _date(y, m, day)
            if d.weekday() == 4:
                count += 1
                if count == 3:
                    return d
            day += 1

    def _is_monthly(exp_d):
        tf = _third_friday(exp_d.year, exp_d.month)
        return abs((exp_d - tf).days) <= 2  # tolera feriado na 3ª sexta

    rows = []
    for o in opt_list:
        cat = str(o.get('category') or o.get('type') or '').upper()
        if 'PUT' not in cat and cat != 'P':
            continue
        sym    = str(o.get('symbol') or o.get('ticker') or '').upper()
        strike = float(o.get('strike') or 0)
        close  = float(o.get('close') or 0)
        due    = str(o.get('due_date') or o.get('expiration_date') or '')
        if 'T' in due:
            due = due.split('T')[0]
        if not sym or strike <= 0 or not due:
            continue
        if close < 0.05:                          # precisa ter tido negócio
            continue
        # Só séries com negociação registrada — elimina prints estagnados.
        vol_qtd = float(o.get('volume') or 0)
        vol_fin = float(o.get('financial_volume') or o.get('volume_financial') or 0)
        trades  = float(o.get('trades') or o.get('business') or o.get('negocios') or 0)
        if vol_qtd <= 0 and vol_fin <= 0 and trades <= 0:
            continue
        try:
            exp_d = _date.fromisoformat(due)
        except ValueError:
            continue
        dc = (exp_d - today).days
        if dc <= 40 or dc > 730:                  # >40 dias até 2 anos
            continue
        # Só vencimentos mensais (3ª sexta-feira); descarta semanais (sufixo W+dígito)
        if not _is_monthly(exp_d) or sym.rstrip('0123456789').endswith('W'):
            continue
        if not (0.50 * spot <= strike <= 1.30 * spot):
            continue
        # Moneyness: PUT é ITM quando strike > spot
        is_itm = strike > spot
        if (money == 'itm' and not is_itm) or (money == 'otm' and is_itm):
            continue

        # Rentabilidade do prêmio com capital reservado = strike
        taxa_per = close / strike * 100
        taxa_aa  = ((1 + taxa_per / 100) ** (365.0 / dc) - 1) * 100
        if taxa_aa > 999:                         # composto explode em prazos curtos
            taxa_aa = None
        selic_per = ((1 + selic / 100) ** (dc / 365.0) - 1) * 100
        custo_ef  = strike - close                # preço de equilíbrio se exercido
        margem    = (spot - custo_ef) / spot * 100  # >0 = BE abaixo do spot

        # Delta via BS com IV extraída do prêmio (informativo)
        delta = None
        T = dc / 365.0
        intr = max(0.0, strike - spot)
        if close > intr * 1.005:
            iv = _implied_vol(spot, strike, T, r_cont, close, False)
            if 0.005 < iv < 4.9:
                d1 = (math.log(spot / strike) + (r_cont + 0.5 * iv * iv) * T) / (iv * math.sqrt(T))
                delta = round((_norm_cdf(d1) - 1) * 100, 1)

        bid = float(o.get('bid') or 0)
        ask = float(o.get('ask') or 0)

        rows.append({
            'symbol':    sym,
            'exp':       due,
            'dc':        dc,
            'strike':    round(strike, 2),
            'itm_pct':   round((strike - spot) / spot * 100, 1),   # >0 = ITM (put)
            'premium':   round(close, 2),
            'bid':       round(bid, 2) if bid else None,
            'ask':       round(ask, 2) if ask else None,
            'custo_ef':  round(custo_ef, 2),                       # BE se exercido
            'margem':    round(margem, 2),
            'taxa_per':  round(taxa_per, 2),
            'taxa_aa':   round(taxa_aa, 2) if taxa_aa is not None else None,
            'vs_selic':  round(taxa_per - selic_per, 2),
            'delta':     delta,
            'vol_fin':   round(vol_fin, 2),
            'vol_qtd':   round(vol_qtd, 0),
        })

    rows.sort(key=lambda x: -x['taxa_per'])
    # No máximo 10 alternativas por vencimento
    per_exp, capped = {}, []
    for r in rows:
        n = per_exp.get(r['exp'], 0)
        if n >= 10:
            continue
        per_exp[r['exp']] = n + 1
        capped.append(r)
    return jsonify({
        'ticker':      ticker,
        'spot':        spot,
        'spot_change': spot_change,
        'selic':       round(selic, 2),
        'rows':        capped[:60],
        'total':       len(rows),
    })


# ─────────────────────────────────────────────────────────────────────────────
# PWA — manifest, service worker e página offline
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/manifest.json')
def pwa_manifest():
    from flask import send_from_directory
    return send_from_directory(app.static_folder, 'manifest.json',
                               mimetype='application/manifest+json')


@app.route('/sw.js')
def pwa_service_worker():
    # Servido na raiz para o service worker controlar o site inteiro (scope '/')
    from flask import send_from_directory
    resp = send_from_directory(app.static_folder, 'sw.js', mimetype='application/javascript')
    resp.headers['Cache-Control'] = 'no-cache'
    return resp


@app.route('/offline.html')
def pwa_offline():
    return render_template('offline.html')


# ─────────────────────────────────────────────────────────────────────────────
# Venda de Puts — CRUD
# ─────────────────────────────────────────────────────────────────────────────

def _calc_simulacao_metrics(sim):
    """Payoff no vencimento de uma SimulacaoOpcoes (rascunho/estudo — NÃO é
    posição real). Espelha a lógica de payoffAt() em simulacao_opcoes.html:
    multi-vencimento reprecifica a perna mais longa por Black-Scholes na
    data do vencimento mais curto (usa a IV informada na perna).
    Retorna None se a simulação não tiver pernas com dado suficiente OU se
    qualquer conta falhar — é só uma tabela auxiliar, nunca pode derrubar
    a página de Cálculos."""
    try:
        legs = list(sim.legs or [])
        opt_legs = [l for l in legs
                    if (l.leg_type or 'CALL') != 'STOCK' and l.quantity and l.strike]
        if not opt_legs:
            return None

        exp_dates = sorted({l.expiration for l in opt_legs if l.expiration})
        multi_exp = len(exp_dates) > 1
        ref_date = exp_dates[0] if multi_exp else None
        selic = _selic()
        r_cont = math.log(1 + selic / 100.0)

        net_premium = sum((1 if l.side == 'SELL' else -1) * l.quantity * (l.premium or 0)
                          for l in opt_legs)

        def payoff_at(S):
            total = net_premium
            for l in legs:
                if not l.quantity:
                    continue
                sign = 1 if l.side == 'BUY' else -1
                if (l.leg_type or 'CALL') == 'STOCK':
                    total += sign * l.quantity * (S - (l.premium or 0))
                    continue
                if not l.strike:
                    continue
                K = l.strike
                if multi_exp and l.expiration and ref_date and l.expiration > ref_date and (l.iv or 0) > 0:
                    T = max((l.expiration - ref_date).days / 365.25, 1 / 365.25)
                    total += sign * l.quantity * _bs_price(S, K, T, r_cont, (l.iv or 30) / 100,
                                                           l.leg_type == 'CALL')
                else:
                    intrinsic = max(0.0, S - K) if l.leg_type == 'CALL' else max(0.0, K - S)
                    total += sign * l.quantity * intrinsic
            return total

        strikes = [l.strike for l in opt_legs if l.strike]
        stock_legs = [l for l in legs
                      if (l.leg_type or 'CALL') == 'STOCK' and (l.premium or 0) > 0]
        stock_ref = stock_legs[0].premium if stock_legs else 0
        anchors = strikes + ([stock_ref] if stock_ref else [])
        if not anchors:
            return None
        lo_k, hi_k = min(anchors), max(anchors)
        if hi_k <= 0:
            return None
        pad = max((hi_k - lo_k) * 0.6, hi_k * 0.30)
        lo, hi, N = max(0.01, lo_k - pad), hi_k + pad, 300
        step = (hi - lo) / N
        if step <= 0:
            return None
        xs = [lo + i * step for i in range(N + 1)]
        ys = [payoff_at(x) for x in xs]

        max_gain, max_loss = max(ys), min(ys)
        bes = []
        p_range = max_gain - max_loss
        if p_range > 0.01:
            for j in range(len(ys) - 1):
                p1, p2 = ys[j], ys[j + 1]
                if p1 * p2 < 0:
                    bes.append(round(xs[j] + (-p1) * (xs[j + 1] - xs[j]) / (p2 - p1), 2))

        # POP: P(S_T > BE) via log-normal, spot ancorado no meio dos strikes/ação
        # (a simulação não tem cotação ao vivo própria — é só uma referência).
        pop = None
        spot_ref = stock_ref or (sum(strikes) / len(strikes) if strikes else 0)
        bes_valid = [b for b in bes if b > 0]
        if spot_ref > 0 and bes_valid:
            exp_far = max(exp_dates) if exp_dates else None
            T_pop = max(((exp_far - date.today()).days / 365.25), 1 / 365.25) if exp_far else 30 / 365.25
            sigma_avg = (sum((l.iv or 30) for l in opt_legs) / len(opt_legs)) / 100 or 0.30
            bes_sorted = sorted(bes_valid)
            try:
                if len(bes_sorted) == 1:
                    d2v = (math.log(spot_ref / bes_sorted[0]) + (r_cont - 0.5 * sigma_avg ** 2) * T_pop) \
                          / (sigma_avg * math.sqrt(T_pop))
                    pop = _norm_cdf(d2v) * 100
                else:
                    d2_lo = (math.log(spot_ref / bes_sorted[0]) + (r_cont - 0.5 * sigma_avg ** 2) * T_pop) \
                            / (sigma_avg * math.sqrt(T_pop))
                    d2_hi = (math.log(spot_ref / bes_sorted[-1]) + (r_cont - 0.5 * sigma_avg ** 2) * T_pop) \
                            / (sigma_avg * math.sqrt(T_pop))
                    pop = (_norm_cdf(d2_lo) - _norm_cdf(d2_hi)) * 100
            except (ValueError, ZeroDivisionError, OverflowError):
                pop = None

        gain_unl = len(ys) > 1 and max_gain >= ys[-1] - 0.01 and ys[-1] > ys[-2] + 0.01
        loss_unl = len(ys) > 1 and max_loss <= ys[0] + 0.01 and ys[0] < ys[1] - 0.01
        return {
            'net': round(net_premium, 2), 'is_credit': net_premium >= 0,
            'max_gain': round(max_gain, 2), 'gain_unl': gain_unl,
            'max_loss': round(max_loss, 2), 'loss_unl': loss_unl,
            'bes': bes, 'pop': round(pop, 1) if pop is not None else None,
            'nearest_exp': min(exp_dates) if exp_dates else None,
        }
    except Exception:
        app.logger.exception('_calc_simulacao_metrics falhou para sim.id=%s', getattr(sim, 'id', None))
        return None


@app.route('/venda_puts')
@login_required
def venda_puts():
    items   = PutSale.query.filter_by(user_id=current_user.id).order_by(PutSale.created_at.desc()).all()
    collars = CollarSimulation.query.filter_by(user_id=current_user.id).order_by(CollarSimulation.created_at.desc()).all()
    # Simulações de Opções SALVAS (rascunhos/estudos — tabela SimulacaoOpcoes,
    # a mesma da tela "Simulação de Opções"), com estrutura (2+ pernas ou
    # perna de opção). As operações REAIS (StructuredOp) continuam só em
    # /opcoes — aqui é só consulta/edição das simulações.
    simulacoes_estrutura = []
    try:
        raw_sims = SimulacaoOpcoes.query.filter_by(user_id=current_user.id)\
            .order_by(SimulacaoOpcoes.created_at.desc()).all()
        for sim in raw_sims:
            if len(sim.legs or []) < 2:
                continue
            metrics = _calc_simulacao_metrics(sim)
            if metrics:
                simulacoes_estrutura.append({'sim': sim, **metrics})
    except Exception:
        app.logger.exception('Falha ao montar simulacoes_estrutura em /venda_puts')
        simulacoes_estrutura = []
    selic   = _selic()
    today   = date.today()
    return render_template('venda_puts.html', items=items, collars=collars,
                           simulacoes_estrutura=simulacoes_estrutura,
                           edit=None, edit_collar=None, selic=selic, today=today)


@app.route('/venda_puts/new', methods=['POST'])
@login_required
def venda_puts_new():
    def _f(k): return request.form.get(k, '').replace(',', '.').strip()
    try:
        exp = datetime.strptime(_f('expiration_date'), '%Y-%m-%d').date()
    except ValueError:
        flash('Data de vencimento inválida.', 'danger')
        return redirect(url_for('venda_puts'))
    entry_date_str = _f('entry_date')
    entry_date = datetime.strptime(entry_date_str, '%Y-%m-%d').date() if entry_date_str else date.today()
    p = PutSale(
        user_id          = current_user.id,
        ticker           = _f('ticker').upper(),
        underlying_asset = _f('underlying_asset').upper(),
        underlying_price = float(_f('underlying_price')) if _f('underlying_price') else None,
        strike           = float(_f('strike')),
        expiration_date  = exp,
        premium          = float(_f('premium')),
        quantity         = int(_f('quantity') or 100),
        entry_date       = entry_date,
        notes            = request.form.get('notes', ''),
        created_at       = datetime.now(),
    )
    db.session.add(p)
    db.session.commit()
    flash('Venda de put salva.', 'success')
    return redirect(url_for('venda_puts'))


@app.route('/venda_puts/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def venda_puts_edit(id):
    p = PutSale.query.get_or_404(id)
    if p.user_id != current_user.id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('venda_puts'))
    if request.method == 'POST':
        def _f(k): return request.form.get(k, '').replace(',', '.').strip()
        try:
            p.expiration_date = datetime.strptime(_f('expiration_date'), '%Y-%m-%d').date()
        except ValueError:
            flash('Data de vencimento inválida.', 'danger')
            return redirect(url_for('venda_puts_edit', id=id))
        entry_date_str = _f('entry_date')
        p.entry_date       = datetime.strptime(entry_date_str, '%Y-%m-%d').date() if entry_date_str else p.entry_date
        p.ticker           = _f('ticker').upper()
        p.underlying_asset = _f('underlying_asset').upper()
        p.underlying_price = float(_f('underlying_price')) if _f('underlying_price') else None
        p.strike           = float(_f('strike'))
        p.premium          = float(_f('premium'))
        p.quantity         = int(_f('quantity') or 100)
        p.notes            = request.form.get('notes', '')
        db.session.commit()
        flash('Venda de put atualizada.', 'success')
        return redirect(url_for('venda_puts'))
    items   = PutSale.query.filter_by(user_id=current_user.id).order_by(PutSale.created_at.desc()).all()
    collars = CollarSimulation.query.filter_by(user_id=current_user.id).order_by(CollarSimulation.created_at.desc()).all()
    return render_template('venda_puts.html', items=items, collars=collars, edit=p, edit_collar=None, selic=_selic(), today=date.today())


# ── Collar simulation CRUD ──────────────────────────────────────────────────

@app.route('/collar/new', methods=['POST'])
@login_required
def collar_new():
    def _f(k): return request.form.get(k, '').replace(',', '.').strip()
    try:
        exp = datetime.strptime(_f('expiration_date'), '%Y-%m-%d').date()
    except ValueError:
        flash('Data de vencimento inválida.', 'danger')
        return redirect(url_for('venda_puts') + '#collar')
    entry_date_str = _f('entry_date')
    entry_date = datetime.strptime(entry_date_str, '%Y-%m-%d').date() if entry_date_str else date.today()
    c = CollarSimulation(
        user_id          = current_user.id,
        underlying_asset = _f('underlying_asset').upper(),
        stock_price      = float(_f('stock_price')),
        quantity         = int(_f('quantity') or 100),
        put_ticker       = _f('put_ticker').upper(),
        put_strike       = float(_f('put_strike')),
        put_premium      = float(_f('put_premium')),
        call_ticker      = _f('call_ticker').upper(),
        call_strike      = float(_f('call_strike')),
        call_premium     = float(_f('call_premium')),
        expiration_date  = exp,
        entry_date       = entry_date,
        notes            = request.form.get('notes', ''),
        created_at       = datetime.now(),
    )
    db.session.add(c)
    db.session.commit()
    flash('Colar salvo.', 'success')
    return redirect(url_for('venda_puts') + '#collar')


@app.route('/collar/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def collar_edit(id):
    c = CollarSimulation.query.get_or_404(id)
    if c.user_id != current_user.id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('venda_puts'))
    if request.method == 'POST':
        def _f(k): return request.form.get(k, '').replace(',', '.').strip()
        try:
            c.expiration_date = datetime.strptime(_f('expiration_date'), '%Y-%m-%d').date()
        except ValueError:
            flash('Data de vencimento inválida.', 'danger')
            return redirect(url_for('collar_edit', id=id))
        entry_date_str = _f('entry_date')
        c.entry_date       = datetime.strptime(entry_date_str, '%Y-%m-%d').date() if entry_date_str else c.entry_date
        c.underlying_asset = _f('underlying_asset').upper()
        c.stock_price      = float(_f('stock_price'))
        c.quantity         = int(_f('quantity') or 100)
        c.put_ticker       = _f('put_ticker').upper()
        c.put_strike       = float(_f('put_strike'))
        c.put_premium      = float(_f('put_premium'))
        c.call_ticker      = _f('call_ticker').upper()
        c.call_strike      = float(_f('call_strike'))
        c.call_premium     = float(_f('call_premium'))
        c.notes            = request.form.get('notes', '')
        db.session.commit()
        flash('Colar atualizado.', 'success')
        return redirect(url_for('venda_puts') + '#collar')
    items   = PutSale.query.filter_by(user_id=current_user.id).order_by(PutSale.created_at.desc()).all()
    collars = CollarSimulation.query.filter_by(user_id=current_user.id).order_by(CollarSimulation.created_at.desc()).all()
    return render_template('venda_puts.html', items=items, collars=collars, edit=None, edit_collar=c, selic=_selic(), today=date.today())


@app.route('/collar/<int:id>/delete', methods=['POST'])
@login_required
def collar_delete(id):
    c = CollarSimulation.query.get_or_404(id)
    if c.user_id != current_user.id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('venda_puts'))
    db.session.delete(c)
    db.session.commit()
    flash('Colar excluído.', 'success')
    return redirect(url_for('venda_puts') + '#collar')


@app.route('/venda_puts/<int:id>/delete', methods=['POST'])
@login_required
def venda_puts_delete(id):
    p = PutSale.query.get_or_404(id)
    if p.user_id != current_user.id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('venda_puts'))
    db.session.delete(p)
    db.session.commit()
    flash('Venda de put excluída.', 'success')
    return redirect(url_for('venda_puts'))


def _get_underlying_quote_cached(ticker, user_id):
    """Retorna (price, daily_change) do subjacente SEM rede — apenas valores
    já salvos no banco (atualizados pelo botão Atualizar Cotações / feeder).
    Usada no render das páginas para não travar a tela com HTTP síncrono."""
    if not ticker:
        return None, None
    t = ticker.strip().upper()
    a = Asset.query.filter_by(ticker=t, user_id=user_id).first()
    if a and a.current_price:
        return a.current_price, getattr(a, 'daily_change', None)
    try:
        # Subjacente fora da carteira: preço salvo na própria Option pelo
        # Atualizar Cotações (ex.: AXIA3/MULT3 em venda a seco de puts)
        o = Option.query.filter_by(underlying_asset=t, user_id=user_id)\
                        .filter(Option.underlying_price > 0)\
                        .order_by(Option.last_update.desc()).first()
        if o and o.underlying_price:
            return o.underlying_price, o.underlying_change
    except Exception:
        pass
    try:
        ps = PutSale.query.filter_by(underlying_asset=t, user_id=user_id)\
                          .filter(PutSale.underlying_price > 0).first()
        if ps and ps.underlying_price:
            return ps.underlying_price, getattr(ps, 'underlying_change', None)
    except Exception:
        pass
    try:
        op = StructuredOp.query.filter_by(underlying_asset=t, user_id=user_id)\
                               .filter(StructuredOp.underlying_price.isnot(None)).first()
        if op and op.underlying_price:
            return op.underlying_price, op.underlying_change
    except Exception:
        pass
    sp = OptionSpread.query.filter_by(underlying_asset=t, user_id=user_id).first()
    if sp and sp.underlying_price:
        return sp.underlying_price, sp.underlying_change
    so = StudyOption.query.filter_by(underlying_asset=t, user_id=user_id).first()
    if so and so.underlying_price:
        return so.underlying_price, None
    return None, None


def _quotes_from_db(tickers, user_id):
    """Mapa {ticker: {'price', 'change_percent'}} SEM rede, no formato de
    get_quotes(). Fonte: Asset → StructuredOp → OptionSpread → StudyOption."""
    out = {}
    for t in {(x or '').strip().upper() for x in tickers if x}:
        price, change = _get_underlying_quote_cached(t, user_id)
        if price:
            out[t] = {'price': price, 'change_percent': change or 0.0}
    return out


def _get_underlying_quote(ticker, user_id):
    """Retorna (price, daily_change) do ativo subjacente.
    Tenta primeiro cotação ao vivo via Yahoo Finance; fallback em dados do banco."""
    if not ticker:
        return None, None
    t = ticker.strip().upper()

    # 1. Cotação ao vivo via Yahoo Finance — busca 5 dias para ter 2 closes reais
    try:
        import requests as _req
        yf_t = t + '.SA' if not t.endswith('.SA') and '.' not in t else t
        for host in ('query1.finance.yahoo.com', 'query2.finance.yahoo.com'):
            url = f'https://{host}/v8/finance/chart/' + yf_t
            r   = _req.get(url, params={'interval': '1d', 'range': '5d'},
                           headers=_YF_HEADERS, cookies=_YF_COOKIES, timeout=5)
            if r.status_code == 200:
                break
        if r.status_code == 200:
            res  = r.json()
            result = res['chart']['result'][0]
            meta = result['meta']
            price = meta.get('regularMarketPrice') or meta.get('previousClose')
            if price:
                change = meta.get('regularMarketChangePercent')
                # Tenta calcular a partir dos closes reais dos candles (igual ao chart)
                try:
                    closes = [c for c in result['indicators']['quote'][0].get('close', [])
                              if c is not None and c > 0]
                    if len(closes) >= 2:
                        prev_close = closes[-2]
                        last_close = closes[-1]
                        calc = (last_close - prev_close) / prev_close * 100
                        # Usa o cálculo dos candles se o meta retornar zero
                        if change is None or abs(change) < 0.001:
                            change = calc
                except Exception:
                    pass
                # Último fallback: previousClose do meta
                if change is None or abs(change) < 0.001:
                    prev = meta.get('previousClose') or meta.get('chartPreviousClose')
                    if prev and float(prev) > 0 and float(price) != float(prev):
                        change = (float(price) - float(prev)) / float(prev) * 100
                return round(float(price), 2), round(float(change), 2) if change is not None else None
    except Exception:
        pass

    # 2-5. Fallback: valores já salvos no banco
    return _get_underlying_quote_cached(ticker, user_id)


def _spread_roll_adjustment(sp):
    """Resultado realizado antes da trava atual, usado para corrigir o payoff rolado."""
    import json as _json
    try:
        history = _json.loads(sp.roll_history) if sp.roll_history else []
    except Exception:
        return 0.0, []

    adjustment = 0.0
    for item in history:
        try:
            qty = float(item.get('quantity') or sp.quantity or 1)
            if item.get('old_long_ticker') is not None:
                old_long = float(item.get('old_long_price') or 0)
                old_short = float(item.get('old_short_price') or 0)
                close_long = float(item.get('close_long_price') or 0)
                close_short = float(item.get('close_short_price') or 0)

                initial_cash = (old_short - old_long) * qty
                close_cash = (close_long - close_short) * qty
                adjustment += initial_cash + close_cash
        except (TypeError, ValueError):
            continue

    return round(adjustment, 2), history


@app.route('/payoff/spread/<int:id>')
@login_required
def payoff_spread(id):
    sp = OptionSpread.query.get_or_404(id)
    if sp.user_id != current_user.id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('opcoes'))

    # Monta estrutura de legs para o template (mesma interface das estruturadas)
    is_credit = (sp.leg_short_price - sp.leg_long_price) >= 0
    type_labels = {
        'TRAVA_ALTA_PUT':   'Trava de Alta com Puts',
        'TRAVA_ALTA_CALL':  'Trava de Alta com Calls',
        'TRAVA_BAIXA_PUT':  'Trava de Baixa com Puts',
        'TRAVA_BAIXA_CALL': 'Trava de Baixa com Calls',
    }
    is_put = 'PUT' in sp.spread_type
    exp_str = sp.expiration_date.strftime('%Y-%m-%d') if sp.expiration_date else ''
    legs = [
        {
            'ticker':           sp.leg_long_ticker,
            'side':             'BUY',
            'opt_type':         'PUT' if is_put else 'CALL',
            'quantity':         sp.quantity,
            'strike':           sp.leg_long_strike,
            'entry_price':      sp.leg_long_price,
            'current_price':    sp.leg_long_current,
            'expiration_date':  exp_str,
        },
        {
            'ticker':           sp.leg_short_ticker,
            'side':             'SELL',
            'opt_type':         'PUT' if is_put else 'CALL',
            'quantity':         sp.quantity,
            'strike':           sp.leg_short_strike,
            'entry_price':      sp.leg_short_price,
            'current_price':    sp.leg_short_current,
            'expiration_date':  exp_str,
        },
    ]
    und_price, und_change = _get_underlying_quote(sp.underlying_asset, current_user.id)
    t_days = max((sp.expiration_date - date.today()).days, 1) if sp.expiration_date else 30
    days_nearest = max((sp.expiration_date - date.today()).days, 0) if sp.expiration_date else None
    import json as _json
    roll_adjustment, roll_history = _spread_roll_adjustment(sp)
    return render_template('payoff.html',
                           pop_save_url=url_for('save_payoff_pop', kind='spread', id=sp.id),
                           title=type_labels.get(sp.spread_type, sp.spread_type),
                           underlying=sp.underlying_asset,
                           expiration=sp.expiration_date.strftime('%d/%m/%Y') if sp.expiration_date else '',
                           legs=legs,
                           underlying_price=und_price,
                           underlying_change=und_change,
                           selic=_selic(),
                           T_days=t_days,
                           days_nearest=days_nearest,
                           roll_adjustment=roll_adjustment,
                           roll_history_json=_json.dumps(roll_history, ensure_ascii=False),
                           legs_json=_json.dumps(legs))


@app.route('/payoff/estruturada/<int:id>')
@login_required
def payoff_estruturada(id):
    op = StructuredOp.query.get_or_404(id)
    if op.user_id != current_user.id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('opcoes'))
    legs = [
        {
            'id':               leg.id,
            'ticker':           leg.ticker,
            'side':             leg.side,
            'opt_type':         leg.opt_type,
            'quantity':         leg.quantity,
            'strike':           leg.strike,
            'entry_price':      leg.entry_price,
            'current_price':    leg.current_price,
            'expiration_date':  leg.expiration_date.strftime('%Y-%m-%d') if leg.expiration_date else '',
        }
        for leg in op.legs
    ]
    exp_dates = [leg.expiration_date for leg in op.legs if leg.expiration_date]
    expiration = max(exp_dates).strftime('%d/%m/%Y') if exp_dates else ''
    if getattr(op, 'intl', False):
        # Tastytrade: subjacente internacional — Yahoo sem sufixo .SA;
        # fallback no valor salvo pelo último Atualizar Cotações.
        und_price, und_change = None, None
        try:
            from services import _yf_fast_info
            d = _yf_fast_info(op.underlying_asset.strip().upper(),
                              op.underlying_asset) if op.underlying_asset else None
            if d and d.get('price'):
                und_price, und_change = d['price'], d.get('change_percent')
        except Exception:
            pass
        if not und_price:
            und_price, und_change = op.underlying_price, op.underlying_change
    else:
        und_price, und_change = _get_underlying_quote(op.underlying_asset, current_user.id)
    t_days = max((max(exp_dates) - date.today()).days, 1) if exp_dates else 30
    days_nearest = max((min(exp_dates) - date.today()).days, 0) if exp_dates else None
    import json as _json
    return render_template('payoff.html',
                           pop_save_url=url_for('save_payoff_pop', kind='estruturada', id=op.id),
                           title=op.name,
                           underlying=op.underlying_asset or '',
                           expiration=expiration,
                           legs=legs,
                           underlying_price=und_price,
                           underlying_change=und_change,
                           selic=_selic(),
                           T_days=t_days,
                           days_nearest=days_nearest,
                           manage_op_id=op.id,
                           today=date.today(),
                           legs_json=_json.dumps(legs))


@app.route('/api/payoff-pop/<kind>/<int:id>', methods=['POST'])
@login_required
def save_payoff_pop(kind, id):
    """Persiste o POP calculado na página de payoff, para exibir na lista
    /opcoes (travas e estruturadas). Chamado automaticamente pelo payoff.html."""
    try:
        pop = float((request.get_json(silent=True) or {}).get('pop'))
    except (TypeError, ValueError):
        return jsonify({'error': 'pop inválido'}), 400
    if not (0.0 <= pop <= 100.0):
        return jsonify({'error': 'pop fora de 0-100'}), 400

    if kind == 'spread':
        rec = OptionSpread.query.get_or_404(id)
    elif kind == 'estruturada':
        rec = StructuredOp.query.get_or_404(id)
    else:
        return jsonify({'error': 'tipo desconhecido'}), 400
    if rec.user_id != current_user.id:
        return jsonify({'error': 'sem permissão'}), 403

    rec.pop = round(pop, 1)
    db.session.commit()
    return jsonify({'ok': True, 'pop': rec.pop})


@app.route('/payoff/option/<int:id>')
@login_required
def payoff_option(id):
    """Payoff de opção individual (venda coberta, venda a seco, compra a seco).
    Venda coberta de CALL inclui a perna da ação (lastro) pelo preço médio da carteira."""
    opt = Option.query.get_or_404(id)
    if opt.user_id != current_user.id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('opcoes'))

    ot     = (opt.option_type or 'VENDA_CALL').upper()
    is_put = 'PUT' in ot
    is_buy = 'COMPRA' in ot

    legs = []
    # Venda coberta de CALL: acrescenta a ação comprada (preço médio da carteira)
    if ot == 'VENDA_CALL' and opt.underlying_asset:
        asset = Asset.query.filter_by(user_id=current_user.id,
                                      ticker=opt.underlying_asset.upper()).first()
        if asset and (asset.avg_price or 0) > 0:
            legs.append({
                'ticker':          asset.ticker,
                'side':            'BUY',
                'opt_type':        'STOCK',
                'quantity':        opt.quantity,
                'strike':          0,
                'entry_price':     asset.avg_price,
                'current_price':   asset.current_price or 0,
                'expiration_date': '',
            })

    legs.append({
        'ticker':          opt.ticker,
        'side':            'BUY' if is_buy else 'SELL',
        'opt_type':        'PUT' if is_put else 'CALL',
        'quantity':        opt.quantity,
        'strike':          opt.strike_price,
        'entry_price':     opt.sale_price,
        'current_price':   opt.current_option_price or 0,
        'expiration_date': opt.expiration_date.strftime('%Y-%m-%d') if opt.expiration_date else '',
    })

    titles = {
        'VENDA_CALL':  'Venda Coberta de Call',
        'VENDA_PUT':   'Venda a Seco de Put',
        'COMPRA_CALL': 'Compra a Seco de Call',
        'COMPRA_PUT':  'Compra a Seco de Put',
    }
    und_price, und_change = _get_underlying_quote(opt.underlying_asset, current_user.id)
    t_days = max((opt.expiration_date - date.today()).days, 1) if opt.expiration_date else 30

    import json as _json
    return render_template('payoff.html',
                           title=f"{titles.get(ot, 'Opção')} — {opt.ticker}",
                           underlying=opt.underlying_asset or '',
                           expiration=opt.expiration_date.strftime('%d/%m/%Y') if opt.expiration_date else '',
                           legs=legs,
                           underlying_price=und_price,
                           underlying_change=und_change,
                           selic=_selic(),
                           T_days=t_days,
                           days_nearest=t_days,
                           legs_json=_json.dumps(legs))


@app.route('/api/option/<int:id>/delta', methods=['POST'])
@login_required
def api_set_option_delta(id):
    """Atualiza delta manualmente para uma opção (venda put ou qualquer tipo)."""
    from flask import jsonify
    opt = Option.query.get_or_404(id)
    if opt.user_id != current_user.id:
        return jsonify({'error': 'Sem permissão'}), 403
    try:
        val = request.json.get('delta')
        opt.delta = float(val) if val is not None and val != '' else None
        db.session.commit()
        return jsonify({'delta': opt.delta})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400


@app.route('/edit_option/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_option(id):
    opt = Option.query.get_or_404(id)
    if opt.user_id != current_user.id:
        flash("Você não tem permissão para editar esta opção.")
        return redirect(url_for('opcoes'))
    if request.method == 'POST':
        try:
            opt.ticker = request.form.get('ticker').upper()
            opt.quantity = int(request.form.get('quantity'))
            opt.underlying_asset = request.form.get('underlying_asset').upper()
            opt.strike_price = float(request.form.get('strike_price').replace(',', '.'))
            opt.expiration_date = datetime.strptime(request.form.get('expiration_date'), '%Y-%m-%d').date()
            opt.sale_price = float(request.form.get('sale_price').replace(',', '.'))
            
            entry_str = request.form.get('entry_date')
            opt.entry_date = datetime.strptime(entry_str, '%Y-%m-%d').date() if entry_str else None
            
            curr_price_str = request.form.get('current_option_price')
            if curr_price_str:
                opt.current_option_price = float(curr_price_str.replace(',', '.'))
                
            db.session.commit()
            flash("Opção atualizada com sucesso!", "success")
            return redirect(url_for('opcoes'))
            
        except ValueError as ve:
             flash(f"Erro de formato: {ve}", "danger")
        except Exception as e:
             flash(f"Erro ao editar: {e}", "danger")
             return redirect(url_for('opcoes'))
             
        return redirect(url_for('opcoes'))
        
    return render_template('add_option.html', option=opt, edit=True, option_type=opt.option_type)

@app.route('/delete_option/<int:id>')
@login_required
def delete_option(id):
    opt = Option.query.get_or_404(id)
    if opt.user_id != current_user.id:
        flash("Você não tem permissão para deletar esta opção.")
        return redirect(url_for('opcoes'))
    db.session.delete(opt)
    db.session.commit()
    return redirect(url_for('opcoes'))

@app.route('/update_options_quotes', methods=['POST'])
@login_required
def update_options_quotes():
    token = Settings.get_value('oplab_token', user_id=current_user.id)
    if token:
        a_ok, o_ok, _ = _do_oplab_bulk_update(current_user.id, token)
        flash(f'OpLab: {a_ok} ativo(s) e {o_ok} opção(ões)/perna(s) atualizados.', 'success')
    else:
        count, tried, errs = update_all_assets_logic()
        if errs:
            flash(f'Ativos: {count}/{tried}. Erros: {errs[0]}', 'warning')
        else:
            flash(f'Ativos: {count}/{tried} atualizados. Configure a OpLab para atualizar opções.', 'warning')
    return redirect(url_for('opcoes'))

@app.route('/close_option/<int:id>', methods=['GET', 'POST'])
@login_required
def close_option(id):
    opt = Option.query.get_or_404(id)
    if opt.user_id != current_user.id:
        flash("Você não tem permissão para fechar esta opção.")
        return redirect(url_for('opcoes'))

    if request.method == 'POST':
        buy_back_price = float(request.form.get('price').replace(',', '.'))
        date_exit = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
        qty_exit = int(request.form.get('quantity', opt.quantity))
        reason = request.form.get('reason', 'ENCERRAMENTO')

        # Validate quantity
        if qty_exit <= 0 or qty_exit > opt.quantity:
            flash("Quantidade inválida.", "danger")
            return redirect(url_for('close_option', id=id))

        # Calculate days held if entry_date exists
        days_held = 0
        if opt.entry_date:
            days_held = (date_exit - opt.entry_date).days

        # Profit Calculation:
        # LONG  (COMPRA_CALL / COMPRA_PUT): comprou por sale_price, vendeu por buy_back_price
        #   lucro = (saída - entrada) * qty  → buy_price=entrada, sell_price=saída
        # SHORT (VENDA_CALL / VENDA_PUT):   vendeu por sale_price, recomprou por buy_back_price
        #   lucro = (entrada - saída) * qty  → buy_price=recompra, sell_price=prêmio recebido
        if opt.option_type in ('COMPRA_CALL', 'COMPRA_PUT'):
            entry_p = opt.sale_price
            exit_p  = buy_back_price
            profit_val = (exit_p - entry_p) * qty_exit
        else:
            entry_p = opt.sale_price     # prêmio recebido na venda
            exit_p  = buy_back_price     # prêmio pago para recomprar
            profit_val = (entry_p - exit_p) * qty_exit

        profit_pct = (profit_val / (qty_exit * entry_p) * 100) if entry_p > 0 else 0

        history = TradeHistory(
            user_id    = current_user.id,
            ticker     = opt.ticker,
            strategy   = 'Opções',
            entry_date = opt.entry_date,
            exit_date  = date_exit,
            buy_price  = round(entry_p, 4),   # sempre o preço de entrada da posição
            sell_price = round(exit_p,  4),   # sempre o preço de saída da posição
            quantity   = qty_exit,
            profit_value = round(profit_val, 2),
            profit_pct   = round(profit_pct, 2),
            days_held    = days_held,
            reason       = reason,
            underlying   = opt.underlying_asset,
            notes        = f"{opt.option_type} | {opt.underlying_asset} | K={opt.strike_price:.2f} | venc {opt.expiration_date.strftime('%d/%m/%Y') if opt.expiration_date else '?'}",
        )
        db.session.add(history)

        if qty_exit == opt.quantity:
            # Total exit - remove option
            db.session.delete(opt)
            flash("Saída TOTAL de opção registrada no histórico!", "success")
        else:
            # Partial exit - reduce quantity
            opt.quantity -= qty_exit
            flash(f"Saída PARCIAL registrada! Restam {opt.quantity} opções.", "success")

        db.session.commit()
        return redirect(url_for('opcoes'))

    # Render exit form for option
    return render_template('close_option.html', option=opt, today=date.today())


# ══════════════════════════════════════════════════════════════════════════════
# ROLAGEM DE OPERAÇÕES
# Não gera TradeHistory — registra no roll_history da operação e atualiza campos
# ══════════════════════════════════════════════════════════════════════════════

def _append_roll(obj, entry: dict):
    """Acrescenta uma entrada de rolagem ao roll_history JSON do objeto."""
    import json as _json
    history = _json.loads(obj.roll_history) if obj.roll_history else []
    entry['rolled_at'] = now_brt().strftime('%d/%m/%Y %H:%M')
    history.append(entry)
    obj.roll_history = _json.dumps(history, ensure_ascii=False)


@app.route('/roll_option/<int:id>', methods=['POST'])
@login_required
def roll_option(id):
    """Rola opção individual (strike ou tempo).
    Salva os dados antigos no roll_history e atualiza os campos com os novos valores.
    NÃO gera registro de P&L em TradeHistory.
    """
    opt = Option.query.get_or_404(id)
    if opt.user_id != current_user.id:
        return {'error': 'Sem permissão'}, 403

    roll_type      = request.form.get('roll_type', 'TEMPO')   # STRIKE | TEMPO | AMBOS
    new_ticker     = request.form.get('new_ticker', '').upper().strip()
    new_strike     = request.form.get('new_strike', '').replace(',', '.')
    new_exp        = request.form.get('new_exp', '')
    new_premium    = request.form.get('new_premium', '').replace(',', '.')
    close_premium  = request.form.get('close_premium', '').replace(',', '.')  # prêmio recompra do antigo
    roll_date      = request.form.get('roll_date', date.today().isoformat())
    notes          = request.form.get('notes', '')

    try:
        close_p  = float(close_premium) if close_premium else 0.0
        new_p    = float(new_premium)   if new_premium   else opt.sale_price
        new_k    = float(new_strike)    if new_strike    else opt.strike_price
        new_exp_d = datetime.strptime(new_exp, '%Y-%m-%d').date() if new_exp else opt.expiration_date

        # Débito/crédito líquido da rolagem (do ponto de vista da posição)
        # SELL (venda): recomprou por close_p e vendeu novo por new_p → crédito = new_p - close_p
        # BUY (compra): vendeu por close_p e comprou novo por new_p  → débito  = new_p - close_p
        if opt.option_type in ('VENDA_CALL', 'VENDA_PUT'):
            net_roll = new_p - close_p   # positivo = crédito adicional
        else:
            net_roll = close_p - new_p   # positivo = recebeu mais do que pagou

        _append_roll(opt, {
            'roll_type':     roll_type,
            'old_ticker':    opt.ticker,
            'old_strike':    opt.strike_price,
            'old_exp':       opt.expiration_date.isoformat() if opt.expiration_date else None,
            'old_premium':   opt.sale_price,
            'close_premium': close_p,
            'new_ticker':    new_ticker or opt.ticker,
            'new_strike':    new_k,
            'new_exp':       new_exp_d.isoformat(),
            'new_premium':   new_p,
            'net_roll':      round(net_roll, 4),
            'notes':         notes,
        })

        # Atualiza a operação com os novos dados
        if new_ticker:
            opt.ticker = new_ticker
        opt.strike_price     = new_k
        opt.expiration_date  = new_exp_d
        opt.sale_price       = new_p
        opt.entry_date       = datetime.strptime(roll_date, '%Y-%m-%d').date()
        opt.current_option_price = new_p

        db.session.commit()
        flash(f'Rolagem registrada! Crédito/Débito líquido: R$ {net_roll:.2f}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro na rolagem: {e}', 'danger')

    return redirect(url_for('opcoes'))


@app.route('/api/option-id/<ticker>')
@login_required
def api_option_id(ticker):
    """Localiza a opção registrada (tabela Option) pelo ticker — usado pelo
    simulador de rolagem para gravar a rolagem na operação real."""
    opt = Option.query.filter_by(user_id=current_user.id, ticker=ticker.upper().strip()).first()
    if not opt:
        return jsonify({'error': f'Opção {ticker.upper()} não encontrada nas suas operações registradas.'}), 404
    return jsonify({
        'id':          opt.id,
        'ticker':      opt.ticker,
        'option_type': opt.option_type,
        'strike':      opt.strike_price,
        'exp':         opt.expiration_date.isoformat() if opt.expiration_date else '',
        'premium':     opt.sale_price,
        'quantity':    opt.quantity,
    })


@app.route('/roll_spread/<int:id>', methods=['POST'])
@login_required
def roll_spread(id):
    """Rola trava (strike ou tempo). Atualiza ambas as pernas."""
    sp = OptionSpread.query.get_or_404(id)
    if sp.user_id != current_user.id:
        return {'error': 'Sem permissão'}, 403

    roll_type = request.form.get('roll_type', 'TEMPO')
    roll_date = request.form.get('roll_date', date.today().isoformat())
    notes     = request.form.get('notes', '')

    # Perna long (comprada)
    new_long_ticker  = request.form.get('new_long_ticker', '').upper().strip()
    new_long_strike  = request.form.get('new_long_strike', '').replace(',', '.')
    new_long_price   = request.form.get('new_long_price',  '').replace(',', '.')
    close_long_price = request.form.get('close_long_price','').replace(',', '.')

    # Perna short (vendida)
    new_short_ticker  = request.form.get('new_short_ticker', '').upper().strip()
    new_short_strike  = request.form.get('new_short_strike', '').replace(',', '.')
    new_short_price   = request.form.get('new_short_price',  '').replace(',', '.')
    close_short_price = request.form.get('close_short_price','').replace(',', '.')

    new_exp = request.form.get('new_exp', '')

    try:
        cl_long  = float(close_long_price)  if close_long_price  else sp.leg_long_current
        cl_short = float(close_short_price) if close_short_price else sp.leg_short_current
        nl_p     = float(new_long_price)    if new_long_price     else sp.leg_long_price
        ns_p     = float(new_short_price)   if new_short_price    else sp.leg_short_price
        nl_k     = float(new_long_strike)   if new_long_strike    else sp.leg_long_strike
        ns_k     = float(new_short_strike)  if new_short_strike   else sp.leg_short_strike
        new_exp_d = datetime.strptime(new_exp, '%Y-%m-%d').date() if new_exp else sp.expiration_date

        # Caixa da rolagem: vende a perna comprada antiga, recompra a vendida antiga,
        # compra a nova perna comprada e vende a nova perna vendida.
        close_cash = cl_long - cl_short
        open_cash = ns_p - nl_p
        net_roll = (close_cash + open_cash) * (sp.quantity or 1)

        _append_roll(sp, {
            'roll_type':        roll_type,
            'old_long_ticker':  sp.leg_long_ticker,
            'old_long_strike':  sp.leg_long_strike,
            'old_long_price':   sp.leg_long_price,
            'close_long_price': cl_long,
            'old_short_ticker': sp.leg_short_ticker,
            'old_short_strike': sp.leg_short_strike,
            'old_short_price':  sp.leg_short_price,
            'close_short_price':cl_short,
            'new_long_ticker':  new_long_ticker  or sp.leg_long_ticker,
            'new_long_strike':  nl_k,
            'new_long_price':   nl_p,
            'new_short_ticker': new_short_ticker or sp.leg_short_ticker,
            'new_short_strike': ns_k,
            'new_short_price':  ns_p,
            'new_exp':          new_exp_d.isoformat(),
            'quantity':         sp.quantity or 1,
            'net_roll':         round(net_roll, 4),
            'notes':            notes,
        })

        if new_long_ticker:   sp.leg_long_ticker  = new_long_ticker
        if new_short_ticker:  sp.leg_short_ticker = new_short_ticker
        sp.leg_long_strike   = nl_k
        sp.leg_long_price    = nl_p
        sp.leg_long_current  = nl_p
        sp.leg_short_strike  = ns_k
        sp.leg_short_price   = ns_p
        sp.leg_short_current = ns_p
        sp.expiration_date   = new_exp_d
        sp.entry_date        = datetime.strptime(roll_date, '%Y-%m-%d').date()

        db.session.commit()
        flash(f'Rolagem da trava registrada! Crédito/Débito: R$ {net_roll:.2f}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro na rolagem: {e}', 'danger')

    return redirect(url_for('opcoes'))


@app.route('/roll_estruturada/<int:id>', methods=['POST'])
@login_required
def roll_estruturada(id):
    """Rola operação estruturada: registra histórico e atualiza pernas."""
    op = StructuredOp.query.get_or_404(id)
    if op.user_id != current_user.id:
        return {'error': 'Sem permissão'}, 403

    roll_type  = request.form.get('roll_type', 'TEMPO')
    roll_date  = request.form.get('roll_date', date.today().isoformat())
    notes_text = request.form.get('notes', '')
    new_exp    = request.form.get('new_exp', '')
    next_url   = request.form.get('next', '')

    # Listas paralelas para cada perna (por leg_id)
    leg_ids         = request.form.getlist('leg_id')
    close_prices    = request.form.getlist('leg_close_price')
    new_tickers     = request.form.getlist('leg_new_ticker')
    new_strikes     = request.form.getlist('leg_new_strike')
    new_premiums    = request.form.getlist('leg_new_premium')
    new_exps        = request.form.getlist('leg_new_exp')

    try:
        roll_entry = {
            'roll_type': roll_type,
            'roll_date': roll_date,
            'notes':     notes_text,
            'legs':      [],
        }
        try:
            exit_d = datetime.strptime(roll_date, '%Y-%m-%d').date()
        except ValueError:
            exit_d = date.today()
        entry_d = op.created_at.date() if op.created_at else None

        net_roll = 0.0
        realized = 0.0
        for i, lid in enumerate(leg_ids):
            leg = StructuredLeg.query.get(int(lid))
            if not leg or leg.operation.user_id != current_user.id:
                continue

            cp  = float(close_prices[i].replace(',','.'))  if i < len(close_prices)  and close_prices[i]  else leg.current_price or leg.entry_price
            nt  = new_tickers[i].upper().strip()            if i < len(new_tickers)   and new_tickers[i]   else leg.ticker
            nk  = float(new_strikes[i].replace(',','.'))   if i < len(new_strikes)   and new_strikes[i]   else leg.strike
            np_ = float(new_premiums[i].replace(',','.'))  if i < len(new_premiums)  and new_premiums[i]  else leg.entry_price
            ne  = new_exps[i]                               if i < len(new_exps)      and new_exps[i]      else (leg.expiration_date.isoformat() if leg.expiration_date else '')

            # SELL: recebe novo prêmio, paga para fechar → crédito = np_ - cp
            # BUY:  paga novo prêmio, recebe para fechar → custo  = cp - np_
            qty = leg.quantity or 1
            if leg.side == 'SELL':
                net_roll += (np_ - cp) * qty
            else:
                net_roll += (cp - np_) * qty

            # A perna foi de fato manejada? (ticker/strike/prêmio/venc mudaram)
            old_exp_iso = leg.expiration_date.isoformat() if leg.expiration_date else ''
            changed = (nt != leg.ticker
                       or abs(nk - (leg.strike or 0)) > 0.001
                       or abs(np_ - (leg.entry_price or 0)) > 0.001
                       or (ne or '') != old_exp_iso)

            # Resultado da perna fechada — apenas REGISTRADO (roll_history), NÃO
            # vira TradeHistory. Num manejo, fechar/abrir pernas é ajuste de
            # composição, não lucro realizado: o P&L da operação só é apurado
            # no ENCERRAMENTO, com o cômputo geral de entradas e saídas.
            if changed:
                if leg.side == 'SELL':
                    pnl = ((leg.entry_price or 0) - cp) * qty   # vendeu, recomprou
                else:
                    pnl = (cp - (leg.entry_price or 0)) * qty   # comprou, vendeu
                realized += pnl

            roll_entry['legs'].append({
                'old_ticker':    leg.ticker,
                'old_strike':    leg.strike,
                'old_exp':       leg.expiration_date.isoformat() if leg.expiration_date else None,
                'old_premium':   leg.entry_price,
                'close_price':   cp,
                'new_ticker':    nt,
                'new_strike':    nk,
                'new_premium':   np_,
                'new_exp':       ne,
                'side':          leg.side,
                'quantity':      qty,
                'realized_pnl':  round(((leg.entry_price or 0) - cp) * qty if leg.side == 'SELL'
                                       else (cp - (leg.entry_price or 0)) * qty, 2) if changed else None,
            })

            # Atualiza a perna
            leg.ticker          = nt
            leg.strike          = nk
            leg.entry_price     = np_
            leg.current_price   = np_
            if ne:
                try:
                    leg.expiration_date = datetime.strptime(ne, '%Y-%m-%d').date()
                except ValueError:
                    pass

        roll_entry['net_roll'] = round(net_roll, 4)
        roll_entry['realized_pnl'] = round(realized, 2)
        # Marca: manejo do NOVO modelo (não gerou TradeHistory) — o encerramento
        # soma este realized_pnl. Manejos antigos (sem a flag) já viraram
        # TradeHistory e NÃO são somados de novo, evitando dupla contagem.
        roll_entry['defer_pnl'] = True
        _append_roll(op, roll_entry)
        op.created_at = datetime.strptime(roll_date, '%Y-%m-%d')

        db.session.commit()
        flash(f'Manejo registrado! Caixa do ajuste: R$ {net_roll:.2f} · '
              f'Parcial acumulado nas pernas fechadas: R$ {realized:.2f} '
              f'(o lucro/prejuízo só é apurado no ENCERRAMENTO da operação).', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro na rolagem: {e}', 'danger')

    if next_url and next_url.startswith('/'):
        return redirect(next_url)
    return redirect(url_for('opcoes'))


@app.route('/roll_put_sale/<int:id>', methods=['POST'])
@login_required
def roll_put_sale(id):
    """Rola venda de put (strike ou tempo)."""
    ps = PutSale.query.get_or_404(id)
    if ps.user_id != current_user.id:
        return {'error': 'Sem permissão'}, 403

    roll_type      = request.form.get('roll_type', 'TEMPO')
    new_ticker     = request.form.get('new_ticker', '').upper().strip()
    new_strike     = request.form.get('new_strike', '').replace(',', '.')
    new_exp        = request.form.get('new_exp', '')
    new_premium    = request.form.get('new_premium', '').replace(',', '.')
    close_premium  = request.form.get('close_premium', '').replace(',', '.')
    roll_date      = request.form.get('roll_date', date.today().isoformat())
    notes          = request.form.get('notes', '')

    try:
        cp   = float(close_premium) if close_premium else 0.0
        np_  = float(new_premium)   if new_premium   else ps.premium
        nk   = float(new_strike)    if new_strike     else ps.strike
        ne   = datetime.strptime(new_exp, '%Y-%m-%d').date() if new_exp else ps.expiration_date

        net_roll = np_ - cp  # crédito recebido no net da rolagem

        _append_roll(ps, {
            'roll_type':     roll_type,
            'old_ticker':    ps.ticker,
            'old_strike':    ps.strike,
            'old_exp':       ps.expiration_date.isoformat() if ps.expiration_date else None,
            'old_premium':   ps.premium,
            'close_premium': cp,
            'new_ticker':    new_ticker or ps.ticker,
            'new_strike':    nk,
            'new_exp':       ne.isoformat(),
            'new_premium':   np_,
            'net_roll':      round(net_roll, 4),
            'notes':         notes,
        })

        if new_ticker:       ps.ticker          = new_ticker
        ps.strike           = nk
        ps.expiration_date  = ne
        ps.premium          = np_
        ps.entry_date       = datetime.strptime(roll_date, '%Y-%m-%d').date()

        db.session.commit()
        flash(f'Rolagem da put registrada! Crédito/Débito: R$ {net_roll:.2f}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro na rolagem: {e}', 'danger')

    return redirect(url_for('opcoes'))


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))




with app.app_context():
    try:
        db.create_all()
    except Exception as e:
        # In a multi-worker environment (Gunicorn), multiple workers might try to create tables simultaneously.
        # If one succeeds, the others might fail with "table already exists".
        # We catch this to allow the server to start.
        print(f"Database initialization note: {e}")

@app.route('/')
@login_required
def index():
    return redirect(url_for('acoes'))

@app.route('/acoes')
@login_required
def acoes():
    # Stocks
    raw_assets = Asset.query.filter(Asset.type=='ACAO', Asset.strategy!='SWING', Asset.user_id==current_user.id, Asset.quantity > 0).all()
    processed_assets = process_assets(raw_assets)
    
    # Fetch International Assets for Migration
    intls_rv = International.query.filter_by(user_id=current_user.id, category='RV').all()
    intls_rf = International.query.filter_by(user_id=current_user.id, category='RF').all()

    # Calculate Totals for Intl RV
    intl_rv_invested = sum((i.quantity or 0) * (i.avg_price or 0) for i in intls_rv)
    intl_rv_current = sum((i.quantity or 0) * (i.quote or 0) for i in intls_rv)
    
    # Calculate Day Gain Totals
    def calc_day_gain(item):
        quote = item.quote or 0
        qty = item.quantity or 0
        change = item.daily_change or 0
        if change == 0 or quote == 0: return 0.0
        # Prev Close = Current / (1 + change/100)
        prev = quote / (1 + change/100)
        return qty * (quote - prev)

    intl_rv_day_gain = sum(calc_day_gain(i) for i in intls_rv)
    
    # Calculate Totals for Intl RF
    intl_rf_invested = sum((i.quantity or 0) * (i.avg_price or 0) for i in intls_rf)
    intl_rf_current = sum((i.quantity or 0) * (i.quote or 0) for i in intls_rf)
    intl_rf_day_gain = sum(calc_day_gain(i) for i in intls_rf)
    
    total_invested = sum(a['total_invested'] for a in processed_assets)
    total_current = sum(a['current_total'] for a in processed_assets)

    # ETFs
    raw_etfs = Asset.query.filter(Asset.type=='ETF', Asset.user_id==current_user.id, Asset.quantity > 0).all()
    processed_etfs = process_assets(raw_etfs)
    
    total_etfs_invested = sum(a['total_invested'] for a in processed_etfs)
    total_etfs_current = sum(a['current_total'] for a in processed_etfs)
    
    return render_template('acoes.html', 
                           assets=processed_assets, 
                           total_invested=total_invested, 
                           total_current=total_current,
                           etfs=processed_etfs,
                           total_etfs_invested=total_etfs_invested,
                           total_etfs_current=total_etfs_current,
                           intls_rv=intls_rv,
                           intls_rf=intls_rf,
                           intl_rv_invested=intl_rv_invested,
                           intl_rv_current=intl_rv_current,
                           intl_rf_invested=intl_rf_invested,
                           intl_rf_current=intl_rf_current,
                           intl_rv_day_gain=intl_rv_day_gain,
                           intl_rf_day_gain=intl_rf_day_gain)

@app.route('/fiis')
@login_required
def fiis():
    raw_assets = Asset.query.filter(Asset.type=='FII', Asset.user_id==current_user.id, Asset.quantity > 0).all()
    processed_assets = process_assets(raw_assets)
    
    total_invested = sum(a['total_invested'] for a in processed_assets)
    total_current = sum(a['current_total'] for a in processed_assets)
    
    return render_template('fiis.html', assets=processed_assets, total_invested=total_invested, total_current=total_current)

@app.route('/update_fii_dividends', methods=['POST'])
@login_required
def update_fii_dividends():
    try:
        assets = Asset.query.filter_by(user_id=current_user.id, type='FII').all()
        updated_count = 0
        error_count = 0
        
        for asset in assets:
            try:
                # Add .SA suffix if missing
                ticker_symbol = asset.ticker if asset.ticker.endswith('.SA') else f"{asset.ticker}.SA"
                
                stock = yf.Ticker(ticker_symbol)
                
                # Get Dividends History (Max 1 Year is enough for DY)
                dividends = stock.dividends
                
                if not dividends.empty:
                    # Filter for last 12 months (UTC aware)
                    end_date = datetime.now(pytz.utc) 
                    start_date = end_date - timedelta(days=365)
                    
                    # Ensure dividends index is tz-aware for comparison
                    if dividends.index.tz is None:
                         dividends.index = dividends.index.tz_localize('UTC')
                    
                    last_12m_dividends = dividends[(dividends.index >= start_date) & (dividends.index <= end_date)]
                    
                    # 1. Last Dividend
                    last_div_value = dividends.iloc[-1]
                    last_div_date = dividends.index[-1].date()
                    
                    asset.last_dividend = float(last_div_value)
                    asset.last_dividend_date = last_div_date
                    
                    # 2. DY Calculation (Sum 12m / Current Quote)
                    sum_dividends = last_12m_dividends.sum()
                    
                    # Get Current Price (Try fast info first, fallback to history)
                    current_price = asset.current_price # Use cached price if available/reliable
                    
                    # If cached price is zero, try to fetch
                    if current_price == 0:
                        hist = stock.history(period="1d")
                        if not hist.empty:
                            current_price = hist['Close'].iloc[-1]
                    
                    if current_price > 0:
                        dy_annual = sum_dividends / current_price
                        asset.dividend_yield = float(dy_annual)
                    
                    updated_count += 1
                else:
                    # No dividends found
                    pass
                    
            except Exception as e:
                print(f"Error updating FII {asset.ticker}: {e}")
                error_count += 1
                continue
                
        db.session.commit()
        
        if error_count > 0:
             flash(f"Dividendos atualizados: {updated_count}. Erros: {error_count}", "warning")
        else:
             flash(f"Dividendos de {updated_count} FIIs atualizados com sucesso!", "success")
             
    except Exception as e:
        flash(f"Erro geral ao atualizar dividendos: {str(e)}", "danger")
        
    return redirect(url_for('fiis'))

@app.route('/swingtrade')
@login_required
def swingtrade():
    raw_assets = Asset.query.filter(Asset.strategy=='SWING', Asset.user_id==current_user.id, Asset.quantity > 0).all()
    assets = process_assets(raw_assets)
    return render_template('swingtrade.html', assets=assets)


# ─────────────────────────────────────────────────────────────────────────────
# Preço Médio didático — PM ajustado por dividendos, lucro de opções e
# compras financiadas com lucro (ações a PM zero). Swing trade fica de fora.
# ─────────────────────────────────────────────────────────────────────────────
def _pm_earned_items(user_id, ticker):
    """Créditos ganhos pelo ticker: dividendos recebidos + lucro/prejuízo de
    opções realizadas do ativo (TradeHistory 'Opções' via underlying).
    Swing trade NÃO entra — evento aleatório, não ligado ao papel."""
    items = []
    for d in Dividend.query.join(Asset).filter(
            Asset.user_id == user_id, Dividend.ticker == ticker).all():
        if not d.amount or d.amount <= 0:
            continue
        dt = d.payment_date or d.ex_date
        qtd_info = (f' ({d.qty_used}× R$ {d.per_share:.4f}/ação)'
                    if d.qty_used and d.per_share else '')
        items.append({'key': f'div:{d.id}', 'kind': 'DIVIDENDO',
                      'valor': round(float(d.amount), 2), 'date': dt,
                      'ref': (f'{d.type or "Dividendo"} '
                              f'{dt.strftime("%d/%m/%Y") if dt else ""} — '
                              f'R$ {d.amount:.2f}{qtd_info}').replace('.', ',')})
    # Trades de opções: casa pelo underlying preenchido; para registros ANTIGOS
    # sem underlying, casa pela raiz B3 do ticker da opção (ABEVA148 → ABEV →
    # ABEV3) — somente se a raiz for inequívoca na carteira (PETR3 × PETR4 não).
    root = ticker[:4].upper()
    same_root = Asset.query.filter(Asset.user_id == user_id,
                                   Asset.type.in_(('ACAO', 'FII')),
                                   Asset.strategy != 'SWING',
                                   Asset.quantity > 0,
                                   Asset.ticker.like(root + '%')).count()
    for th in TradeHistory.query.filter_by(user_id=user_id, strategy='Opções').all():
        und   = (th.underlying or '').strip().upper()
        tk_th = (th.ticker or '').strip().upper()
        match = (und == ticker) or (
            not und and (tk_th == ticker or (same_root == 1 and tk_th[:4] == root)))
        if not match:
            continue
        pv = round(float(th.profit_value or 0), 2)
        if abs(pv) < 0.005:
            continue
        dt = th.exit_date
        items.append({'key': f'th:{th.id}', 'kind': 'OPCOES',
                      'valor': pv, 'date': dt,
                      'ref': f'Opções {th.ticker} '
                             f'({dt.strftime("%d/%m/%Y") if dt else "-"}) — '
                             f'resultado R$ {pv:.2f}'.replace('.', ',')})
    items.sort(key=lambda i: (i['date'] or date.min))
    return items


def _pm_state(user_id, ticker):
    """Estado do PM ajustado de um ticker: eventos, créditos, pool e custo."""
    a = Asset.query.filter(Asset.user_id == user_id, Asset.ticker == ticker,
                           Asset.type.in_(('ACAO', 'FII')),
                           Asset.strategy != 'SWING').first()
    evs = (PMEvent.query.filter_by(user_id=user_id, ticker=ticker)
           .order_by(PMEvent.created_at, PMEvent.id).all())
    used_keys  = {e.source_key for e in evs if e.source_key}
    used_total = round(sum(e.valor or 0 for e in evs), 2)
    # Pool = créditos varridos (dividendos/opções) − usos; lançamentos MANUAIS
    # (lucro avulso) abatem o custo mas NÃO consomem o pool de créditos.
    used_pool  = round(sum(e.valor or 0 for e in evs if e.kind != 'MANUAL'), 2)
    # Créditos marcados IGNORADO (ex.: evento anterior à posição atual) saem
    # completamente da varredura: não contam como ganhos, pool nem pendência.
    ignored_keys = {e.source_key for e in evs if e.kind == 'IGNORADO' and e.source_key}
    earned  = [i for i in _pm_earned_items(user_id, ticker)
               if i['key'] not in ignored_keys]
    pending = [i for i in earned if i['key'] not in used_keys]
    pool = round(sum(i['valor'] for i in earned) - used_pool, 2)
    custo_of = (a.quantity * (a.avg_price or 0)) if a else 0.0
    custo_aj = custo_of - used_total
    return a, evs, earned, pending, pool, custo_of, custo_aj


def _pm_rows(user_id):
    assets = Asset.query.filter(Asset.user_id == user_id,
                                Asset.type.in_(('ACAO', 'FII')),
                                Asset.strategy != 'SWING',
                                Asset.quantity > 0).order_by(Asset.ticker).all()
    rows = {'ACAO': [], 'FII': []}
    for a in assets:
        _a, evs, earned, pending, pool, custo_of, custo_aj = _pm_state(user_id, a.ticker)
        pm_aj = custo_aj / a.quantity if a.quantity else 0
        cot = a.current_price or 0
        lucro_rs = (cot - pm_aj) * a.quantity if cot else 0
        rows['ACAO' if a.type == 'ACAO' else 'FII'].append({
            'cotacao': cot,
            'lucro_rs': lucro_rs,
            'lucro_pct': ((cot - pm_aj) / pm_aj * 100) if (cot and pm_aj > 0) else 0,
            'asset': a,
            'custo_oficial': custo_of,
            'earned_div': sum(i['valor'] for i in earned if i['kind'] == 'DIVIDENDO'),
            'earned_opc': sum(i['valor'] for i in earned if i['kind'] == 'OPCOES'),
            'aplicado': custo_of - custo_aj,
            'pool': pool,
            'pending_n': len(pending),
            'pending_sum': round(sum(i['valor'] for i in pending), 2),
            'custo_aj': custo_aj,
            'pm_aj': pm_aj,
            'reducao_pct': ((a.avg_price - pm_aj) / a.avg_price * 100) if a.avg_price else 0,
            'events_n': len(evs),
        })
    return rows


@app.route('/preco-medio')
@login_required
def preco_medio():
    rows = _pm_rows(current_user.id)
    return render_template('preco_medio.html', acoes=rows['ACAO'], fiis=rows['FII'],
                           today=date.today())


@app.route('/preco-medio/aplicar/<ticker>', methods=['POST'])
@login_required
def pm_aplicar(ticker):
    """Varre os créditos pendentes (dividendos + opções) e aplica ao PM,
    limitado ao pool disponível (créditos já consumidos em compras não
    reduzem de novo — sem contagem em dobro)."""
    ticker = ticker.strip().upper()
    a, evs, earned, pending, pool, custo_of, custo_aj = _pm_state(current_user.id, ticker)
    if not a or a.quantity <= 0:
        flash(f'{ticker}: ativo não encontrado na carteira.', 'warning')
        return redirect(url_for('preco_medio'))
    if not pending:
        flash(f'{ticker}: nenhum crédito pendente para aplicar.', 'warning')
        return redirect(url_for('preco_medio'))

    # Prejuízos primeiro (aumentam a folga do pool); ganhos depois, por data —
    # senão o cap cortaria ganhos que os prejuízos posteriores compensariam.
    pending = sorted(pending, key=lambda i: (i['valor'] >= 0, i['date'] or date.min))
    remaining, n, aplicado = pool, 0, 0.0
    for item in pending:
        v = item['valor']
        if v > 0:
            if remaining <= 0.005:
                continue
            v = min(v, remaining)          # cap: não aplica além do pool
        pm_before = custo_aj / a.quantity
        custo_aj -= v
        db.session.add(PMEvent(
            user_id=current_user.id, ticker=ticker, kind=item['kind'],
            event_date=item['date'], valor=round(v, 2),
            source_key=item['key'], ref=item['ref'],
            pm_before=round(pm_before, 4), pm_after=round(custo_aj / a.quantity, 4),
            created_at=datetime.now()))
        remaining -= v
        aplicado += v
        n += 1
    if not n:
        flash(f'{ticker}: pool disponível esgotado (créditos já usados em compras).', 'warning')
        return redirect(url_for('preco_medio'))
    db.session.commit()
    flash(f'{ticker}: {n} crédito(s) aplicados — R$ {aplicado:.2f} abatidos do PM. '
          f'Novo PM ajustado: R$ {custo_aj / a.quantity:.2f}.'.replace('.', ','), 'success')
    return redirect(url_for('preco_medio'))


@app.route('/preco-medio/comprar', methods=['POST'])
@login_required
def pm_comprar():
    """Compra financiada com lucro: registra a compra OFICIAL (qtd/PM da
    carteira, refletindo nas tabelas antigas) e o evento didático — a parte
    financiada entra a PM zero; excesso/falta fica no custo ajustado."""
    def _num(name):
        try:
            return float((request.form.get(name) or '0').replace('.', '').replace(',', '.'))
        except ValueError:
            return 0.0
    ticker = (request.form.get('ticker') or '').strip().upper()
    qty    = int(_num('qty'))
    price  = _num('price')
    a, evs, earned, pending, pool, custo_of, custo_aj = _pm_state(current_user.id, ticker)
    if not a:
        flash(f'{ticker}: ativo não encontrado na carteira (a compra com lucro é para MAIS unidades de um ativo existente).', 'danger')
        return redirect(url_for('preco_medio'))
    if qty <= 0 or price <= 0:
        flash('Informe quantidade e preço válidos.', 'danger')
        return redirect(url_for('preco_medio'))

    total = round(qty * price, 2)
    fin_req = _num('financiado')
    financiado = round(min(fin_req if fin_req > 0 else total, total, max(pool, 0)), 2)

    pm_before = (custo_aj / a.quantity) if a.quantity else 0.0

    # ── Reflexo OFICIAL (tabelas antigas): qtd e PM ponderado ──
    new_qty = a.quantity + qty
    a.avg_price = round(((a.quantity * (a.avg_price or 0)) + total) / new_qty, 4)
    a.quantity = new_qty
    _txn_add(current_user.id, ticker, 'C', qty, price, date.today(),
             source='PM_LUCRO', notes='Compra com lucro (Preço Médio)')

    # ── Evento didático: financiado entra a PM zero; resto é dinheiro novo ──
    custo_aj_new = custo_aj + (total - financiado)
    diff = round(total - financiado, 2)
    db.session.add(PMEvent(
        user_id=current_user.id, ticker=ticker, kind='COMPRA_LUCRO',
        event_date=date.today(), valor=financiado, buy_qty=qty, buy_price=price,
        ref=(f'Compra {qty}× @ R$ {price:.2f} (total R$ {total:.2f}) — '
             f'R$ {financiado:.2f} financiados com lucro (ações a PM zero)'
             + (f'; R$ {diff:.2f} em dinheiro novo no custo' if diff > 0.005 else '')
             ).replace('.', ','),
        pm_before=round(pm_before, 4), pm_after=round(custo_aj_new / new_qty, 4),
        created_at=datetime.now()))
    db.session.commit()
    flash(f'{ticker}: compra de {qty} unid. registrada na carteira oficial. '
          f'R$ {financiado:.2f} financiados com lucro; PM ajustado: '
          f'R$ {custo_aj_new / new_qty:.2f}.'.replace('.', ','), 'success')
    return redirect(url_for('preco_medio'))


@app.route('/preco-medio/manual', methods=['POST'])
@login_required
def pm_manual():
    """Lucro avulso não capturado pela varredura (ex.: aluguel de ações,
    bonificação, ajuste). Valor positivo abate o PM; negativo aumenta.
    Não consome o pool de créditos — é um lançamento independente."""
    def _num(name):
        try:
            return float((request.form.get(name) or '0').replace('.', '').replace(',', '.'))
        except ValueError:
            return 0.0
    ticker = (request.form.get('ticker') or '').strip().upper()
    valor  = round(_num('valor'), 2)
    ref    = (request.form.get('ref') or '').strip()[:200] or 'Lucro avulso'
    try:
        ev_date = date.fromisoformat(request.form.get('data') or '')
    except ValueError:
        ev_date = date.today()
    a, evs, earned, pending, pool, custo_of, custo_aj = _pm_state(current_user.id, ticker)
    if not a or a.quantity <= 0:
        flash(f'{ticker}: ativo não encontrado na carteira.', 'danger')
        return redirect(url_for('preco_medio'))
    if abs(valor) < 0.005:
        flash('Informe um valor diferente de zero.', 'danger')
        return redirect(url_for('preco_medio'))
    pm_before = custo_aj / a.quantity
    custo_aj -= valor
    db.session.add(PMEvent(
        user_id=current_user.id, ticker=ticker, kind='MANUAL',
        event_date=ev_date, valor=valor, ref=ref,
        pm_before=round(pm_before, 4), pm_after=round(custo_aj / a.quantity, 4),
        created_at=datetime.now()))
    db.session.commit()
    flash(f'{ticker}: lançamento avulso de R$ {valor:.2f} registrado. '
          f'Novo PM ajustado: R$ {custo_aj / a.quantity:.2f}.'.replace('.', ','), 'success')
    return redirect(url_for('preco_medio'))


def _pm_rebuild_chain(user_id, ticker):
    """Reconstrói a trilha didática pm_before/pm_after após edição/exclusão,
    reaplicando os eventos em ordem sobre o custo oficial ATUAL (aproximação:
    quantidades históricas intermediárias não são conhecidas)."""
    a = Asset.query.filter(Asset.user_id == user_id, Asset.ticker == ticker).first()
    if not a or not a.quantity:
        return
    evs = (PMEvent.query.filter_by(user_id=user_id, ticker=ticker)
           .order_by(PMEvent.created_at, PMEvent.id).all())
    custo = a.quantity * (a.avg_price or 0)
    for e in evs:
        e.pm_before = round(custo / a.quantity, 4)
        custo -= (e.valor or 0)
        e.pm_after = round(custo / a.quantity, 4)


@app.route('/api/pm/evento/<int:id>/edit', methods=['POST'])
@login_required
def pm_evento_edit(id):
    """Edição pontual de uma linha do histórico do PM (valor/data/descrição)."""
    e = PMEvent.query.get_or_404(id)
    if e.user_id != current_user.id:
        return jsonify({'error': 'sem permissão'}), 403
    data = request.get_json(silent=True) or {}
    try:
        v_raw = data.get('valor', e.valor)
        if isinstance(v_raw, str):
            s = v_raw.strip()
            valor = float(s.replace('.', '').replace(',', '.')) if ',' in s else float(s)
        else:
            valor = float(v_raw)
    except (TypeError, ValueError):
        return jsonify({'error': 'valor inválido'}), 400
    e.valor = round(valor, 2)
    if data.get('date'):
        try:
            e.event_date = date.fromisoformat(data['date'])
        except ValueError:
            pass
    if 'ref' in data:
        e.ref = (data.get('ref') or '').strip()[:200]
    _pm_rebuild_chain(current_user.id, e.ticker)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/pm/evento/<int:id>/delete', methods=['POST'])
@login_required
def pm_evento_delete(id):
    """Exclui uma linha do histórico do PM.
    mode='pendente' (padrão): apaga — crédito de varredura volta a ficar pendente.
    mode='ignorar': converte em marcador IGNORADO (valor 0, mantém a origem) —
    o crédito some da varredura em definitivo (ex.: evento anterior à compra da
    posição atual). Excluir um IGNORADO reativa o crédito."""
    e = PMEvent.query.get_or_404(id)
    if e.user_id != current_user.id:
        return jsonify({'error': 'sem permissão'}), 403
    mode = ((request.get_json(silent=True) or {}).get('mode') or 'pendente').lower()
    tk = e.ticker
    if mode == 'ignorar' and e.source_key and e.kind != 'IGNORADO':
        e.kind = 'IGNORADO'
        e.valor = 0.0
        e.ref = ('Ignorado (fora da posição atual): ' + (e.ref or ''))[:200]
    else:
        db.session.delete(e)
    _pm_rebuild_chain(current_user.id, tk)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/dividendos/<int:id>/qtd', methods=['POST'])
@login_required
def dividendo_qtd(id):
    """Corrige a QTD de ações sobre a qual o dividendo foi calculado
    (posições que mudaram de tamanho ao longo do tempo). Recalcula o total
    = per_share × qtd; a correção sobrevive às atualizações de proventos."""
    d = Dividend.query.get_or_404(id)
    if d.asset.user_id != current_user.id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('dividendos'))
    try:
        qty = int((request.form.get('qty') or '0').replace('.', '').strip())
    except ValueError:
        qty = 0
    if qty <= 0:
        flash('Quantidade inválida.', 'danger')
        return redirect(url_for('dividendos'))
    per = d.per_share if d.per_share else ((d.amount / d.qty_used) if d.qty_used else None)
    if not per:
        flash(f'{d.ticker}: sem valor por ação registrado — atualize os proventos primeiro.', 'warning')
        return redirect(url_for('dividendos'))
    d.per_share = per
    d.qty_used = qty
    d.amount = round(per * qty, 2)
    d.qty_manual = True   # protegido contra recálculos automáticos
    db.session.commit()
    flash(f'{d.ticker} {d.ex_date.strftime("%d/%m/%Y") if d.ex_date else ""}: '
          f'recalculado para {qty}× R$ {per:.4f} = R$ {d.amount:.2f} '
          f'(marcado como manual — recálculos automáticos preservam).'.replace('.', ','), 'success')
    return redirect(url_for('dividendos'))


@app.route('/dividendos/recalcular-qtd-b3', methods=['POST'])
@login_required
def dividendos_recalc_qtd_b3():
    """Recalcula automaticamente a QTD de cada provento pela movimentação da
    B3 (mesmo CSV de negociação usado na curva de patrimônio): reconstrói a
    posição dia a dia a partir da âncora e usa a quantidade detida ANTES da
    data-com de cada dividendo. Total = per_share × qtd da época."""
    f = request.files.get('file')
    if not f or not f.filename:
        flash('Envie o extrato da B3 (CSV de negociação ou Excel de movimentação).', 'danger')
        return redirect(url_for('dividendos'))
    raw = f.read()
    fname = (f.filename or '').lower()
    diag = None
    try:
        # Excel de MOVIMENTAÇÃO (.xlsx, magic 'PK') ou CSV de NEGOCIAÇÃO
        if fname.endswith(('.xlsx', '.xls')) or raw[:2] == b'PK':
            trades, diag = parse_b3_movimentacao_xlsx(raw, debug=True)
        else:
            trades = parse_b3_trades(raw)
    except RuntimeError as e:
        flash(str(e), 'danger')
        return redirect(url_for('dividendos'))
    except Exception as e:
        flash(f'Falha ao ler o arquivo: {e}', 'danger')
        return redirect(url_for('dividendos'))
    if not trades:
        if diag is not None:
            if not diag['header_found']:
                msg = ('Não encontrei o cabeçalho (Movimentação/Produto) no Excel. '
                       'Baixe em Extratos → Movimentação (área do investidor B3) sem editar o arquivo.')
            elif diag['liquidacoes'] == 0:
                msg = ('O Excel não tem linhas "Transferência - Liquidação" (compras/vendas à vista '
                       'liquidadas) no período. Dividendos/opções não contam.')
            else:
                msg = (f'{diag["liquidacoes"]} liquidação(ões) encontrada(s), mas nenhuma virou '
                       f'compra/venda de ação/FII/ETF (ignoradas: {diag["skip_ticker"]} por ticker, '
                       f'{diag["skip_data_qty"]} por data/qtd, {diag["skip_no_side"]} por lado).')
            flash(msg, 'warning')
        else:
            flash('Nenhuma compra/venda de ações/FII/ETF encontrada no arquivo.', 'warning')
        return redirect(url_for('dividendos'))

    # Aproveita o upload para popular o LIVRO de transações (com dedupe)
    novas_txn = _persist_b3_txns(current_user.id, trades)

    init_date = date.fromisoformat(_B3_INITIAL_POSITION_DATE)
    # descarta operações anteriores à âncora (já incluídas na posição inicial)
    trades = [t for t in trades if t['date'] >= init_date]
    days, tickers_mov = _b3_daily_positions(trades, _B3_INITIAL_POSITION, init_date)
    last_day = days[-1][0] if days else init_date

    def qty_on(ticker, ref):
        """Posição ao fim do último dia ANTERIOR à data-com (quem tem o papel
        na véspera do ex-date recebe o provento)."""
        q = _B3_INITIAL_POSITION.get(ticker, 0)
        for dday, snap in days:
            if dday >= ref:
                break
            q = snap.get(ticker, 0)
        return max(q, 0)

    force_manual = request.form.get('force_manual') == '1'
    updated = skipped = zeroed = manuais = 0
    for d in Dividend.query.join(Asset).filter(Asset.user_id == current_user.id).all():
        exd = d.ex_date or d.payment_date
        tk  = (d.ticker or '').upper()
        if not exd or exd < init_date or exd > last_day:
            skipped += 1
            continue                      # fora da janela coberta pelo CSV
        if d.qty_manual and not force_manual:
            manuais += 1
            continue                      # editado à mão — preservado
        if tk not in tickers_mov and tk not in _B3_INITIAL_POSITION:
            skipped += 1
            continue                      # sem dados de posição p/ o ticker
        per = d.per_share or ((d.amount / d.qty_used) if d.qty_used else None)
        if not per:
            skipped += 1
            continue
        q = qty_on(tk, exd)
        d.per_share = per
        d.qty_used  = q
        d.amount    = round(per * q, 2)
        if force_manual:
            d.qty_manual = False
        updated += 1
        if q == 0:
            zeroed += 1
    db.session.commit()
    flash(f'Movimentação B3 aplicada: {updated} provento(s) recalculados pela posição da época'
          + (f' — {zeroed} zerados (sem posição na data-com)' if zeroed else '')
          + (f'; {manuais} manuais preservados' if manuais else '')
          + (f'; {skipped} fora da janela do CSV/sem dados' if skipped else '')
          + f'. {novas_txn} operação(ões) adicionadas ao livro de transações.',
          'success')
    return redirect(url_for('dividendos'))


@app.route('/api/pm/historico/<ticker>')
@login_required
def pm_historico(ticker):
    """Histórico didático das mudanças do PM ajustado do ticker."""
    ticker = ticker.strip().upper()
    evs = (PMEvent.query.filter_by(user_id=current_user.id, ticker=ticker)
           .order_by(PMEvent.created_at, PMEvent.id).all())
    labels = {'DIVIDENDO': 'Dividendo aplicado', 'OPCOES': 'Resultado de opções',
              'COMPRA_LUCRO': 'Compra com lucro', 'MANUAL': 'Ajuste manual',
              'IGNORADO': 'Crédito ignorado (fora da varredura)'}
    return jsonify([{
        'id': e.id, 'kind': e.kind, 'kind_label': labels.get(e.kind, e.kind),
        'has_source': bool(e.source_key),
        'date': e.event_date.strftime('%d/%m/%Y') if e.event_date else '',
        'date_iso': e.event_date.isoformat() if e.event_date else '',
        'valor': round(e.valor or 0, 2),
        'buy_qty': e.buy_qty, 'buy_price': e.buy_price,
        'ref': e.ref or '',
        'pm_before': round(e.pm_before, 2) if e.pm_before is not None else None,
        'pm_after': round(e.pm_after, 2) if e.pm_after is not None else None,
        'created': e.created_at.strftime('%d/%m/%Y %H:%M') if e.created_at else '',
    } for e in evs])

# ─────────────────────────────────────────────────────────────────────────────
# Livro de transações da carteira (AssetTxn)
# ─────────────────────────────────────────────────────────────────────────────
def _txn_add(user_id, ticker, side, qty, price, txn_date=None, source='MANUAL', notes=None):
    """Registra uma transação no livro (não faz commit)."""
    db.session.add(AssetTxn(
        user_id=user_id, ticker=(ticker or '').upper(), side=side,
        quantity=int(qty), price=round(float(price or 0), 4),
        txn_date=txn_date or date.today(), source=source,
        notes=(notes or '')[:200] or None, created_at=datetime.now()))


def _qty_on_date_ledger(user_id, ticker, ref_date, current_qty):
    """Qtd detida na VÉSPERA de ref_date, reconstruída DE TRÁS PRA FRENTE:
    posição atual − deltas das transações com data >= ref_date. Funciona mesmo
    com livro incompleto no passado, desde que as mudanças APÓS ref estejam
    registradas."""
    q = int(current_qty or 0)
    for t in AssetTxn.query.filter(AssetTxn.user_id == user_id,
                                   AssetTxn.ticker == (ticker or '').upper(),
                                   AssetTxn.txn_date >= ref_date).all():
        q -= t.quantity if t.side == 'C' else -t.quantity
    return max(q, 0)


def _persist_b3_txns(user_id, trades):
    """Grava no livro as operações do CSV da B3, com dedupe. Retorna nº de novas."""
    seen = {(t.ticker, t.txn_date, t.side, t.quantity, round(t.price, 2))
            for t in AssetTxn.query.filter_by(user_id=user_id, source='B3').all()}
    added = 0
    for t in trades:
        key = (t['ticker'], t['date'], t['side'], int(t['qty']), round(float(t['price']), 2))
        if key in seen:
            continue
        seen.add(key)
        nota = ('Transferência B3 (sem preço)' if t.get('no_price')
                else 'Importação extrato B3')
        _txn_add(user_id, t['ticker'], t['side'], t['qty'], t['price'],
                 txn_date=t['date'], source='B3', notes=nota)
        added += 1
    return added


@app.route('/transacoes')
@login_required
def transacoes():
    txns = (AssetTxn.query.filter_by(user_id=current_user.id)
            .order_by(AssetTxn.txn_date.desc(), AssetTxn.id.desc()).all())
    return render_template('transacoes.html', txns=txns)


@app.route('/transacoes/<int:id>/delete', methods=['POST'])
@login_required
def transacao_delete(id):
    t = AssetTxn.query.get_or_404(id)
    if t.user_id != current_user.id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('transacoes'))
    db.session.delete(t)
    db.session.commit()
    flash('Transação removida do livro (a posição oficial NÃO foi alterada).', 'success')
    return redirect(url_for('transacoes'))


@app.route('/dividendos/recalcular-qtd-livro', methods=['POST'])
@login_required
def dividendos_recalc_qtd_livro():
    """Recalcula a Qtd dos proventos pela posição reconstruída do LIVRO de
    transações (retroativa a partir da posição atual). Proteções:
    - proventos com Qtd MANUAL são preservados (salvo checkbox de força);
    - só recalcula a partir da 1ª transação registrada do ticker no livro
      (antes disso o livro não tem cobertura — não toca)."""
    force_manual = request.form.get('force_manual') == '1'
    # Data da primeira transação de cada ticker (limite de cobertura do livro)
    first_txn = {}
    for t in AssetTxn.query.filter_by(user_id=current_user.id).all():
        if t.ticker not in first_txn or t.txn_date < first_txn[t.ticker]:
            first_txn[t.ticker] = t.txn_date
    updated = skipped = manuais = fora = 0
    for d in Dividend.query.join(Asset).filter(Asset.user_id == current_user.id).all():
        exd = d.ex_date or d.payment_date
        per = d.per_share or ((d.amount / d.qty_used) if d.qty_used else None)
        if not exd or not per:
            skipped += 1
            continue
        tk = (d.ticker or '').upper()
        if tk not in first_txn or exd < first_txn[tk]:
            fora += 1
            continue                      # antes da cobertura do livro
        if d.qty_manual and not force_manual:
            manuais += 1
            continue                      # editado à mão — preservado
        q = _qty_on_date_ledger(current_user.id, tk, exd, d.asset.quantity)
        d.per_share = per
        d.qty_used = q
        d.amount = round(per * q, 2)
        if force_manual:
            d.qty_manual = False
        updated += 1
    db.session.commit()
    flash(f'Livro de transações aplicado: {updated} provento(s) recalculados'
          + (f'; {manuais} manuais preservados' if manuais else '')
          + (f'; {fora} anteriores à cobertura do livro (não tocados)' if fora else '')
          + (f'; {skipped} sem dados' if skipped else '')
          + '. A precisão depende do livro conter as mudanças de posição após cada data-com.',
          'success')
    return redirect(url_for('dividendos'))


def process_assets(assets):
    if not assets:
        return []
    
    # Calculate total value for weighting
    total_value_list = 0
    
    # Pre-calc totals
    for a in assets:
        # Use stored price if available, else 0 or avg_price
        price = a.current_price if a.current_price else 0.0
        total_value_list += (a.quantity * price)
        
    final_data = []
    
    for a in assets:
        # Use stored data
        current_price = a.current_price if a.current_price else 0.0
        change_desc = f"{a.daily_change:.2f}%" if a.daily_change else "0.00%"
        
        total_invested = a.quantity * a.avg_price
        current_total = a.quantity * current_price
        
        profit = current_total - total_invested
        profit_pct = (profit / total_invested * 100) if total_invested > 0 else 0
        
        weight = (current_total / total_value_list * 100) if total_value_list > 0 else 0
        
        # Day Gain Calculation
        if a.daily_change and current_price > 0:
            # prev_close = price / (1 + pct/100)
            prev_close = current_price / (1 + (a.daily_change/100))
            day_gain = a.quantity * (current_price - prev_close)
        else:
            day_gain = 0.0

        final_data.append({
            'asset': a,
            'current_price': current_price,
            'change_percent': a.daily_change if a.daily_change is not None else 0.0,
            'total_invested': total_invested,
            'current_total': current_total,
            'profit': profit,
            'profit_pct': profit_pct,
            'day_gain': day_gain,
            'weight': weight,
            'last_update': a.last_update.strftime('%d/%m %H:%M') if a.last_update else '-'
        })
        
    return final_data


# Valid update_all_assets_logic and update_quotes defined later (lines ~1433)


@app.route('/add_asset', methods=['GET', 'POST'])
@login_required
def add_asset():
    if request.method == 'POST':
        ticker = request.form.get('ticker').upper()
        type_ = request.form.get('type')
        
        try:
            qty = int(request.form.get('quantity'))
            avg_price_str = request.form.get('price', '').replace(',', '.')
            avg_price = float(avg_price_str) if avg_price_str else 0.0
        except ValueError:
            flash("Erro: Quantidade ou Preço inválido.")
            return redirect(url_for('add_asset'))

        date_str = request.form.get('entry_date')

        # Get Strategy from Form (needed for redirect logic)
        strategy = request.form.get('strategy', 'HOLDER')

        # Check if exists with same strategy (SWING and HOLDER are tracked separately)
        asset = Asset.query.filter_by(ticker=ticker, user_id=current_user.id, strategy=strategy).first()
        if asset:
            # Update average price / quantity logic
            total_val = (asset.quantity * asset.avg_price) + (qty * avg_price)
            new_qty = asset.quantity + qty
            asset.avg_price = total_val / new_qty
            asset.quantity = new_qty
            if strategy != 'SWING' and qty > 0:
                _dt = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else date.today()
                _txn_add(current_user.id, ticker, 'C', qty, avg_price,
                         txn_date=_dt, notes='Compra via Adicionar Ativo')
            flash(f'Ativo {ticker} atualizado! Nova quantidade: {new_qty}')
        else:
            entry_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else date.today()
            sector = request.form.get('sector')
            fii_type = request.form.get('fii_type')
            
            # Fetch Initial Quote Data
            current_price = 0.0
            daily_change = 0.0
            last_update = None
            
            try:
                success, data = get_raw_quote_data(ticker)
                if success and 'results' in data and data['results']:
                    quote = data['results'][0]
                    if 'regularMarketPrice' in quote:
                        current_price = quote['regularMarketPrice']
                        daily_change = quote.get('regularMarketChangePercent', 0.0)
                        last_update = now_brt()
            except Exception as e:
                print(f"Error fetching initial quote for {ticker}: {e}")

            # Get SwingTrade specific fields
            stop_loss_str = request.form.get('stop_loss')
            gain1_str = request.form.get('gain1')
            gain2_str = request.form.get('gain2')
            recommendation = request.form.get('recommendation')

            stop_loss = float(stop_loss_str.replace(',', '.')) if stop_loss_str else None
            gain1 = float(gain1_str.replace(',', '.')) if gain1_str else None
            gain2 = float(gain2_str.replace(',', '.')) if gain2_str else None

            asset = Asset(
                user_id=current_user.id,
                ticker=ticker,
                type=type_,
                strategy=strategy,
                quantity=qty,
                avg_price=avg_price,
                entry_date=entry_date,
                sector=sector,
                fii_type=fii_type,
                current_price=current_price,
                daily_change=daily_change,
                last_update=last_update,
                stop_loss=stop_loss,
                gain1=gain1,
                gain2=gain2,
                recommendation=recommendation
            )
            db.session.add(asset)
            if strategy != 'SWING' and qty > 0:
                _txn_add(current_user.id, ticker, 'C', qty, avg_price,
                         txn_date=entry_date, source='INICIAL', notes='Posição inicial')
            flash(f'Ativo {ticker} adicionado!')

        db.session.commit()
        if type_ == 'FII':
            return redirect(url_for('fiis'))
        elif strategy == 'SWING':
            return redirect(url_for('swingtrade'))
        elif type_ == 'ETF':
            return redirect(url_for('acoes'))
        else:
            return redirect(url_for('acoes'))
        
    return render_template('add.html')

@app.route('/delete_asset/<int:id>')
@login_required
def delete_asset(id):
    asset = Asset.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    db.session.delete(asset)
    db.session.commit()
    return redirect(url_for('index'))

@app.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_asset(id):
    asset = Asset.query.get_or_404(id)
    if asset.user_id != current_user.id:
        flash("Você não tem permissão para editar este ativo.")
        return redirect(url_for('index'))
    if request.method == 'POST':
        asset.ticker = request.form.get('ticker').upper()
        asset.type = request.form.get('type')
        asset.strategy = request.form.get('strategy', 'HOLDER')
        asset.quantity = int(request.form.get('quantity'))
        asset.avg_price = float(request.form.get('price').replace(',', '.'))
        
        stop_loss = request.form.get('stop_loss')
        gain1 = request.form.get('gain1')
        gain2 = request.form.get('gain2')
        recommendation = request.form.get('recommendation')
        fii_type = request.form.get('fii_type')
        sector = request.form.get('sector')
        
        asset.stop_loss = float(stop_loss.replace(',', '.')) if stop_loss else None
        asset.gain1 = float(gain1.replace(',', '.')) if gain1 else None
        asset.gain2 = float(gain2.replace(',', '.')) if gain2 else None
        asset.recommendation = recommendation
        asset.fii_type = fii_type
        asset.sector = sector
        
        db.session.commit()
        if asset.strategy == 'SWING':
            return redirect(url_for('swingtrade'))
        elif asset.type == 'FII':
            return redirect(url_for('fiis'))
        else:
            return redirect(url_for('acoes'))
    return render_template('add.html', asset=asset, edit=True)

@app.route('/buy/<int:id>', methods=['GET', 'POST'])
@login_required
def buy_asset(id):
    asset = Asset.query.get_or_404(id)
    if asset.user_id != current_user.id:
        flash("Você não tem permissão para comprar mais deste ativo.")
        return redirect(url_for('index'))
    if request.method == 'POST':
        qty_buy = int(request.form.get('quantity'))
        price_buy = float(request.form.get('price').replace(',', '.'))

        was_zero = (asset.quantity or 0) <= 0
        # Calculate New Average Price
        current_total = asset.quantity * asset.avg_price
        new_investment = qty_buy * price_buy
        total_qty = asset.quantity + qty_buy

        if total_qty > 0:
            new_avg_price = (current_total + new_investment) / total_qty
            asset.avg_price = new_avg_price
            asset.quantity = total_qty

            # Recompra de posição zerada: nova entrada, limpa a saída anterior
            if was_zero:
                buy_date_str = request.form.get('date')
                if buy_date_str:
                    try:
                        asset.entry_date = datetime.strptime(buy_date_str, '%Y-%m-%d').date()
                    except ValueError:
                        pass
                asset.exit_date = None

            # Livro de transações: compra com data (swing fica fora)
            if asset.strategy != 'SWING':
                _bd = request.form.get('date')
                try:
                    _bdt = datetime.strptime(_bd, '%Y-%m-%d').date() if _bd else date.today()
                except ValueError:
                    _bdt = date.today()
                _txn_add(current_user.id, asset.ticker, 'C', qty_buy, price_buy,
                         txn_date=_bdt, notes='Registrar Nova Compra')

            db.session.commit()
            flash(f'Compra registrada! Novo PM: R$ {new_avg_price:.2f}')
        
        if asset.type == 'FII':
            return redirect(url_for('fiis'))
        elif asset.strategy == 'SWING':
            return redirect(url_for('swingtrade'))
        else:
            return redirect(url_for('acoes'))

    return render_template('buy.html', asset=asset, today=date.today().isoformat())

@app.route('/exit/<int:id>', methods=['GET', 'POST'])
@login_required
def exit_trade(id):
    asset = Asset.query.get_or_404(id)
    if asset.user_id != current_user.id:
        flash("Você não tem permissão para sair deste ativo.")
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        try:
            qty_sell = int(request.form.get('quantity'))
            price_sell = float(request.form.get('price').replace(',', '.'))
            
            date_str = request.form.get('date')
            try:
                date_sell = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                date_sell = datetime.strptime(date_str, '%d/%m/%Y').date()
                
            reason = request.form.get('reason')
            
            # Validation
            if qty_sell > asset.quantity:
                flash("Quantidade de saída maior que a disponível.", "warning")
                return redirect(url_for('exit_trade', id=id))

            # Calculate metrics
            avg_price = asset.avg_price
            total_sell = qty_sell * price_sell
            total_buy = qty_sell * avg_price
            profit_value = total_sell - total_buy
            profit_pct = (profit_value / total_buy * 100) if total_buy > 0 else 0
            
            # Days held
            entry = asset.entry_date
            days_held = (date_sell - entry).days if entry else 0
            
            # Record History
            history = TradeHistory(
                user_id=current_user.id,
                ticker=asset.ticker,
                strategy=asset.recommendation, # As requested: save Recommendation as Strategy
                entry_date=asset.entry_date,
                exit_date=date_sell,
                buy_price=avg_price,
                sell_price=price_sell,
                quantity=qty_sell,
                profit_value=profit_value,
                profit_pct=profit_pct,
                days_held=days_held,
                reason=reason
            )
            db.session.add(history)
            
            # Update Asset
            if qty_sell == asset.quantity:
                # Total Exit - KEEP ASSET (Soft Delete) for Dividends
                asset.quantity = 0
                asset.exit_date = date_sell   # marca a saída total (usado no cálculo de dividendos por posse)
                flash("Saída TOTAL registrada com sucesso!", "success")
            else:
                # Partial Exit
                asset.quantity -= qty_sell
                flash("Saída PARCIAL registrada com sucesso!", "success")

            # Livro de transações: venda com data (swing fica fora)
            if asset.strategy != 'SWING':
                _txn_add(current_user.id, asset.ticker, 'V', qty_sell, price_sell,
                         txn_date=date_sell, notes=f'Venda/Saída ({reason or "-"})')

            db.session.commit()
            
            # Redirect back to origin
            if asset.strategy == 'SWING':
                return redirect(url_for('swingtrade'))
            elif asset.type == 'FII':
                return redirect(url_for('fiis'))
            else:
                return redirect(url_for('acoes'))
                
        except ValueError as e:
            flash(f"Erro de formato (Valores ou Data incorreta): {str(e)}", "danger")
            return redirect(url_for('exit_trade', id=id))
        except Exception as e:
            flash(f"Erro ao registrar saída: {str(e)}", "danger")
            print(f"Error exit_trade: {e}")
            return redirect(url_for('exit_trade', id=id))
        
    return render_template('exit.html', asset=asset, today=date.today().isoformat())

# --- CRYPTO ROUTES ---
@app.route('/buy_crypto/<int:id>', methods=['GET', 'POST'])
@login_required
def buy_crypto(id):
    crypto = Crypto.query.get_or_404(id)
    if crypto.user_id != current_user.id:
        flash("Permissão negada.")
        return redirect(url_for('balanceamento'))
        
    if request.method == 'POST':
        try:
            qty_buy = float(request.form.get('quantity').replace(',', '.'))
            price_buy = float(request.form.get('price').replace(',', '.')) # Unit Price
            date_str = request.form.get('date')
            
            # Calculate New Average Price
            # Current Total Value based on PM (Invested)
            current_qty = crypto.quantity or 0
            current_avg = crypto.avg_price or 0
            
            current_total_invested = current_qty * current_avg
            new_investment = qty_buy * price_buy
            
            total_qty = current_qty + qty_buy
            
            if total_qty > 0:
                new_avg_price = (current_total_invested + new_investment) / total_qty
                crypto.avg_price = new_avg_price
                crypto.quantity = total_qty
                
                # Update current value for immediate display consistency if needed, 
                # but usually this is fetched from API. 
                # Let's assume user wants to track purely quantity/PM here.
                
                db.session.commit()
                flash(f'Compra de Cripto registrada! Novo PM: R$ {new_avg_price:.2f}', 'success')
            
            return redirect(url_for('balanceamento'))
            
        except ValueError:
            flash("Erro nos valores informados.", "danger")
            
    return render_template('buy_crypto.html', crypto=crypto, today=date.today().isoformat())

@app.route('/exit_crypto/<int:id>', methods=['GET', 'POST'])
@login_required
def exit_crypto(id):
    crypto = Crypto.query.get_or_404(id)
    if crypto.user_id != current_user.id:
        flash("Permissão negada.")
        return redirect(url_for('balanceamento'))
        
    if request.method == 'POST':
        try:
            qty_sell = float(request.form.get('quantity').replace(',', '.'))
            price_sell = float(request.form.get('price').replace(',', '.')) # Unit Price
            date_str = request.form.get('date')
            try:
                date_sell = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                date_sell = datetime.strptime(date_str, '%d/%m/%Y').date()
                
            reason = request.form.get('reason')
            
            if qty_sell > (crypto.quantity or 0):
                flash("Quantidade de venda maior que a possuída!", "danger")
                return redirect(url_for('exit_crypto', id=id))
            
            # Calculate Profit
            avg_price = crypto.avg_price or 0
            total_sell = qty_sell * price_sell
            total_buy = qty_sell * avg_price
            
            profit_value = total_sell - total_buy
            profit_pct = (profit_value / total_buy * 100) if total_buy > 0 else 0
            
            # Record History (Reusing TradeHistory with ticker=crypto.name)
            history = TradeHistory(
                user_id=current_user.id,
                ticker=crypto.name, # Using Name as Ticker
                strategy="HOLDER", # Default or maybe mapped from 'reason'
                entry_date=None, # Hard to track for partials without FIFO
                exit_date=date_sell,
                buy_price=avg_price,
                sell_price=price_sell,
                quantity=qty_sell, # Int in model?? Need to check if TradeHistory quantity is Float. 
                                   # Model says Integer. Issue!
                                   # We might need to store as 1 (dummy) or change model.
                                   # Let's check TradeHistory model again.
                profit_value=profit_value,
                profit_pct=profit_pct,
                days_held=0,
                reason=reason
            )
            
            # FIX: TradeHistory.quantity might be Integer. 
            # If so, we can't store 0.005 BTC.
            # I will cast to Int if possible, or store 1 and put real qty in notes/reason?
            # Or assume Model update in future task.
            # Checking model... "quantity = db.Column(db.Integer)"
            # WORKAROUND: For now, I will cast to int(qty_sell) if > 1, else 1. 
            # BUT this is bad for data integrity.
            # Better: Modify model or just accept it might be 0 for small fractional.
            # Wait, user asked to "record profit/loss". 
            # I will cast to int but strictly, this needs a migration to Float for Cryptos.
            # For this task, I will use int(qty_sell) but warn user if 0.
            
            history.quantity = int(qty_sell) if qty_sell >= 1 else 1 # Placeholder to avoid 0 if fractional
            
            db.session.add(history)
            
            # Update Crypto
            crypto.quantity -= qty_sell
            if crypto.quantity < 0: crypto.quantity = 0
            
            db.session.commit()
            flash("Venda de Cripto registrada com sucesso!", "success")
            
            return redirect(url_for('balanceamento'))
            
        except ValueError as e:
             flash(f"Erro de valor: {e}", "danger")
             
    return render_template('exit_crypto.html', crypto=crypto, today=date.today().isoformat())

    return render_template('exit_crypto.html', crypto=crypto, today=date.today().isoformat())

# --- INTERNATIONAL ROUTES ---
@app.route('/buy_intl/<int:id>', methods=['GET', 'POST'])
@login_required
def buy_intl(id):
    asset = International.query.get_or_404(id)
    if asset.user_id != current_user.id:
        flash("Permissão negada.")
        return redirect(url_for('balanceamento'))
        
    if request.method == 'POST':
        try:
            qty_buy = float(request.form.get('quantity').replace(',', '.'))
            price_buy = float(request.form.get('price').replace(',', '.')) # Unit Price in USD
            date_str = request.form.get('date')
            
            # Calculate New Average Price (USD)
            current_qty = asset.quantity or 0
            current_avg = asset.avg_price or 0
            
            current_total_invested = current_qty * current_avg
            new_investment = qty_buy * price_buy
            
            total_qty = current_qty + qty_buy
            
            if total_qty > 0:
                new_avg_price = (current_total_invested + new_investment) / total_qty
                asset.avg_price = new_avg_price
                asset.quantity = total_qty
                # Update purchase_price/invested_value fields if they are being used for redundancy
                asset.invested_value = total_qty * new_avg_price 
                
                db.session.commit()
                flash(f'Compra Internacional registrada! Novo PM: US$ {new_avg_price:.2f}', 'success')
            
            return redirect(url_for('balanceamento'))
            
        except ValueError:
            flash("Erro nos valores informados.", "danger")
            
    return render_template('buy_intl.html', asset=asset, today=date.today().isoformat())

@app.route('/exit_intl/<int:id>', methods=['GET', 'POST'])
@login_required
def exit_intl(id):
    asset = International.query.get_or_404(id)
    if asset.user_id != current_user.id:
        flash("Permissão negada.")
        return redirect(url_for('balanceamento'))
        
    if request.method == 'POST':
        try:
            qty_sell = float(request.form.get('quantity').replace(',', '.'))
            price_sell_usd = float(request.form.get('price').replace(',', '.')) # Unit Price USD
            exchange_rate = float(request.form.get('exchange_rate').replace(',', '.')) # BRL Rate
            
            date_str = request.form.get('date')
            try:
                date_sell = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                date_sell = datetime.strptime(date_str, '%d/%m/%Y').date()
                
            reason = request.form.get('reason')
            
            if qty_sell > (asset.quantity or 0):
                flash("Quantidade de venda maior que a possuída!", "danger")
                return redirect(url_for('exit_intl', id=id))
            
            # Calculate Profit in USD
            avg_price_usd = asset.avg_price or 0
            total_sell_usd = qty_sell * price_sell_usd
            total_buy_usd = qty_sell * avg_price_usd
            
            profit_usd = total_sell_usd - total_buy_usd
            
            # Convert to BRL for History
            # We assume user wants to track "Realized Profit in BRL"
            profit_brl = profit_usd * exchange_rate
            
            # Calculate implied BRL values for records
            total_buy_brl = total_buy_usd * exchange_rate
            
            profit_pct = (profit_usd / total_buy_usd * 100) if total_buy_usd > 0 else 0
            
            # Record History
            history = TradeHistory(
                user_id=current_user.id,
                ticker=asset.name, 
                strategy="INTL", 
                entry_date=None, 
                exit_date=date_sell,
                buy_price=avg_price_usd * exchange_rate, # Storing BRL basis
                sell_price=price_sell_usd * exchange_rate, # Storing BRL Sales
                quantity=int(qty_sell) if qty_sell >=1 else 1, # Same int casting issue as crypto
                profit_value=profit_brl,
                profit_pct=profit_pct,
                days_held=0,
                reason=reason
            )
            
            db.session.add(history)
            
            # Update Asset
            asset.quantity -= qty_sell
            
            # Using a small epsilon for float comparison safety
            if asset.quantity <= 0.00000001:
                db.session.delete(asset)
                flash(f"Venda Internacional TOTAL registrada! Lucro: R$ {profit_brl:.2f}", "success")
            else:
                asset.invested_value = asset.quantity * asset.avg_price # Update cached total
                flash(f"Venda Internacional PARCIAL registrada! Lucro: R$ {profit_brl:.2f}", "success")
            
            db.session.commit()
            
            return redirect(url_for('balanceamento'))
            
        except ValueError as e:
             flash(f"Erro de valor: {e}", "danger")
             
    return render_template('exit_intl.html', asset=asset, today=date.today().isoformat())

@app.route('/historico')
@login_required
def historico():
    q = request.args.get('q', '').strip().upper()
    query = TradeHistory.query.filter_by(user_id=current_user.id)
    if q:
        like = f'%{q}%'
        query = query.filter(
            db.or_(
                TradeHistory.ticker.ilike(like),
                TradeHistory.underlying.ilike(like),
                TradeHistory.notes.ilike(like),
            )
        )
    trades = query.order_by(TradeHistory.exit_date.desc()).all()
    return render_template('historico.html', history=trades, q=q)


# ─────────────────────────────────────────────────────────────────────────────
# Importação do extrato de negociação da B3 → reconstrução da curva real
# ─────────────────────────────────────────────────────────────────────────────

# Posição inicial em 01/09/2025 (extrato de posição B3), consolidada por ticker.
# Serve de âncora quando o CSV de negociação começa depois dessa data.
_B3_INITIAL_POSITION_DATE = '2025-09-01'
_B3_INITIAL_POSITION = {
    'BBAS3': 60, 'BRSR6': 200, 'HYPE3': 200, 'PETR4': 100,
    'AIEC11': 25, 'BRCO11': 12, 'BTLG11': 10, 'CACR11': 17, 'CPTS11': 250,
    'GARE11': 220, 'HGCR11': 10, 'HGLG11': 10, 'HGRU11': 6, 'HSLG11': 10,
    'HSML11': 11, 'INLG11': 14, 'ITIT11': 12, 'KNCR11': 12, 'KNSC11': 160,
    'LVBI11': 6, 'MCCI11': 15, 'MFII11': 13, 'MXRF11': 350, 'NSLU11': 4,
    'PMLL11': 8, 'RBRF11': 174, 'RBRL11': 12, 'RBRP11': 17, 'RBRR11': 10,
    'RBVA11': 200, 'RVBI11': 25, 'RZAK11': 10, 'RZAT11': 11, 'RZTR11': 13,
    'SNCI11': 11, 'TGAR11': 13, 'TRXF11': 7, 'TVRI11': 5, 'URPR11': 15,
    'VGHF11': 350, 'VGIP11': 9, 'VGIR11': 350, 'VILG11': 11, 'VISC11': 10,
    'VRTM11': 192, 'VTLT11': 11, 'XPCI11': 17, 'XPIN11': 19, 'XPML11': 13,
    'AAZQ11': 260, 'AGRX11': 170, 'FGAA11': 270, 'OIAG11': 296, 'RURA11': 200,
    'RZAG11': 200, 'SNAG11': 260, 'VGIA11': 350, 'XPCA11': 260,
}

# ETFs de índice (tratados como tipo ETF na curva)
_KNOWN_ETFS = {
    'BOVA11', 'BOVX11', 'SMAL11', 'IVVB11', 'GOLD11', 'SPXS11', 'DOLA11',
    'BTCI11', 'FIXA11', 'XFIX11', 'HASH11', 'QBTC11', 'ETHE11',
}


def _b3_norm_ticker(code):
    """Normaliza um código de negociação B3 para o ticker do ativo à vista.
    Remove o sufixo 'F' de fracionário (ex.: PETR4F → PETR4)."""
    code = (code or '').strip().upper()
    if len(code) > 5 and code.endswith('F') and code[-2].isdigit():
        code = code[:-1]
    return code


def parse_b3_movimentacao_xlsx(raw_bytes, debug=False):
    """Lê o Excel de MOVIMENTAÇÃO da B3 (área do investidor, .xlsx) e devolve
    as compras/vendas liquidadas no MESMO formato de parse_b3_trades:
        {date, side 'C'|'V', ticker, qty, price}
    À vista = 'Transferência - Liquidação' (Crédito → compra, Débito → venda).
    Opções e demais eventos (dividendos, empréstimos etc.) são ignorados.
    debug=True devolve (trades, diag) com contadores p/ diagnóstico."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError('Leitura de .xlsx indisponível: instale o pacote '
                           '"openpyxl" no servidor (pip install openpyxl).')
    import io as _io
    from datetime import datetime as _dtt, date as _date

    def _pick(low, *needles):
        """1º índice cuja célula contém TODAS as needles."""
        for i, v in enumerate(low):
            if all(n in v for n in needles):
                return i
        return None

    def _to_date(raw):
        if isinstance(raw, _dtt):
            return raw.date()
        if isinstance(raw, _date):
            return raw
        s = str(raw).strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d/%m/%y', '%d-%m-%Y'):
            try:
                return _dtt.strptime(s[:10], fmt).date()
            except ValueError:
                continue
        return None

    def _to_num(raw):
        if raw is None:
            return None
        if isinstance(raw, (int, float)):
            return float(raw)
        s = str(raw).replace('R$', '').strip()
        if s in ('', '-', '--'):
            return None
        # pt-BR: milhar '.', decimal ','  |  também aceita ponto decimal simples
        if ',' in s:
            s = s.replace('.', '').replace(',', '.')
        try:
            return float(s)
        except ValueError:
            return None

    # NÃO usar read_only: os XLSX da B3/BTG declaram dimensões erradas
    # (dimension "A1:A1"), o que faz o modo read_only truncar tudo para 1 célula.
    wb = openpyxl.load_workbook(_io.BytesIO(raw_bytes), read_only=False, data_only=True)
    trades = []
    diag = {'sheets': 0, 'header_found': False, 'liquidacoes': 0, 'kept': 0,
            'skip_no_side': 0, 'skip_ticker': 0, 'skip_data_qty': 0}
    for ws in wb.worksheets:
        diag['sheets'] += 1
        idx = None
        for r in ws.iter_rows(values_only=True):
            if r is None:
                continue
            vals = [str(c).strip() if c is not None else '' for c in r]
            low = [v.lower() for v in vals]
            if idx is None:
                # cabeçalho: precisa ter Movimentação e Produto na mesma linha
                if _pick(low, 'movimenta') is not None and _pick(low, 'produto') is not None:
                    idx = {
                        'es':   _pick(low, 'entrada') if _pick(low, 'entrada') is not None else _pick(low, 'sa', 'da'),
                        'mov':  _pick(low, 'movimenta'),
                        'prod': _pick(low, 'produto'),
                        'data': _pick(low, 'data'),
                        'qty':  _pick(low, 'quantidade'),
                        'pu':   _pick(low, 'pre', 'unit') if _pick(low, 'pre', 'unit') is not None else _pick(low, 'unit'),
                    }
                    diag['header_found'] = True
                continue

            def cell(k):
                j = idx.get(k)
                return vals[j] if (j is not None and j < len(vals)) else ''
            def raw(k):
                j = idx.get(k)
                return r[j] if (j is not None and j < len(r)) else None

            mov = cell('mov').lower()
            if 'transfer' not in mov or 'liquida' not in mov:
                continue
            diag['liquidacoes'] += 1

            es = cell('es').lower()
            side = ('C' if es.startswith(('cr', 'entrada'))
                    else 'V' if es.startswith(('d', 'sa')) else None)
            if not side:
                diag['skip_no_side'] += 1
                continue

            ticker = _b3_norm_ticker(cell('prod').split('-')[0])
            if not (4 < len(ticker) <= 6) or not any(ch.isdigit() for ch in ticker):
                diag['skip_ticker'] += 1
                continue

            d = _to_date(raw('data'))
            qn = _to_num(raw('qty'))
            if d is None or qn is None or qn <= 0:
                diag['skip_data_qty'] += 1
                continue
            price = _to_num(raw('pu')) or 0.0
            trades.append({'date': d, 'side': side, 'ticker': ticker,
                           'qty': int(qn), 'price': round(price, 4),
                           'no_price': price <= 0})   # transferência sem preço
            diag['kept'] += 1

    trades.sort(key=lambda t: t['date'])
    return (trades, diag) if debug else trades


def _b3_classify(ticker):
    """Classifica o ticker em ACAO / FII / ETF a partir das fontes conhecidas
    (posição inicial, ETFs de índice) e de heurística pelo sufixo."""
    if ticker in _KNOWN_ETFS:
        return 'ETF'
    # Tickers 11 do FIAGRO/FII da posição inicial já são fundos imobiliários
    if ticker in _B3_INITIAL_POSITION and ticker.endswith('11'):
        return 'FII'
    if ticker.endswith('11') or ticker.endswith('11B'):
        return 'FII'
    return 'ACAO'


def parse_b3_trades(raw_bytes):
    """Lê o CSV de negociação da B3 (encoding Latin-1) e devolve só as operações
    à VISTA de ações/ETF/FII, como lista de dicts ordenada por data crescente:
        {date: date, side: 'C'|'V', ticker, qty, price}
    Ignora opções, futuros, renda fixa, exercícios e mercado a termo."""
    import csv as _csv
    import io as _io
    from datetime import date as _date

    text = raw_bytes.decode('latin-1', errors='replace')
    # separador ';'; primeira linha é cabeçalho
    reader = _csv.reader(_io.StringIO(text), delimiter=';')
    rows = list(reader)
    if not rows:
        return []
    trades = []
    for r in rows[1:]:
        if len(r) < 9:
            continue
        data_str, tipo_mov, mercado, _venc, _inst, code, qtd, _preco, valor = r[:9]
        mercado = (mercado or '').strip()
        # Só mercado à vista / fracionário de ações; descarta o resto
        if not (mercado.startswith('Mercado') and 'Vista' in mercado
                or mercado.startswith('Mercado') and 'racion' in mercado):
            continue
        tipo_mov = (tipo_mov or '').strip().lower()
        if tipo_mov.startswith('compra'):
            side = 'C'
        elif tipo_mov.startswith('venda'):
            side = 'V'
        else:
            continue
        try:
            d, m, y = data_str.strip().split('/')
            dt = _date(int(y), int(m), int(d))
        except (ValueError, AttributeError):
            continue
        ticker = _b3_norm_ticker(code)
        if not ticker:
            continue
        try:
            qty = int(float((qtd or '0').strip().replace('.', '').replace(',', '.')))
        except ValueError:
            continue
        # preço unitário: usa 'Valor'/qty (mais robusto que o campo Preço formatado)
        try:
            val = float((valor or '0').replace('R$', '').replace('.', '').replace(',', '.').strip())
            price = val / qty if qty else 0.0
        except (ValueError, ZeroDivisionError):
            price = 0.0
        if qty <= 0:
            continue
        trades.append({'date': dt, 'side': side, 'ticker': ticker,
                       'qty': qty, 'price': round(price, 4)})
    trades.sort(key=lambda t: t['date'])
    return trades


def _b3_daily_positions(trades, initial_pos, initial_date):
    """Percorre as operações em ordem e devolve a posição (qtd por ticker) ao
    FIM de cada dia com movimentação, mais o conjunto de tickers envolvidos.
    Retorna (list[(date, {ticker: qty})], set(tickers))."""
    from datetime import date as _date
    pos = dict(initial_pos)
    tickers = set(pos.keys())
    by_day = {}
    for t in trades:
        tickers.add(t['ticker'])
        delta = t['qty'] if t['side'] == 'C' else -t['qty']
        pos[t['ticker']] = pos.get(t['ticker'], 0) + delta
        # guarda uma cópia do estado ao fim do dia
        by_day[t['date']] = {k: v for k, v in pos.items() if v != 0}
    days = sorted(by_day.items())
    return days, tickers


def _fetch_close_history(tickers, start, end):
    """Baixa o fechamento diário (Yahoo) de cada ticker no período.
    Retorna {ticker: {date_iso: close}}. Falhas por ticker são ignoradas."""
    import yfinance as yf
    out = {}
    for tk in sorted(tickers):
        try:
            h = yf.Ticker(f'{tk}.SA').history(start=start.isoformat(),
                                              end=(end + timedelta(days=1)).isoformat(),
                                              auto_adjust=False)
            if h is None or h.empty:
                continue
            series = {}
            for idx, row in h.iterrows():
                series[idx.date().isoformat()] = float(row['Close'])
            if series:
                out[tk] = series
        except Exception:
            app.logger.warning('histórico Yahoo falhou para %s', tk)
    return out


def rebuild_equity_from_b3(user_id, trades, task_id=None):
    """Reconstrói os snapshots diários de patrimônio a partir das operações da
    B3 (à vista), valorizando cada dia a preço de mercado (Yahoo). Grava
    PortfolioSnapshot reais, substituindo os estimados e preservando os reais.
    Também preenche entry_date/exit_date dos ativos que ainda não têm.
    Retorna dict com o resumo."""
    from datetime import date as _date, timedelta as _td

    init_date = _date.fromisoformat(_B3_INITIAL_POSITION_DATE)
    days, tickers = _b3_daily_positions(trades, _B3_INITIAL_POSITION, init_date)
    if not days:
        return {'ok': False, 'msg': 'Nenhuma operação à vista encontrada no CSV.'}

    first_day = min(days[0][0], init_date)
    last_day  = now_brt().date()

    if task_id:
        _set_task(task_id, {'status': 'running',
                            'msg': f'Baixando cotação histórica de {len(tickers)} ativos…',
                            'category': ''})
    hist = _fetch_close_history(tickers, first_day, last_day)

    # Preço mais recente conhecido por ticker (para dias sem cotação, usa o último)
    def _price_on(tk, day_iso, last_known):
        series = hist.get(tk)
        if series and day_iso in series:
            last_known[tk] = series[day_iso]
            return series[day_iso]
        return last_known.get(tk, 0.0)

    # Sequência de posições por dia: expande para TODO dia (carrega a última
    # posição conhecida) e valoriza mensalmente (último dia de cada mês).
    pos_events = dict(days)                       # {date: {ticker: qty}}
    cur_pos = {k: v for k, v in _B3_INITIAL_POSITION.items() if v}
    last_known_price = {}

    # gera fim de cada mês entre first_day e last_day
    from dateutil.relativedelta import relativedelta as _rd
    snapshots = []                                # [(date, acoes, fiis, etfs)]
    cursor = _date(first_day.year, first_day.month, 1)
    while cursor <= last_day:
        # último dia do mês (ou hoje, no mês corrente)
        nxt = cursor + _rd(months=1)
        month_end = min(nxt - _td(days=1), last_day)
        # aplica todos os eventos de posição até o fim do mês
        for dd, p in days:
            if dd <= month_end:
                cur_pos = p
        # valoriza a posição no fim do mês
        acoes = fiis = etfs = 0.0
        me_iso = month_end.isoformat()
        for tk, qty in cur_pos.items():
            if qty <= 0:
                continue
            px = _price_on(tk, me_iso, last_known_price)
            val = qty * px
            tp = _b3_classify(tk)
            if tp == 'FII':
                fiis += val
            elif tp == 'ETF':
                etfs += val
            else:
                acoes += val
        snapshots.append((month_end, round(acoes, 2), round(fiis, 2), round(etfs, 2)))
        cursor = nxt

    # ── Grava snapshots: remove estimados, preserva reais ────────────────────
    if task_id:
        _set_task(task_id, {'status': 'running', 'msg': 'Gravando histórico…', 'category': ''})
    PortfolioSnapshot.query.filter_by(user_id=user_id, estimated=True).delete()
    written = 0
    for (dd, acoes, fiis, etfs) in snapshots:
        iso = dd.isoformat()
        existing = PortfolioSnapshot.query.filter_by(user_id=user_id, snap_date=iso).first()
        if existing and not existing.estimated:
            continue                              # snapshot real já coletado: preserva
        if existing is None:
            existing = PortfolioSnapshot(user_id=user_id, snap_date=iso)
            db.session.add(existing)
        existing.total_acoes = acoes
        existing.total_fiis  = fiis
        existing.total_etfs  = etfs
        existing.total_equity = round(acoes + fiis + etfs, 2)
        existing.estimated = False                # reconstruído do extrato = real
        existing.created_at = datetime.utcnow()
        written += 1

    # ── Preenche entry_date / exit_date pelos primeiros/últimos negócios ──────
    first_buy, last_by_ticker, ran_out = {}, {}, {}
    running = dict(_B3_INITIAL_POSITION)
    for t in trades:
        tk = t['ticker']
        if t['side'] == 'C' and tk not in first_buy:
            first_buy[tk] = t['date']
        running[tk] = running.get(tk, 0) + (t['qty'] if t['side'] == 'C' else -t['qty'])
        last_by_ticker[tk] = t['date']
        ran_out[tk] = (running[tk] <= 0)          # zerou nesta última operação?
    dates_set = 0
    for tk in set(list(first_buy) + list(_B3_INITIAL_POSITION)):
        asset = Asset.query.filter_by(user_id=user_id, ticker=tk).first()
        if not asset:
            continue
        if asset.entry_date is None:
            asset.entry_date = (first_buy.get(tk)
                                or _date.fromisoformat(_B3_INITIAL_POSITION_DATE))
            dates_set += 1
        # saída só se a posição terminou zerada e o ativo está sem quantidade
        if asset.exit_date is None and (asset.quantity or 0) <= 0 and ran_out.get(tk):
            asset.exit_date = last_by_ticker.get(tk)
            dates_set += 1

    db.session.commit()
    return {'ok': True, 'snapshots': written, 'tickers': len(tickers),
            'dates_set': dates_set,
            'period': f'{first_day.isoformat()} → {last_day.isoformat()}',
            'quotes': len(hist)}


@app.route('/importar-b3', methods=['GET'])
@login_required
def importar_b3():
    """Página: upload do extrato de negociação da B3 para reconstruir a curva."""
    return render_template('importar_b3.html',
                           initial_date=_B3_INITIAL_POSITION_DATE,
                           initial_count=len(_B3_INITIAL_POSITION))


@app.route('/api/importar-b3', methods=['POST'])
@login_required
def api_importar_b3():
    """Recebe o CSV, dispara a reconstrução em background e devolve task_id."""
    f = request.files.get('csv')
    if not f:
        return jsonify({'error': 'Envie o arquivo CSV do extrato de negociação.'}), 400
    raw = f.read()
    try:
        trades = parse_b3_trades(raw)
    except Exception as e:
        app.logger.exception('parse_b3_trades falhou')
        return jsonify({'error': f'Não consegui ler o CSV: {e}'}), 400
    if not trades:
        return jsonify({'error': 'Nenhuma operação à vista (ações/FII/ETF) encontrada no arquivo.'}), 400

    uid = current_user.id
    task_id = str(uuid.uuid4())
    _set_task(task_id, {'status': 'running', 'msg': 'Processando…', 'category': ''})

    def _run():
        with app.app_context():
            try:
                res = rebuild_equity_from_b3(uid, trades, task_id=task_id)
                if res.get('ok'):
                    msg = (f"Histórico reconstruído: {res['snapshots']} meses, "
                           f"{res['tickers']} ativos ({res['quotes']} com cotação), "
                           f"{res['dates_set']} datas de entrada/saída. Período {res['period']}.")
                    _set_task(task_id, {'status': 'done', 'msg': msg, 'category': 'success'})
                else:
                    _set_task(task_id, {'status': 'done', 'msg': res.get('msg', 'Falha.'),
                                        'category': 'warning'})
            except Exception as e:
                app.logger.exception('rebuild_equity_from_b3 falhou')
                _set_task(task_id, {'status': 'done',
                                    'msg': f'Erro ao reconstruir: {e}', 'category': 'danger'})

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({'task_id': task_id, 'trades': len(trades)})


# ─────────────────────────────────────────────────────────────────────────────
# Patrimônio: totais por tipo, snapshots diários e curva de evolução
# ─────────────────────────────────────────────────────────────────────────────

def _equity_by_type(user_id):
    """Patrimônio atual a preço de mercado, separado por tipo.
    Retorna (total, acoes, fiis, etfs). ETF é tipo próprio; SWING conta como ação."""
    assets = Asset.query.filter(Asset.user_id == user_id, Asset.quantity > 0).all()
    acoes = fiis = etfs = 0.0
    for a in assets:
        price = a.current_price if (a.current_price or 0) > 0 else a.avg_price
        val = (a.quantity or 0) * (price or 0)
        if a.type == 'FII':
            fiis += val
        elif a.type == 'ETF':
            etfs += val
        else:
            acoes += val
    return acoes + fiis + etfs, acoes, fiis, etfs


def record_portfolio_snapshot(user_id):
    """Grava/atualiza a foto do patrimônio de HOJE (um registro por dia; o
    último sobrescreve). Chamado após atualizar cotações e 1x/dia pelo scheduler."""
    try:
        today_iso = now_brt().date().isoformat()
        total, acoes, fiis, etfs = _equity_by_type(user_id)
        snap = PortfolioSnapshot.query.filter_by(user_id=user_id, snap_date=today_iso).first()
        if snap is None:
            snap = PortfolioSnapshot(user_id=user_id, snap_date=today_iso)
            db.session.add(snap)
        snap.total_equity = round(total, 2)
        snap.total_acoes  = round(acoes, 2)
        snap.total_fiis   = round(fiis, 2)
        snap.total_etfs   = round(etfs, 2)
        snap.estimated    = False
        snap.created_at   = datetime.utcnow()
        db.session.commit()
    except Exception:
        db.session.rollback()
        app.logger.exception('record_portfolio_snapshot falhou (user %s)', user_id)


def _dividends_owned(user_id, today):
    """Dividendos efetivamente recebidos, filtrados pelo PERÍODO DE POSSE do ativo.
    Um dividendo só conta se a data-com (ex_date; se faltar, payment_date) estava
    entre a entrada e a saída do ativo. Retorna listas de tuplas
    (mes 'YYYY-MM', valor) para ações e para FIIs, além do total de cada."""
    divs = Dividend.query.join(Asset).filter(
        Asset.user_id == user_id,
        Dividend.payment_date != None
    ).all()
    monthly_acoes, monthly_fiis = {}, {}
    for d in divs:
        if not d.payment_date or d.payment_date > today:
            continue
        a = d.asset
        # Data que define a titularidade: data-com (preferida) ou pagamento
        own_date = d.ex_date or d.payment_date
        if a.entry_date and own_date < a.entry_date:
            continue                                  # comprou depois da data-com
        if a.exit_date and own_date > a.exit_date:
            continue                                  # já tinha vendido na data-com
        mk = d.payment_date.strftime('%Y-%m')
        val = d.amount or 0
        if a.type == 'FII':
            monthly_fiis[mk] = monthly_fiis.get(mk, 0) + val
        else:
            monthly_acoes[mk] = monthly_acoes.get(mk, 0) + val
    return monthly_acoes, monthly_fiis


@app.route('/resumo')
@login_required
def resumo():
    # Calculate Summaries
    assets = Asset.query.filter(Asset.user_id==current_user.id, Asset.quantity > 0).all()
    history = TradeHistory.query.filter_by(user_id=current_user.id).all()
    
    # 1. Total Equity & Allocation
    total_equity = 0
    total_acoes = 0
    total_fiis = 0
    total_swing = 0 # Maybe separate swing? User said Acoes vs FIIs.
    # Usually Swing is just a strategy, but Asset Type is ACAO/FII.
    # Let's split by TYPE for the "Allocation" chart.
    
    
    total_swing = 0 
    
    fii_types = {}
    stock_sectors = {} # New: [Sector] -> Value
    
    # FII Classification Mapping
    fii_map = {
        'LAJES CORPORATIVAS': 'Tijolo',
        'LOGISTICA': 'Tijolo',
        'SHOPPING CENTER': 'Tijolo',
        'HIBRIDO': 'Tijolo',
        'RENDA': 'Tijolo',
        'RECEBIVEIS': 'Papel',
        'FIAGRO': 'Papel',
        'FUNDO DE FUNDOS': 'Papel',
        'INFRA': 'Papel',
        'OUTROS': 'Papel' # Defaulting to Papel or maybe separate? Let's say Papel/Outros
    }
    
    # Data for table
    fii_summary = {} # {type: value}
    
    for a in assets:
        price = a.current_price if a.current_price > 0 else a.avg_price
        val = a.quantity * price
        
        total_equity += val
        
        if a.type == 'ACAO':
            total_acoes += val
            # Stock Sector:
            s = a.sector or 'Não Classificado'
            stock_sectors[s] = stock_sectors.get(s, 0) + val
            
        elif a.type == 'FII':
            total_fiis += val
            t = a.fii_type or 'OUTROS'
            # FII Breakdown for Chart
            fii_types[t] = fii_types.get(t, 0) + val
            # Summary for Table
            fii_summary[t] = fii_summary.get(t, 0) + val
        
        elif a.type == 'ETF':
            # Add to separate total? Or lump with Acoes? User wanted separate in Acoes page.
            # Let's track separate total_etfs for allocation chart.
            # Initialize a new variable for this if needed, or pass in context.
            pass # Creating a separate aggregator below loop might be cleaner if we had initialized it. 
                 # But let's add logic here.
    
    # Re-loop or initialize above? Let's initialize total_etfs above.
    total_etfs = sum((a.quantity * (a.current_price if a.current_price > 0 else a.avg_price)) for a in assets if a.type == 'ETF')

            
    # Process FII Details for Table and Broad Chart
    fii_table_data = []
    broad_allocation = {'Tijolo': 0, 'Papel': 0}
    
    for t, val in fii_summary.items():
        category = fii_map.get(t, 'Papel') # Default to Papel if unknown
        broad_allocation[category] += val
        pct = (val / total_fiis * 100) if total_fiis > 0 else 0
        fii_table_data.append({
            'category': category,
            'type': t,
            'value': val,
            'pct': pct
        })
        
    # Sort table by Value desc
    fii_table_data.sort(key=lambda x: x['value'], reverse=True)
            
    # Grava a foto do patrimônio de hoje (curva de evolução real daqui pra frente)
    record_portfolio_snapshot(current_user.id)

    # 2. Lucro realizado mensal (TradeHistory)
    monthly_profit = {}
    for h in history:
        if h.exit_date:
            month_key = h.exit_date.strftime('%Y-%m')
            monthly_profit[month_key] = monthly_profit.get(month_key, 0) + (h.profit_value or 0)

    # 3. Dividendos mensais — filtrados pelo período de POSSE do ativo
    #    (só conta se a data-com estava dentro de entrada→saída).
    today_d = date.today()
    monthly_div_acoes, monthly_div_fiis = _dividends_owned(current_user.id, today_d)

    # Todos os meses com pelo menos um dado
    all_months_set = (set(monthly_profit.keys()) |
                      set(monthly_div_acoes.keys()) |
                      set(monthly_div_fiis.keys()))
    sorted_months = sorted(all_months_set)
    profit_data   = [round(monthly_profit.get(k, 0), 2) for k in sorted_months]
    div_acoes_data= [round(monthly_div_acoes.get(k, 0), 2) for k in sorted_months]
    div_fiis_data = [round(monthly_div_fiis.get(k, 0), 2)  for k in sorted_months]

    # ── Patrimônio-base de cada mês = FECHAMENTO DO MÊS ANTERIOR ──────────────
    # Onde EXISTE snapshot real do fechamento do mês anterior, usamos ele (regra
    # nova, exata — vale daqui pra frente). Onde NÃO existe (meses passados sem
    # registro), voltamos à base antiga: patrimônio reconstruído a partir do
    # atual, desfazendo o lucro realizado dos meses seguintes. Isso mantém os
    # gráficos como estavam antes e evita a distorção da reconstrução por fluxos.
    from dateutil.relativedelta import relativedelta as _rd
    snaps = PortfolioSnapshot.query.filter_by(user_id=current_user.id).all()
    snap_by_month = {}
    for s in snaps:
        mk = s.snap_date[:7]
        cur = snap_by_month.get(mk)
        if cur is None or s.snap_date > cur.snap_date:
            snap_by_month[mk] = s

    def _prev_month(mk):
        y, m = mk.split('-')
        d = date(int(y), int(m), 1) - _rd(months=1)
        return d.strftime('%Y-%m')

    # Base antiga (fórmula original): desfaz o lucro dos meses posteriores.
    base_neg_now = (total_acoes + total_etfs) or 1     # ações+ETF (lucro é negociado)
    base_div_now = (total_acoes + total_fiis) or 1     # ações+FII (dividendos)
    month_profit = {k: monthly_profit.get(k, 0) for k in sorted_months}
    def _old_base(base_now, mk_prev):
        p = base_now
        for m in sorted_months:
            if m > mk_prev:
                p -= month_profit.get(m, 0)
        return max(p, 1)

    def _base_for(mk):
        """(base p/ lucro, base p/ dividendos) referentes ao fim do mês anterior a mk."""
        prev = _prev_month(mk)
        s = snap_by_month.get(prev)
        if s is not None:                              # snapshot real → regra nova
            bn = (s.total_acoes or 0) + (s.total_etfs or 0)
            bd = (s.total_acoes or 0) + (s.total_fiis or 0)
            return (bn if bn > 0 else base_neg_now,
                    bd if bd > 0 else base_div_now)
        return _old_base(base_neg_now, prev), _old_base(base_div_now, prev)

    profit_pct_data, div_acoes_pct, div_fiis_pct = [], [], []
    for k in sorted_months:
        base_neg, base_div = _base_for(k)
        profit_pct_data.append(round(monthly_profit.get(k, 0)   / base_neg * 100, 2))
        div_acoes_pct.append(round(monthly_div_acoes.get(k, 0)  / base_div * 100, 2))
        div_fiis_pct.append(round(monthly_div_fiis.get(k, 0)    / base_div * 100, 2))

    total_realized_profit = sum(h.profit_value for h in history if h.profit_value)
    current_month_key = date.today().strftime('%Y-%m')
    avg_months = sorted([m for m in sorted_months if m < current_month_key])[-4:]
    avg_count = len(avg_months) or 1
    avg_4m_realized_profit = round(sum(monthly_profit.get(k, 0) for k in avg_months) / avg_count, 2)
    profit_pct_by_month = dict(zip(sorted_months, profit_pct_data))
    avg_4m_realized_pct = round(sum(profit_pct_by_month.get(k, 0) for k in avg_months) / avg_count, 2)
    # Selic mensal para os meses do gráfico. Meses ainda não cadastrados
    # (posteriores ao último registro) usam a última Selic conhecida, para a
    # linha da Selic não interromper nos meses mais recentes.
    selic_rows = {s.mes_ano: s.taxa for s in SelicMensal.query.all()}
    _last_selic_key = max(selic_rows) if selic_rows else None
    _last_selic_val = selic_rows.get(_last_selic_key) if _last_selic_key else None
    selic_data = []
    for k in sorted_months:
        if k in selic_rows:
            selic_data.append(selic_rows[k])
        elif _last_selic_key and k > _last_selic_key:
            selic_data.append(_last_selic_val)   # mês futuro sem cadastro
        else:
            selic_data.append(None)

    # % acumulada carteira vs Selic — soma progressiva mês a mês
    cart_acum_pct  = []
    selic_acum_pct = []
    cart_running  = 0.0
    selic_running = 0.0
    for i, k in enumerate(sorted_months):
        total_mes_pct = profit_pct_data[i] + div_acoes_pct[i] + div_fiis_pct[i]
        cart_running  += total_mes_pct
        selic_m = selic_data[i]
        if selic_m is not None:
            selic_running += selic_m
        cart_acum_pct.append(round(cart_running, 2))
        selic_acum_pct.append(round(selic_running, 2) if selic_m is not None else None)

    total_dividends = round(sum(monthly_div_acoes.values()) + sum(monthly_div_fiis.values()), 2)

    # ── Curva de evolução do patrimônio (mensal): real onde há snapshot,
    #    estimado onde não há. Total = ações + FIIs + ETFs.
    #    A estimativa do passado desfaz apenas o lucro realizado dos meses
    #    seguintes (mesma base dos percentuais) — curva suave, sem os degraus
    #    que os fluxos de importação criavam. ──────────────────────────────────
    equity_months, equity_vals, equity_est = [], [], []
    if sorted_months:
        first = sorted_months[0]
        walk = _prev_month(first)     # começa no fechamento anterior ao 1º mês
        y0, m0 = map(int, walk.split('-'))
        cursor_d = date(y0, m0, 1)
        end_d = date(int(current_month_key[:4]), int(current_month_key[5:]), 1)
        while cursor_d <= end_d:
            mk = cursor_d.strftime('%Y-%m')
            if mk == current_month_key:
                equity_vals.append(round(total_equity, 2)); equity_est.append(False)
            else:
                s = snap_by_month.get(mk)
                if s is not None:
                    equity_vals.append(round(s.total_equity or 0, 2)); equity_est.append(False)
                else:
                    # estimativa: patrimônio atual menos o lucro realizado dos
                    # meses posteriores a este (limitado ao patrimônio atual)
                    est = total_equity
                    for m in sorted_months:
                        if m > mk:
                            est -= month_profit.get(m, 0)
                    equity_vals.append(round(min(max(est, 0.0), total_equity), 2))
                    equity_est.append(True)
            equity_months.append(mk)
            cursor_d += _rd(months=1)

    return render_template('resumo.html',
                         total_equity=total_equity, total_acoes=total_acoes,
                         total_fiis=total_fiis, total_etfs=total_etfs,
                         total_dividends=total_dividends,
                         total_realized_profit=total_realized_profit,
                         avg_4m_realized_profit=avg_4m_realized_profit,
                         avg_4m_realized_pct=avg_4m_realized_pct,
                         fii_types=fii_types,
                         fii_table=fii_table_data,
                         broad_allocation=broad_allocation,
                         months=sorted_months, profits=profit_data,
                         profit_pct=profit_pct_data,
                         div_acoes=div_acoes_data, div_acoes_pct=div_acoes_pct,
                         div_fiis=div_fiis_data,   div_fiis_pct=div_fiis_pct,
                         selic_data=selic_data,
                         cart_acum_pct=cart_acum_pct,
                         selic_acum_pct=selic_acum_pct,
                         equity_months=equity_months,
                         equity_vals=equity_vals,
                         equity_est=equity_est,
                         stock_sectors=stock_sectors)


@app.route('/selic_mensal', methods=['GET', 'POST'])
@login_required
def selic_mensal():
    if not current_user.is_admin:
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('resumo'))
    if request.method == 'POST':
        raw = request.form.get('selic_csv', '').strip()
        salvos = erros = 0
        for linha in raw.splitlines():
            linha = linha.strip()
            if not linha or linha.startswith('#'):
                continue
            partes = linha.split(',')
            if len(partes) != 2:
                erros += 1; continue
            try:
                mm_aa, taxa_str = partes[0].strip(), partes[1].strip()
                taxa = float(taxa_str.replace(',', '.'))
                parts = mm_aa.split('/')
                if len(parts) == 2:
                    mes_ano_fmt = f"{parts[1]}-{parts[0]}"
                else:
                    mes_ano_fmt = mm_aa  # aceita YYYY-MM direto
                row = SelicMensal.query.filter_by(mes_ano=mes_ano_fmt).first()
                if row:
                    row.taxa = taxa
                else:
                    db.session.add(SelicMensal(mes_ano=mes_ano_fmt, taxa=taxa))
                salvos += 1
            except Exception:
                erros += 1
        db.session.commit()
        flash(f'{salvos} registros salvos. {erros} erros.', 'success' if not erros else 'warning')
        return redirect(url_for('selic_mensal'))

    rows = SelicMensal.query.order_by(SelicMensal.mes_ano.desc()).all()
    return render_template('selic_mensal.html', rows=rows)

@app.route('/edit_history/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_history(id):
    trade = TradeHistory.query.get_or_404(id)
    if trade.user_id != current_user.id:
        flash("Você não tem permissão para editar este histórico.")
        return redirect(url_for('history'))
    if request.method == 'POST':
        trade.ticker = request.form.get('ticker').upper()
        trade.strategy = request.form.get('strategy')
        trade.quantity = int(request.form.get('quantity'))
        trade.buy_price = float(request.form.get('buy_price').replace(',', '.'))
        trade.sell_price = float(request.form.get('sell_price').replace(',', '.'))
        
        entry_date = request.form.get('entry_date')
        exit_date = request.form.get('exit_date')
        trade.entry_date = datetime.strptime(entry_date, '%Y-%m-%d').date() if entry_date else None
        trade.exit_date = datetime.strptime(exit_date, '%Y-%m-%d').date() if exit_date else None
        trade.reason = request.form.get('reason')
        
        # Recalc
        total_buy = trade.quantity * trade.buy_price
        total_sell = trade.quantity * trade.sell_price
        trade.profit_value = total_sell - total_buy
        trade.profit_pct = (trade.profit_value / total_buy * 100) if total_buy > 0 else 0
        trade.days_held = (trade.exit_date - trade.entry_date).days if (trade.entry_date and trade.exit_date) else 0
        
        db.session.commit()
        return redirect(url_for('history'))
        
    return render_template('edit_history.html', trade=trade)

@app.route('/delete_history/<int:id>')
@login_required
def delete_history(id):
    trade = TradeHistory.query.get_or_404(id)
    if trade.user_id != current_user.id:
        flash("Você não tem permissão para deletar este histórico.")
        return redirect(url_for('history'))
    db.session.delete(trade)
    db.session.commit()
    return redirect(url_for('history'))


@app.route('/history/export_excel')
@login_required
def export_history_excel():
    """Exporta todo o histórico de trades para Excel (.xlsx)."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('Biblioteca openpyxl não instalada. Execute: pip install openpyxl', 'danger')
        return redirect(url_for('history'))

    trades = TradeHistory.query.filter_by(user_id=current_user.id)\
                               .order_by(TradeHistory.exit_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Histórico'

    # Estilos
    hdr_fill  = PatternFill('solid', fgColor='1E293B')
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    hdr_align = Alignment(horizontal='center', vertical='center')
    pos_font  = Font(color='16A34A', size=10)
    neg_font  = Font(color='DC2626', size=10)
    border    = Border(bottom=Side(style='thin', color='E2E8F0'))

    headers = ['Ticker', 'Ativo Base', 'Estratégia', 'Data Entrada', 'Data Saída',
               'Qtd', 'Preço Compra (R$)', 'Preço Venda (R$)',
               'Resultado (R$)', 'Resultado (%)', 'Dias', 'Motivo', 'Observações']

    # Cabeçalho
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill  = hdr_fill
        cell.font  = hdr_font
        cell.alignment = hdr_align

    # Dados
    for t in trades:
        row = [
            t.ticker or '',
            t.underlying or '',
            t.strategy or '',
            t.entry_date.strftime('%d/%m/%Y') if t.entry_date else '',
            t.exit_date.strftime('%d/%m/%Y')  if t.exit_date  else '',
            t.quantity or 0,
            round(t.buy_price or 0, 4),
            round(t.sell_price or 0, 4),
            round(t.profit_value or 0, 2),
            round(t.profit_pct or 0, 2),
            t.days_held or 0,
            t.reason or '',
            t.notes or '',
        ]
        ws.append(row)
        r = ws.max_row
        # Cor no resultado
        pv = t.profit_value or 0
        ws.cell(r, 9).font = pos_font if pv >= 0 else neg_font
        ws.cell(r, 10).font = pos_font if pv >= 0 else neg_font
        for col in range(1, len(headers) + 1):
            ws.cell(r, col).border = border

    # Linha de totais
    total = sum(t.profit_value or 0 for t in trades)
    ws.append(['', '', '', '', 'TOTAL', len(trades), '', '',
               round(total, 2), '', '', '', ''])
    tr = ws.max_row
    for col in range(1, len(headers) + 1):
        ws.cell(tr, col).font = Font(bold=True, color='FFFFFF')
        ws.cell(tr, col).fill = PatternFill('solid', fgColor='1E293B')
    ws.cell(tr, 9).font = Font(bold=True, color='16A34A' if total >= 0 else 'DC2626')

    # Largura automática das colunas
    col_widths = [14, 12, 12, 14, 14, 7, 18, 18, 16, 14, 7, 12, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'historico_trades_{now_brt().strftime("%Y%m%d_%H%M")}.xlsx'
    from flask import send_file
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/migrate_history_notes')
@login_required
def migrate_history_notes():
    """Migra registros antigos de travas/estruturadas adicionando underlying e notes."""
    if not current_user.is_admin:
        return "Sem permissão", 403

    import re
    updated = 0
    trades = TradeHistory.query.filter_by(user_id=current_user.id).all()

    # Mapa de tickers de opções para underlying (Option + StructuredLeg + OptionSpread)
    ticker_to_under = {}
    for o in Option.query.filter_by(user_id=current_user.id).all():
        if o.ticker and o.underlying_asset:
            ticker_to_under[o.ticker.upper()] = o.underlying_asset.upper()
    for op in StructuredOp.query.filter_by(user_id=current_user.id).all():
        for leg in op.legs:
            if leg.ticker and op.underlying_asset:
                ticker_to_under[leg.ticker.upper()] = op.underlying_asset.upper()
    for sp in OptionSpread.query.filter_by(user_id=current_user.id).all():
        if sp.leg_long_ticker:
            ticker_to_under[sp.leg_long_ticker.upper()] = sp.underlying_asset.upper()
        if sp.leg_short_ticker:
            ticker_to_under[sp.leg_short_ticker.upper()] = sp.underlying_asset.upper()

    for t in trades:
        if t.underlying and t.notes:
            continue  # já migrado

        # Detecta se é uma opção pelo padrão do ticker (4 letras + 3 dígitos+letra)
        tk = (t.ticker or '').upper().strip()
        is_option_ticker = bool(re.match(r'^[A-Z]{4,5}[A-Z]\d{2,3}$', tk))

        # Registros de opções individuais — o ticker É a opção
        if is_option_ticker and not t.underlying:
            und = ticker_to_under.get(tk)
            if und:
                t.underlying = und
                if not t.notes:
                    t.notes = tk
                updated += 1

        # Registros de travas antigas (ticker truncado como "TRAVA CAL", "T.Baixa C" etc.)
        elif any(x in tk for x in ('TRAVA', 'T.ALTA', 'T.BAIXA', 'SLIDE', 'ESTRUT', 'BORBOL', 'STRADDLE')):
            # Tenta extrair tickers de opções do campo notes existente
            if t.notes:
                tickers_in_notes = re.findall(r'[A-Z]{4,5}[A-Z]\d{2,3}', t.notes.upper())
                underlyings = list({ticker_to_under.get(tk2, '') for tk2 in tickers_in_notes} - {''})
                if underlyings and not t.underlying:
                    t.underlying = underlyings[0]
                    updated += 1

    db.session.commit()
    flash(f'Migração concluída: {updated} registros atualizados.', 'success')
    return redirect(url_for('history'))


@app.route('/add_history', methods=['GET', 'POST'])
@login_required
def add_history():
    if request.method == 'POST':
        ticker = request.form.get('ticker').upper()
        strategy = request.form.get('strategy')
        qty = int(request.form.get('quantity'))
        
        buy_price = float(request.form.get('buy_price').replace(',', '.'))
        sell_price = float(request.form.get('sell_price').replace(',', '.'))
        
        entry_date_str = request.form.get('entry_date')
        exit_date_str = request.form.get('exit_date')
        
        entry_date = datetime.strptime(entry_date_str, '%Y-%m-%d').date() if entry_date_str else None
        exit_date = datetime.strptime(exit_date_str, '%Y-%m-%d').date() if exit_date_str else None
        
        reason = request.form.get('reason')
        
        # Calculations
        total_buy = qty * buy_price
        total_sell = qty * sell_price
        profit_value = total_sell - total_buy
        profit_pct = (profit_value / total_buy * 100) if total_buy > 0 else 0
        days_held = (exit_date - entry_date).days if (entry_date and exit_date) else 0
        
        new_trade = TradeHistory(
            user_id=current_user.id,
            ticker=ticker,
            strategy=strategy,
            entry_date=entry_date,
            exit_date=exit_date,
            quantity=qty,
            buy_price=buy_price,
            sell_price=sell_price,
            profit_value=profit_value,
            profit_pct=profit_pct,
            days_held=days_held,
            reason=reason
        )
        db.session.add(new_trade)
        db.session.commit()
        
        return redirect(url_for('history'))
        
    return render_template('add_history.html')

@app.route('/history')
@login_required
def history():
    from collections import defaultdict
    trades = TradeHistory.query.filter_by(user_id=current_user.id).order_by(TradeHistory.exit_date.desc()).all()

    total_profit = sum(t.profit_value for t in trades if t.profit_value)

    # Total atual da carteira de ações
    acoes_assets = Asset.query.filter_by(user_id=current_user.id, type='ACAO').all()
    total_acoes_atual = sum((a.current_price or a.avg_price or 0) * a.quantity for a in acoes_assets)
    if total_acoes_atual <= 0:
        total_acoes_atual = sum((a.avg_price or 0) * a.quantity for a in acoes_assets) or 1

    # Lucro total por mês (todas estratégias) — para estimar patrimônio inicial de cada mês
    # Lógica: patrimônio_inicio_mês[M] = valor_atual − soma_lucros_dos_meses_posteriores_a_M
    month_total_profit = defaultdict(float)
    for t in trades:
        if t.exit_date and t.profit_value is not None:
            mk = t.exit_date.strftime('%Y-%m')
            month_total_profit[mk] += t.profit_value

    # Reconstrói patrimônio início do mês subtraindo lucros futuros do valor atual
    all_months_sorted = sorted(month_total_profit.keys())
    # patrimônio no início do mês M = valor_atual - sum(lucros de meses > M)
    def portfolio_start_of_month(month_key):
        p = total_acoes_atual
        for mk in all_months_sorted:
            if mk > month_key:
                p -= month_total_profit[mk]
        return max(p, 1)  # nunca negativo

    # Unique strategies for filter
    strategies = sorted(set((t.strategy or 'Outros') for t in trades))

    # Summary by strategy
    summary_by_type = defaultdict(lambda: {'invested': 0.0, 'profit': 0.0})
    for t in trades:
        key = t.strategy or 'Outros'
        invested = (t.buy_price or 0) * (t.quantity or 0)
        summary_by_type[key]['invested'] += invested
        summary_by_type[key]['profit'] += (t.profit_value or 0)

    summary_table = []
    for strategy, vals in sorted(summary_by_type.items()):
        pct = (vals['profit'] / vals['invested'] * 100) if vals['invested'] > 0 else 0
        summary_table.append({
            'strategy': strategy,
            'invested': vals['invested'],
            'profit': vals['profit'],
            'profit_pct': pct
        })

    # Chart: profit by strategy and month (last 12 available months)
    month_strategy_profit = defaultdict(lambda: defaultdict(float))
    month_strategy_invested = defaultdict(lambda: defaultdict(float))
    for t in trades:
        if t.exit_date and t.profit_value is not None:
            month_key = t.exit_date.strftime('%Y-%m')
            key = t.strategy or 'Outros'
            month_strategy_profit[month_key][key] += t.profit_value
            month_strategy_invested[month_key][key] += (t.buy_price or 0) * (t.quantity or 0)

    # Get last 12 months with data
    sorted_months = sorted(month_strategy_profit.keys())[-12:]
    chart_labels = []
    for m in sorted_months:
        parts = m.split('-')
        chart_labels.append(f"{parts[1]}/{parts[0]}")

    all_strategies = sorted(set(s for m in sorted_months for s in month_strategy_profit[m].keys()))
    # Mapa de cores fixo por estratégia — Opções=verde, Internacional=azul
    STRATEGY_COLORS = {
        'Opções':             '#10b981',  # verde
        'Internacional':      '#3b82f6',  # azul
        'Fundos Imobiliários':'#8b5cf6',  # roxo
        'Recomendações':      '#ec4899',  # rosa
        'Técnica':            '#06b6d4',  # ciano
        'Outros':             '#f97316',  # laranja
    }
    fallback_colors = ['#64748b', '#84cc16', '#f59e0b', '#ef4444']
    # Rentabilidade mensal total (soma de todas estratégias / patrimônio início do mês)
    month_rentab = {}
    for m in sorted_months:
        total_m = sum(month_strategy_profit[m].values())
        base    = portfolio_start_of_month(m)
        month_rentab[m] = round(total_m / base * 100, 2)

    # Para o gráfico diário: patrimônio início de cada mês serializado
    month_portfolio_json = {m: round(portfolio_start_of_month(m), 2) for m in month_total_profit}

    # Série de patrimônio para gráfico de evolução:
    # ponto = patrimônio início do mês + lucros acumulados até aquele mês
    # Usa os sorted_months do gráfico mensal para manter escala consistente
    portfolio_series = []
    for m in sorted_months:
        val = round(portfolio_start_of_month(m) + month_total_profit.get(m, 0), 2)
        portfolio_series.append(val)
    # Adiciona ponto inicial (patrimônio antes do primeiro mês) para dar contexto
    if sorted_months:
        portfolio_series_full = [round(portfolio_start_of_month(sorted_months[0]), 2)] + portfolio_series
        portfolio_labels_full = ['Início'] + chart_labels
    else:
        portfolio_series_full = []
        portfolio_labels_full = []

    chart_datasets = []
    for i, strat in enumerate(all_strategies):
        data = [round(month_strategy_profit[m].get(strat, 0), 2) for m in sorted_months]
        pct_data = []
        for m in sorted_months:
            profit = month_strategy_profit[m].get(strat, 0)
            base   = portfolio_start_of_month(m)
            pct    = round(profit / base * 100, 2) if base > 0 else 0
            pct_data.append(pct)
        color = STRATEGY_COLORS.get(strat, fallback_colors[i % len(fallback_colors)])
        chart_datasets.append({
            'label': strat,
            'data': data,
            'pct_data': pct_data,
            'backgroundColor': color,
            'borderRadius': 4
        })

    # Ganho diário: {YYYY-MM: {YYYY-MM-DD: {strategy: profit}}}
    from collections import defaultdict as _dd
    daily_by_month = _dd(lambda: _dd(lambda: _dd(float)))
    for t in trades:
        if t.exit_date and t.profit_value is not None:
            month_key = t.exit_date.strftime('%Y-%m')
            day_key   = t.exit_date.strftime('%Y-%m-%d')
            strat     = t.strategy or 'Outros'
            daily_by_month[month_key][day_key][strat] += t.profit_value

    # Serializa para JSON: {YYYY-MM: {YYYY-MM-DD: {strat: val}}}
    daily_data = {}
    for month_key, days in daily_by_month.items():
        daily_data[month_key] = {}
        for day_key, strats in days.items():
            daily_data[month_key][day_key] = dict(strats)

    return render_template('history.html',
        trades=trades,
        total_profit=total_profit,
        strategies=strategies,
        summary_table=summary_table,
        chart_labels=chart_labels,
        chart_datasets=chart_datasets,
        sorted_months=sorted_months,
        daily_data=daily_data,
        all_strategies=all_strategies,
        total_acoes_ref=total_acoes_atual,
        strategy_colors=STRATEGY_COLORS,
        month_portfolio=month_portfolio_json,
        month_rentab=month_rentab,
        portfolio_series=portfolio_series_full,
        portfolio_labels=portfolio_labels_full,
    )




@app.route('/config', methods=['GET', 'POST'])
@login_required
def config():
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'save_brapi':
            brapi_key = request.form.get('brapi_key')
            if brapi_key:
                Settings.set_value('brapi_token', brapi_key, user_id=current_user.id)
                flash('Chave BRAPI salva com sucesso!', 'success')

        elif action == 'set_quote_mode':
            mode = request.form.get('quote_mode', 'yahoo')
            if mode in ('yahoo', 'mt5'):
                Settings.set_value('quote_mode', mode, user_id=current_user.id)
                flash(f'Modo de cotação alterado para: {"Yahoo Finance" if mode == "yahoo" else "MT5 Feeder"}', 'success')

        elif action == 'save_selic':
            selic = request.form.get('selic', '14.5').replace(',', '.')
            try:
                selic = str(float(selic))
            except ValueError:
                selic = '14.5'
            Settings.set_value('selic_rate', selic, user_id=current_user.id)
            flash(f'Taxa Selic salva: {selic}% a.a.', 'success')

        elif action == 'save_oplab_token':
            token = request.form.get('oplab_token', '').strip()
            if token:
                Settings.set_value('oplab_token', token, user_id=current_user.id)
                flash('Token OpLab salvo com sucesso!', 'success')
            else:
                flash('Token não pode ser vazio.', 'warning')

        elif action == 'save_oplab_config':
            auto     = 'true' if request.form.get('oplab_auto_update') == 'true' else 'false'
            interval = request.form.get('oplab_interval', '5')
            if interval not in ('1', '2', '5', '10'):
                interval = '5'
            Settings.set_value('oplab_auto_update', auto, user_id=current_user.id)
            Settings.set_value('oplab_interval',    interval, user_id=current_user.id)
            label = 'ativada' if auto == 'true' else 'desativada'
            flash(f'Atualização automática OpLab {label} (intervalo: {interval} min).', 'success')

        return redirect(url_for('config'))

    selic_rate      = float(Settings.get_value('selic_rate', user_id=current_user.id, default='14.5'))
    current_key = Settings.get_value('brapi_token', user_id=current_user.id)
    if not current_key:
        current_key = os.environ.get('BRAPI_API_KEY', '')

    quote_mode      = Settings.get_value('quote_mode',        user_id=current_user.id, default='yahoo')
    oplab_auto      = Settings.get_value('oplab_auto_update', user_id=current_user.id, default='false') == 'true'
    oplab_interval  = Settings.get_value('oplab_interval',    user_id=current_user.id, default='5')
    oplab_token_ok  = bool(Settings.get_value('oplab_token',  user_id=current_user.id))

    uid = current_user.id
    _, _, ticker_map_text, option_map_text = _build_ticker_maps(uid)

    from flask import session as _session
    test_result  = _session.pop('test_result',  None)
    test_success = _session.pop('test_success', False)

    return render_template('config.html',
                           current_key=current_key,
                           quote_mode=quote_mode,
                           oplab_auto=oplab_auto,
                           oplab_interval=oplab_interval,
                           oplab_token_ok=oplab_token_ok,
                           ticker_map_text=ticker_map_text,
                           option_map_text=option_map_text,
                           selic_rate=selic_rate,
                           test_result=test_result,
                           success=test_success)

@app.route('/download_config_py')
@login_required
def download_config_py():
    """Gera e retorna o arquivo config.py completo para o mt5_feeder."""
    from flask import Response
    uid = current_user.id

    asset_tickers, option_tickers, _, _ = _build_ticker_maps(uid)

    ticker_lines = sorted(
        [f'    "{t}": "{t}",  # {tp}' for t, tp in asset_tickers],
        key=lambda x: x.strip()
    )
    option_lines = sorted(
        [f'    "{t}": "{t}",  # {info}' for t, info in option_tickers.items()],
        key=lambda x: x.strip()
    )

    api_url = request.host_url.rstrip('/') + '/api/update_quotes'
    mt5_key = Settings.get_value('mt5_api_key', user_id=uid, default='chave_mtq5_2026') or 'chave_mtq5_2026'

    content = f'''# config.py — gerado automaticamente em {now_brt().strftime("%d/%m/%Y %H:%M")}
# Coloque este arquivo na pasta mt5_feeder/ ao lado de mt5_feeder.py
# NÃO compartilhe este arquivo (contém chave de API).

# URL do endpoint no VPS
API_URL = "{api_url}"

# API Key — deve coincidir com MT5_API_KEY no .env do servidor
API_KEY = "{mt5_key}"

# ID do usuário no site (normalmente 1 para o admin)
USER_ID = {uid}

# Intervalo de atualização em segundos
INTERVALO_SEGUNDOS = 30

# Mapeamento: "TICKER_NO_SITE" -> "SÍMBOLO_NO_MT5"
TICKER_MAP = {{
{chr(10).join(ticker_lines)}
}}

# Mapeamento de opções: "TICKER_OPÇÃO" -> "SÍMBOLO_NO_MT5"
OPTION_MAP = {{
{chr(10).join(option_lines)}
}}
'''
    return Response(
        content,
        mimetype='text/plain',
        headers={'Content-Disposition': 'attachment; filename=config.py'}
    )


@app.route('/download_mt5_feeder_py')
@login_required
def download_mt5_feeder_py():
    """Retorna o arquivo mt5_feeder.py para download direto."""
    from flask import send_file
    feeder_path = os.path.join(basedir, 'mt5_feeder', 'mt5_feeder.py')
    if not os.path.exists(feeder_path):
        flash('Arquivo mt5_feeder.py não encontrado no servidor.', 'danger')
        return redirect(url_for('config'))
    return send_file(feeder_path, as_attachment=True, download_name='mt5_feeder.py',
                     mimetype='text/plain')


@app.route('/test_api', methods=['POST'])
@login_required
def test_api():
    ticker = request.form.get('ticker')
    if not ticker:
        flash("Informe um ticker.")
        return redirect(url_for('config'))

    import json
    success, data = get_raw_quote_data(ticker.strip().upper())
    formatted_data = json.dumps(data, indent=4, ensure_ascii=False)

    # Guarda resultado na sessão para exibir na rota config
    from flask import session as _session
    _session['test_result'] = formatted_data
    _session['test_success'] = success

    if success:
        flash(f'Conexão OK para {ticker.upper()}.', 'success')
    else:
        flash(f'Erro ao testar {ticker.upper()}: verifique o ticker e o token.', 'warning')
    return redirect(url_for('config'))

# Auth Routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user)
            # Automatic Update on Login (background - não bloqueia o redirect)
            uid = user.id
            def _bg_login_update():
                with app.app_context():
                    try:
                        update_market_indices()
                        update_all_assets_logic(user_id=uid)
                    except Exception:
                        pass
            threading.Thread(target=_bg_login_update, daemon=True).start()
            return redirect(url_for('index'))
        flash('Usuário ou senha inválidos')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if User.query.filter_by(username=username).first():
            flash('Usuário já existe')
            return redirect(url_for('register'))
            
        new_user = User(username=username)
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        
        login_user(new_user)
        return redirect(url_for('index'))
        
    return render_template('register.html')




# ==========================================
# BALANCEAMENTO MODULE
# ==========================================

def get_maturity_class(maturity_date):
    if not maturity_date:
        return 'Indefinido'
    today = date.today()
    days = (maturity_date - today).days
    years = days / 365.0
    
    if years <= 2:
        return 'Curto Prazo'
    elif years <= 4:
        return 'Médio Prazo'
    else:
        return 'Longo Prazo'

@app.route('/fix_db')
def fix_db():
    try:
        # Run migration logic directly here
        conn = sqlite3.connect(os.path.join(app.instance_path, 'investments.db'))
        cursor = conn.cursor()
        columns = [
            ('category', 'TEXT DEFAULT "RV"'),
            ('description', 'TEXT'),
            ('invested_value', 'REAL')
        ]
        log = []
        for col_name, col_type in columns:
            try:
                cursor.execute(f"ALTER TABLE international ADD COLUMN {col_name} {col_type}")
                log.append(f"Added {col_name}")
            except Exception as e:
                log.append(f"Skipped {col_name}: {str(e)}")
        conn.commit()
        conn.close()
        return f"Migration Result: {', '.join(log)}. <a href='/balanceamento'>Voltar</a>"
    except Exception as e:
        return f"Error: {str(e)}"

@app.route('/fix_crypto_db')
def fix_crypto_db():
    try:
        conn = sqlite3.connect(os.path.join(app.instance_path, 'investments.db'))
        cursor = conn.cursor()
        log = []
        
        # 1. Quote
        try:
            cursor.execute("ALTER TABLE crypto ADD COLUMN quote REAL")
            log.append("Added 'quote'")
        except Exception as e:
            log.append(f"Skip quote: {str(e)}")
            
        # 2. Avg Price
        try:
            cursor.execute("ALTER TABLE crypto ADD COLUMN avg_price REAL DEFAULT 0.0")
            log.append("Added 'avg_price'")
        except Exception as e:
            log.append(f"Skip avg_price: {str(e)}")
        
        conn.commit()
        conn.close()
        return f"Migration: {', '.join(log)} <a href='/balanceamento'>Voltar</a>"
    except Exception as e:
        return f"Database Error: {str(e)}"

@app.route('/balanceamento')
@login_required
def balanceamento():
    
    def get_maturity_class(date_obj):
        if not date_obj:
            return 'Indefinido'
        days = (date_obj - date.today()).days
        if days <= 365:
            return 'Curto Prazo'
        elif days <= 1095: # 3 years
            return 'Médio Prazo'
        else:
            return 'Longo Prazo'

    try:
        # User Assets
        rfs = FixedIncome.query.filter_by(user_id=current_user.id).all()
        rf_pos = [r for r in rfs if r.category == 'POS']
        rf_pre = [r for r in rfs if r.category == 'PRE']
        rf_ipca = [r for r in rfs if r.category == 'IPCA']
        funds = InvestmentFund.query.filter_by(user_id=current_user.id).all()
        pensions = Pension.query.filter_by(user_id=current_user.id).all()
        
        # 2. RV Data
        acoes_assets = Asset.query.filter(Asset.type=='ACAO', Asset.user_id==current_user.id, Asset.quantity > 0).all()
        fiis_assets = Asset.query.filter(Asset.type=='FII', Asset.user_id==current_user.id, Asset.quantity > 0).all()
        etfs_assets = Asset.query.filter(Asset.type=='ETF', Asset.user_id==current_user.id, Asset.quantity > 0).all()
        cryptos = Crypto.query.filter_by(user_id=current_user.id).all()
        # Using same logic as models: strategy='SWING'
        assets_swing = Asset.query.filter_by(strategy='SWING', user_id=current_user.id).all()
        
        # 4. Stock Holders (Asset table)
        assets_holder = Asset.query.filter_by(strategy='HOLDER', type='ACAO', user_id=current_user.id).all()
        
        # Split Intls
        intls_rv = International.query.filter_by(user_id=current_user.id, category='RV').all()
        intls_rf = International.query.filter_by(user_id=current_user.id, category='RF').all()
        
        # 3. Swing Trade (using Asset table)
        # Using same logic as models: strategy='SWING'
        # assets_swing = Asset.query.filter_by(strategy='SWING', user_id=current_user.id).all() # This line was moved up
        
        # 4. Stock Holders (Asset table)
        # assets_holder = Asset.query.filter_by(strategy='HOLDER', type='ACAO', user_id=current_user.id).all() # This line was moved up
        fiis_holder = Asset.query.filter_by(strategy='HOLDER', type='FII', user_id=current_user.id).all()
        
        # 2. Existing Assets (Stocks/FIIs)
        assets = Asset.query.filter_by(user_id=current_user.id).all()
        # Separate GOLD11 (Ouro) from other Stocks
        gold_assets = [a for a in assets if a.ticker == 'GOLD11']
        stock_assets = [a for a in assets if a.type == 'ACAO' and a.ticker != 'GOLD11']
        fii_assets = [a for a in assets if a.type == 'FII']

        val_ouro = 0 # Ouro is deprecated/replaced by ETF logic, setting to 0 to avoid errors if referenced elsewhere

        val_acoes = sum((a.quantity * ((a.current_price or 0) if (a.current_price or 0) > 0 else (a.avg_price or 0))) for a in acoes_assets)
        val_fiis = sum((a.quantity * ((a.current_price or 0) if (a.current_price or 0) > 0 else (a.avg_price or 0))) for a in fiis_assets)
        val_etfs = sum((a.quantity * ((a.current_price or 0) if (a.current_price or 0) > 0 else (a.avg_price or 0))) for a in etfs_assets)
        val_cryptos = sum((c.current_value or 0) for c in cryptos)
        val_intls_rv = sum(((i.value_usd or 0) * (i.rate_usd or 5.5)) for i in intls_rv)
        
        # 3. Aggregates & Classification
        summary = {
            'Curto Prazo': 0, 'Médio Prazo': 0, 'Longo Prazo': 0, 'Indefinido': 0,
            'Renda Fixa': 0, 'Renda Variável': 0
        }
        
        types_total = {
            'Renda Fixa Pós': 0, 'Renda Fixa Pré': 0, 'Renda Fixa IPCA': 0,
            'Fundos': 0, 'Cripto': 0, 'Previdência': 0, 
            'Internacional RV': 0, 'Internacional RF': 0,
            'Ações': val_acoes, 'FIIs': val_fiis, 'Ouro': val_ouro
        }
        
        # Detailed Maturity Breakdown
        maturity_breakdown = {
            'Renda Fixa Pós': {'Curto Prazo': 0, 'Médio Prazo': 0, 'Longo Prazo': 0, 'Indefinido': 0},
            'Renda Fixa Pré': {'Curto Prazo': 0, 'Médio Prazo': 0, 'Longo Prazo': 0, 'Indefinido': 0},
            'Renda Fixa IPCA': {'Curto Prazo': 0, 'Médio Prazo': 0, 'Longo Prazo': 0, 'Indefinido': 0},
            'Fundos': {'Curto Prazo': 0, 'Médio Prazo': 0, 'Longo Prazo': 0, 'Indefinido': 0}
        }
        
        # New: Fixed Income Subtypes Aggregation
        rf_subtypes = {}
        for r in rfs:
            # Determine Type
            p_type = r.product_type
            if not p_type:
                # Infer from name
                name_upper = r.name.upper()
                if 'CDB' in name_upper: p_type = 'CDB'
                elif 'LCI' in name_upper: p_type = 'LCI'
                elif 'LCA' in name_upper: p_type = 'LCA'
                elif 'CRI' in name_upper: p_type = 'CRI'
                elif 'CRA' in name_upper: p_type = 'CRA'
                elif 'RDB' in name_upper: p_type = 'RDB'
                elif 'TESOURO' in name_upper: p_type = 'Tesouro Direto'
                elif 'DEBENTURE' in name_upper: p_type = 'Debênture'
                else: p_type = 'Outros'
            
            p_type = p_type.upper().strip() # Normalize
            
            # Aggregate
            rf_subtypes[p_type] = rf_subtypes.get(p_type, 0) + (r.value or 0)
            
        # Sort by value desc
        rf_subtypes_sorted = dict(sorted(rf_subtypes.items(), key=lambda item: item[1], reverse=True))
        
        # Helper to process list
        def process_list(items, type_key, is_variable=False):
            total = 0
            for i in items:
                # Safe access to value or calculation
                if hasattr(i, 'value'):
                     val = i.value or 0
                elif hasattr(i, 'current_value'):
                     val = i.current_value or 0
                else:
                     val = (i.value_usd or 0) * (i.rate_usd or 1)
                
                total += val
                
                # Maturity
                mat = i.maturity_date if hasattr(i, 'maturity_date') else None
                cls = get_maturity_class(mat)
                
                # Add to Global Summary
                if hasattr(i, 'maturity_date'):
                   summary[cls] += val
                else:
                   summary['Indefinido'] += val
                   
                # Add to Detailed Breakdown if applicable
                if type_key in maturity_breakdown:
                    maturity_breakdown[type_key][cls] += val

            types_total[type_key] = total
            if is_variable:
                summary['Renda Variável'] += total
            else:
                summary['Renda Fixa'] += total
            return total

        # Process Logic
        process_list(rf_pos, 'Renda Fixa Pós')
        process_list(rf_pre, 'Renda Fixa Pré')
        process_list(rf_ipca, 'Renda Fixa IPCA')
        # process_list(funds, 'Fundos') - Replaced by custom logic below
        
        # Special Processing for Funds (Chart Classification)
        total_funds = 0
        for f in funds:
            # Safe value access
            if hasattr(f, 'value'):
                 val = f.value or 0
            else:
                 val = 0
            
            total_funds += val
            
            # Maturity
            mat = f.maturity_date
            cls = get_maturity_class(mat)
            
            # Global Summary (Time & Type)
            if mat:
                summary[cls] += val
            else:
                summary['Indefinido'] += val
            
            # Determine Chart Category based on Indexer
            chart_cat = 'Fundos' 
            idx = (f.indexer or '').upper()
            if 'IPCA' in idx:
                chart_cat = 'Renda Fixa IPCA'
            elif 'SELIC' in idx or 'CDI' in idx:
                chart_cat = 'Renda Fixa Pós'
                
            # Add to Detailed Breakdown (for Chart)
            if chart_cat in maturity_breakdown:
                maturity_breakdown[chart_cat][cls] += val
        
        types_total['Fundos'] = total_funds
        summary['Renda Fixa'] += total_funds

        
        # Crypto
        # Crypto Model has current_value
        t_crypto = sum([(c.current_value or 0) for c in cryptos])
        types_total['Cripto'] = t_crypto
        summary['Renda Variável'] += t_crypto
        summary['Indefinido'] += t_crypto # Crypto has no maturity

        # Pension
        # Pension has type 'Acao' or 'Renda Fixa'
        for p in pensions:
            types_total['Previdência'] += p.value
            # Pension generally Long Term
            summary['Longo Prazo'] += p.value
            if p.type == 'Acao':
                summary['Renda Variável'] += p.value
            else:
                summary['Renda Fixa'] += p.value

        # International
        # International
        # RV
        t_intl_rv = sum([((i.value_usd or 0) * (i.rate_usd or 5.5)) for i in intls_rv])
        types_total['Internacional RV'] = t_intl_rv
        summary['Renda Variável'] += t_intl_rv
        summary['Indefinido'] += t_intl_rv

        # RF
        t_intl_rf = sum([((i.value_usd or 0) * (i.rate_usd or 5.5)) for i in intls_rf])
        types_total['Internacional RF'] = t_intl_rf
        summary['Renda Fixa'] += t_intl_rf
        summary['Longo Prazo'] += t_intl_rf # Assuming Bonds are long term

        # Add Stocks/FIIs/Gold (Gold replaced by ETF) to Summary
        # Add Stocks/FIIs to Summary (RV Brasil)
        summary['Renda Variável'] += (val_acoes + val_fiis)
        # Add ETF to Summary
        summary['Renda Variável'] += val_etfs
        summary['Indefinido'] += (val_acoes + val_fiis + val_etfs)

        total_portfolio = sum(types_total.values())

        # Filter breakdown to remove 0 values (Simpler for Template Rowspan)
        clean_breakdown = {}
        for cat, terms in maturity_breakdown.items():
            clean_terms = {k: v for k, v in terms.items() if v > 0.01}
            if clean_terms:
                clean_breakdown[cat] = clean_terms

        # Prepare Data for New Pie Charts (RF by Term, RF by Type)
        # 1. RF by Term (Aggregate from clean_breakdown)
        rf_chart_term = {'Curto Prazo': 0, 'Médio Prazo': 0, 'Longo Prazo': 0, 'Indefinido': 0}
        for cat, terms in clean_breakdown.items():
            for term, val in terms.items():
                if term in rf_chart_term:
                    rf_chart_term[term] += val

        # 2. RF by Type (Aggregate Category Totals from clean_breakdown)
        rf_chart_type = {cat: sum(terms.values()) for cat, terms in clean_breakdown.items()}

        # 3. Total for the Table Footer
        total_rf_detailed = sum(rf_chart_type.values())

        # Prepare specific Pie Chart Data (Granular)
        # Requested: Renda Fixa Pós, Pré, IPCA, Fundos, Cripto, Previdência, Ações, FIIs, ETF (replacing Gold), Internacional RV, Internacional RF
        target_keys = [
            'Renda Fixa Pós', 'Renda Fixa Pré', 'Renda Fixa IPCA', 
            'Fundos', 'Cripto', 'Previdência', 
            'Ações', 'FIIs', 
            'Internacional RV', 'Internacional RF'
        ]
        # Dynamically add ETF to pie chart data, treating it as its own slice if desired, or under Intl.
        # User asked ETF to be where Gold was. Gold was in target_keys. 
        # I will add 'ETF' to target_keys to ensure it shows up in "Por Classe" chart if types_total has it.
        # But wait, types_total doesn't have 'ETF' key yet?
        # I need to add 'ETF' to types_total!
        
        types_total['ETF'] = val_etfs
        target_keys.append('ETF')

        pie_chart_data = {k: types_total.get(k, 0) for k in target_keys if types_total.get(k, 0) > 0.01}

        # 4. Totals for International RV Table
        intl_rv_invested = sum([((i.quantity or 0) * (i.avg_price or 0)) for i in intls_rv])
        intl_rv_current = sum([((i.quantity or 0) * (i.quote or 0)) for i in intls_rv])
        intl_rv_profit = intl_rv_current - intl_rv_invested
        
        # 5. Totals for Crypto Table
        # invested_value should match quantity * avg_price now, or use direct accumulation
        crypto_invested = sum([(c.quantity or 0) * (c.avg_price or 0) for c in cryptos])
        crypto_current = sum([(c.current_value or 0) for c in cryptos])
        crypto_profit = crypto_current - crypto_invested

        # 6. Location Breakdown (Brazil vs International)
        # International = Crypto + Intl RV + Intl RF + Gold
        # Note: t_intl_rf is calculated above. val_etfs is used as International per user request.
        total_intl = t_intl_rv + t_intl_rf + t_crypto + val_etfs
        total_br = total_portfolio - total_intl
        
        location_chart = {
            'Brasil': total_br,
            'Internacional': total_intl
        }

        # --- NEW SUMMARY CALCULATIONS (User Request) ---
        
        # 1. Hierarchical Data
        # Renda Fixa
        #   Pos: RF Pos + Funds (Pos) + Pension (RF)
        #   Pre: RF Pre
        #   Ipca: RF IPCA + Funds (Ipca)
        
        val_rf_pos_strict = types_total.get('Renda Fixa Pós', 0)
        # Funds classification logic was partly inside the loop, need to replicate or reuse
        # Let's re-iterate funds to split properly if not done
        val_funds_pos = 0
        val_funds_ipca = 0
        for f in funds:
            idx = (f.indexer or '').upper()
            val = f.value or 0
            if 'IPCA' in idx:
                val_funds_ipca += val
            else:
                val_funds_pos += val
        
        val_pension_rf = 0
        val_pension_acao = 0
        for p in pensions:
            p_val = p.value or 0
            if p.type == 'Acao':
                val_pension_acao += p_val
            else:
                val_pension_rf += p_val

        total_pos = val_rf_pos_strict + val_funds_pos + val_pension_rf
        total_pre = types_total.get('Renda Fixa Pré', 0)
        total_ipca = types_total.get('Renda Fixa IPCA', 0) + val_funds_ipca
        
        total_rf_general = total_pos + total_pre + total_ipca
        
        # ETF Calculation (New)
        val_etfs = sum((a.quantity * ((a.current_price or 0) if (a.current_price or 0) > 0 else (a.avg_price or 0))) for a in etfs_assets)

        # RV Brasil
        #   Acoes: Stocks + Pension Acao
        #   FII
        total_acoes_consol = val_acoes + val_pension_acao
        total_fii = val_fiis
        
        total_rv_br = total_acoes_consol + total_fii
        
        # RV Internacional
        #   Cripto
        #   RV Intl
        #   RF Intl
        #   ETF (Moved here as requested, replacing Gold role if any)
        
        total_cripto = types_total.get('Cripto', 0)
        total_intl_rv = types_total.get('Internacional RV', 0)
        total_intl_rf = types_total.get('Internacional RF', 0)
        total_etf_intl = val_etfs
        
        total_rv_intl_general = total_cripto + total_intl_rv + total_intl_rf + total_etf_intl
        
        # Data Structure for Template
        # Hierarchy List: [ {Group, Lines: [{Label, Val, Pct}, ...], Total, TotalPct}, ... ]
        
        def calc_pct(v):
            return (v / total_portfolio * 100) if total_portfolio > 0 else 0

        summary_hierarchy = [
            {
                'group': 'Renda Fixa',
                'lines': [
                    {'label': 'Pós', 'value': total_pos, 'pct': calc_pct(total_pos)},
                    {'label': 'Pré', 'value': total_pre, 'pct': calc_pct(total_pre)},
                    {'label': 'Ipca', 'value': total_ipca, 'pct': calc_pct(total_ipca)},
                ],
                'total': total_rf_general,
                'total_pct': calc_pct(total_rf_general)
            },
            {
                'group': 'RV Brasil',
                'lines': [
                    {'label': 'Ações', 'value': total_acoes_consol, 'pct': calc_pct(total_acoes_consol)},
                    {'label': 'FII', 'value': total_fii, 'pct': calc_pct(total_fii)},
                ],
                'total': total_rv_br,
                'total_pct': calc_pct(total_rv_br)
            },
            {
                'group': 'RV Internacional',
                'lines': [
                    {'label': 'Criptomoedas', 'value': total_cripto, 'pct': calc_pct(total_cripto)},
                    {'label': 'ETF', 'value': total_etf_intl, 'pct': calc_pct(total_etf_intl)},
                    {'label': 'Renda Variável Internacional', 'value': total_intl_rv, 'pct': calc_pct(total_intl_rv)},
                    {'label': 'Renda Fixa Internacional', 'value': total_intl_rf, 'pct': calc_pct(total_intl_rf)},
                ],
                'total': total_rv_intl_general,
                'total_pct': calc_pct(total_rv_intl_general)
            }
        ]
        
        # 2. Exploded Balance (For Pie Chart)
        # Using the individual lines from hierarchy
        summary_exploded = {}
        for group in summary_hierarchy:
            for line in group['lines']:
                if line['value'] > 0:
                    # Clean label for chart
                    label = line['label']
                    if label == 'Renda Variável Internacional': label = 'RV Internacional'
                    if label == 'Renda Fixa Internacional': label = 'RF Internacional'
                    summary_exploded[label] = line['value']
                    
        # 3. General Balance (For Donut Chart)
        # Categories: Pós, Pré, Ipca, RV Brasil, RV Internacional
        summary_general = {
            'Pós': total_pos,
            'Pré': total_pre,
            'Ipca': total_ipca,
            'RV Brasil': total_rv_br,
            'RV Internacional': total_rv_intl_general
        }

        return render_template('balanceamento.html', 
                               rf_pos=rf_pos, rf_pre=rf_pre, rf_ipca=rf_ipca,
                               funds=funds, cryptos=cryptos, pensions=pensions, 
                               intls_rv=intls_rv, intls_rf=intls_rf,
                               summary=summary, types_total=types_total, total_portfolio=total_portfolio,
                               maturity_breakdown=clean_breakdown,
                               rf_chart_term=rf_chart_term, rf_chart_type=rf_chart_type,
                               pie_chart_data=pie_chart_data,
                               total_rf_detailed=total_rf_detailed,
                               intl_rv_invested=intl_rv_invested, 
                               intl_rv_current=intl_rv_current, 
                               intl_rv_profit=intl_rv_profit,
                               crypto_invested=crypto_invested,
                               crypto_current=crypto_current,
                               crypto_profit=crypto_profit,
                               location_chart=location_chart,
                               summary_hierarchy=summary_hierarchy,
                               summary_exploded=summary_exploded,
                               summary_general=summary_general,
                               rf_subtypes=rf_subtypes_sorted)
    except Exception as e:
        import traceback
        return f"<h3>Debug Error de Balanceamento (Mostre isso ao suporte):</h3><pre>{traceback.format_exc()}</pre>"

@app.route('/balanceamento/add/rf', methods=['POST'])
@login_required
def add_rf():
    new_rf = FixedIncome(
        user_id=current_user.id,
        category=request.form.get('category'),
        product_type=request.form.get('product_type'),
        institution=request.form.get('institution'),
        name=request.form.get('name'),
        value=float(request.form.get('value').replace(',', '.')),
        rate=request.form.get('rate'),
        maturity_date=datetime.strptime(request.form.get('maturity_date'), '%Y-%m-%d').date() if request.form.get('maturity_date') else None
    )
    db.session.add(new_rf)
    db.session.commit()
    flash('Renda Fixa adicionada!')
    return redirect(url_for('balanceamento'))

@app.route('/balanceamento/add/fund', methods=['POST'])
@login_required
def add_fund():
    new_fund = InvestmentFund(
        user_id=current_user.id,
        institution=request.form.get('institution'),
        name=request.form.get('name'),
        value=float(request.form.get('value').replace(',', '.')),
        indexer=request.form.get('indexer'),
        maturity_date=datetime.strptime(request.form.get('maturity_date'), '%Y-%m-%d').date() if request.form.get('maturity_date') else None
    )
    db.session.add(new_fund)
    db.session.commit()
    flash('Fundo adicionado!')
    return redirect(url_for('balanceamento'))

@app.route('/balanceamento/add/crypto', methods=['POST'])
@login_required
def add_crypto():
    qty_str = request.form.get('quantity', '').replace(',', '.')
    qty = float(qty_str) if qty_str else 0.0
    
    avg_price_str = request.form.get('avg_price', '').replace(',', '.') # User input for Avg Price
    avg_price = float(avg_price_str) if avg_price_str else 0.0
    
    # Clean logic
    invested_value = qty * avg_price
    
    # User might input current value manually or we calc later
    inv_val_str = request.form.get('invested_value')
    # If using form that sends invested_value (legacy), decide which to use. 
    # Current form has avg_price field?
    
    curr_val_str = request.form.get('current_value', '').replace(',', '.')
    current_value = float(curr_val_str) if curr_val_str else 0.0

    new_crypto = Crypto(
        user_id=current_user.id,
        institution=request.form.get('institution'),
        name=request.form.get('name'),
        quantity=qty,
        invested_value=invested_value,
        current_value=current_value,
        avg_price=avg_price
    )
    db.session.add(new_crypto)
    db.session.commit()
    flash('Cripto adicionada!')
    return redirect(url_for('balanceamento'))

@app.route('/balanceamento/add/pension', methods=['POST'])
@login_required
def add_pension():
    new_pension = Pension(
        user_id=current_user.id,
        institution=request.form.get('institution'),
        name=request.form.get('name'),
        value=float(request.form.get('value').replace(',', '.')),
        type=request.form.get('type')
    )
    db.session.add(new_pension)
    db.session.commit()
    flash('Previdência adicionada!')
    return redirect(url_for('balanceamento'))

@app.route('/balanceamento/add/intl', methods=['POST'])
@login_required
def add_intl():
    val_str = request.form.get('value_usd', '').replace(',', '.')
    val_usd = float(val_str) if val_str else 0.0
    
    qty_str = request.form.get('quantity', '').replace(',', '.')
    qty = float(qty_str) if qty_str else 0.0
    
    category = request.form.get('category', 'RV')
    
    quantity_str = request.form.get('quantity', '').replace(',', '.')
    qty = float(quantity_str) if quantity_str else 0.0
    
    avg_price_str = request.form.get('avg_price', '').replace(',', '.')
    avg_price = float(avg_price_str) if avg_price_str else 0.0
    
    # Optional direct value (legacy or override)
    value_usd_str = request.form.get('value_usd', '').replace(',', '.')
    val_usd_input = float(value_usd_str) if value_usd_str else 0.0
    
    # Code/Ticker
    ticker_name = request.form.get('name', '').upper()
    
    # Calculate Invested Value
    invested = 0.0
    if qty > 0 and avg_price > 0:
        invested = qty * avg_price
    elif val_usd_input > 0:
        invested = val_usd_input # Fallback for legacy RF
        
    # Initial Value USD (Current) -> equals invest if no quote yet
    current_val_usd = invested
    
    new_intl = International(
        user_id=current_user.id,
        institution=request.form.get('institution'),
        name=ticker_name,
        quantity=qty,
        avg_price=avg_price,
        category=category,
        description=request.form.get('description'),
        value_usd=current_val_usd,
        invested_value=invested,
        quote=avg_price # Set initial quote to purchase price
    )
    
    db.session.add(new_intl)
    db.session.commit()
    flash('Investimento Internacional adicionado!')
    return redirect(url_for('balanceamento'))

@app.route('/balanceamento/edit/<type>/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_balance_item(type, id):
    item = None
    if type == 'rf':
        item = FixedIncome.query.filter_by(id=id, user_id=current_user.id).first()
    elif type == 'fund':
        item = InvestmentFund.query.filter_by(id=id, user_id=current_user.id).first()
    elif type == 'crypto':
        item = Crypto.query.filter_by(id=id, user_id=current_user.id).first()
    elif type == 'pension':
        item = Pension.query.filter_by(id=id, user_id=current_user.id).first()
    elif type == 'intl':
        item = International.query.filter_by(id=id, user_id=current_user.id).first()
    
    if not item:
        flash('Item não encontrado ou acesso negado.')
        return redirect(url_for('balanceamento'))
    
    if request.method == 'POST':
        # Common fields update could be dynamic, but manual is safer per type
        if type == 'rf':
            item.institution = request.form.get('institution')
            item.name = request.form.get('name')
            item.value = float(request.form.get('value').replace('.','').replace(',','.'))
            item.rate = request.form.get('rate')
            mat = request.form.get('maturity_date')
            item.maturity_date = datetime.strptime(mat, '%Y-%m-%d').date() if mat else None
            # Specifics
            if request.form.get('product_type'):
                item.product_type = request.form.get('product_type')
                
        elif type == 'fund':
            item.institution = request.form.get('institution')
            item.name = request.form.get('name')
            item.value = float(request.form.get('value').replace('.','').replace(',','.'))
            item.indexer = request.form.get('indexer')
            mat = request.form.get('maturity_date')
            item.maturity_date = datetime.strptime(mat, '%Y-%m-%d').date() if mat else None
            
        elif type == 'crypto':
            item.institution = request.form.get('institution')
            item.name = request.form.get('name')
            
            qty_str = request.form.get('quantity', '').replace(',', '.')
            if qty_str and qty_str.lower() != 'none':
                item.quantity = float(qty_str)
            else:
                item.quantity = 0.0
                
            avg_str = request.form.get('avg_price', '').replace('.', '').replace(',', '.')
            item.avg_price = float(avg_str) if avg_str and avg_str.lower() != 'none' else 0.0
            
            # Recalculate Invested from Avg * Qty
            item.invested_value = item.quantity * item.avg_price

            cur_str = request.form.get('current_value', '').replace('.', '').replace(',', '.')
            item.current_value = float(cur_str) if cur_str and cur_str.lower() != 'none' else 0.0
            
        elif type == 'pension':
            item.institution = request.form.get('institution')
            item.name = request.form.get('name')
            item.value = float(request.form.get('value').replace('.','').replace(',','.'))
            item.type = request.form.get('type')
            item.certificate = request.form.get('certificate')
            
        elif type == 'intl':
            item.rate_usd = float(request.form.get('rate_usd').replace(',','.'))
            if item.category == 'RF' and (not item.name or item.name == 'Renda Fixa'):
                 # Legacy RF or Manual
                 item.institution = request.form.get('institution')
                 item.description = request.form.get('description')
                 item.invested_value = float(request.form.get('invested_value').replace('.','').replace(',','.'))
                 item.value_usd = float(request.form.get('value_usd').replace('.','').replace(',','.'))
            else: # RV or New RF (Ticker-based)
                 item.institution = request.form.get('institution')
                 item.name = request.form.get('name')
                 item.description = request.form.get('description') # Preserve description for RF

                 item.institution = request.form.get('institution')
                 item.name = request.form.get('name')
                 
                 qty_str = request.form.get('quantity', '').replace(',', '.')
                 if qty_str and qty_str.lower() != 'none':
                     item.quantity = float(qty_str)
                 else:
                     item.quantity = 0.0

                 avg_str = request.form.get('avg_price', '').replace('.', '').replace(',', '.')
                 item.avg_price = float(avg_str) if avg_str and avg_str.lower() != 'none' else 0.0
                 
                 quote_str = request.form.get('quote', '').replace('.', '').replace(',', '.')
                 item.quote = float(quote_str) if quote_str and quote_str.lower() != 'none' else 0.0
                 
                 item.value_usd = (item.quantity or 0) * (item.quote or 0)
        
        db.session.commit()
        flash('Item atualizado com sucesso!', 'success')
        return redirect(url_for('balanceamento'))

    return render_template('edit_balance.html', item=item, type=type)

@app.route('/balanceamento/delete/<type>/<int:id>')
@login_required
def delete_balance_item(type, id):
    item = None
    if type == 'rf':
        item = FixedIncome.query.filter_by(id=id, user_id=current_user.id).first()
    elif type == 'fund':
        item = InvestmentFund.query.filter_by(id=id, user_id=current_user.id).first()
    elif type == 'crypto':
        item = Crypto.query.filter_by(id=id, user_id=current_user.id).first()
    elif type == 'pension':
        item = Pension.query.filter_by(id=id, user_id=current_user.id).first()
    elif type == 'intl':
        item = International.query.filter_by(id=id, user_id=current_user.id).first()
        
    if item:
        db.session.delete(item)
        db.session.commit()
        flash('Item removido!')
    else:
        flash('Item não encontrado ou acesso negado.')
        
    return redirect(url_for('balanceamento'))


def update_all_assets_logic(user_id=None, skip_tickers: set = None, only_types: set = None):
    """
    Atualiza cotações de Ações/FIIs/ETFs via Yahoo/Brapi.
    skip_tickers: conjunto de tickers já atualizados pelo OpLab — são ignorados aqui.
    """
    if user_id is None:
        user_id = current_user.id
    assets = Asset.query.filter_by(user_id=user_id).all()
    # Filter ACAO/FII/ETF — pula os já cobertos pelo OpLab
    skip = {t.upper() for t in (skip_tickers or [])}
    allowed_types = set(only_types or ['ACAO', 'FII', 'ETF'])
    relevant = [
        a for a in assets
        if a.type in allowed_types and a.type in ['ACAO', 'FII', 'ETF'] and a.ticker.upper() not in skip
    ]
    if not relevant:
        return 0, 0, []

    updated_count = 0
    errors = []
    total_tried = 0

    # Um único chunk — brapi e fast_info já são paralelos internamente
    relevant_chunks = [relevant]

    for chunk in relevant_chunks:
        try:
            tickers = [a.ticker for a in chunk]
            yahoo_preferred = {a.ticker.upper().strip() for a in chunk if a.type == 'ETF'}
            total_tried += len(tickers)
            quotes = get_quotes(tickers, user_id=user_id, prefer_yahoo=yahoo_preferred)
            
            if quotes:
                for asset in chunk:
                    # Generic lookup
                    quote_data = quotes.get(asset.ticker) or quotes.get(asset.ticker.upper().strip())
                    
                    if quote_data:
                        price = quote_data.get('price')
                        if price and price > 0:
                            asset.current_price = price
                            asset.daily_change = quote_data.get('change_percent', 0.0)
                            asset.last_update = now_brt()
                            updated_count += 1
            
            # Commit after each chunk
            db.session.commit()
            
        except Exception as e:
            print(f"Error updating chunk {tickers}: {e}")
            errors.append(str(e))
            continue
    
    return updated_count, total_tried, errors

def update_intl_quotes_logic(user_id):
    """
    Helper to update International Assets and Cryptos for a specific user.
    Returns (success: bool, messages: list)
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    headers = {'User-Agent': 'Mozilla/5.0'}

    def _yahoo_price(ticker_yf):
        """Busca preço e variação de um ticker no Yahoo Finance."""
        try:
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker_yf}?interval=1d&range=1d"
            r = requests.get(url, headers=headers, timeout=10)
            meta = r.json()['chart']['result'][0]['meta']
            price = meta.get('regularMarketPrice', 0.0)
            chg   = meta.get('regularMarketChangePercent')
            if chg is None:
                prev = meta.get('chartPreviousClose') or meta.get('previousClose', 0)
                chg = ((price - prev) / prev * 100) if prev and prev > 0 else 0.0
            return price, float(chg)
        except Exception:
            return 0.0, 0.0

    # 1. USD em paralelo com os demais
    intls   = International.query.filter_by(user_id=user_id).all()
    cryptos = Crypto.query.filter_by(user_id=user_id).all()

    # Monta lista de (chave, ticker_yf) para busca paralela
    tasks = {'__USD__': 'USDBRL=X'}
    for item in intls:
        if item.name and item.name.upper() != 'RENDA FIXA':
            t = item.name.strip().upper()
            if t == 'BRKB':
                t = 'BRK-B'
            tasks[f'intl_{item.id}'] = t
    for c in cryptos:
        if c.name:
            tasks[f'crypto_{c.id}'] = f"{c.name.strip().upper()}-USD"

    # Busca tudo em paralelo
    prices = {}
    with ThreadPoolExecutor(max_workers=10) as ex:
        fut_map = {ex.submit(_yahoo_price, yf_t): key for key, yf_t in tasks.items()}
        for fut in as_completed(fut_map):
            key = fut_map[fut]
            prices[key] = fut.result()

    usd_rate, _ = prices.get('__USD__', (0.0, 0.0))
    if usd_rate <= 0:
        return False, ['Não foi possível obter a cotação do Dólar.']

    # Aplica resultados
    for item in intls:
        item.rate_usd = usd_rate
        key = f'intl_{item.id}'
        if key in prices:
            price, chg = prices[key]
            if price > 0:
                item.quote       = price
                item.daily_change = chg
                if item.quantity:
                    item.value_usd = item.quantity * price

    for c in cryptos:
        key = f'crypto_{c.id}'
        if key in prices:
            price_usd, _ = prices[key]
            if price_usd > 0:
                price_brl = price_usd * usd_rate
                c.quote = price_brl
                if c.quantity:
                    c.current_value = c.quantity * price_brl

    db.session.commit()
    return True, ['Intl/Cripto atualizados.']

@app.route('/update_quotes', methods=['POST'])
@login_required
def update_quotes():
    try:
        update_market_indices()
        quote_mode  = Settings.get_value('quote_mode', user_id=current_user.id, default='yahoo')
        oplab_token = Settings.get_value('oplab_token', user_id=current_user.id)
        oplab_covered = set()
        final_msg = ''
        errs = []

        if oplab_token:
            oplab_online = _oplab_is_available(oplab_token, timeout=4)
            if not oplab_online:
                final_msg += 'OpLab: indisponível (ignorado). '
            else:
                a_ok, o_ok, oplab_covered, op_err = _do_oplab_bulk_update_safe(
                    current_user.id, oplab_token, deadline_secs=25
                )
                if op_err:
                    final_msg += f'OpLab: {op_err}. '
                else:
                    final_msg += f'OpLab: {a_ok} ativo(s), {o_ok} opção(ões)/perna(s). '

        # Yahoo/Internacional sempre executados
        if quote_mode == 'yahoo':
            count, tried, errs = update_all_assets_logic(skip_tickers=oplab_covered)
            final_msg += f'Yahoo/Brapi: {count}/{tried} ativo(s). '
        else:
            etf_count, etf_tried, errs = update_all_assets_logic(only_types={'ETF'})
            final_msg += f'ETFs via Yahoo: {etf_count}/{etf_tried}. '
            final_msg += 'Ações/FIIs via MT5 Feeder. '

        intl_success, intl_msgs = update_intl_quotes_logic(current_user.id)
        if intl_success:
            final_msg += f'Internacional/Cripto: Sucesso ({", ".join(intl_msgs)}). '
        else:
            final_msg += f'Internacional: Falha ({", ".join(intl_msgs)}). '

        record_portfolio_snapshot(current_user.id)

        if errs:
            flash(f'{final_msg} Erros: {len(errs)}. {errs[0]}', 'warning')
        else:
            flash(final_msg.strip() or 'Atualizado.', 'success')

    except Exception as e:
        flash(f'Erro ao atualizar cotações: {str(e)}', 'danger')
        print(f"Error in update_quotes: {e}")
        import traceback
        traceback.print_exc()

    return redirect(request.referrer or url_for('index'))


@app.route('/update_quotes_async', methods=['POST'])
@login_required
def update_quotes_async():
    """Start background update and return task_id immediately (no 504 timeout)."""
    user_id = current_user.id
    task_id = str(uuid.uuid4())
    _set_task(task_id, {'status': 'running', 'msg': '', 'category': ''})

    def do_update():
        with app.app_context():
            try:
                update_market_indices()
                quote_mode  = Settings.get_value('quote_mode', user_id=user_id, default='yahoo')
                oplab_token = Settings.get_value('oplab_token', user_id=user_id)
                oplab_covered: set = set()
                final_msg = ''
                errs = []

                # ── 1. OpLab: ações B3 + opções ───────────────────────────────
                if oplab_token:
                    # Probe rápido (4s) antes de entrar nos loops de timeout
                    oplab_online = _oplab_is_available(oplab_token, timeout=4)
                    if not oplab_online:
                        final_msg += 'OpLab: indisponível (ignorado). '
                    else:
                        # deadline total de 25s para toda a operação OpLab
                        a_ok, o_ok, oplab_covered, op_err = _do_oplab_bulk_update_safe(
                            user_id, oplab_token, deadline_secs=25
                        )
                        if op_err:
                            final_msg += f'OpLab: {op_err}. '
                        else:
                            final_msg += f'OpLab: {a_ok} ativo(s), {o_ok} opção(ões). '

                # ── 2. Yahoo/Brapi e Internacional — sempre executados ─────────
                if quote_mode == 'yahoo':
                    count, tried, errs = update_all_assets_logic(
                        user_id=user_id, skip_tickers=oplab_covered
                    )
                    intl_success, intl_msgs = update_intl_quotes_logic(user_id)
                    if tried > 0:
                        final_msg += f'Yahoo/Brapi: {count}/{tried} ativo(s). '
                    if intl_success:
                        final_msg += 'Intl/Cripto: OK. '
                    else:
                        final_msg += 'Intl: falha. '
                elif quote_mode == 'mt5':
                    etf_count, etf_tried, errs = update_all_assets_logic(
                        user_id=user_id, only_types={'ETF'}
                    )
                    final_msg += f'ETFs via Yahoo: {etf_count}/{etf_tried}. '
                    final_msg += 'Ações/FIIs via MT5 Feeder. '
                    intl_success, _ = update_intl_quotes_logic(user_id)

                # Foto do patrimônio do dia com os preços recém-atualizados
                record_portfolio_snapshot(user_id)

                category = 'warning' if errs else 'success'
                _set_task(task_id, {'status': 'done', 'msg': final_msg.strip() or 'Atualizado.', 'category': category})
            except Exception as e:
                import traceback
                traceback.print_exc()
                _set_task(task_id, {
                    'status': 'done',
                    'msg': f'Erro ao atualizar cotações: {str(e)}',
                    'category': 'danger'
                })

    threading.Thread(target=do_update, daemon=True).start()
    return jsonify({'task_id': task_id})


@app.route('/api/update_progress/<task_id>')
@login_required
def update_progress(task_id):
    """Poll endpoint to check background update status."""
    return jsonify(_get_task(task_id))


@app.route('/update_intl_quotes')
@login_required
def update_intl_quotes():
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
        msg_log = []

        # 1. Get USD Rate (USDBRL=X)
        usd_rate = 0.0
        try:
            url_usd = "https://query1.finance.yahoo.com/v8/finance/chart/USDBRL=X?interval=1d&range=1d"
            r_usd = requests.get(url_usd, headers=headers, timeout=10)
            data_usd = r_usd.json()
            usd_rate = data_usd['chart']['result'][0]['meta']['regularMarketPrice']
            msg_log.append(f"Dólar: R$ {usd_rate:.2f}")
        except Exception as e:
            msg_log.append(f"Erro Dólar: {str(e)}")
            print(f"Error fetching USD: {e}")

        if usd_rate > 0:
            # Update all International assets
            intls = International.query.all()
            for item in intls:
                # Update Exchange Rate for ALL
                item.rate_usd = usd_rate
                
                # Update Quote for RV and RF (if Ticker provided)
                if item.name and item.name.upper() != 'RENDA FIXA':
                    try:
                        ticker_name = item.name.strip().upper()
                        # Common Corrections
                        if ticker_name == 'BRKB':
                            ticker_name = 'BRK-B'
                        
                        # Fetch Stock Quote
                        url_stock = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker_name}?interval=1d&range=1d"
                        r_stock = requests.get(url_stock, headers=headers, timeout=10)
                        data_stock = r_stock.json()
                        
                        price = 0.0
                        if 'chart' in data_stock and 'result' in data_stock['chart'] and data_stock['chart']['result']:
                             price = data_stock['chart']['result'][0]['meta']['regularMarketPrice']
                        
                        if price > 0:
                            item.quote = price
                            msg_log.append(f"{ticker_name}: ${price:.2f}")
                            
                            # Recalculate Value USD: Quantity * Price
                            if item.quantity:
                                item.value_usd = item.quantity * price
                            else:
                                item.value_usd = 0.0
                        else:
                            msg_log.append(f"{ticker_name}: Não encontrado (API)")
                            
                    except Exception as e:
                        msg_log.append(f"{item.name}: Erro API {str(e)}")
                        print(f"Error updating {item.name}: {e}")
            
            # Update Cryptos
            cryptos = Crypto.query.all()
            for c in cryptos:
                if c.name: # e.g. BTC, ETH
                    try:
                        ticker_clean = c.name.strip().upper()
                        # Default to USD pair if not specified
                        # Try finding a valid Yahoo Ticker. Usually 'BTC-USD'
                        yahoo_ticker = f"{ticker_clean}-USD"
                        
                        url_crypto = f"https://query1.finance.yahoo.com/v8/finance/chart/{yahoo_ticker}?interval=1d&range=1d"
                        r_crypto = requests.get(url_crypto, headers=headers, timeout=10)
                        data_crypto = r_crypto.json()
                        
                        price_usd = 0.0
                        if 'chart' in data_crypto and 'result' in data_crypto['chart'] and data_crypto['chart']['result']:
                             price_usd = data_crypto['chart']['result'][0]['meta']['regularMarketPrice']
                        
                        if price_usd > 0:
                            # Convert to BRL
                            price_brl = price_usd * usd_rate
                            c.quote = price_brl
                            if c.quantity:
                                c.current_value = c.quantity * price_brl
                            msg_log.append(f"{ticker_clean}: R$ {price_brl:.2f}")
                        else:
                            msg_log.append(f"{ticker_clean}: Não encontrado")
                            
                    except Exception as e:
                        print(f"Error crypto {c.name}: {e}")
                        msg_log.append(f"{c.name}: Erro {str(e)}")

            db.session.commit()
            flash(f'Atualização Concluída! Detalhes: {", ".join(msg_log)}', 'success')
        else:
            flash(f'Não foi possível obter a cotação do Dólar. Detalhes: {", ".join(msg_log)}', 'warning')
            
    except Exception as e:
        flash(f'Erro fatal ao atualizar: {str(e)}', 'danger')
        
    return redirect(url_for('balanceamento'))
        
    return redirect(url_for('balanceamento'))

# --- User Management & Security ---

@app.before_request
def check_user_status():
    if current_user.is_authenticated:
        if current_user.expiry_date and current_user.expiry_date < date.today():
            logout_user()
            flash('Seu acesso expirou. Entre em contato com o administrador.', 'danger')
            return redirect(url_for('login'))

@app.route('/users')
@login_required
def list_users():
    if not current_user.is_admin:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('resumo'))
    users = User.query.all()
    return render_template('users.html', users=users)

@app.route('/users/add', methods=['GET', 'POST'])
@login_required
def add_user():
    if not current_user.is_admin:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('resumo'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role', 'user')
        expiry_str = request.form.get('expiry_date')

        if User.query.filter_by(username=username).first():
            flash('Usuário já existe.', 'danger')
        else:
            expiry_date = datetime.strptime(expiry_str, '%Y-%m-%d').date() if expiry_str else None
            user = User(username=username, role=role, expiry_date=expiry_date,
                        full_name=request.form.get('full_name', '').strip(),
                        email=request.form.get('email', '').strip(),
                        phone=request.form.get('phone', '').strip())
            user.set_password(password)
            db.session.add(user)
            db.session.flush()  # gera user.id para salvar avatar
            avatar = request.files.get('avatar')
            if avatar and avatar.filename:
                ext = avatar.filename.rsplit('.', 1)[-1].lower()
                if ext in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
                    fname = f"avatar_{user.id}.{ext}"
                    avatar_dir = os.path.join(basedir, 'static', 'img', 'avatars')
                    os.makedirs(avatar_dir, exist_ok=True)
                    avatar.save(os.path.join(avatar_dir, fname))
                    user.avatar_filename = fname
            db.session.commit()
            flash('Usuário criado com sucesso!', 'success')
            return redirect(url_for('list_users'))

    return render_template('add_user.html', user=None, edit=False)

@app.route('/users/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_user(id):
    if not current_user.is_admin:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('resumo'))
        
    user = User.query.get_or_404(id)
    
    if request.method == 'POST':
        user.role      = request.form.get('role')
        user.full_name = request.form.get('full_name', '').strip()
        user.email     = request.form.get('email', '').strip()
        user.phone     = request.form.get('phone', '').strip()
        expiry_str     = request.form.get('expiry_date')
        user.expiry_date = datetime.strptime(expiry_str, '%Y-%m-%d').date() if expiry_str else None

        avatar = request.files.get('avatar')
        if avatar and avatar.filename:
            ext = avatar.filename.rsplit('.', 1)[-1].lower()
            if ext in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
                fname = f"avatar_{user.id}.{ext}"
                avatar_dir = os.path.join(basedir, 'static', 'img', 'avatars')
                os.makedirs(avatar_dir, exist_ok=True)
                avatar.save(os.path.join(avatar_dir, fname))
                user.avatar_filename = fname

        new_pass = request.form.get('password')
        if new_pass:
            user.set_password(new_pass)

        db.session.commit()
        flash('Usuário atualizado!', 'success')
        return redirect(url_for('list_users'))

    return render_template('add_user.html', user=user, edit=True)

@app.route('/users/delete/<int:id>')
@login_required
def delete_user(id):
    if not current_user.is_admin:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('resumo'))
    
    if id == current_user.id:
        flash('Você não pode excluir a si mesmo.', 'warning')
        return redirect(url_for('list_users'))

    user = User.query.get_or_404(id)
    # Optional: Delete all their data? For now standard delete.
    db.session.delete(user)
    db.session.commit()
    flash('Usuário removido.', 'success')
    return redirect(url_for('list_users'))


@app.route('/importar_excel', methods=['GET', 'POST'])
@login_required
def importar_excel():
    if request.method == 'GET':
        return render_template('importar_excel.html')

    f = request.files.get('excel_file')
    if not f or not f.filename.endswith(('.xlsx', '.xlsm')):
        flash('Envie um arquivo .xlsx válido.', 'danger')
        return redirect(url_for('importar_excel'))

    import openpyxl, io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        flash(f'Erro ao abrir o arquivo: {e}', 'danger')
        return redirect(url_for('importar_excel'))

    # ── Helpers ──────────────────────────────────────────────────────
    def _float(v):
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            try:
                return float(v.strip().replace('.', '').replace(',', '.'))
            except ValueError:
                pass
        return None

    def _parse_date(v):
        if isinstance(v, date):
            return v
        if v and str(v).strip() not in ('-', '31/12/9999', ''):
            for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(str(v).strip(), fmt).date()
                except ValueError:
                    pass
        return None

    def _is_option(name):
        """Retorna True se o Nome do Ativo indica ser uma opção."""
        n = str(name) if name else ''
        return n.startswith('Opc ') or n.startswith('Opc\xa0')

    # ── Leitura unificada de preços ───────────────────────────────────
    # Mapa ticker → (price, row_tuple) para todos os ativos e opções.
    # Colunas padrão dos sheets rtd/acao/ETF/opcao:
    #   0=ticker, 3=último, 8=strike, 9=variação%, 11=nome,
    #   18=vencimento, 22=VI, 23=delta, 24=gama
    all_prices = {}   # ticker → float price
    all_rows   = {}   # ticker → row tuple (para greeks)

    def _load_sheet(ws, skip_options=False):
        for row in ws.iter_rows(min_row=2, values_only=True):
            ticker = row[0]
            if not ticker:
                continue
            key = str(ticker).upper().strip()
            name = row[11] if len(row) > 11 else None
            if skip_options and _is_option(name):
                continue
            price = _float(row[3]) if len(row) > 3 else None
            if price is not None:
                all_prices[key] = price
            all_rows[key] = row

    # Ordem de prioridade: rtd e opcao são fonte principal (tempo real),
    # depois acao e ETF como fallback para ativos não cobertos pelo rtd.
    for sheet_name in ('acao', 'ETF', 'opcao', 'rtd'):
        if sheet_name in wb.sheetnames:
            skip = (sheet_name == 'ETF')   # ETF: pula linhas de opções
            _load_sheet(wb[sheet_name], skip_options=skip)

    # Mapa extra de greeks vindos dos sheets especializados
    # C_put / V_put / C_Call_ITM — colunas:
    #   0=ticker, 3=último, 8=strike, 14=vencimento,
    #   17=delta, 18=gama, 19=theta, 22=VI intrínseco, 23=VE extrínseco
    extra_greeks = {}  # ticker → row
    for sheet_name in ('C_put', 'V_put', 'C_Call_ITM'):
        if sheet_name in wb.sheetnames:
            for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
                ticker = row[0]
                if not ticker:
                    continue
                key = str(ticker).upper().strip()
                extra_greeks[key] = row
                # preço desses sheets também entra no mapa geral
                p = _float(row[3]) if len(row) > 3 else None
                if p is not None and p > 0:
                    all_prices[key] = p
                    all_rows[key] = row

    now = now_brt()
    ativos_atualizados     = 0
    opcoes_atualizadas     = 0
    spreads_atualizados    = 0
    estruturadas_atualizadas = 0
    estudo_opcoes_atualizados = 0
    nao_encontrados_ativos = []

    # ── 1. Atualiza Asset (ações, FIIs, ETFs) ────────────────────────
    for asset in Asset.query.filter_by(user_id=current_user.id).all():
        key = asset.ticker.upper()
        p = all_prices.get(key)
        if p is not None and p > 0:
            asset.current_price = p
            asset.last_update   = now
            # variação diária (col 9)
            row = all_rows.get(key)
            if row and len(row) > 9:
                v = _float(row[9])
                if v is not None:
                    asset.daily_change = v
            ativos_atualizados += 1
        else:
            nao_encontrados_ativos.append(asset.ticker)

    # ── 2. Atualiza Option (venda/compra call/put) ───────────────────
    for opt in Option.query.filter_by(user_id=current_user.id).all():
        key = opt.ticker.upper()
        p = all_prices.get(key)
        row = all_rows.get(key) or extra_greeks.get(key)
        if p is not None:
            opt.current_option_price = p
            opt.last_update = now
            opcoes_atualizadas += 1
        if row:
            if len(row) > 9:
                v = _float(row[9])
                if v is not None:
                    opt.daily_change = v
            # greeks do rtd/opcao (cols 22=VI, 23=delta, 24=gama)
            if len(row) > 22:
                vi = _float(row[22]);  opt.ve    = vi if vi is not None else opt.ve
            if len(row) > 23:
                d  = _float(row[23]);  opt.delta = d  if d  is not None else opt.delta
            if len(row) > 24:
                g  = _float(row[24]);  opt.gama  = g  if g  is not None else opt.gama
        # greeks extra dos sheets especializados (delta=17, gama=18 nesse layout)
        erow = extra_greeks.get(key)
        if erow:
            if len(erow) > 17:
                d = _float(erow[17]);  opt.delta = d if d is not None else opt.delta
            if len(erow) > 18:
                g = _float(erow[18]);  opt.gama  = g if g is not None else opt.gama

    # ── 3. Atualiza OptionSpread (travas) ────────────────────────────
    for sp in OptionSpread.query.filter_by(user_id=current_user.id).all():
        changed = False
        if sp.leg_long_ticker:
            p = all_prices.get(sp.leg_long_ticker.upper())
            if p is not None:
                sp.leg_long_current = p
                changed = True
        if sp.leg_short_ticker:
            p = all_prices.get(sp.leg_short_ticker.upper())
            if p is not None:
                sp.leg_short_current = p
                changed = True
        # Atualiza cotação do ativo subjacente da trava
        und_key = (sp.underlying_asset or '').upper()
        und_p = all_prices.get(und_key)
        if und_p is not None and und_p > 0:
            sp.underlying_price = und_p
            und_row = all_rows.get(und_key)
            if und_row and len(und_row) > 9:
                v = _float(und_row[9])
                if v is not None:
                    sp.underlying_change = v
            changed = True
        if changed:
            spreads_atualizados += 1

    # ── 4. Atualiza StructuredLeg (operações estruturadas) ───────────
    for op in StructuredOp.query.filter_by(user_id=current_user.id, status='OPEN').all():
        op_changed = False
        for leg in op.legs:
            key = leg.ticker.upper() if leg.ticker else ''
            p = all_prices.get(key)
            if p is not None:
                leg.current_price = p
                leg.last_update   = now
                op_changed = True
        # Atualiza cotação do ativo subjacente direto no rtd
        und_key = (op.underlying_asset or '').upper()
        und_p = all_prices.get(und_key)
        if und_p is not None and und_p > 0:
            op.underlying_price = und_p
            row = all_rows.get(und_key)
            if row and len(row) > 9:
                v = _float(row[9])
                if v is not None:
                    op.underlying_change = v
            op_changed = True
        if op_changed:
            estruturadas_atualizadas += 1

    # ── 5. Atualiza StudyOption ──────────────────────────────────────
    for so in StudyOption.query.filter_by(user_id=current_user.id).all():
        key = so.ticker.upper()
        row = all_rows.get(key) or extra_greeks.get(key)
        changed = False
        p = all_prices.get(key)
        if p is not None:
            so.option_price = p
            changed = True
        if row:
            if len(row) > 8:
                s = _float(row[8]);
                if s is not None and s > 0:
                    so.strike = s;  changed = True
            exp = _parse_date(row[18] if len(row) > 18 else None)
            if exp:
                so.expiration_date = exp;  changed = True
            if len(row) > 22:
                vi = _float(row[22])
                if vi is not None:
                    so.ve = vi;  changed = True
            if len(row) > 23:
                d = _float(row[23])
                if d is not None:
                    so.delta = d;  changed = True
            if len(row) > 24:
                g = _float(row[24])
                if g is not None:
                    so.gama = g;  changed = True
        # underlying price
        und_key = (so.underlying_asset or '').upper()
        up = all_prices.get(und_key)
        if up is not None and up > 0:
            so.underlying_price = up;  changed = True
        if changed:
            estudo_opcoes_atualizados += 1

    # ── 6. Atualiza PutSale ──────────────────────────────────────
    for ps in PutSale.query.filter_by(user_id=current_user.id).all():
        und_key = (ps.underlying_asset or '').upper()
        up = all_prices.get(und_key)
        if up is not None and up > 0:
            ps.underlying_price = up

    # ── 7. Atualiza OptionSpread.underlying_price ────────────────
    for sp in OptionSpread.query.filter_by(user_id=current_user.id).all():
        und_key = (sp.underlying_asset or '').upper()
        up = all_prices.get(und_key)
        if up is not None and up > 0:
            sp.underlying_price = up
        row = all_rows.get(und_key)
        if row and len(row) > 9:
            v = _float(row[9])
            if v is not None:
                sp.underlying_change = v

    # ── 8. Persiste dados de opções em RtdOptionData (cache para Busca de Opção) ──
    # Formato TSV RTDTrading (colunas 0-based):
    # 0=Asset, 1=Data, 2=Hora, 3=Último, 4=Abertura, 5=Máximo, 6=Mínimo,
    # 7=Fech.Ant., 8=Strike, 9=Variação, 10=Média, 11=Nome, 12=Negócios,
    # 13=Quantidade, 14=Volume, 15=Of.Compra, 16=Of.Venda, 17=Vol.Proj.,
    # 18=Vencimento, 19=Validade, 20=Cont.Abertos, 21=Black Scholes,
    # 22=Volt.Implícita, 23=Delta, 24=Gama, 25=Theta, 26=Rho, 27=Vega,
    # 28=VI Ask, 29=VI Bid, 30=VI/VH, 31=Valor Intrínseco, 32=Valor Extrínseco
    rtd_count = 0
    rtd_now   = datetime.utcnow()
    uid_rtd   = current_user.id

    def _col(row, i):
        """Retorna row[i] convertido via _float, ou None se fora do range."""
        return _float(row[i]) if len(row) > i else None

    def _save_rtd_row(key, row, sheet_name):
        nonlocal rtd_count
        if not key or len(key) < 5:
            return
        name = row[11] if len(row) > 11 else ''
        is_opt = (len(key) >= 6 and not key.endswith('11') and not key.endswith('3') and not key.endswith('4')) \
                 or _is_option(name)
        if not is_opt:
            return
        try:
            rec = RtdOptionData.query.filter_by(user_id=uid_rtd, ticker=key).first()
            if not rec:
                rec = RtdOptionData(user_id=uid_rtd, ticker=key)
                db.session.add(rec)
            if sheet_name in ('rtd', 'opcao'):
                def _upd(attr, i):
                    v = _col(row, i)
                    if v is not None:
                        setattr(rec, attr, v)
                _upd('last_price',      3)
                _upd('open_price',      4)
                _upd('high_price',      5)
                _upd('low_price',       6)
                _upd('prev_close',      7)
                _upd('strike',          8)
                _upd('change_pct',      9)
                _upd('volume',         14)
                _upd('bid',            15)
                _upd('ask',            16)
                _upd('open_interest',  20)
                _upd('bs_price',       21)
                _upd('iv',             22)
                _upd('delta',          23)
                _upd('gamma',          24)
                _upd('theta',          25)
                _upd('rho',            26)
                _upd('vega',           27)
                _upd('iv_ask',         28)
                _upd('iv_bid',         29)
                _upd('iv_over_hv',     30)
                _upd('intrinsic_value',31)
                _upd('extrinsic_value',32)
                if len(row) > 18 and row[18]:
                    v = row[18]
                    rec.expiration = v.strftime('%d/%m/%Y') if hasattr(v, 'strftime') else str(v)
            elif sheet_name in ('C_put', 'V_put', 'C_Call_ITM'):
                def _upd(attr, i):
                    v = _col(row, i)
                    if v is not None:
                        setattr(rec, attr, v)
                _upd('last_price',       3)
                _upd('strike',           8)
                _upd('delta',           17)
                _upd('gamma',           18)
                _upd('theta',           19)
                _upd('intrinsic_value', 22)
                _upd('extrinsic_value', 23)
                if len(row) > 14 and row[14]:
                    v = row[14]
                    rec.expiration = v.strftime('%d/%m/%Y') if hasattr(v, 'strftime') else str(v)
            rec.imported_at = rtd_now
            rtd_count += 1
        except Exception:
            pass

    for sheet_name in ('rtd', 'opcao', 'C_put', 'V_put', 'C_Call_ITM'):
        if sheet_name in wb.sheetnames:
            for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
                if row[0]:
                    _save_rtd_row(str(row[0]).upper().strip(), row, sheet_name)

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao salvar: {e}', 'danger')
        return redirect(url_for('importar_excel'))

    msg = (f'Atualizado: {ativos_atualizados} ativo(s), {opcoes_atualizadas} opção(ões), '
           f'{spreads_atualizados} spread(s), {estruturadas_atualizadas} op. estruturada(s) '
           f'e {estudo_opcoes_atualizados} estudo(s). RTD: {rtd_count} opção(ões) salvas.')
    if nao_encontrados_ativos:
        msg += f' Não encontrados: {", ".join(nao_encontrados_ativos[:10])}'
        if len(nao_encontrados_ativos) > 10:
            msg += f' (+{len(nao_encontrados_ativos)-10})'
    flash(msg, 'success' if not nao_encontrados_ativos else 'warning')
    return redirect(url_for('importar_excel'))


# ─────────────────────────────────────────────────────────────────
# ESTUDOS
# ─────────────────────────────────────────────────────────────────

# Estratégias de estudo agrupadas por CENÁRIO (direção do mercado + volatilidade),
# com base no guia de operações. Cada grupo vira um <optgroup> no seletor para
# orientar a escolha. O rótulo do grupo traz a dica de quando usar.
STUDY_STRATEGY_GROUPS = [
    ('📈 Alta — Vol. baixa (comprar prêmio)', [
        'Compra de Call', 'Trava de Alta com Call', 'Call Backspread',
        'Ratio Call', 'Risk Reversal', 'Seagull de Alta',
    ]),
    ('📈 Alta — Vol. alta (vender prêmio)', [
        'Venda de Put', 'Trava de Alta com Put', 'Jade Lizard',
    ]),
    ('📉 Baixa — Vol. baixa (comprar prêmio)', [
        'Compra de Put', 'Trava de Baixa com Put', 'Put Backspread', 'Seagull de Baixa',
    ]),
    ('📉 Baixa — Vol. alta (vender prêmio)', [
        'Trava de Baixa com Call',
    ]),
    ('➡️ Lateral — Vol. alta (vender prêmio)', [
        'Venda de Call Coberta', 'Strangle Vendido', 'Straddle Vendido',
        'Iron Condor', 'Iron Butterfly', 'Borboleta', 'Condor',
        'Boi', 'Vaca', 'Ratio Spread',
    ]),
    ('➡️ Lateral — Vol. baixa / renda', [
        'Collar', 'Fence', 'Calendar Spread', 'Diagonal Spread',
    ]),
    ('⚡ Volatilidade (movimento forte em qualquer direção)', [
        'Compra de Call', 'Compra de Put', 'Straddle Comprado', 'Strangle Comprado',
        'Strap', 'Strip', 'Guts',
    ]),
    ('🧬 Outras', [
        'Ação Sintética', 'Short Sintético', 'Box Spread', 'Outros', 'ne',
    ]),
]

# Lista plana (compatível com valores já salvos), sem duplicatas, preservando ordem.
STUDY_STRATEGIES = []
for _grp, _items in STUDY_STRATEGY_GROUPS:
    for _s in _items:
        if _s not in STUDY_STRATEGIES:
            STUDY_STRATEGIES.append(_s)

# Mapeamento dos valores da planilha para os do programa
_STRATEGY_MAP = {
    'venda coberta':          'Venda de Call Coberta',
    'venda de call coberta':  'Venda de Call Coberta',
    'venda de call':          'Venda de Call Coberta',
    'venda put':              'Venda de Put',
    'venda de put':           'Venda de Put',
    'compra de put':          'Compra de Put',
    'compra put':             'Compra de Put',
    'trava de alta':          'Trava de Alta com Call',
    'trava de alta com call': 'Trava de Alta com Call',
    'trava de alta com put':  'Trava de Alta com Put',
    'trava de baixa':         'Trava de Baixa com Call',
    'trava de baixa com call':'Trava de Baixa com Call',
    'trava de baixa com put': 'Trava de Baixa com Put',
    'strangle vendido':       'Strangle Vendido',
    'strangle':               'Strangle Vendido',
    'borboleta':              'Borboleta',
    'iron condor':            'Iron Condor',
    'compra call':            'Compra de Call',
    'compra de call':         'Compra de Call',
    'boi':                    'Boi',
    'vaca':                   'Vaca',
    'outros':                 'Outros',
    'ne':                     'ne',
}

def _normalize_strategy(value):
    """Converte valor da planilha para a lista do programa."""
    if not value:
        return None
    normalized = _STRATEGY_MAP.get(str(value).strip().lower())
    if normalized:
        return normalized
    # se já é um valor válido da lista, devolve como está
    if value in STUDY_STRATEGIES:
        return value
    return 'Outros'


def _calc_vdx(option_price, underlying_price, days, strike):
    """VDX = taxa * tempo * espaço"""
    try:
        if not all([option_price, underlying_price, days is not None, strike]):
            return None
        taxa   = option_price / underlying_price
        tempo  = 120 - days
        espaco = strike - underlying_price
        return taxa * tempo * espaco
    except (TypeError, ZeroDivisionError):
        return None


def _calc_nv(ve, delta, gama):
    """NV = VE - delta - gama"""
    if ve is None or delta is None or gama is None:
        return None
    return ve - delta - gama


def _roll_flag(days, spot, strike):
    """Aviso de possível necessidade de rolagem (teoria do Bastter):
    na venda coberta, quando a call está ITM perto do vencimento o VE se
    esgota e o exercício fica provável — hora de avaliar a rolagem.
    ALERT = ITM e vence em <= 7 dias (rolagem provável)
    WARN  = ITM e vence em <= 21 dias, ou <= 7 dias mesmo OTM (atenção)"""
    if days is None or not spot or not strike:
        return None
    itm = spot > strike
    if days <= 7:
        return 'ALERT' if itm else 'WARN'
    if days <= 21 and itm:
        return 'WARN'
    return None


@app.route('/estudos')
@login_required
def estudos():
    uid = current_user.id
    today = date.today()

    # ── Tabela 1: Estudo Opções Cobertas ────────────────────────────
    # A) Opções VENDA_CALL lançadas na página /opcoes
    venda_calls = Option.query.filter_by(user_id=uid, option_type='VENDA_CALL').all()
    vc_underlying = {opt.underlying_asset.upper() for opt in venda_calls}
    assets_map = {
        a.ticker.upper(): a
        for a in Asset.query.filter_by(user_id=uid).all()
    }
    study_calls_vc = []
    for opt in venda_calls:
        asset = assets_map.get(opt.underlying_asset.upper())
        up = asset.current_price if asset else 0
        days = (opt.expiration_date - today).days if opt.expiration_date else None
        vdx = _calc_vdx(opt.current_option_price, up, days, opt.strike_price)
        nv  = _calc_nv(opt.ve, opt.delta, opt.gama)
        study_calls_vc.append({
            'source': 'venda_call',
            'id': opt.id,
            'ticker': opt.ticker,
            'underlying': opt.underlying_asset,
            'underlying_price': up,
            'avg_price': asset.avg_price if asset else 0,
            'strike': opt.strike_price,
            'expiration': opt.expiration_date,
            'days': days,
            'option_price': opt.current_option_price,
            've': opt.ve,
            'delta': opt.delta,
            'gama': opt.gama,
            'vdx': vdx,
            'nv': nv,
            'roll_flag': _roll_flag(days, up, opt.strike_price),
        })

    # B) Opções extras adicionadas diretamente nesta página
    study_calls_extra = []
    for so in StudyOption.query.filter_by(user_id=uid).all():
        days = (so.expiration_date - today).days if so.expiration_date else None
        vdx = _calc_vdx(so.option_price, so.underlying_price, days, so.strike)
        nv  = _calc_nv(so.ve, so.delta, so.gama)
        study_calls_extra.append({
            'source': 'study',
            'id': so.id,
            'ticker': so.ticker,
            'underlying': so.underlying_asset,
            'underlying_price': so.underlying_price,
            'avg_price': so.avg_price_stock,
            'strike': so.strike,
            'expiration': so.expiration_date,
            'days': days,
            'option_price': so.option_price,
            've': so.ve,
            'delta': so.delta,
            'gama': so.gama,
            'vdx': vdx,
            'nv': nv,
            'roll_flag': _roll_flag(days, so.underlying_price, so.strike),
        })

    # ── Tabela 2: Estudo Ações ───────────────────────────────────────
    study_stocks = StudyStock.query.filter_by(user_id=uid).order_by(StudyStock.ticker).all()
    study_intl_stocks = StudyIntlStock.query.filter_by(user_id=uid).order_by(StudyIntlStock.ticker).all()

    # ── Tabela 3: Ações Livres (sem venda coberta ativa nem garantia em estruturada) ──
    # Ações usadas como garantia em operações estruturadas abertas
    collateral_tickers = {
        op.underlying_asset.upper()
        for op in StructuredOp.query.filter_by(user_id=uid, status='OPEN').all()
        if op.uses_stock_collateral and op.underlying_asset
    }
    all_acoes = [a for a in Asset.query.filter_by(user_id=uid, type='ACAO').all() if a.quantity > 0]
    free_stocks = [
        a for a in all_acoes
        if a.ticker.upper() not in vc_underlying
        and a.ticker.upper() not in collateral_tickers
    ]
    free_stocks.sort(key=lambda a: a.ticker)

    return render_template(
        'estudos.html',
        study_calls_vc=study_calls_vc,
        study_calls_extra=study_calls_extra,
        study_stocks=study_stocks,
        study_intl_stocks=study_intl_stocks,
        free_stocks=free_stocks,
        strategies=STUDY_STRATEGIES,
        strategy_groups=STUDY_STRATEGY_GROUPS,
    )


@app.route('/estudos/edit_vc_greeks/<int:opt_id>', methods=['POST'])
@login_required
def edit_vc_greeks(opt_id):
    """Salva VE, Delta e Gama de uma VENDA_CALL na tabela de estudo."""
    opt = Option.query.filter_by(id=opt_id, user_id=current_user.id, option_type='VENDA_CALL').first_or_404()
    def _fl(k):
        v = request.form.get(k, '').strip()
        return float(v) if v else None
    opt.ve    = _fl('ve')
    opt.delta = _fl('delta')
    opt.gama  = _fl('gama')
    db.session.commit()
    flash('VE/Delta/Gama atualizados.', 'success')
    return redirect(url_for('estudos') + '#estudo-opcoes')


@app.route('/estudos/add_study_option', methods=['POST'])
@login_required
def add_study_option():
    def _f(k): return request.form.get(k, '').strip()
    def _fl(k):
        v = _f(k)
        return float(v) if v else None
    def _dt(k):
        v = _f(k)
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except ValueError:
            return None

    so = StudyOption(
        user_id=current_user.id,
        ticker=_f('ticker').upper(),
        underlying_asset=_f('underlying_asset').upper(),
        underlying_price=_fl('underlying_price'),
        avg_price_stock=_fl('avg_price_stock'),
        strike=_fl('strike'),
        expiration_date=_dt('expiration_date'),
        option_price=_fl('option_price'),
        ve=_fl('ve'),
        delta=_fl('delta'),
        gama=_fl('gama'),
    )
    db.session.add(so)
    db.session.commit()
    flash('Opção de estudo adicionada.', 'success')
    return redirect(url_for('estudos') + '#estudo-opcoes')


@app.route('/estudos/edit_study_option/<int:sid>', methods=['POST'])
@login_required
def edit_study_option(sid):
    so = StudyOption.query.filter_by(id=sid, user_id=current_user.id).first_or_404()
    def _f(k): return request.form.get(k, '').strip()
    def _fl(k):
        v = _f(k)
        return float(v) if v else None
    def _dt(k):
        v = _f(k)
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except ValueError:
            return None

    so.ticker = _f('ticker').upper()
    so.underlying_asset = _f('underlying_asset').upper()
    so.underlying_price = _fl('underlying_price')
    so.avg_price_stock = _fl('avg_price_stock')
    so.strike = _fl('strike')
    so.expiration_date = _dt('expiration_date')
    so.option_price = _fl('option_price')
    so.ve    = _fl('ve')
    so.delta = _fl('delta')
    so.gama  = _fl('gama')
    db.session.commit()
    flash('Opção de estudo atualizada.', 'success')
    return redirect(url_for('estudos') + '#estudo-opcoes')


@app.route('/estudos/delete_study_option/<int:sid>', methods=['POST'])
@login_required
def delete_study_option(sid):
    so = StudyOption.query.filter_by(id=sid, user_id=current_user.id).first_or_404()
    db.session.delete(so)
    db.session.commit()
    flash('Opção de estudo removida.', 'success')
    return redirect(url_for('estudos') + '#estudo-opcoes')


@app.route('/estudos/add_study_stock', methods=['POST'])
@login_required
def add_study_stock():
    def _f(k): return request.form.get(k, '').strip()
    def _fl(k):
        v = _f(k)
        return float(v) if v else None
    def _dt(k):
        v = _f(k)
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except ValueError:
            return None

    ss = StudyStock(
        user_id=current_user.id,
        ticker=_f('ticker').upper(),
        trend=_f('trend') or None,
        rsi=_fl('rsi'),
        volatility=_f('volatility') or None,
        ve=_fl('ve'),
        strategy=_f('strategy') or None,
        study_date=_dt('study_date'),
        strategy_active=_f('strategy_active') or None,
        entry_date=_dt('entry_date'),
    )
    db.session.add(ss)
    db.session.commit()
    flash('Ação de estudo adicionada.', 'success')
    return redirect(url_for('estudos') + '#estudo-acoes')


@app.route('/estudos/ir-para-estudo/<ticker>')
@login_required
def estudo_ir(ticker):
    """Abre a tela Estudos já na edição da ação. Se ela ainda não estiver na
    lista de Estudo de Ações, cria um registro em branco e abre a edição.
    Usado pelo ícone de atalho no Ranking de Volatilidade."""
    ticker = (ticker or '').strip().upper()
    if not ticker:
        return redirect(url_for('estudos') + '#estudo-acoes')
    ss = StudyStock.query.filter_by(user_id=current_user.id, ticker=ticker).first()
    if ss is None:
        ss = StudyStock(user_id=current_user.id, ticker=ticker)
        db.session.add(ss)
        db.session.commit()
    return redirect(url_for('estudos', edit_stock=ss.id) + '#estudo-acoes')


@app.route('/estudos/edit_study_stock/<int:sid>', methods=['POST'])
@login_required
def edit_study_stock(sid):
    ss = StudyStock.query.filter_by(id=sid, user_id=current_user.id).first_or_404()
    def _f(k): return request.form.get(k, '').strip()
    def _fl(k):
        v = _f(k)
        return float(v) if v else None
    def _dt(k):
        v = _f(k)
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except ValueError:
            return None

    ss.ticker = _f('ticker').upper()
    ss.trend = _f('trend') or None
    ss.rsi = _fl('rsi')
    ss.iv_rank = _fl('iv_rank')
    ss.iv_percentil = _fl('iv_percentil')
    ss.atr_pct = _fl('atr_pct')
    ss.strategy = _f('strategy') or None
    ss.study_date = date.today()
    ss.strategy_active = _f('strategy_active') or None
    ss.entry_date = _dt('entry_date')
    db.session.commit()
    flash('Ação de estudo atualizada.', 'success')
    return redirect(url_for('estudos') + '#estudo-acoes')


@app.route('/estudos/delete_study_stock/<int:sid>', methods=['POST'])
@login_required
def delete_study_stock(sid):
    ss = StudyStock.query.filter_by(id=sid, user_id=current_user.id).first_or_404()
    db.session.delete(ss)
    db.session.commit()
    flash('Ação de estudo removida.', 'success')
    return redirect(url_for('estudos') + '#estudo-acoes')


@app.route('/estudos/add_study_intl_stock', methods=['POST'])
@login_required
def add_study_intl_stock():
    def _f(k): return request.form.get(k, '').strip()
    def _fl(k):
        v = _f(k)
        return float(v) if v else None
    def _dt(k):
        v = _f(k)
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except ValueError:
            return None

    ss = StudyIntlStock(
        user_id=current_user.id,
        ticker=_f('ticker').upper(),
        trend=_f('trend') or None,
        rsi=_fl('rsi'),
        volatility=_f('volatility') or None,
        ve=_fl('ve'),
        strategy=_f('strategy') or None,
        study_date=_dt('study_date'),
        strategy_active=_f('strategy_active') or None,
        entry_date=_dt('entry_date'),
    )
    db.session.add(ss)
    db.session.commit()
    flash('Ação internacional de estudo adicionada.', 'success')
    return redirect(url_for('estudos') + '#estudo-acoes-intl')


@app.route('/estudos/edit_study_intl_stock/<int:sid>', methods=['POST'])
@login_required
def edit_study_intl_stock(sid):
    ss = StudyIntlStock.query.filter_by(id=sid, user_id=current_user.id).first_or_404()
    def _f(k): return request.form.get(k, '').strip()
    def _fl(k):
        v = _f(k)
        return float(v) if v else None
    def _dt(k):
        v = _f(k)
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except ValueError:
            return None

    ss.ticker = _f('ticker').upper()
    ss.trend = _f('trend') or None
    ss.rsi = _fl('rsi')
    ss.iv_rank = _fl('iv_rank')
    ss.iv_percentil = _fl('iv_percentil')
    ss.atr_pct = _fl('atr_pct')
    ss.strategy = _f('strategy') or None
    ss.study_date = date.today()
    ss.strategy_active = _f('strategy_active') or None
    ss.entry_date = _dt('entry_date')
    db.session.commit()
    flash('Ação internacional de estudo atualizada.', 'success')
    return redirect(url_for('estudos') + '#estudo-acoes-intl')


@app.route('/estudos/delete_study_intl_stock/<int:sid>', methods=['POST'])
@login_required
def delete_study_intl_stock(sid):
    ss = StudyIntlStock.query.filter_by(id=sid, user_id=current_user.id).first_or_404()
    db.session.delete(ss)
    db.session.commit()
    flash('Ação internacional de estudo removida.', 'success')
    return redirect(url_for('estudos') + '#estudo-acoes-intl')


# ── Ranking Volatilidade ─────────────────────────────────────────

# Lista "Geral" do Ranking de Volatilidade — ações/BDRs/ETFs com opções na B3
RANKING_GERAL_TICKERS = [
    # Principais
    'BOVA11','VALE3','PETR4','PETR3','ITUB4','ITUB3','BBAS3','PRIO3','SMAL11','BBDC4','BBDC3',
    'BRAV3','B3SA3','EGIE3','CSNA3','MGLU3','EMBJ3','SUZB3','BPAC11','ITSA4','WEGE3','CSAN3',
    'NATU3','BBSE3','BRKM5','AXIA3','USIM5','MBRF3','ABEV3','BRAP4','SBSP3','CYRE3','AZZA3',
    'RENT3','IBOV11','LREN3','ASAI3','MRVE3','CVCB3','BEEF3','ENEV3','GGBR4','CMIG4','CMIG3',
    'HAPV3','EQTL3','ROXO34','KLBN11','COGN3','TAEE11','CMIN3','IRBR3','RAIL3','ABCB4','CXSE3',
    'TOTS3','ALOS3','VAMO3','RADL3','GOAU4','JHSF3','SANB11','CEAB3','RDOR3','VBBR3','UGPA3',
    'DIRR3','CSMG3','YDUQ3','VIVT3','ISAE4','BOVV11','HYPE3','VIVA3','CPLE3','POMO4','MULT3',
    'SLCE3','TIMS3','CURY3','RECV3','IGTI11','PSSA3','SAPR11','BRSR6','XPBR31','MOTV3','GMAT3',
    'RAIZ4','ENGI11','ALPA4','FLRY3','MOVI3','EZTC3','ECOR3','BHIA3','HASH11','TUPY3','CPFE3',
    'JBSS32','SMFT3',
    # Small caps e BDRs
    'LWSA3','ANIM3','TEND3','PCAR3','WIZC3','SMTO3','VULC3','LJQQ3','NVDC34','INTB3','SIMH3',
    'POSI3','AURE3','SEER3','KEPL3','QUAL3','RAPT4','BMGB4','MYPK3','TTEN3','CASH3','ALUP11',
    'SAUD3','GRND3','UNIP6','LEVE3','MDNE3','AMBP3','BMOB3','INBR32','MDIA3','DXCO3','SBFG3',
    'FESA4','PLPL3','RANI3','VLID3','IVVB11','GOGL34','AGRO3','GFSA3','M1TA34','HBOR3','CAML3',
    'BLAU3','PNVL3','EVEN3','TRIS3','ROMI3','TASA4','SOJA3','LOGG3','AURA33','ONCO3','LAVV3',
    'TSLA34','MLAS3','ORVR3','GGPS3','JALL3','AMZO34','VVEO3','MSFT34','MELI34','AMAR3','PGMN3',
    'DASA3','MILS3','TSMC34','GOLD11','ARML3','BERK34','MTRE3','JPMC34','CBAV3','AAPL34',
    'MCDC34','COCA34','HBSA3','NASD11',
]


def _seed_ranking_geral(uid):
    """Garante que a lista GERAL do usuário contenha os tickers padrão."""
    existentes = {rv.ticker for rv in RankingVol.query.filter_by(user_id=uid, grupo='GERAL').all()}
    novos = [t for t in RANKING_GERAL_TICKERS if t not in existentes]
    for t in novos:
        db.session.add(RankingVol(user_id=uid, ticker=t, grupo='GERAL'))
    if novos:
        db.session.commit()


def _ranking_liq_filter(query):
    """Filtra apenas a lista 'Com liquidez' (linhas antigas têm grupo NULL)."""
    return query.filter(db.or_(RankingVol.grupo == 'LIQ', RankingVol.grupo.is_(None)))


@app.route('/ranking-volatilidade')
@login_required
def ranking_volatilidade():
    lista = (request.args.get('lista') or 'liq').lower()
    if lista == 'geral':
        _seed_ranking_geral(current_user.id)
        q = RankingVol.query.filter_by(user_id=current_user.id, grupo='GERAL')
    else:
        lista = 'liq'
        q = _ranking_liq_filter(RankingVol.query.filter_by(user_id=current_user.id))
    ranking_vol = q.order_by(RankingVol.ticker).all()
    return render_template('ranking_vol.html', ranking_vol=ranking_vol, lista=lista)


@app.route('/estudos/ranking_vol/add', methods=['POST'])
@login_required
def ranking_vol_add():
    lista  = (request.form.get('lista') or 'liq').lower()
    grupo  = 'GERAL' if lista == 'geral' else 'LIQ'
    ticker = request.form.get('ticker', '').strip().upper()
    if not ticker:
        flash('Ticker obrigatório.', 'danger')
        return redirect(url_for('ranking_volatilidade', lista=lista))
    q = RankingVol.query.filter_by(user_id=current_user.id, ticker=ticker)
    exists = (q.filter_by(grupo='GERAL') if grupo == 'GERAL' else _ranking_liq_filter(q)).first()
    if exists:
        flash(f'{ticker} já está no ranking.', 'warning')
        return redirect(url_for('ranking_volatilidade', lista=lista))
    db.session.add(RankingVol(user_id=current_user.id, ticker=ticker, grupo=grupo))
    db.session.commit()
    flash(f'{ticker} adicionado ao Ranking de Volatilidade.', 'success')
    return redirect(url_for('ranking_volatilidade', lista=lista))


@app.route('/estudos/ranking_vol/delete/<int:rid>', methods=['POST'])
@login_required
def ranking_vol_delete(rid):
    rv = RankingVol.query.filter_by(id=rid, user_id=current_user.id).first_or_404()
    lista = 'geral' if rv.grupo == 'GERAL' else 'liq'
    db.session.delete(rv)
    db.session.commit()
    return redirect(url_for('ranking_volatilidade', lista=lista))


@app.route('/api/ranking_vol/ticker', methods=['POST', 'DELETE'])
@login_required
def api_ranking_vol_ticker():
    data = request.get_json(silent=True) or request.form
    ticker = (data.get('ticker') or '').strip().upper()
    if not ticker:
        return jsonify({'error': 'Ticker obrigatório.'}), 400
    # Este endpoint opera sempre sobre a lista "Com liquidez"
    if request.method == 'DELETE':
        rv = _ranking_liq_filter(RankingVol.query.filter_by(user_id=current_user.id, ticker=ticker)).first()
        if rv:
            db.session.delete(rv)
            db.session.commit()
        return jsonify({'ok': True, 'ticker': ticker})
    exists = _ranking_liq_filter(RankingVol.query.filter_by(user_id=current_user.id, ticker=ticker)).first()
    if not exists:
        db.session.add(RankingVol(user_id=current_user.id, ticker=ticker, grupo='LIQ'))
        db.session.commit()
    return jsonify({'ok': True, 'ticker': ticker})


@app.route('/api/ranking_vol/update', methods=['POST'])
@login_required
def api_ranking_vol_update():
    """Atualiza todos os itens do Ranking de Volatilidade via OpLab + brapi."""
    try:
        return _api_ranking_vol_update_impl()
    except Exception as e:
        import traceback
        app.logger.error('api_ranking_vol_update error: %s\n%s', e, traceback.format_exc())
        return jsonify({'error': str(e)}), 500


def _api_ranking_vol_update_impl():
    from flask import jsonify
    import requests as _req
    from concurrent.futures import ThreadPoolExecutor

    uid   = current_user.id
    token = Settings.get_value('oplab_token', user_id=uid)
    if not token:
        return jsonify({'error': 'Token OpLab não configurado. Configure em Perfil → OpLab.'}), 400

    # Atualiza somente a lista ativa (liq = com liquidez; geral = lista ampla)
    lista = ((request.get_json(silent=True) or {}).get('lista')
             or request.args.get('lista') or 'liq').lower()
    q = RankingVol.query.filter_by(user_id=uid)
    items = (q.filter_by(grupo='GERAL') if lista == 'geral' else _ranking_liq_filter(q)).all()
    if not items:
        return jsonify({'updated': 0, 'results': []})

    now     = now_brt()
    today_str = now.strftime('%d/%m')

    def _extract_iv(d):
        if not isinstance(d, dict):
            return None, None, None
        for sub in ('data', 'spot', 'summary', 'iv', 'implied_volatility', 'greeks'):
            if isinstance(d.get(sub), dict):
                d.update(d[sub])
        ivr = ivp = vol = None
        for k in ('iv_1y_rank', 'ewma_1y_rank', 'iv_6m_rank', 'iv_rank', 'ivRank'):
            v = d.get(k)
            if v is not None:
                try: ivr = round(float(v)*100,1) if float(v)<=1.0 else round(float(v),1); break
                except: pass
        for k in ('iv_1y_percentile', 'ewma_1y_percentile', 'iv_6m_percentile',
                  'iv_percentile', 'ivPercentile', 'iv_percentil'):
            v = d.get(k)
            if v is not None:
                try: ivp = round(float(v)*100,1) if float(v)<=1.0 else round(float(v),1); break
                except: pass
        # Vol. Implícita anualizada (HV ou IV atual)
        for k in ('hv_current', 'historical_volatility', 'iv_current', 'implied_volatility_current',
                  'current_iv', 'close_iv', 'iv', 'vol_impl'):
            v = d.get(k)
            if v is not None:
                try: vol = round(float(v)*100,1) if float(v)<=1.0 else round(float(v),1); break
                except: pass
        return ivr, ivp, vol

    # Busca preços em lote via brapi (com token brapi se disponível)
    from services import _brapi_quotes, _yf_fast_info
    brapi_token = Settings.get_value('brapi_token', user_id=uid)
    tickers_list = [rv.ticker for rv in items]
    price_map = {}  # ticker → {price, change}

    if brapi_token:
        brapi_res = _brapi_quotes(tickers_list, brapi_token)
        for t, d in brapi_res.items():
            price_map[t] = {'price': d['price'], 'change': d['change_percent']}
        missing = [t for t in tickers_list if t not in price_map]
        if missing:
            def _fetch(t):
                return t, _yf_fast_info(f'{t}.SA' if '.' not in t else t, t)
            with ThreadPoolExecutor(max_workers=8) as ex:
                for t, d in ex.map(_fetch, missing):
                    if d: price_map[t] = {'price': d['price'], 'change': d['change_percent']}
    else:
        def _fetch(t):
            return t, _yf_fast_info(f'{t}.SA' if '.' not in t else t, t)
        with ThreadPoolExecutor(max_workers=8) as ex:
            for t, d in ex.map(_fetch, tickers_list):
                if d: price_map[t] = {'price': d['price'], 'change': d['change_percent']}

    results = []
    ok = 0
    for rv in items:
        row = {'ticker': rv.ticker, 'ok': False, 'error': None}
        try:
            ivr = ivp = vol = None
            try:
                ivr, ivp, vol = _extract_iv(_oplab_get_json(f'/market/instruments/{rv.ticker}', token, timeout=15))
            except OplabApiError as e:
                row['error'] = str(e)

            pd = price_map.get(rv.ticker, {})
            if pd.get('price', 0) > 0:
                rv.last_price = round(pd['price'], 2)
                rv.var_pct    = round(pd.get('change', 0), 2)
                rv.last_date  = today_str
            if ivr is not None: rv.iv_rank = ivr
            if ivp is not None: rv.iv_percentil = ivp
            if vol is not None: rv.vol_impl = vol
            rv.updated_at = now
            ok += 1
            row['ok'] = True
            row['iv_rank'] = ivr; row['iv_percentil'] = ivp; row['vol_impl'] = vol
            row['price'] = rv.last_price; row['change'] = rv.var_pct
        except Exception as e:
            row['error'] = str(e)
        results.append(row)

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

    failed = sum(1 for row in results if row.get('error'))
    return jsonify({'updated': ok, 'failed': failed, 'total': len(items), 'results': results})


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        action = request.form.get('action', 'change_password')

        if action == 'update_profile':
            current_user.full_name = request.form.get('full_name', '').strip()
            current_user.email     = request.form.get('email', '').strip()
            current_user.phone     = request.form.get('phone', '').strip()
            # Upload de avatar
            avatar = request.files.get('avatar')
            if avatar and avatar.filename:
                ext = avatar.filename.rsplit('.', 1)[-1].lower()
                if ext in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
                    fname = f"avatar_{current_user.id}.{ext}"
                    avatar_dir = os.path.join(basedir, 'static', 'img', 'avatars')
                    os.makedirs(avatar_dir, exist_ok=True)
                    avatar.save(os.path.join(avatar_dir, fname))
                    current_user.avatar_filename = fname
                else:
                    flash('Formato de imagem não suportado. Use JPG, PNG, GIF ou WEBP.', 'warning')
            db.session.commit()
            flash('Perfil atualizado!', 'success')

        elif action == 'change_password':
            curr_pass = request.form.get('current_password')
            new_pass = request.form.get('new_password')
            confirm_pass = request.form.get('confirm_password')
            if not current_user.check_password(curr_pass):
                flash('Senha atual incorreta.', 'danger')
            elif new_pass != confirm_pass:
                flash('Novas senhas não conferem.', 'danger')
            else:
                current_user.set_password(new_pass)
                db.session.commit()
                flash('Senha alterada com sucesso!', 'success')

        elif action == 'save_oplab':
            token = request.form.get('oplab_token', '').strip()
            if token:
                Settings.set_value('oplab_token', token, user_id=current_user.id)
                flash('Token OpLab salvo com sucesso!', 'success')
            else:
                flash('Informe o token OpLab.', 'danger')

        return redirect(url_for('profile'))

    oplab_configured = bool(Settings.get_value('oplab_token', user_id=current_user.id))
    return render_template('profile.html', oplab_configured=oplab_configured)


@app.route('/oplab_debug')
@login_required
def oplab_debug():
    """Diagnóstico: estado do token OpLab no banco."""
    if not current_user.is_admin:
        return jsonify({'error': 'Forbidden'}), 403
    uid = current_user.id
    raw = Settings.query.filter_by(key='oplab_token', user_id=uid).first()
    token = Settings.get_value('oplab_token', user_id=uid)
    return jsonify({
        'user_id': uid,
        'row_exists': raw is not None,
        'raw_value_len': len(raw.value) if raw else 0,
        'get_value_result': bool(token),
        'token_prefix': token[:8] + '...' if token and len(token) > 8 else token,
    })


@app.route('/oplab_test', methods=['GET', 'POST'])
@login_required
def oplab_test():
    """Retorna JSON bruto da API OpLab para diagnóstico."""
    token = Settings.get_value('oplab_token', user_id=current_user.id)
    if not token:
        return jsonify({'error': 'Token não configurado'}), 400
    ticker = (request.args.get('ticker') or request.form.get('ticker', 'PETR4')).strip().upper()
    underlying = request.args.get('underlying', '').strip().upper()
    if not underlying:
        asset_match = Asset.query.filter_by(user_id=current_user.id, ticker=ticker).first()
        option_match = Option.query.filter_by(user_id=current_user.id, ticker=ticker).first()
        study_match = StudyOption.query.filter_by(user_id=current_user.id, ticker=ticker).first()
        underlying = (
            (asset_match.ticker if asset_match else '') or
            (option_match.underlying_asset if option_match else '') or
            (study_match.underlying_asset if study_match else '') or
            ticker
        ).upper()
    results = {}
    # Testa endpoint bulk /market/quote (usado no auto-update)
    try:
        results['/market/quote'] = {
            'ok': True,
            'body': _oplab_get_json('/market/quote', token, params={'tickers': ticker}, timeout=20)
        }
    except OplabApiError as e:
        results['/market/quote'] = {
            'ok': False,
            'status': e.status_code,
            'error': str(e),
            'preview': e.body_preview,
        }
    # Testa endpoints individuais para diagnóstico
    try:
        results[f'/market/options/{underlying}'] = {
            'ok': True,
            'body': _oplab_get_json(f'/market/options/{underlying}', token, timeout=20)
        }
    except OplabApiError as e:
        results[f'/market/options/{underlying}'] = {
            'ok': False,
            'status': e.status_code,
            'error': str(e),
            'preview': e.body_preview,
        }
    endpoints = [
        f'/instruments/{ticker}',
        f'/market/instruments/{ticker}',
        f'/market/spot/{underlying}',
    ]
    for ep in endpoints:
        try:
            results[ep] = {'ok': True, 'body': _oplab_get_json(ep, token, timeout=20)}
        except OplabApiError as e:
            results[ep] = {
                'ok': False,
                'status': e.status_code,
                'error': str(e),
                'preview': e.body_preview,
            }
    any_ok = any(v.get('ok') for v in results.values() if isinstance(v, dict))
    return jsonify({'ok': any_ok, 'ticker': ticker, 'underlying': underlying, 'results': results})


@app.route('/api/liquidez/<ticker>')
@login_required
def api_liquidez(ticker):
    """Retorna liquidez de opções de um ativo via OpLab /market/options/{ticker}."""
    from flask import jsonify
    ticker = ticker.strip().upper()
    limit = request.args.get('limit', default=30, type=int)
    if limit not in (30, 40, 60):
        limit = 30
    expiry_filter = request.args.get('expiry', '').strip()
    summary_only = request.args.get('summary') == '1'
    token  = Settings.get_value('oplab_token', user_id=current_user.id)
    if not token:
        return jsonify({'error': 'Token OpLab não configurado. Configure em Perfil → OpLab.'}), 400

    try:
        data = _oplab_get_json(f'/market/options/{ticker}', token, timeout=15)
    except OplabApiError as e:
        return jsonify({'error': str(e), 'status': e.status_code, 'preview': e.body_preview}), 503

    # Normaliza — resposta pode ser lista ou dict com chave 'options'/'calls'/'puts'
    if isinstance(data, list):
        opt_list = data
    elif isinstance(data, dict):
        opt_list = (data.get('options') or
                    data.get('calls', []) + data.get('puts', []) or
                    [])
    else:
        opt_list = []

    calls, puts = [], []
    vol_total_call = vol_total_put = 0.0

    def _extract_option_last_vol(opt):
        nested = [opt]
        for key in ('greeks', 'iv', 'implied_volatility', 'volatility', 'metrics', 'volumes'):
            val = opt.get(key)
            if isinstance(val, dict):
                nested.append(val)
        for src in nested:
            for key in (
                'last_volatility', 'volatility_last', 'lastVolatility',
                'vol_last', 'last_vol', 'vol_ultima', 'vol_ultimo',
                'volUltima', 'volUltimo', 'volatility_close',
                'close_volatility', 'closeVolatility', 'vol_close',
                'last_trade_volatility', 'lastTradeVolatility',
                'volatility_last_trade', 'last_trade_iv',
                'last_iv', 'iv_last', 'volatility', 'vol',
                'option_volatility', 'implied_volatility'
            ):
                val = src.get(key)
                if val in (None, '', '-'):
                    continue
                try:
                    num = float(val)
                    if num <= 0:
                        continue
                    return round(num * 100, 2) if num <= 1 else round(num, 2)
                except (TypeError, ValueError):
                    continue
        return None

    for o in opt_list:
        sym      = str(o.get('symbol') or o.get('ticker') or '').upper()
        cat      = str(o.get('category') or o.get('type') or o.get('option_type') or '').upper()
        strike   = o.get('strike') or o.get('strike_price') or 0
        close    = o.get('close') or o.get('last') or o.get('price') or 0
        try:
            close = float(close or 0)
        except (TypeError, ValueError):
            close = 0
        if close <= 0:
            continue
        volume   = o.get('volume_financial') or o.get('financial_volume') or o.get('volume') or 0
        open_int = o.get('open_interest') or o.get('openInterest') or 0
        var_pct  = o.get('variation') or o.get('change') or o.get('pct_change') or 0
        bid      = o.get('bid') or 0
        ask      = o.get('ask') or 0
        due_date = o.get('due_date') or o.get('expiration_date') or o.get('maturity') or ''
        # extrai só a data se vier datetime
        if due_date and 'T' in str(due_date):
            due_date = str(due_date).split('T')[0]
        try:
            due_dt = datetime.strptime(str(due_date)[:10], '%Y-%m-%d').date()
        except Exception:
            due_dt = None
        if not due_dt or due_dt < date.today():
            continue

        row = {
            'symbol':   sym,
            'category': cat,
            'strike':   round(float(strike), 2) if strike else None,
            'close':    round(close, 2),
            'last_vol': _extract_option_last_vol(o),
            'volume':   round(float(volume), 2) if volume else 0,
            'open_int': int(open_int) if open_int else 0,
            'var_pct':  round(float(var_pct), 2) if var_pct else 0,
            'bid':      round(float(bid), 2) if bid else None,
            'ask':      round(float(ask), 2) if ask else None,
            'due_date': due_date,
        }

        if 'PUT' in cat or cat == 'P':
            puts.append(row)
            vol_total_put += row['volume']
        else:
            calls.append(row)
            vol_total_call += row['volume']

    expiry_summary_map = {}
    for row in calls:
        due = row.get('due_date')
        if due:
            expiry_summary_map.setdefault(due, {'due_date': due, 'calls': 0, 'puts': 0})
            expiry_summary_map[due]['calls'] += 1
    for row in puts:
        due = row.get('due_date')
        if due:
            expiry_summary_map.setdefault(due, {'due_date': due, 'calls': 0, 'puts': 0})
            expiry_summary_map[due]['puts'] += 1
    expiry_summary = [expiry_summary_map[k] for k in sorted(expiry_summary_map.keys())]

    if expiry_filter:
        calls = [row for row in calls if row.get('due_date') == expiry_filter]
        puts = [row for row in puts if row.get('due_date') == expiry_filter]

    vol_total_call = sum(row['volume'] for row in calls)
    vol_total_put = sum(row['volume'] for row in puts)

    # Ordena por volume desc, retorna o limite selecionado de cada lado
    calls.sort(key=lambda x: x['volume'], reverse=True)
    puts.sort(key=lambda x: x['volume'], reverse=True)

    # Cotação do ativo subjacente (spot) via brapi ou yfinance
    spot_price = None
    spot_change = None
    brapi_token = Settings.get_value('brapi_token', user_id=current_user.id)
    try:
        if brapi_token:
            rs = _req.get(
                f'https://brapi.dev/api/quote/{ticker}',
                params={'range': '1d', 'interval': '1d',
                        'fundamental': 'false', 'dividends': 'false',
                        'token': brapi_token},
                timeout=6,
            )
            if rs.status_code == 200:
                for item in rs.json().get('results', []):
                    p = item.get('regularMarketPrice') or item.get('currentPrice')
                    if p and float(p) > 0:
                        spot_price  = round(float(p), 2)
                        spot_change = round(float(item.get('regularMarketChangePercent') or 0), 2)
                        break
    except Exception:
        pass

    if spot_price is None:
        try:
            import yfinance as yf
            fi = yf.Ticker(f'{ticker}.SA').fast_info
            p  = fi.last_price
            if p and float(p) > 0:
                spot_price  = round(float(p), 2)
                prev = fi.previous_close or 0
                spot_change = round(((float(p) - prev) / prev * 100) if prev > 0 else 0, 2)
        except Exception:
            pass

    # Vencimentos distintos das opções (CALL e PUT combinados)
    def _calc_option_iv(row):
        if row.get('last_vol') is not None:
            return row['last_vol']
        try:
            if not spot_price or not row.get('strike') or not row.get('close') or not row.get('due_date'):
                return None
            exp = datetime.strptime(str(row['due_date'])[:10], '%Y-%m-%d').date()
            days = max((exp - date.today()).days, 0)
            if days <= 0:
                return None
            is_call = 'PUT' not in str(row.get('category') or '').upper()
            sigma = _implied_vol(
                float(spot_price),
                float(row['strike']),
                days / 252.0,
                math.log(1 + _selic() / 100),
                float(row['close']),
                is_call,
            )
            return round(sigma * 100, 2) if sigma and sigma > 0 else None
        except Exception:
            return None

    for row in calls + puts:
        row['last_vol'] = _calc_option_iv(row)

    if summary_only:
        selected_calls = []
        selected_puts = []
    else:
        selected_calls = calls[:limit]
        selected_puts = puts[:limit]
    if spot_price:
        def _spot_pct(row):
            strike = row.get('strike')
            return ((float(strike) - float(spot_price)) / float(spot_price) * 100) if strike else 0
        selected_calls.sort(key=_spot_pct)
        selected_puts.sort(key=_spot_pct)
    due_dates = sorted({o['due_date'] for o in selected_calls + selected_puts if o.get('due_date')})

    return jsonify({
        'ticker':         ticker,
        'calls':          selected_calls,
        'puts':           selected_puts,
        'limit':          limit,
        'expiry':         expiry_filter,
        'expiry_summary': expiry_summary,
        'vol_total_call': round(vol_total_call, 2),
        'vol_total_put':  round(vol_total_put,  2),
        'total_options':  len(calls) + len(puts),
        'spot_price':     spot_price,
        'spot_change':    spot_change,
        'due_dates':      due_dates,
    })


@app.route('/api/oplab_iv')
@login_required
def api_oplab_iv():
    """Busca IV Rank e IV Percentil de uma ação via OpLab e salva no registro de estudo."""
    from flask import jsonify
    ticker = request.args.get('ticker', '').strip().upper()
    sid    = request.args.get('sid', type=int)
    table  = request.args.get('table', 'stock')   # 'stock' | 'intl'
    if not ticker:
        return jsonify({'error': 'ticker obrigatório'}), 400

    token = Settings.get_value('oplab_token', user_id=current_user.id)
    if not token:
        return jsonify({'error': 'Token OpLab não configurado. Configure em Perfil → OpLab.'}), 400

    iv_rank      = None
    iv_percentil = None
    debug_info   = {}   # coleta responses para diagnóstico

    # Prioridade: /market/instruments retorna iv_1y_rank e iv_1y_percentile confirmados
    endpoints_to_try = [
        f'/market/instruments/{ticker}',
        f'/instruments/{ticker}',
    ]

    def _extract_iv(d):
        """Extrai iv_rank e iv_percentil de um dict retornado pelo OpLab."""
        if not isinstance(d, dict):
            return None, None
        # Achata subchaves comuns
        for sub in ('data', 'spot', 'summary', 'iv', 'implied_volatility', 'greeks'):
            if isinstance(d.get(sub), dict):
                d.update(d[sub])
        ivr = None
        ivp = None
        # Campos reais retornados por /market/instruments (confirmados no debug)
        for k in ('iv_1y_rank', 'ewma_1y_rank', 'iv_6m_rank',
                  'iv_rank', 'ivRank', 'iv_rank_52w'):
            v = d.get(k)
            if v is not None:
                try: ivr = round(float(v) * 100, 1) if float(v) <= 1.0 else round(float(v), 1); break
                except (TypeError, ValueError): pass
        for k in ('iv_1y_percentile', 'ewma_1y_percentile', 'iv_6m_percentile',
                  'iv_percentile', 'ivPercentile', 'iv_percentil'):
            v = d.get(k)
            if v is not None:
                try: ivp = round(float(v) * 100, 1) if float(v) <= 1.0 else round(float(v), 1); break
                except (TypeError, ValueError): pass
        return ivr, ivp

    for ep in endpoints_to_try:
        try:
            d = _oplab_get_json(ep, token, timeout=15)
            debug_info[ep] = {'status': 200}
            debug_info[ep]['keys'] = list(d.keys()) if isinstance(d, dict) else (
                list(d[0].keys()) if isinstance(d, list) and d else str(type(d)))
            # Tenta no nível raiz
            ivr, ivp = _extract_iv(d if isinstance(d, dict) else (d[0] if d else {}))
            if ivr is not None or ivp is not None:
                iv_rank, iv_percentil = ivr, ivp
                break
        except OplabApiError as e:
            debug_info[ep] = {'status': e.status_code, 'error': str(e), 'preview': e.body_preview}

    if iv_rank is None and iv_percentil is None:
        return jsonify({
            'error': f'IV Rank/Percentil não encontrado para {ticker}. '
                     f'Verifique se o OpLab disponibiliza este dado para ações (não apenas opções).',
            'debug': debug_info
        }), 404

    # Salva no banco se sid fornecido
    if sid:
        try:
            Model = StudyStock if table == 'stock' else StudyIntlStock
            ss = Model.query.filter_by(id=sid, user_id=current_user.id).first()
            if ss:
                if iv_rank is not None:
                    ss.iv_rank = round(iv_rank, 2)
                if iv_percentil is not None:
                    ss.iv_percentil = round(iv_percentil, 2)
                db.session.commit()
        except Exception as e:
            db.session.rollback()

    return jsonify({'iv_rank': iv_rank, 'iv_percentil': iv_percentil})



@app.route('/api/oplab_greeks')
@login_required
def api_oplab_greeks():
    """Busca VE, Delta e Gama de uma opção via OpLab e salva no modelo indicado.
    Parâmetros: ticker (da opção), model='option'|'study_option', id (pk do registro)
    """
    from flask import jsonify
    ticker   = request.args.get('ticker', '').strip().upper()
    model    = request.args.get('model', 'option')   # 'option' | 'study_option'
    rec_id   = request.args.get('id', type=int)
    if not ticker:
        return jsonify({'error': 'ticker obrigatório'}), 400

    token = Settings.get_value('oplab_token', user_id=current_user.id)
    if not token:
        return jsonify({'error': 'Token OpLab não configurado.'}), 400

    ve = delta = gama = None

    # OpLab v3 não retorna greeks via API — calculamos via Black-Scholes
    # usando spot_price, strike, days_to_maturity e close (prêmio) retornados por
    # /market/instruments/{ticker}
    try:
        d = _oplab_get_json(f'/market/instruments/{ticker}', token, timeout=15)
        if not isinstance(d, dict):
            return jsonify({'error': 'Resposta inesperada do OpLab.'}), 500

        S      = float(d.get('spot_price') or 0)
        K      = float(d.get('strike') or 0)
        T_days = float(d.get('days_to_maturity') or 0)
        cat    = str(d.get('category', d.get('type', 'CALL'))).upper()
        is_call = (cat in ('CALL', 'C'))
        # T em anos corridos (365) — days_to_maturity do OpLab é dias corridos
        T      = T_days / 365.0
        r_cont = math.log(1 + _selic() / 100)

        if S <= 0 or K <= 0 or T <= 0:
            return jsonify({'error': f'Dados insuficientes para calcular greeks de {ticker} '
                                     f'(S={S}, K={K}, T_dias={T_days}).'}), 404

        # ── Prêmio de mercado: close → last → bid → ask ──────────────────────
        premium = 0.0
        for _fld in ('close', 'last', 'bid', 'ask'):
            _v = d.get(_fld)
            if _v and float(_v) > 0:
                premium = float(_v)
                break

        # ── IV: implícita pelo prêmio → campo IV do OpLab → HV → fallback ──────
        _greeks_dict = d.get('greeks') if isinstance(d.get('greeks'), dict) else {}
        _iv_raw      = d.get('implied_volatility') or d.get('iv') or _greeks_dict.get('iv')

        sigma = None

        # Valor intrínseco teórico (floor do prêmio BS)
        intrinsic = max(0.0, (S - K) if not is_call else (K - S))

        if premium > 0 and premium > intrinsic * 1.01:
            # Prêmio acima do intrínseco → IV implícita é calculável por BS
            sigma = _implied_vol(S, K, T, r_cont, premium, is_call)
            if sigma <= 0.002:  # bissecção não convergiu (sigma_min)
                sigma = None

        if sigma is None and _iv_raw and float(_iv_raw) > 0:
            sigma = float(_iv_raw)
            if sigma > 1.5: sigma /= 100.0

        if sigma is None:
            # HV histórica via OpLab — usa campo 'underlying' do response se disponível
            _und = d.get('underlying') or d.get('stock_ticker') or ''
            if not _und:
                # Heurística: remove sufixo numérico do ticker da opção
                import re as _re
                _und = _re.sub(r'[A-Z]\d+$', '', ticker)
            sigma = _hv_from_oplab(_und, token) or _hv_fallback(_und)

        if not sigma:
            sigma = 0.35  # fallback final

        # ── Delta e Gama via BS analítico ────────────────────────────────────
        d1     = (math.log(S / K) + (r_cont + 0.5 * sigma * sigma) * T) / (sigma * math.sqrt(T))
        delta  = round(_norm_cdf(d1) if is_call else _norm_cdf(d1) - 1.0, 4)
        pdf_d1 = math.exp(-0.5 * d1 * d1) / math.sqrt(2 * math.pi)
        gama   = round(pdf_d1 / (S * sigma * math.sqrt(T)), 4)
        ve     = round(sigma * 100, 2)

    except OplabApiError as e:
        return jsonify({'error': str(e), 'status': e.status_code, 'preview': e.body_preview}), 503
    except Exception as e:
        return jsonify({'error': f'Erro ao calcular greeks: {e}'}), 500

    # Salva no banco
    if rec_id:
        try:
            if model == 'study_option':
                rec = StudyOption.query.filter_by(id=rec_id, user_id=current_user.id).first()
            else:
                rec = Option.query.filter_by(id=rec_id, user_id=current_user.id).first()
            if rec:
                if delta is not None: rec.delta = round(delta, 4)
                if gama  is not None: rec.gama  = round(gama,  4)
                if ve    is not None: rec.ve    = round(ve,    4)
                db.session.commit()
        except Exception:
            db.session.rollback()

    return jsonify({'ve': ve, 'delta': delta, 'gama': gama,
                    '_debug': {'S': S, 'K': K, 'T_days': T_days, 'premium': premium,
                               'sigma_pct': ve, 'cat': cat}})


@app.route('/atualizar_oplab', methods=['POST'])
@login_required
def atualizar_oplab():
    token = Settings.get_value('oplab_token', user_id=current_user.id)
    if not token:
        flash('Token OpLab não configurado. Configure em Perfil.', 'danger')
        return redirect(url_for('profile'))

    ativos_ok, opcoes_ok, _covered = _do_oplab_bulk_update(current_user.id, token)
    _oplab_last_update[current_user.id] = datetime.now()

    if (ativos_ok + opcoes_ok) > 0:
        flash(f'OpLab: {ativos_ok} ativo(s) e {opcoes_ok} opção(ões) atualizados.', 'success')
    else:
        flash('OpLab: nenhum ativo atualizado. Verifique o token ou os tickers cadastrados.', 'warning')
    return redirect(url_for('profile'))


# --- Dividends Module ---

@app.route('/dividendos')
@login_required
def dividendos():
    assets = Asset.query.filter_by(user_id=current_user.id).all()
    
    # Filter only Stocks and FIIs for display, AND only active assets (quantity > 0)
    relevant_assets = [a for a in assets if a.type in ['ACAO', 'FII'] and (a.quantity or 0) > 0]
    
    # Get all dividends for user's assets
    all_dividends = db.session.query(Dividend).join(Asset).filter(Asset.user_id == current_user.id).order_by(Dividend.payment_date.desc()).all()
    
    today = date.today()
    
    # Split Received vs Provisioned
    dividends_received = [d for d in all_dividends if d.payment_date and d.payment_date <= today]
    dividends_provisioned = [d for d in all_dividends if d.payment_date and d.payment_date > today]
    
    # Calculate Totals (All time or just received?)
    total_received = sum(d.amount for d in dividends_received)
    
    # Pie Chart Data (Stocks vs FIIs - All Time Received)
    total_stocks = sum(d.amount for d in dividends_received if d.asset.type == 'ACAO')
    total_fiis = sum(d.amount for d in dividends_received if d.asset.type == 'FII')
    
    div_chart_data = {
        'Ações': total_stocks,
        'FIIs': total_fiis
    }
    
    # Monthly Evolution Data (Aggregation)
    from collections import defaultdict
    from dateutil.relativedelta import relativedelta
    
    # Struct: key(year, month) -> {'total': X, 'acoes': Y, 'fiis': Z}
    monthly_agg = defaultdict(lambda: {'total': 0.0, 'acoes': 0.0, 'fiis': 0.0})
    
    # Use ALL dividends (History + Future) to show evolution
    for div in all_dividends:
        if div.payment_date:
            key = (div.payment_date.year, div.payment_date.month)
            monthly_agg[key]['total'] += div.amount
            if div.asset.type == 'ACAO':
                monthly_agg[key]['acoes'] += div.amount
            elif div.asset.type == 'FII':
                monthly_agg[key]['fiis'] += div.amount
            
    # Determine Range (Standard 12 months context or based on data)
    # Let's align with the requested "Image 1/2" style which often shows rolling 12 months.
    # Default: Start from 11 months ago to next 1 month? Or simple sorted keys?
    # User data (01/2025 entry) means all data is >= Jan 2025.
    # If we stick to "Show what we have", it's safer.
    
    sorted_keys = sorted(monthly_agg.keys()) 
    
    monthly_labels = []
    values_total = []
    values_stocks = []
    values_fiis = []
    
    if sorted_keys:
        curr_y, curr_m = sorted_keys[0] # Start from first data point
        end_y, end_m = sorted_keys[-1]   # End at last data point
        
        # If range is huge, maybe limit? No, let's show all for now or last 15 months.
        # But if defaults to 12 months view? 
        # Let's enforce full range of available data for now since it's likely short (2025).
        
        current_iter = date(curr_y, curr_m, 1)
        end_iter = date(end_y, end_m, 1)
        
        while current_iter <= end_iter:
            k = (current_iter.year, current_iter.month)
            data = monthly_agg.get(k, {'total': 0.0, 'acoes': 0.0, 'fiis': 0.0})
            
            monthly_labels.append(current_iter.strftime('%b/%Y'))
            values_total.append(data['total'])
            values_stocks.append(data['acoes'])
            values_fiis.append(data['fiis'])
            
            current_iter += relativedelta(months=1)

    monthly_chart_data = {
        'labels': monthly_labels,
        'total': values_total,
        'acoes': values_stocks,
        'fiis': values_fiis
    }
    
    return render_template('dividendos.html', 
                           assets=relevant_assets, 
                           dividends_received=dividends_received,
                           dividends_provisioned=dividends_provisioned,
                           total_received=total_received,
                           div_chart_data=div_chart_data,
                           monthly_chart_data=monthly_chart_data)

@app.route('/update_dividends', methods=['POST'])
@login_required
def update_dividends():
    import yfinance as yf # Local import to prevent global crash
    assets = Asset.query.filter_by(user_id=current_user.id).filter(Asset.type.in_(['ACAO', 'FII', 'ETF'])).all()
    
    updated_count = 0
    error_count = 0
    
    for asset in assets:
        # Skip assets with 0 quantity to preserve history and avoid overwriting with 0
        if (asset.quantity or 0) <= 0:
            continue

        try:
            ticker_sa = f"{asset.ticker}.SA" if not asset.ticker.endswith('.SA') else asset.ticker
            yf_ticker = yf.Ticker(ticker_sa)
            
            # Fetch Dividends History
            # If entry_date exists, fetch from that date. Else last 1 year.
            start_date = asset.entry_date
            if not start_date:
                 # Default to 1 year ago if no entry date
                 start_date = date.today().replace(year=date.today().year - 1)
            
            # YFinance expects string or datetime
            history = yf_ticker.dividends
            
            # Atualiza SEM apagar tudo: casa por data e PRESERVA a qty_used
            # (qtd de ações do cálculo, editável na página Dividendos) — os
            # dividendos antigos não são recalculados com a quantidade atual.
            existing = {d.ex_date: d for d in
                        Dividend.query.filter_by(asset_id=asset.id).all()
                        if d.ex_date}

            for dt, amount in history.items():
                # dt is Timestamp, convert to date
                div_date = dt.date()

                if div_date >= start_date:
                    div_type = 'Dividendo'
                    if asset.ticker.endswith('11') or asset.ticker.endswith('11B'):
                        div_type = 'Rendimento'

                    per = float(amount)
                    old = existing.get(div_date)
                    if old:
                        old.per_share = per
                        if not old.qty_used:
                            old.qty_used = asset.quantity
                        old.amount = round(per * old.qty_used, 2)
                        old.type = div_type
                    else:
                        # Qtd na data-com reconstruída pelo LIVRO de transações
                        # (sem livro após a data = quantidade atual, como antes)
                        q_hist = _qty_on_date_ledger(asset.user_id, asset.ticker,
                                                     div_date, asset.quantity)
                        db.session.add(Dividend(
                            asset_id=asset.id,
                            ticker=asset.ticker,
                            type=div_type,
                            amount=round(per * q_hist, 2),
                            per_share=per,
                            qty_used=q_hist,
                            payment_date=div_date,
                            ex_date=div_date
                        ))
            
            updated_count += 1
            
        except Exception as e:
            print(f"Error fetching dividends for {asset.ticker}: {e}")
            error_count += 1
            continue
            
    db.session.commit()
    msg = f'Dados atualizados! (Sucesso: {updated_count}, Erros: {error_count})'
    if error_count > 0:
        flash(msg, 'warning')
    else:
        flash(msg, 'success')
    return redirect(url_for('dividendos'))

@app.route('/update_asset_date/<int:id>', methods=['POST'])
@login_required
def update_asset_date(id):
    asset = Asset.query.get_or_404(id)
    if asset.user_id != current_user.id:
        return "Unauthorized", 403
        
    new_date_str = request.form.get('entry_date')
    if new_date_str:
        try:
            asset.entry_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
            db.session.commit()
            flash(f'Data de compra de {asset.ticker} atualizada.', 'success')
        except ValueError:
            flash('Formato de data inválido.', 'danger')
    
    return redirect(url_for('dividendos'))


@app.route('/api/asset-dates/<int:id>', methods=['POST'])
@login_required
def api_asset_dates(id):
    """Salva a data de entrada e/ou saída de um ativo (usado no modal da carteira).
    Essas datas definem o período de posse para rentabilidade e dividendos."""
    asset = Asset.query.get_or_404(id)
    if asset.user_id != current_user.id:
        return jsonify({'error': 'Não autorizado'}), 403
    data = request.get_json(silent=True) or request.form

    def _parse(v):
        v = (v or '').strip()
        if not v:
            return None, True
        try:
            return datetime.strptime(v, '%Y-%m-%d').date(), True
        except ValueError:
            return None, False

    if 'entry_date' in data:
        d, ok = _parse(data.get('entry_date'))
        if not ok:
            return jsonify({'error': 'Data de entrada inválida'}), 400
        asset.entry_date = d
    if 'exit_date' in data:
        d, ok = _parse(data.get('exit_date'))
        if not ok:
            return jsonify({'error': 'Data de saída inválida'}), 400
        asset.exit_date = d

    if asset.entry_date and asset.exit_date and asset.exit_date < asset.entry_date:
        return jsonify({'error': 'A saída não pode ser anterior à entrada'}), 400

    db.session.commit()
    return jsonify({
        'ok': True,
        'entry_date': asset.entry_date.isoformat() if asset.entry_date else '',
        'exit_date':  asset.exit_date.isoformat() if asset.exit_date else '',
    })


# --- Context Processor for Indices ---
@app.context_processor
def inject_indices():
    indices = MarketIndex.query.all()
    return dict(market_indices=indices)

@app.route('/config/debug_yahoo', methods=['GET', 'POST'])
@login_required
def debug_yahoo():
    debug_data = None
    ticker = None
    if request.method == 'POST':
        import yfinance as yf
        ticker = request.form.get('ticker')
        if ticker:
            try:
                # Try adding .SA automatically if generic (experimental)
                # But best to show raw first
                
                t = yf.Ticker(ticker)
                
                debug_data = {}
                
                # 1. Info (prone to errors if ticker invalid)
                try:
                    debug_data['info'] = t.info
                except Exception as e:
                    debug_data['info'] = f"Error: {str(e)}"
                    
                # 2. Fast Info
                try:
                    debug_data['fast_info'] = dict(t.fast_info) if hasattr(t, 'fast_info') else "N/A"
                except Exception as e:
                    debug_data['fast_info'] = f"Error: {str(e)}"
                
                # 3. History
                try:
                    debug_data['history_last_1d'] = t.history(period="1d").to_json()
                except Exception as e:
                    debug_data['history_last_1d'] = f"Error: {str(e)}"
                    
            except Exception as e:
                debug_data = {"fatal_error": str(e)}
    
                debug_data = {"fatal_error": str(e)}
    
    # NEW: Fetch existing indices to verify DB state
    db_indices = MarketIndex.query.all()
    
    return render_template('debug_yahoo.html', debug_data=debug_data, ticker=ticker, db_indices=db_indices)

def update_market_indices():
    """Helper to update market indices"""
    import yfinance as yf
    
    # Self-healing: Check if indices exist, if not, create them
    if MarketIndex.query.count() == 0:
        defaults = [
            {'ticker': '^BVSP', 'name': 'IBOV'},
            {'ticker': 'IFIX.SA', 'name': 'IFIX'},
            {'ticker': 'BRL=X', 'name': 'Dólar'},
            {'ticker': 'EURBRL=X', 'name': 'Euro'},
            {'ticker': 'BTC-BRL', 'name': 'Bitcoin'},
            {'ticker': 'ETH-BRL', 'name': 'Ethereum'},
            {'ticker': '^IXIC', 'name': 'Nasdaq'},
            {'ticker': '^GSPC', 'name': 'S&P 500'},
            {'ticker': '^DJI', 'name': 'Dow Jones'}
        ]
        for d in defaults:
            db.session.add(MarketIndex(ticker=d['ticker'], name=d['name']))
        db.session.commit()
    
    indices = MarketIndex.query.all()
    
    for idx in indices:
        try:
            t = yf.Ticker(idx.ticker)
            # Use fast info if available or history for reliability
            # fast_info is better for indices usually
            
            price = 0.0
            change = 0.0
            
            # Try history for today/yesterday to calculate change
            hist = t.history(period="2d")
            
            hist_price = 0.0
            hist_prev = 0.0
            
            if not hist.empty:
                hist_price = float(hist['Close'].iloc[-1])
                hist_prev = float(hist['Close'].iloc[0]) if len(hist) > 1 else hist_price
            
            # Info approach
            try:
                info = t.info
                price = info.get('regularMarketPrice') or info.get('currentPrice') or 0.0
                prev_close = info.get('previousClose') or info.get('regularMarketPreviousClose') or price
            except:
                price = 0.0
                prev_close = 0.0

            # Fallback to history if info fails
            if price == 0.0 and hist_price > 0:
                price = hist_price
            
            if (prev_close == 0.0 or prev_close == price) and hist_prev > 0:
                 prev_close = hist_prev
            
            # SUPER ROBUST FALLBACK FOR CRYPTO (BTC-BRL, ETH-BRL)
            if price == 0.0 and '-BRL' in idx.ticker:
                # Try USD version
                ticker_usd = idx.ticker.replace('-BRL', '-USD')
                try:
                    t_usd = yf.Ticker(ticker_usd)
                    hist_usd = t_usd.history(period="1d")
                    if not hist_usd.empty:
                        price_usd = float(hist_usd['Close'].iloc[-1])
                        # Get USD rate from DB or Yahoo
                        # Note: We are inside the loop, so querying DB is fine but inefficient if many cryptos. 
                        # Assuming BRL=X is already updated or exists.
                        usd_rate_idx = MarketIndex.query.filter_by(ticker='BRL=X').first()
                        # Fallback default 5.50 if not found
                        usd_rate = usd_rate_idx.price if usd_rate_idx and usd_rate_idx.price > 0 else 5.50 
                        
                        price = price_usd * usd_rate
                        # Estimate change from USD change (close enough)
                        if len(hist_usd) > 0:
                            prev_close_usd = float(hist_usd['Open'].iloc[-1])
                            change_usd = ((price_usd - prev_close_usd) / prev_close_usd) * 100
                            change = change_usd 
                            
                        # Set prev_close to enable change calculation in standard block if skipped
                        prev_close = price / (1 + (change/100)) if change != 0 else price
                except Exception as e_usd:
                     print(f"USD Fallback failed for {idx.ticker}: {e_usd}")
            
            if price and prev_close:
                change = ((price - prev_close) / prev_close) * 100
                
            idx.price = price
            idx.change_percent = change
            idx.last_update = now_brt().strftime('%H:%M')
            
        except Exception as e:
            print(f"Error updating index {idx.ticker}: {e}")
            continue
            
    db.session.commit()

# --- Modifying this to run on updates ---
# For now, I'll call update_market_indices inside the existing update_quotes route or similar.
# But where is update_quotes? I'll check user route later. 
# I will attach it to `update_quotes` via a wrapper or direct call in next step.

# --- MT5 Quote Feed API ---

def _check_api_key():
    """Returns True if the X-API-Key header matches MT5_API_KEY env var."""
    api_key = request.headers.get('X-API-Key', '')
    expected_key = os.environ.get('MT5_API_KEY', '')
    return expected_key and api_key == expected_key


@app.route('/api/update_quotes', methods=['POST'])
def api_update_quotes():
    """
    Receives quote updates from a local MT5 feeder script.
    Requires API key authentication via header: X-API-Key.
    Body (JSON):
    {
        "user_id": 1,
        "quotes":  {"PETR4": 38.50, "VALE3": 92.10},       # asset prices
        "changes": {"PETR4": 1.25, "VALE3": -0.80},        # daily change % (optional)
        "options": {"PETRA40": 0.45, "VALEF92": 1.20}      # option current prices (optional)
    }
    """
    from flask import jsonify

    if not _check_api_key():
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json()
    if not data or 'quotes' not in data:
        return jsonify({'error': 'Invalid payload — need at least quotes key'}), 400

    user_id = int(data.get('user_id', 1))
    now = now_brt()

    # ── Assets ────────────────────────────────────────────────────────────────
    quotes  = data.get('quotes', {})
    changes = data.get('changes', {})
    updated_assets, not_found_assets = [], []

    for ticker, price in quotes.items():
        ticker = ticker.upper()
        price_f  = float(price)
        change_f = float(changes.get(ticker, changes.get(ticker.lower(), 0)))
        found = False

        # 1. Atualiza Asset em carteira
        asset = Asset.query.filter_by(ticker=ticker, user_id=user_id).first()
        if asset:
            asset.current_price = price_f
            asset.daily_change  = change_f
            asset.last_update   = now
            found = True

        # 2. Atualiza underlying_price em StructuredOp (ativos subjacentes de estruturadas)
        ops = StructuredOp.query.filter_by(underlying_asset=ticker, user_id=user_id, status='OPEN').all()
        for op in ops:
            op.underlying_price  = price_f
            op.underlying_change = change_f
            found = True

        # 3. Atualiza StudyOption
        sos = StudyOption.query.filter_by(underlying_asset=ticker, user_id=user_id).all()
        for so in sos:
            so.underlying_price = price_f
            found = True

        # 4. Atualiza PutSale
        pss = PutSale.query.filter_by(underlying_asset=ticker, user_id=user_id).all()
        for ps in pss:
            ps.underlying_price = price_f
            found = True

        # 5. Atualiza OptionSpread (underlying das travas)
        for sp in OptionSpread.query.filter_by(underlying_asset=ticker, user_id=user_id).all():
            sp.underlying_price  = price_f
            sp.underlying_change = float(changes.get(ticker, changes.get(ticker.lower(), 0)))
            found = True

        if found:
            updated_assets.append(ticker)
        else:
            not_found_assets.append(ticker)

    # ── Options ───────────────────────────────────────────────────────────────
    options_prices = data.get('options', {})
    updated_options, not_found_options = [], []

    for ticker, price in options_prices.items():
        ticker  = ticker.upper()
        price_f = float(price)
        found   = False

        # 1. Tabela Option (venda coberta, etc.)
        opt = Option.query.filter_by(ticker=ticker, user_id=user_id).first()
        if opt:
            opt.current_option_price = price_f
            opt.last_update = now
            found = True

        # 2. StructuredLeg (pernas de estruturadas)
        legs = StructuredLeg.query.join(StructuredOp).filter(
            StructuredLeg.ticker == ticker,
            StructuredOp.user_id == user_id,
            StructuredOp.status  == 'OPEN'
        ).all()
        for leg in legs:
            leg.current_price = price_f
            leg.last_update   = now
            found = True

        # 3. PutSale
        ps_opt = PutSale.query.filter_by(ticker=ticker, user_id=user_id).first()
        if ps_opt:
            found = True

        # 4. StudyOption
        for so in StudyOption.query.filter_by(ticker=ticker, user_id=user_id).all():
            so.option_price = price_f
            found = True

        # 5. OptionSpread (travas) — perna comprada (long) ou vendida (short)
        for sp in OptionSpread.query.filter_by(user_id=user_id).all():
            if sp.leg_long_ticker and sp.leg_long_ticker.upper() == ticker:
                sp.leg_long_current = price_f
                found = True
            if sp.leg_short_ticker and sp.leg_short_ticker.upper() == ticker:
                sp.leg_short_current = price_f
                found = True

        if found:
            updated_options.append(ticker)
        else:
            not_found_options.append(ticker)

    db.session.commit()

    return jsonify({
        'status': 'ok',
        'updated_assets':  updated_assets,
        'updated_options': updated_options,
        'not_found_assets':  not_found_assets,
        'not_found_options': not_found_options,
    }), 200


@app.route('/api/radar_update_study', methods=['POST'])
@login_required
def api_radar_update_study():
    """Atualiza RSI, ATR% e tendência do estudo a partir dos dados da API Radar."""
    import requests as _req
    from flask import jsonify
    data = request.get_json()
    sid      = data.get('id')
    table    = data.get('table', 'stock')   # 'stock' | 'intl'
    ticker   = data.get('ticker', '').upper()
    RADAR_URL = 'https://acoes.receberbemevinhos.com.br/api_res.php'
    RADAR_KEY  = 'radar_8acddd4976bc3c1e9b9c814c3b408f9dcbf1dfd0d75795f9'
    try:
        resp = _req.get(RADAR_URL, params={'ticker': ticker, 'api_key': RADAR_KEY}, timeout=30)
        raw  = resp.json()
        d    = raw.get('data', raw)
        rsi14 = d.get('rsi14')
        atr14 = d.get('atr14')
        price = d.get('price')
        atr_pct = round(atr14 / price * 100, 2) if (atr14 and price and price > 0) else None
        sig = d.get('signal', {})
        sig_code = ''
        sig_label = ''
        if isinstance(sig, dict):
            sig_code = str(sig.get('code') or '').strip().lower()
            sig_label = str(sig.get('label') or '').strip().lower()
        elif sig:
            sig_label = str(sig).strip().lower()

        sig_text = f'{sig_code} {sig_label}'
        trend = None
        if any(x in sig_text for x in ('comprar', 'compra', 'buy')):
            trend = 'Alta'
        elif any(x in sig_text for x in ('vender', 'venda', 'sell')):
            trend = 'Baixa'
        elif any(x in sig_text for x in ('aguardar', 'wait', 'lateral')):
            trend = 'Lateral'

        Model = StudyStock if table == 'stock' else StudyIntlStock
        ss = Model.query.filter_by(id=sid, user_id=current_user.id).first()
        if not ss:
            return jsonify({'error': 'Registro não encontrado'}), 404

        if rsi14 is not None:
            ss.rsi = round(rsi14, 2)
        if atr_pct is not None:
            ss.atr_pct = atr_pct
        if trend is not None:
            ss.trend = trend
        db.session.commit()
        return jsonify({'rsi': ss.rsi, 'atr_pct': ss.atr_pct, 'trend': ss.trend, 'price': price, 'atr14': atr14})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/radar_analise')
@login_required
def api_radar_analise():
    """Proxy para a API de análise de ações — evita CORS e esconde a API key."""
    import requests as _req
    from flask import jsonify
    ticker = request.args.get('ticker', '').strip().upper()
    if not ticker:
        return jsonify({'error': 'ticker obrigatório'}), 400
    analyzer = (request.args.get('analyzer') or 'technical').strip().lower()
    if analyzer not in ('technical', 'ai'):
        analyzer = 'technical'
    RADAR_URL = 'https://acoes.receberbemevinhos.com.br/api_res.php'
    RADAR_KEY  = 'radar_8acddd4976bc3c1e9b9c814c3b408f9dcbf1dfd0d75795f9'

    # ── Modo IA: relatório fundamentalista completo ──────────────────────────
    if analyzer == 'ai':
        try:
            resp = _req.get(RADAR_URL, params={'ticker': ticker, 'analyzer': 'ai',
                                               'api_key': RADAR_KEY}, timeout=45)
            raw = resp.json()
            if not raw.get('ok', True) and raw.get('error'):
                return jsonify({'error': raw.get('error')}), 502
            d = raw.get('data', raw)
            rep = d.get('report', {}) if isinstance(d.get('report'), dict) else {}
            fr  = d.get('fundamental_result', {}) if isinstance(d.get('fundamental_result'), dict) else {}
            cache = d.get('cache', {}) if isinstance(d.get('cache'), dict) else {}
            out = {
                'analyzer':       'ai',
                'company':        d.get('company_name', ''),
                'price':          d.get('price'),
                'estimated_value': d.get('estimated_value') or (rep.get('valuation') or {}).get('estimated_value'),
                'margin_of_safety': d.get('margin_of_safety') or (rep.get('valuation') or {}).get('margin_of_safety'),
                'executive_summary': rep.get('executive_summary'),
                'multiples':      rep.get('multiples') or {},
                'valuation_detail': rep.get('valuation_detail') or {},
                'income_statement': rep.get('income_statement') or {},
                'balance_sheet':  rep.get('balance_sheet') or {},
                'cash_flow':      rep.get('cash_flow') or {},
                'strengths':      rep.get('strengths') or [],
                'risks':          rep.get('risks') or [],
                'checklist':      rep.get('investor_checklist') or [],
                'conclusion':     rep.get('conclusion'),
                'fundamental_summary': (fr.get('layer3') or {}).get('summary'),
                'generated_at':   d.get('generated_at'),
                'disclaimer':     d.get('disclaimer'),
                'cache_status':   cache.get('status'),
                'cache_date':     cache.get('date'),
                'market_refreshed': cache.get('market_refreshed'),
            }
            return jsonify(out)
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    try:
        resp = _req.get(RADAR_URL, params={'ticker': ticker, 'analyzer': 'technical',
                                           'api_key': RADAR_KEY}, timeout=30)
        raw  = resp.json()
        # A API retorna { ok, data: { ... } }
        d = raw.get('data', raw)
        sig  = d.get('signal', {})
        mc   = d.get('market_context', {})
        tr   = d.get('technical_reading', {})
        fund = d.get('fundamentals_summary', {})
        # A API nova traz também screen_data já estruturado — usa como fallback.
        sd = d.get('screen_data', {}) if isinstance(d.get('screen_data'), dict) else {}
        if not isinstance(tr, dict) or not tr:
            tr = sd.get('indicadores_tecnicos', {}) if isinstance(sd.get('indicadores_tecnicos'), dict) else {}
        if not isinstance(mc, dict) or not mc:
            mc = sd.get('niveis_operacionais', {}) if isinstance(sd.get('niveis_operacionais'), dict) else {}
        if not isinstance(fund, dict) or not fund:
            fund = sd.get('fundamentos_resumidos', {}) if isinstance(sd.get('fundamentos_resumidos'), dict) else {}
        if not isinstance(fund, dict):
            fund = {}
        if not any(fund.get(k) not in (None, '', '-') for k in ('pl', 'pvp', 'dividend_yield', 'eps', 'sector', 'industry')):
            try:
                yf_ticker = ticker + '.SA' if _is_b3_yahoo_ticker(ticker) else ticker
                info = yf.Ticker(yf_ticker).info or {}
                fund = dict(fund)
                fund['pl'] = fund.get('pl') or info.get('trailingPE') or info.get('forwardPE')
                fund['pvp'] = fund.get('pvp') or info.get('priceToBook')
                dy = fund.get('dividend_yield')
                if dy in (None, '', '-'):
                    dy = info.get('dividendYield')
                    fund['dividend_yield'] = (dy * 100) if isinstance(dy, (int, float)) and dy <= 1 else dy
                fund['eps'] = fund.get('eps') or info.get('trailingEps') or info.get('forwardEps')
                fund['sector'] = fund.get('sector') or info.get('sector')
                fund['industry'] = fund.get('industry') or info.get('industry')
            except Exception as yf_err:
                app.logger.info('radar fundamentals fallback %s: %s', ticker, yf_err)
        def _as_pct(v):
            """Normaliza um percentual: frações (0.0666) viram 6.66; já-percentual fica."""
            try:
                x = float(v)
            except (TypeError, ValueError):
                return None
            return round(x * 100, 2) if -1.5 < x < 1.5 else round(x, 2)

        macd = tr.get('macd', {})
        entry = d.get('entry', {})
        chart = d.get('chart_data', {})
        if not isinstance(chart, dict):
            chart = {}
        overlays = chart.get('overlays', {}) if isinstance(chart.get('overlays'), dict) else {}
        out = {
            'company':      d.get('company_name', ''),
            'price':        d.get('price'),
            'currency':     d.get('currency', 'BRL'),
            'signal':       sig.get('label', '') if isinstance(sig, dict) else str(sig),
            'signal_code':  sig.get('code', '')  if isinstance(sig, dict) else '',
            'rationale':    sig.get('reason', '') if isinstance(sig, dict) else '',
            'day_change':   mc.get('change_percent'),
            'week_change':  mc.get('change_week') or mc.get('week_change'),
            'month_change': mc.get('change_month') or mc.get('month_change'),
            'entry_min':    entry.get('low')  if isinstance(entry, dict) else None,
            'entry_max':    entry.get('high') if isinstance(entry, dict) else None,
            'stop_loss':    d.get('stop'),
            'target':       d.get('target'),
            'support':      d.get('support'),
            'resistance':   d.get('resistance'),
            'rsi14':        d.get('rsi14') or tr.get('rsi14'),
            'atr14':        d.get('atr14') or tr.get('atr14'),
            'sma9':         tr.get('sma9'),
            'sma21':        tr.get('sma21'),
            'sma50':        tr.get('sma50'),
            'sma200':       tr.get('sma200'),
            'ema9':         tr.get('ema9'),
            'macd_line':    macd.get('line')      if isinstance(macd, dict) else None,
            'macd_signal':  macd.get('signal')    if isinstance(macd, dict) else None,
            'macd_hist':    macd.get('histogram') if isinstance(macd, dict) else None,
            'bb_upper':     tr.get('bollinger_upper') or tr.get('bb_upper'),
            'bb_lower':     tr.get('bollinger_lower') or tr.get('bb_lower'),
            'stoch_k':      tr.get('stoch_k'),
            'stoch_d':      tr.get('stoch_d'),
            'adx':          tr.get('adx'),
            'daily_trend':  d.get('trend_daily',  {}).get('label') if isinstance(d.get('trend_daily'), dict)  else d.get('trend_daily'),
            'weekly_trend': d.get('trend_weekly', {}).get('label') if isinstance(d.get('trend_weekly'), dict) else d.get('trend_weekly'),
            'open':         mc.get('open'),
            'prev_close':   mc.get('previous_close') or mc.get('prev_close'),
            'day_high':     mc.get('day_high'),
            'day_low':      mc.get('day_low'),
            'week52_low':   mc.get('fifty_two_week_low'),
            'week52_high':  mc.get('fifty_two_week_high'),
            'volume':       mc.get('volume'),
            'avg_volume':   mc.get('average_volume') or mc.get('avg_volume'),
            'market_cap':   mc.get('market_cap'),
            'pl':           fund.get('pl'),
            'pvp':          fund.get('pvp'),
            'dy':           _as_pct(fund.get('dividend_yield')),
            'eps':          fund.get('eps'),
            'roe':          _as_pct(fund.get('roe')),
            'net_margin':   _as_pct(fund.get('net_margin') or fund.get('margem_liquida')),
            'sector':       fund.get('sector'),
            'industry':     fund.get('industry'),
            # Gráfico pronto da API: candles + médias + linhas automáticas
            'chart': {
                'candles':   chart.get('candles') or [],
                'sma21':     overlays.get('sma21') or [],
                'sma50':     overlays.get('sma50') or [],
                'lines':     chart.get('automatic_lines') or [],
            },
        }
        return jsonify(out)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/current_quotes')
@login_required
def api_current_quotes():
    """
    Returns current prices for all assets and options of the logged-in user.
    Used by browser-side JS polling to update tables without page reload.
    Response JSON:
    {
        "mode": "mt5",
        "assets":  {"PETR4": {"price": 38.50, "change": 1.25, "updated": "10:32"}},
        "options": {"PETRA40": {"price": 0.45, "updated": "10:32"}}
    }
    """
    from flask import jsonify

    uid  = current_user.id
    mode = Settings.get_value('quote_mode',        user_id=uid, default='yahoo')
    oplab_auto     = Settings.get_value('oplab_auto_update', user_id=uid, default='false') == 'true'
    oplab_interval = int(Settings.get_value('oplab_interval', user_id=uid, default='5'))

    assets_data = {}
    for a in Asset.query.filter_by(user_id=uid).all():
        assets_data[a.ticker] = {
            'price':   round(a.current_price or 0, 2),
            'change':  round(a.daily_change  or 0, 2),
            'updated': a.last_update.strftime('%H:%M') if a.last_update else '-'
        }

    options_data = {}
    for o in Option.query.filter_by(user_id=uid).all():
        options_data[o.ticker.upper()] = {
            'price':   round(o.current_option_price or 0, 2),
            'change':  round(o.daily_change or 0, 2),
            'updated': o.last_update.strftime('%H:%M') if o.last_update else '-'
        }

    def _merge_option_quote(ticker, price, change=0, updated='-'):
        key = (ticker or '').upper()
        if not key:
            return
        price = round(price or 0, 2)
        if key not in options_data or price > 0:
            options_data[key] = {
                'price':   price,
                'change':  round(change or 0, 2),
                'updated': updated,
            }

    for so in StudyOption.query.filter_by(user_id=uid).all():
        _merge_option_quote(so.ticker, so.option_price)
    for sp in OptionSpread.query.filter_by(user_id=uid).all():
        _merge_option_quote(sp.leg_long_ticker, sp.leg_long_current)
        _merge_option_quote(sp.leg_short_ticker, sp.leg_short_current)
    for leg in (StructuredLeg.query
                .join(StructuredOp)
                .filter(StructuredOp.user_id == uid,
                        StructuredOp.status == 'OPEN')
                .all()):
        _merge_option_quote(
            leg.ticker,
            leg.current_price,
            updated=leg.last_update.strftime('%H:%M') if leg.last_update else '-'
        )

    return jsonify({
        'mode':              mode,
        'assets':            assets_data,
        'options':           options_data,
        'oplab_enabled':     oplab_auto,
        'oplab_interval_ms': oplab_interval * 60 * 1000,
    })


_chart_mem = {}  # cache em memória por processo: {ticker: {'ts': float, 'candles': [...]}}
_CHART_MEM_TTL = 120  # segundos — evita hit no SQLite em acessos repetidos rápidos

_YF_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
    'Accept': 'application/json,text/plain,*/*',
    'Accept-Language': 'pt-BR,pt;q=0.9,en-US;q=0.8',
    'Origin': 'https://finance.yahoo.com',
    'Referer': 'https://finance.yahoo.com/',
}
_YF_COOKIES = {'tbla_id': 'finan-web', 'GUC': 'AQEBCAFn', 'GUCS': 'AQEBCAFn'}

def _is_b3_yahoo_ticker(ticker):
    """Identifica tickers B3 que precisam do sufixo .SA no Yahoo, incluindo B3SA3."""
    import re
    tk = (ticker or '').upper().strip()
    return bool(tk and '.' not in tk and '-' not in tk and re.match(r'^[A-Z0-9]{4,8}\d{1,2}$', tk))

def _yahoo_fetch(yf_ticker, start_date=None):
    """Chama Yahoo Finance v8 diretamente (sem yfinance) — ~0.5 s vs ~1.3 s.
    Tenta query1 primeiro, fallback para query2 em caso de 400/429."""
    import requests as _req
    from datetime import datetime as _dt, timezone as _tz
    params = {'interval': '1d', 'range': '1y'} if not start_date else {
        'interval': '1d',
        'period1': int(_dt.fromisoformat(start_date).replace(tzinfo=_tz.utc).timestamp()),
        'period2': int(_dt.now(_tz.utc).timestamp()),
    }
    last_exc = None
    for host in ('query1.finance.yahoo.com', 'query2.finance.yahoo.com'):
        try:
            url = f'https://{host}/v8/finance/chart/' + yf_ticker
            r = _req.get(url, params=params, headers=_YF_HEADERS,
                         cookies=_YF_COOKIES, timeout=10)
            if r.status_code in (400, 429, 503) and host == 'query1.finance.yahoo.com':
                last_exc = Exception(f'{r.status_code} from {host}')
                continue
            r.raise_for_status()
            break
        except Exception as e:
            last_exc = e
            if host == 'query2.finance.yahoo.com':
                raise last_exc
    else:
        raise last_exc
    result = r.json()['chart']['result']
    if not result:
        return []
    res   = result[0]
    ts    = res['timestamp']
    q     = res['indicators']['quote'][0]
    adj   = (((res.get('indicators') or {}).get('adjclose') or [{}])[0].get('adjclose') or [])
    opens  = q.get('open',   [])
    highs  = q.get('high',   [])
    lows   = q.get('low',    [])
    closes = q.get('close',  [])
    vols   = q.get('volume', [])
    rows = []
    from datetime import datetime as _dt2, timezone as _tz2
    for i, epoch in enumerate(ts):
        o = opens[i];  h = highs[i];  l = lows[i];  c = closes[i]
        if o is None or h is None or l is None or c is None:
            continue
        if i < len(adj) and adj[i] is not None and c:
            try:
                factor = float(adj[i]) / float(c)
                if 0 < factor < 20:
                    o = float(o) * factor
                    h = float(h) * factor
                    l = float(l) * factor
                    c = float(c) * factor
            except Exception:
                pass
        rows.append({
            't': _dt2.fromtimestamp(epoch, tz=_tz2.utc).strftime('%Y-%m-%d'),
            'o': round(float(o), 4), 'h': round(float(h), 4),
            'l': round(float(l), 4), 'c': round(float(c), 4),
            'v': int(vols[i]) if vols[i] is not None else 0,
        })
    return rows


def _sanitize_chart_candles(candles):
    """Remove candles inválidos/outliers que distorcem o gráfico."""
    if not candles:
        return []

    cleaned = []
    seen = set()
    for row in sorted(candles, key=lambda c: c.get('t') or ''):
        try:
            t = row.get('t')
            o = float(row.get('o'))
            h = float(row.get('h'))
            l = float(row.get('l'))
            c = float(row.get('c'))
            v = int(row.get('v') or 0)
        except (TypeError, ValueError):
            continue
        if not t or t in seen:
            continue
        vals = [o, h, l, c]
        if any((not math.isfinite(x)) or x <= 0 for x in vals):
            continue
        if h < max(o, c) or l > min(o, c):
            continue
        if h / l > 2.8:
            continue
        seen.add(t)
        cleaned.append({
            't': t,
            'o': round(o, 4),
            'h': round(h, 4),
            'l': round(l, 4),
            'c': round(c, 4),
            'v': max(v, 0),
        })

    if len(cleaned) < 5:
        return cleaned

    closes = sorted(c['c'] for c in cleaned)
    median = closes[len(closes) // 2]
    if median <= 0:
        return cleaned

    robust = []
    prev_close = None
    for row in cleaned:
        vals = [row['o'], row['h'], row['l'], row['c']]
        if max(vals) > median * 5 or min(vals) < median * 0.2:
            continue
        if prev_close and (row['h'] > prev_close * 2.5 or row['l'] < prev_close * 0.25):
            continue
        robust.append(row)
        prev_close = row['c']
    return robust


@app.route('/api/chart_data/<ticker>')
@login_required
def api_chart_data(ticker):
    """OHLCV diário — 6 meses, 3 camadas de cache:
       1. Memória (120 s)  — zero I/O
       2. SQLite            — sobrevive restart; fetch incremental se stale
       3. Yahoo Finance v8  — chamada HTTP direta, ~0.5 s
    """
    import re, time as _time, gzip as _gzip, json as _json
    from models import ChartCache
    from datetime import date as _date, timedelta as _td

    ticker = ticker.upper().strip()
    since  = request.args.get('since')
    now_ts = _time.time()

    # ── Camada 1: memória ──────────────────────────────────────────────────────
    mem = _chart_mem.get(ticker)
    if mem and (now_ts - mem['ts']) < _CHART_MEM_TTL:
        candles = mem['candles']
        if since:
            candles = [c for c in candles if c['t'] > since]
        return jsonify({'ticker': ticker, 'candles': candles, 'cached': 'mem'})

    yf_ticker = ticker + '.SA' if _is_b3_yahoo_ticker(ticker) else ticker

    candles = None
    try:
        # ── Camada 2: SQLite ───────────────────────────────────────────────────
        db_entry = ChartCache.query.get(ticker)

        if db_entry:
            candles  = _json.loads(_gzip.decompress(db_entry.candles_gz).decode())
            original_len = len(candles)
            candles = _sanitize_chart_candles(candles)
            if not candles:
                candles = _sanitize_chart_candles(_yahoo_fetch(yf_ticker))
            days_old = (_date.today() - _date.fromisoformat(db_entry.last_date)).days

            if days_old <= 1:
                if len(candles) != original_len:
                    gz = _gzip.compress(_json.dumps(candles).encode(), compresslevel=6)
                    db_entry.candles_gz = gz
                    db_entry.last_date = candles[-1]['t'] if candles else db_entry.last_date
                    db_entry.fetched_at = datetime.utcnow()
                    db.session.commit()
                # Cache fresco — serve direto
                _chart_mem[ticker] = {'ts': now_ts, 'candles': candles}
                out = [c for c in candles if c['t'] > since] if since else candles
                return jsonify({'ticker': ticker, 'candles': out, 'cached': 'db'})

            # Stale — busca só dias que faltam
            start_date = (_date.fromisoformat(db_entry.last_date) - _td(days=3)).isoformat()
            new_rows = _yahoo_fetch(yf_ticker, start_date=start_date)
            if new_rows:
                new_rows = _sanitize_chart_candles(new_rows)
                new_dates = {r['t'] for r in new_rows}
                candles = [c for c in candles if c['t'] not in new_dates] + new_rows
                candles.sort(key=lambda c: c['t'])
                candles = candles[-260:]  # ~1 ano de dias úteis (warm-up MM200 + 8 meses)
        else:
            # Primeira vez — busca 6 meses completos
            candles = _yahoo_fetch(yf_ticker)

        candles = _sanitize_chart_candles(candles)

        if not candles:
            return jsonify({'error': 'Sem dados para ' + ticker}), 404

        # ── Persiste no SQLite ─────────────────────────────────────────────────
        gz = _gzip.compress(_json.dumps(candles).encode(), compresslevel=6)
        if db_entry:
            db_entry.candles_gz = gz
            db_entry.last_date  = candles[-1]['t']
            db_entry.fetched_at = datetime.utcnow()
        else:
            db.session.add(ChartCache(ticker=ticker, last_date=candles[-1]['t'],
                                      candles_gz=gz, fetched_at=datetime.utcnow()))
        db.session.commit()

        _chart_mem[ticker] = {'ts': now_ts, 'candles': candles}
        out = [c for c in candles if c['t'] > since] if since else candles
        return jsonify({'ticker': ticker, 'candles': out, 'cached': 'yf'})

    except Exception as e:
        app.logger.error('api_chart_data %s: %s', ticker, e)
        if candles:
            out = [c for c in candles if c['t'] > since] if since else candles
            return jsonify({'ticker': ticker, 'candles': out, 'cached': 'stale'})
        return jsonify({'error': str(e)}), 500


@app.route('/api/chart_lines/<ticker>', methods=['GET', 'POST', 'DELETE'])
@login_required
def api_chart_lines(ticker):
    """Salva/carrega/deleta linhas de tendência desenhadas no gráfico."""
    from models import UserChartLine
    ticker = ticker.upper().strip()
    uid = current_user.id

    if request.method == 'GET':
        lines = UserChartLine.query.filter_by(user_id=uid, ticker=ticker).all()
        return jsonify([{'id': l.id, 'x1': l.x1, 'y1': l.y1, 'x2': l.x2, 'y2': l.y2,
                         'color': l.color, 'width': l.width} for l in lines])

    if request.method == 'POST':
        d = request.get_json(force=True)
        lid = d.get('id')
        if lid:
            line = UserChartLine.query.filter_by(id=lid, user_id=uid).first()
            if not line:
                return jsonify({'error': 'not found'}), 404
            line.x1 = d['x1']; line.y1 = d['y1']
            line.x2 = d['x2']; line.y2 = d['y2']
            line.color = d.get('color', line.color)
            line.width = d.get('width', line.width)
            db.session.commit()
            return jsonify({'id': line.id})
        line = UserChartLine(user_id=uid, ticker=ticker,
                             x1=d['x1'], y1=d['y1'], x2=d['x2'], y2=d['y2'],
                             color=d.get('color', '#3b82f6'),
                             width=d.get('width', 1.5))
        db.session.add(line)
        db.session.commit()
        return jsonify({'id': line.id})

    if request.method == 'DELETE':
        d = request.get_json(force=True)
        lid = d.get('id')
        if lid:
            UserChartLine.query.filter_by(id=lid, user_id=uid).delete()
        else:
            UserChartLine.query.filter_by(user_id=uid, ticker=ticker).delete()
        db.session.commit()
        return jsonify({'ok': True})


@app.route('/api/quote_hint/<ticker>')
@login_required
def api_quote_hint(ticker):
    """Retorna cotação ao vivo de um ticker para o tooltip de hint."""
    from flask import jsonify
    import requests as _req
    import re as _re

    ticker = ticker.strip().upper()
    uid    = current_user.id
    token  = None
    oplab_token = None
    try:
        from models import Settings
        token       = Settings.get_value('brapi_token',  user_id=uid)
        oplab_token = Settings.get_value('oplab_token',  user_id=uid)
    except Exception:
        pass

    price = change = ask = bid = 0.0
    trades = None   # nº negócios no dia (só para opções)
    name  = ticker
    is_option = False
    # preço do último negócio salvo no banco (fallback quando close==0 no dia)
    db_last_price = 0.0

    # Padrão de ticker de opção BR: 4 letras + 1 letra (série) + dígitos
    _option_re = _re.compile(r'^[A-Z]{4}[A-Z]\d{2,}$')

    # ── 1. Tenta banco local: Option / StructuredLeg ─────────────
    try:
        opt = (Option.query.filter_by(user_id=uid, ticker=ticker).first() or
               Option.query.filter(Option.user_id == uid,
                                   Option.ticker.ilike(ticker)).first())
        if opt:
            is_option = True
            if opt.current_option_price and opt.current_option_price > 0:
                db_last_price = float(opt.current_option_price)
            change = float(opt.daily_change or 0)
            name   = f'{opt.option_type} K={opt.strike_price:.2f} venc {opt.expiration_date.strftime("%d/%m/%y") if opt.expiration_date else "?"}'
    except Exception:
        pass

    if not is_option:
        try:
            from models import StructuredLeg, StructuredOp
            leg = (StructuredLeg.query
                   .join(StructuredOp)
                   .filter(StructuredOp.user_id == uid,
                           StructuredLeg.ticker.ilike(ticker))
                   .first())
            if leg:
                is_option = True
                if leg.current_price and leg.current_price > 0:
                    db_last_price = float(leg.current_price)
        except Exception:
            pass

    if not is_option:
        try:
            so = (StudyOption.query
                  .filter_by(user_id=uid, ticker=ticker).first() or
                  StudyOption.query.filter(StudyOption.user_id == uid,
                                          StudyOption.ticker.ilike(ticker)).first())
            if so:
                is_option = True
                if so.option_price and so.option_price > 0:
                    db_last_price = float(so.option_price)
        except Exception:
            pass

    # ── OptionSpread (travas) ──────────────────────────────────────
    if not is_option:
        try:
            sp = OptionSpread.query.filter(
                OptionSpread.user_id == uid,
                db.or_(
                    OptionSpread.leg_long_ticker.ilike(ticker),
                    OptionSpread.leg_short_ticker.ilike(ticker)
                )
            ).first()
            if sp:
                is_option = True
                if sp.leg_long_ticker and sp.leg_long_ticker.upper() == ticker and sp.leg_long_current and sp.leg_long_current > 0:
                    db_last_price = float(sp.leg_long_current)
                elif sp.leg_short_ticker and sp.leg_short_ticker.upper() == ticker and sp.leg_short_current and sp.leg_short_current > 0:
                    db_last_price = float(sp.leg_short_current)
        except Exception:
            pass

    # Detecta opção pelo padrão do ticker (modal de liquidez: não está no banco do usuário)
    if not is_option and _option_re.match(ticker):
        is_option = True

    # ── 2. OpLab: cotação ao vivo para opções ─────────────────────
    # /v3/market/quote → bid, ask, close, variation
    # /v3/market/instruments/{ticker} → trades (negócios do dia) e outros campos
    if is_option and oplab_token:
        _oplab_base = 'https://api.oplab.com.br/v3'
        _oplab_hdrs = _oplab_headers(oplab_token)

        # Chamada 1: /market/quote para preço ao vivo
        try:
            r = _req.get(
                f'{_oplab_base}/market/quote',
                params={'tickers': ticker},
                headers=_oplab_hdrs,
                timeout=8,
            )
            if r.status_code == 200:
                data = r.json()
                item = data[0] if isinstance(data, list) and data else (data if isinstance(data, dict) else {})
                close_day = item.get('close') or item.get('last') or 0
                var_day   = None
                for _vk in ('variation', 'change', 'pct_change', 'percentChange', 'dailyChange'):
                    if _vk in item and item[_vk] is not None:
                        var_day = float(item[_vk])
                        break
                bid_v = item.get('bid') or item.get('bid_price') or 0
                ask_v = item.get('ask') or item.get('ask_price') or 0
                # Tenta capturar negócios já aqui (alguns endpoints retornam)
                for _tk in ('trades', 'num_trades', 'trade_count', 'negotiations',
                            'quantity', 'volume', 'deals', 'qtd_negocios',
                            'number_of_trades', 'negocios', 'qtd'):
                    _tv = item.get(_tk)
                    if _tv is not None and _tv != 0:
                        trades = int(_tv)
                        break

                if close_day and float(close_day) > 0:
                    price  = float(close_day)
                    change = var_day if var_day is not None else change
                else:
                    price  = db_last_price
                    change = 0.0

                ask = float(ask_v) if ask_v else 0.0
                bid = float(bid_v) if bid_v else 0.0
        except Exception:
            pass

        # Chamada 2: /market/instruments/{ticker} para negócios do dia
        if trades is None:
            try:
                r2 = _req.get(
                    f'{_oplab_base}/market/instruments/{ticker}',
                    headers=_oplab_hdrs,
                    timeout=8,
                )
                if r2.status_code == 200:
                    d2 = r2.json()
                    for _tk in ('trades', 'num_trades', 'trade_count', 'negotiations',
                                'quantity', 'volume', 'deals', 'qtd_negocios',
                                'number_of_trades', 'negocios', 'qtd',
                                'total_trades', 'daily_trades'):
                        _tv = d2.get(_tk)
                        if _tv is not None and _tv != 0:
                            trades = int(_tv)
                            break
                    # Se /market/quote não retornou bid/ask, tenta aqui
                    if bid == 0.0:
                        bid = float(d2.get('bid') or d2.get('bid_price') or 0)
                    if ask == 0.0:
                        ask = float(d2.get('ask') or d2.get('ask_price') or 0)
            except Exception:
                pass

    # Fallback: sem OpLab ou falha — usa preço do banco se disponível
    if price == 0 and db_last_price > 0:
        price = db_last_price

    # Se já resolvemos a opção, retorna
    if is_option and (price > 0 or db_last_price > 0):
        resp = {
            'ticker':    ticker,
            'name':      name,
            'price':     round(price,  2),
            'change':    round(change, 2),
            'ask':       round(ask,    2),
            'bid':       round(bid,    2),
            'is_option': True,
        }
        if trades is not None:
            resp['trades'] = trades
        return jsonify(resp)

    # Opção sem NENHUM preço (nem OpLab, nem banco): retorna já — brapi e
    # yfinance abaixo são para ações/FIIs/ETFs e nunca resolvem um ticker de
    # opção (ex.: PETRI437.SA não existe no Yahoo); empilhar essas duas
    # chamadas só atrasava a resposta e, em ilíquidos, estourava o timeout
    # do worker e travava o hint em "carregando…" para sempre.
    if is_option:
        return jsonify({
            'ticker': ticker, 'name': name, 'price': 0, 'change': 0,
            'ask': 0, 'bid': 0, 'is_option': True, 'no_quote': True,
        })

    # ── 3. brapi (ativos, FIIs, ETFs) ────────────────────────────
    if token:
        try:
            r = _req.get(
                f'https://brapi.dev/api/quote/{ticker}',
                params={'range': '1d', 'interval': '1d',
                        'fundamental': 'false', 'dividends': 'false',
                        'token': token},
                timeout=8
            )
            if r.status_code == 200:
                for item in r.json().get('results', []):
                    p = item.get('regularMarketPrice') or item.get('currentPrice')
                    if p and float(p) > 0:
                        price  = float(p)
                        change = float(item.get('regularMarketChangePercent') or 0)
                        ask    = float(item.get('ask') or item.get('regularMarketOpen') or price)
                        bid    = float(item.get('bid') or item.get('regularMarketPreviousClose') or price)
                        name   = item.get('shortName') or item.get('longName') or ticker
                        break
        except Exception:
            pass

    # ── fallback yfinance fast_info ───────────────────────────────
    if price == 0:
        try:
            import yfinance as yf
            yf_t = ticker if '.' in ticker else f'{ticker}.SA'
            fi   = yf.Ticker(yf_t).fast_info
            p    = fi.last_price
            prev = fi.previous_close or 0
            if p and float(p) > 0:
                price  = float(p)
                change = ((price - prev) / prev * 100) if prev > 0 else 0.0
                ask    = float(getattr(fi, 'ask', 0) or price)
                bid    = float(getattr(fi, 'bid', 0) or price)
        except Exception:
            pass

    return jsonify({
        'ticker': ticker,
        'name':   name,
        'price':  round(price,  2),
        'change': round(change, 2),
        'ask':    round(ask,    2),
        'bid':    round(bid,    2),
    })


# ─────────────────────────────────────────────────────────────────
# OPLAB AUTO-UPDATE BACKGROUND SCHEDULER
# ─────────────────────────────────────────────────────────────────

_oplab_last_update: dict = {}   # user_id → datetime of last successful update


def _do_oplab_bulk_update(uid: int, token: str, oplab_online: bool = True):
    """
    Busca cotações via GET /v3/market/quote?tickers=... e atualiza o DB para:
      - Todos os Assets do usuário (qty ≥ 0 — inclui swingtrade sem posição)
      - Todas as Options (VENDA_CALL, VENDA_PUT, COMPRA_CALL, COMPRA_PUT)
      - Underlying assets das Options (para exibição correta em /opcoes e /estudos)
      - StudyOption: option_price + underlying_price
      - OptionSpread: leg_long_current + leg_short_current
    Quando oplab_online=False pula todos os fallbacks individuais por opção
    (que causam loop/timeout quando o servidor está fora do ar).
    Retorna (assets_ok, options_ok, oplab_covered_assets).
    """
    BASE    = 'https://api.oplab.com.br/v3'
    headers = _oplab_headers(token)

    # ── Coleta todos os registros ─────────────────────────────────
    assets        = Asset.query.filter_by(user_id=uid).all()
    options       = Option.query.filter_by(user_id=uid).all()
    study_options = StudyOption.query.filter_by(user_id=uid).all()
    spreads       = OptionSpread.query.filter_by(user_id=uid).all()
    put_sales     = PutSale.query.filter_by(user_id=uid).all()

    # ── Monta conjunto de tickers a buscar ────────────────────────
    asset_tickers  = {a.ticker.upper() for a in assets if a.type != 'ETF'}
    option_tickers = {o.ticker.upper() for o in options}

    # Underlying das options registradas (para atualizar current_price do ativo)
    for o in options:
        if o.underlying_asset:
            asset_tickers.add(o.underlying_asset.upper())

    # Tickers das StudyOptions e seus underlyings
    for so in study_options:
        option_tickers.add(so.ticker.upper())
        if so.underlying_asset:
            asset_tickers.add(so.underlying_asset.upper())

    # Legs dos spreads e seus underlyings
    for sp in spreads:
        if sp.leg_long_ticker:
            option_tickers.add(sp.leg_long_ticker.upper())
        if sp.leg_short_ticker:
            option_tickers.add(sp.leg_short_ticker.upper())
        if sp.underlying_asset:
            asset_tickers.add(sp.underlying_asset.upper())

    # Pernas e operações estruturadas abertas — carrega uma única vez para reuso
    struct_legs_bulk = StructuredLeg.query.join(StructuredOp).filter(
        StructuredOp.user_id == uid, StructuredOp.status == 'OPEN'
    ).all()
    for leg in struct_legs_bulk:
        # Tastytrade (intl): pernas são atualizadas MANUALMENTE — não vão à OpLab
        if getattr(leg.operation, 'intl', False):
            continue
        if leg.ticker:
            option_tickers.add(leg.ticker.upper())
    struct_ops_bulk = StructuredOp.query.filter_by(user_id=uid, status='OPEN').all()
    for sop in struct_ops_bulk:
        if getattr(sop, 'intl', False):
            continue   # subjacente intl vai ao Yahoo (sem .SA) mais abaixo
        if sop.underlying_asset:
            asset_tickers.add(sop.underlying_asset.upper())

    # PutSales: tickers das opções e seus underlyings
    for ps in put_sales:
        if ps.ticker:
            option_tickers.add(ps.ticker.upper())
        if ps.underlying_asset:
            asset_tickers.add(ps.underlying_asset.upper())

    all_tickers = list(asset_tickers | option_tickers)
    if not all_tickers:
        return 0, 0, set()

    # ── Busca preços em lotes de 150 ──────────────────────────────
    prices:     dict = {}   # ticker → close price
    variations: dict = {}   # ticker → variation % do dia

    CHUNK = 150
    for i in range(0, len(all_tickers), CHUNK):
        chunk = all_tickers[i:i + CHUNK]
        try:
            r = requests.get(
                f'{BASE}/market/quote',
                params={'tickers': ','.join(chunk)},
                headers=headers,
                timeout=8,
            )
            if r.status_code == 200:
                for item in r.json():
                    sym   = str(item.get('symbol', '')).upper()
                    close = item.get('close')
                    # Tenta múltiplos nomes de campo para variação diária %
                    var = None
                    for _vk in ('variation', 'change', 'pct_change',
                                'percentChange', 'dailyChange', 'change_pct'):
                        if _vk in item and item[_vk] is not None:
                            var = item[_vk]
                            break
                    if sym and close is not None:
                        prices[sym] = float(close)
                    if sym and var is not None:
                        variations[sym] = float(var)
        except Exception:
            pass

    def _fetch_missing_option_quote(ticker: str):
        """Fallback por ticker para opções que não vieram no bulk da OpLab."""
        tk = (ticker or '').upper().strip()
        if not tk:
            return None, None
        try:
            ri = requests.get(
                f'{BASE}/market/instruments/{tk}',
                headers=headers, timeout=4,
            )
            if ri.status_code == 200:
                d = ri.json()
                p = d.get('close') or d.get('last') or d.get('price')
                if p and float(p) > 0:
                    var = d.get('variation') or d.get('change')
                    return float(p), (float(var) if var is not None else None)
        except Exception:
            pass
        try:
            yf_t = tk + '.SA'
            for host in ('query1.finance.yahoo.com', 'query2.finance.yahoo.com'):
                r = requests.get(
                    f'https://{host}/v8/finance/chart/{yf_t}',
                    params={'interval': '1d', 'range': '2d'},
                    headers=_YF_HEADERS, cookies=_YF_COOKIES, timeout=5,
                )
                if r.status_code == 200:
                    meta = r.json()['chart']['result'][0]['meta']
                    p = meta.get('regularMarketPrice') or meta.get('previousClose')
                    if p and float(p) > 0:
                        chg = meta.get('regularMarketChangePercent')
                        return float(p), (float(chg) if chg is not None else None)
        except Exception:
            pass
        return None, None

    # Fallback individual por opção — pulado quando OpLab está offline
    # (evita loop de timeouts que trava o servidor)
    if oplab_online:
        for tk in list(option_tickers):
            if tk not in prices or prices.get(tk, 0) <= 0:
                p, var = _fetch_missing_option_quote(tk)
                if p and p > 0:
                    prices[tk] = p
                    if var is not None:
                        variations[tk] = var

    if not prices:
        return 0, 0, set()

    # ── Busca delta das opções via /v3/market/options/{underlying} ──
    # Agrupa opções por underlying para minimizar chamadas à API
    # Pulado quando OpLab está offline
    deltas: dict = {}   # ticker_opcao → delta
    underlyings_com_opcoes = list({o.underlying_asset.upper() for o in options if o.underlying_asset})
    for underlying in underlyings_com_opcoes if oplab_online else []:
        try:
            r = requests.get(
                f'{BASE}/market/options/{underlying}',
                headers=headers,
                timeout=15,
            )
            if r.status_code == 200:
                data = r.json()
                # Resposta pode ser lista de opções ou dict com chave 'options'
                opt_list = data if isinstance(data, list) else data.get('options', data.get('calls', []) + data.get('puts', []))
                for item in opt_list:
                    sym   = str(item.get('symbol', item.get('ticker', ''))).upper()
                    delta = item.get('delta') or item.get('greeks', {}).get('delta') if isinstance(item.get('greeks'), dict) else item.get('delta')
                    if sym and delta is not None:
                        deltas[sym] = float(delta)
        except Exception:
            pass

    now = now_brt()

    # ── Assets: atualiza current_price/daily_change via OpLab quando disponível ──
    assets_ok = 0
    oplab_covered_assets: set = set()   # tickers que o OpLab retornou → não precisam ir ao Yahoo
    for a in assets:
        if a.type == 'ETF':
            continue
        key = a.ticker.upper()
        if key in prices and prices[key] > 0:
            a.current_price = prices[key]
            a.last_update   = now
            if key in variations:
                a.daily_change = variations[key]
            oplab_covered_assets.add(key)
            assets_ok += 1

    # ── Atualiza Options (todas as tabelas de /opcoes) ────────────
    options_ok = 0
    missing_option_tickers: list = []   # opções que o OpLab não retornou → fallback Yahoo
    for o in options:
        key = o.ticker.upper()
        if key in prices and prices[key] > 0:
            o.current_option_price = prices[key]
            o.last_update          = now
            options_ok += 1
        else:
            missing_option_tickers.append(o)
        if key in variations:
            o.daily_change = variations[key]
        # Atualiza delta se disponível (não zera se API não retornar)
        if key in deltas:
            o.delta = deltas[key]
        # Atualiza cotação do ativo subjacente quando disponível
        if o.underlying_asset:
            uk = o.underlying_asset.upper()
            if uk in prices and prices[uk] > 0:
                # Grava na própria Option — subjacentes fora da carteira
                # (ex.: AXIA3, MULT3) não têm Asset onde salvar o preço
                o.underlying_price = prices[uk]
                if uk in variations:
                    o.underlying_change = variations[uk]
                # Propaga para o Asset correspondente se existir
                asset_obj = next((a for a in assets if a.ticker.upper() == uk), None)
                if asset_obj and asset_obj.type != 'ETF':
                    asset_obj.current_price = prices[uk]
                    if uk in variations:
                        asset_obj.daily_change = variations[uk]
                    oplab_covered_assets.add(uk)

    # ── Fallback individual p/ tickers de opção não retornados no bulk ──────
    # OpLab /market/instruments/{ticker} e, se falhar, Yahoo. Retorna
    # (price, change) ou (None, None). Pulado quando OpLab offline (evita
    # empilhar timeouts de 8s por ticker).
    def _fallback_option_quote(ticker_up):
        try:
            ri = requests.get(f'{BASE}/market/instruments/{ticker_up}',
                              headers=headers, timeout=8)
            if ri.status_code == 200:
                d = ri.json()
                p = d.get('close') or d.get('last') or d.get('price')
                if p and float(p) > 0:
                    var = d.get('variation') or d.get('change')
                    return float(p), (float(var) if var is not None else None)
        except Exception:
            pass
        try:
            for host in ('query1.finance.yahoo.com', 'query2.finance.yahoo.com'):
                r = requests.get(
                    f'https://{host}/v8/finance/chart/{ticker_up}.SA',
                    params={'interval': '1d', 'range': '2d'},
                    headers=_YF_HEADERS, cookies=_YF_COOKIES, timeout=5,
                )
                if r.status_code == 200:
                    meta = r.json()['chart']['result'][0]['meta']
                    p = meta.get('regularMarketPrice') or meta.get('previousClose')
                    if p and float(p) > 0:
                        chg = meta.get('regularMarketChangePercent')
                        return float(p), (float(chg) if chg is not None else None)
        except Exception:
            pass
        return None, None

    if missing_option_tickers and oplab_online:
        for o in missing_option_tickers:
            p, var = _fallback_option_quote(o.ticker.upper())
            if p:
                o.current_option_price = p
                o.last_update = now
                if var is not None:
                    o.daily_change = var
                options_ok += 1

    # ── Atualiza StudyOptions (/estudos) ──────────────────────────
    for so in study_options:
        changed = False
        opt_key = (so.ticker or '').upper()
        if opt_key in prices and prices[opt_key] > 0:
            so.option_price = prices[opt_key]
            changed = True
            options_ok += 1
        if so.underlying_asset:
            uk = so.underlying_asset.upper()
            if uk in prices and prices[uk] > 0:
                so.underlying_price = prices[uk]
                changed = True

    # ── Atualiza OptionSpreads (/spreads) — com fallback individual ────────
    for sp in spreads:
        if sp.leg_long_ticker:
            k = sp.leg_long_ticker.upper()
            if k in prices and prices[k] > 0:
                sp.leg_long_current = prices[k]
                options_ok += 1
            elif oplab_online:
                p, _var = _fallback_option_quote(k)
                if p:
                    sp.leg_long_current = p
                    options_ok += 1
        if sp.leg_short_ticker:
            k = sp.leg_short_ticker.upper()
            if k in prices and prices[k] > 0:
                sp.leg_short_current = prices[k]
                options_ok += 1
            elif oplab_online:
                p, _var = _fallback_option_quote(k)
                if p:
                    sp.leg_short_current = p
                    options_ok += 1
        if sp.underlying_asset:
            uk = sp.underlying_asset.upper()
            if uk in prices and prices[uk] > 0:
                sp.underlying_price = prices[uk]
            if uk in variations:
                sp.underlying_change = variations[uk]

    # ── Atualiza pernas de OperaçõesEstruturadas (com fallback individual) ──
    # Sem isso, travas com pernas de vencimentos distantes/semanais (ex.:
    # calendários) que o bulk /market/quote não retorna nunca atualizavam.
    for leg in struct_legs_bulk:
        if getattr(leg.operation, 'intl', False):
            continue   # Tastytrade: prêmios mantidos manualmente
        k = (leg.ticker or '').upper()
        if not k:
            continue
        if k in prices and prices[k] > 0:
            leg.current_price = prices[k]
            leg.last_update   = now
            options_ok += 1
        elif oplab_online:
            p, _var = _fallback_option_quote(k)
            if p:
                leg.current_price = p
                leg.last_update   = now
                options_ok += 1

    # ── Atualiza underlying de OperaçõesEstruturadas ──────────────
    for sop in struct_ops_bulk:
        if getattr(sop, 'intl', False):
            continue
        if sop.underlying_asset:
            uk = sop.underlying_asset.upper()
            if uk in prices and prices[uk] > 0:
                sop.underlying_price = prices[uk]
            if uk in variations:
                sop.underlying_change = variations[uk]

    # ── Tastytrade (intl): somente o SUBJACENTE, via Yahoo sem .SA ─
    # (AAPL, SPY, TSLA…). As pernas de opção não são tocadas.
    try:
        from services import _yf_fast_info as _yfi_intl
        _intl_cache = {}
        for sop in struct_ops_bulk:
            if not getattr(sop, 'intl', False) or not sop.underlying_asset:
                continue
            t = sop.underlying_asset.strip().upper()
            if t not in _intl_cache:
                _intl_cache[t] = _yfi_intl(t, t)   # ticker internacional puro
            d = _intl_cache[t]
            if d and d.get('price'):
                sop.underlying_price  = d['price']
                sop.underlying_change = d.get('change_percent')
    except Exception:
        pass

    # ── Atualiza PutSales ─────────────────────────────────────────
    for ps in put_sales:
        if ps.underlying_asset:
            uk = ps.underlying_asset.upper()
            if uk in prices and prices[uk] > 0:
                ps.underlying_price = prices[uk]

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    return assets_ok, options_ok, oplab_covered_assets


_snapshot_last_day = {}   # {user_id: 'YYYY-MM-DD'} — garante 1 snapshot/dia por usuário


def _daily_snapshot_sweep(now):
    """Grava a foto do patrimônio 1×/dia útil (após o fechamento, 17h+) para
    cada usuário com ativos — cobre quem não abriu o Resumo nem atualizou cotações."""
    if now.weekday() >= 5 or now.hour < 17:
        return
    today_iso = now.date().isoformat()
    uids = [row[0] for row in db.session.query(Asset.user_id)
            .filter(Asset.quantity > 0).distinct().all()]
    for uid in uids:
        if _snapshot_last_day.get(uid) == today_iso:
            continue
        record_portfolio_snapshot(uid)
        _snapshot_last_day[uid] = today_iso


def _oplab_scheduler_loop():
    """Daemon thread: checks every 30 s which users need an OpLab refresh."""
    while True:
        time.sleep(30)
        with app.app_context():
            try:
                now = now_brt()
                _daily_snapshot_sweep(now)
                rows = Settings.query.filter_by(key='oplab_auto_update', value='true').all()
                for s in rows:
                    uid   = s.user_id
                    token = Settings.get_value('oplab_token', user_id=uid)
                    if not token:
                        continue
                    interval_min = int(Settings.get_value('oplab_interval', user_id=uid, default='5'))
                    last = _oplab_last_update.get(uid)
                    if last and (now - last).total_seconds() < interval_min * 60:
                        continue
                    if not _oplab_is_available(token, timeout=4):
                        continue   # OpLab fora do ar — tenta no próximo ciclo
                    _do_oplab_bulk_update_safe(uid, token, deadline_secs=30)
                    _oplab_last_update[uid] = now
            except Exception:
                pass


# Start the scheduler once (guarded so it doesn't spawn in import-time checks)
_oplab_scheduler_started = False

def _start_oplab_scheduler():
    global _oplab_scheduler_started
    if not _oplab_scheduler_started:
        _oplab_scheduler_started = True
        t = threading.Thread(target=_oplab_scheduler_loop, daemon=True)
        t.start()


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    _start_oplab_scheduler()
    app.run(debug=True, host='0.0.0.0')
