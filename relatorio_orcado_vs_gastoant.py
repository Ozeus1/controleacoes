import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import pandas as pd # Necessário para manipulação de dados e exportação para Excel
from openpyxl import Workbook # Necessário para exportação para Excel
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

class RelatorioOrcadoVsGasto:
    def __init__(self, parent):
        self.root = tk.Toplevel(parent)
        self.root.title("Relatório Mensal: Orçado vs. Gasto")
        self.root.geometry("900x600")
        self.root.resizable(True, True)
        self.root.transient(parent)
        self.root.grab_set()

        self.db_conn = sqlite3.connect('financas.db')

        self.criar_widgets()
        self.carregar_relatorio()

    def criar_widgets(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Frame para botões (Exportar, Fechar)
        botoes_frame = ttk.Frame(main_frame)
        botoes_frame.pack(fill=tk.X, pady=(0, 10))

        self.btn_exportar = ttk.Button(botoes_frame, text="Exportar para Excel", command=self.exportar_para_excel)
        self.btn_exportar.pack(side=tk.LEFT, padx=(0, 5))

        self.btn_fechar = ttk.Button(botoes_frame, text="Fechar", command=self.root.destroy)
        self.btn_fechar.pack(side=tk.LEFT)
        
        # Frame da Tabela
        tree_frame = ttk.LabelFrame(main_frame, text="Detalhes do Relatório")
        tree_frame.pack(fill=tk.BOTH, expand=True)

        # Treeview para exibir o relatório
        self.tree = ttk.Treeview(tree_frame, columns=('mes_ano', 'categoria', 'orcado', 'gasto', 'saldo'), show='headings')
        
        self.tree.heading('mes_ano', text='Mês/Ano')
        self.tree.heading('categoria', text='Categoria')
        self.tree.heading('orcado', text='Orçado (R$)')
        self.tree.heading('gasto', text='Gasto (R$)')
        self.tree.heading('saldo', text='Saldo (R$)')

        self.tree.column('mes_ano', anchor=tk.W, width=100)
        self.tree.column('categoria', anchor=tk.W, width=200)
        self.tree.column('orcado', anchor=tk.E, width=120)
        self.tree.column('gasto', anchor=tk.E, width=120)
        self.tree.column('saldo', anchor=tk.E, width=120)

        # Scrollbar
        scrollbar_y = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(fill=tk.BOTH, expand=True)

    def carregar_relatorio(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        try:
            # 1. Carregar despesas e agregar por mês/categoria
            df_expenses = pd.read_sql_query("""
                SELECT 
                    strftime('%Y-%m', data_pagamento) as AnoMes, 
                    conta_despesa, 
                    SUM(valor) as Gasto 
                FROM despesas 
                GROUP BY AnoMes, conta_despesa
            """, self.db_conn)

            # 2. Carregar orçamentos
            df_budget = pd.read_sql_query("""
                SELECT 
                    conta_despesa, 
                    valor_orcado as Orcado 
                FROM orcamento
            """, self.db_conn)

            if df_expenses.empty and df_budget.empty:
                messagebox.showinfo("Relatório", "Não há dados de despesas ou orçamentos para gerar o relatório.", parent=self.root)
                return
            if df_expenses.empty:
                messagebox.showinfo("Relatório", "Não há dados de despesas para gerar o relatório.", parent=self.root)
                # Ainda podemos mostrar o orçamento se houver
                all_months = ["N/A"] # Placeholder se não houver despesas
            else:
                all_months = sorted(df_expenses['AnoMes'].unique(), reverse=True)


            all_categories_exp = df_expenses['conta_despesa'].unique() if not df_expenses.empty else []
            all_categories_bud = df_budget['conta_despesa'].unique() if not df_budget.empty else []
            
            all_categories = pd.unique(list(all_categories_exp) + list(all_categories_bud))

            if not all_months or not all_categories.size:
                 messagebox.showinfo("Relatório", "Não há meses com despesas ou categorias definidas para o relatório.", parent=self.root)
                 return

            # 3. Criar base do relatório com todas as combinações de meses e categorias
            # Se não houver despesas, usamos um placeholder para o mês se houver orçamento.
            # O ideal é ter meses, então se só há orçamento, o relatório não faz muito sentido sem gastos.
            # Por simplicidade, vamos focar nos meses onde há despesas.
            if df_expenses.empty and not df_budget.empty: # Caso exista orçamento mas nenhuma despesa
                df_report_base_list = []
                # Para exibir o orçamento mesmo sem despesas, podemos listar as categorias orçadas
                # sem um mês específico, ou para um mês "placeholder".
                # Aqui, vamos simplificar: se não há despesas, o relatório "Orçado vs Gasto" fica limitado.
                # Vamos mostrar o orçamento para cada categoria, sem atrelar a um mês específico dos gastos.
                for cat in all_categories:
                    df_report_base_list.append({'AnoMes': "Orçamento Base", 'conta_despesa': cat})
                if not df_report_base_list:
                    messagebox.showinfo("Relatório", "Não há categorias com orçamento definido.", parent=self.root)
                    return
                df_report_base = pd.DataFrame(df_report_base_list)
            elif not df_expenses.empty:
                 df_report_base = pd.MultiIndex.from_product([all_months, all_categories], names=['AnoMes', 'conta_despesa']).to_frame(index=False)
            else: # Nem despesa nem orçamento
                messagebox.showinfo("Relatório", "Não há dados para gerar o relatório.", parent=self.root)
                return


            # 4. Merge com despesas e orçamentos
            if not df_expenses.empty:
                df_report = pd.merge(df_report_base, df_expenses, on=['AnoMes', 'conta_despesa'], how='left')
            else: # Se não houver despesas, começamos o df_report a partir da base (que pode ter categorias do orçamento)
                df_report = df_report_base.copy()
                df_report['Gasto'] = 0.0 # Adiciona coluna 'Gasto' com 0
            
            if not df_budget.empty:
                df_report = pd.merge(df_report, df_budget, on='conta_despesa', how='left')
            else: # Se não houver orçamento, adicionamos a coluna 'Orcado' com 0
                df_report['Orcado'] = 0.0


            # 5. Preencher NaNs e calcular Saldo
            df_report['Gasto'] = df_report['Gasto'].fillna(0.0)
            df_report['Orcado'] = df_report['Orcado'].fillna(0.0)
            df_report['Saldo'] = df_report['Orcado'] - df_report['Gasto']
            
            # 6. Filtrar linhas onde Orçado e Gasto são ambos zero (opcional, mas limpa a visualização)
            df_report_final = df_report[(df_report['Orcado'] != 0) | (df_report['Gasto'] != 0)]
            
            # 7. Ordenar e popular Treeview
            df_report_final = df_report_final.sort_values(by=['AnoMes', 'conta_despesa'], ascending=[False, True])

            self.report_data_df = df_report_final # Salva para exportação

            for index, row in df_report_final.iterrows():
                saldo_val = row['Saldo']
                cor_saldo = 'red' if saldo_val < 0 else 'darkgreen' if saldo_val > 0 else 'black'
                
                self.tree.insert('', tk.END, values=(
                    row['AnoMes'],
                    row['conta_despesa'],
                    f"R$ {row['Orcado']:.2f}".replace('.', ','),
                    f"R$ {row['Gasto']:.2f}".replace('.', ','),
                    f"R$ {saldo_val:.2f}".replace('.', ',')
                ), tags=(cor_saldo,))
            
            self.tree.tag_configure('red', foreground='red')
            self.tree.tag_configure('darkgreen', foreground='darkgreen')
            self.tree.tag_configure('black', foreground='black')

        except sqlite3.Error as e:
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao consultar dados: {e}", parent=self.root)
        except Exception as e:
            messagebox.showerror("Erro Inesperado", f"Ocorreu um erro ao gerar o relatório: {e}", parent=self.root)
            import traceback
            traceback.print_exc()


    def exportar_para_excel(self):
        if not hasattr(self, 'report_data_df') or self.report_data_df.empty:
            messagebox.showwarning("Sem Dados", "Não há dados para exportar.", parent=self.root)
            return

        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Arquivos Excel", "*.xlsx")],
                title="Salvar Relatório Orçado vs. Gasto",
                initialfile="Relatorio_Orcado_vs_Gasto.xlsx"
            )

            if not file_path:
                return

            # Preparar DataFrame para exportação (formatar valores como números)
            df_export = self.report_data_df.copy()
            df_export.rename(columns={
                'AnoMes': 'Mês/Ano',
                'conta_despesa': 'Categoria',
                'Orcado': 'Valor Orçado (R$)',
                'Gasto': 'Valor Gasto (R$)',
                'Saldo': 'Saldo (R$)'
            }, inplace=True)

            # Usar openpyxl para mais controle de formatação
            wb = Workbook()
            ws = wb.active
            ws.title = "Orcado_vs_Gasto"

            # Título do Relatório
            ws['A1'] = "Relatório Mensal: Orçado vs. Gasto"
            ws.merge_cells('A1:E1')
            ws['A1'].font = Font(size=16, bold=True, color="000080") # Azul Marinho
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.row_dimensions[1].height = 25
            
            # Cabeçalhos
            headers = list(df_export.columns)
            for col_num, header_title in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num, value=header_title)
                cell.font = Font(bold=True, color="FFFFFF") # Branco
                cell.fill = पैटर्नFill(start_color="000080", end_color="000080", fill_type="solid") # Azul Marinho
                cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions[get_column_letter(col_num)].width = 20


            # Dados
            for r_idx, row in enumerate(df_export.values.tolist(), 4): # Começa da linha 4
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if isinstance(value, (int, float)):
                        cell.number_format = 'R$ #,##0.00'
                        if c_idx == headers.index('Saldo (R$)') + 1: # Coluna Saldo
                             if value < 0:
                                cell.font = Font(color="FF0000") # Vermelho para saldo negativo
                             elif value > 0:
                                cell.font = Font(color="008000") # Verde para saldo positivo


            # Ajustar largura das colunas
            for col_num, column_title in enumerate(headers, 1):
                max_length = len(column_title)
                column_letter = get_column_letter(col_num)
                for cell_tuple in ws[column_letter]:
                    try:
                        if len(str(cell_tuple.value)) > max_length:
                            max_length = len(str(cell_tuple.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width if adjusted_width < 50 else 50


            wb.save(file_path)
            messagebox.showinfo("Exportação Concluída", f"Relatório exportado com sucesso para:\n{file_path}", parent=self.root)

        except Exception as e:
            messagebox.showerror("Erro na Exportação", f"Não foi possível exportar o relatório: {e}", parent=self.root)
            import traceback
            traceback.print_exc()

    def __del__(self):
        if self.db_conn:
            self.db_conn.close()

# Função para ser chamada pelo programa principal
def iniciar_relatorio_orcado_vs_gasto(parent):
    RelatorioOrcadoVsGasto(parent)